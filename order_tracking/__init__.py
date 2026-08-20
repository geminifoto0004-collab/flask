"""
訂單流程追蹤系統 - Blueprint入口
包含所有路由定義和業務邏輯
"""
from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, g, current_app, make_response
from werkzeug.exceptions import NotFound
try:
    from werkzeug.security import check_password_hash, generate_password_hash
except ImportError:
    # 如果沒有werkzeug，使用簡單的檢查（僅開發環境）
    def check_password_hash(hashed, password):
        return hashed == f"hash_{password}"
    def generate_password_hash(password):
        return f"hash_{password}"

from datetime import datetime, date, timezone, timedelta
import os
import functools
import sqlite3
import json
import hashlib
import threading
import secrets
import time
import ipaddress
import socket
import base64
import zipfile
import tempfile
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
try:
    import jwt
    HAS_JWT = True
except ImportError:
    jwt = None
    HAS_JWT = False

from .models import get_db, init_db, ensure_factory_visit_tables, ensure_local_guest_link_tables, ensure_public_guest_share_registry, calculate_status_light, get_status_light_hint, update_status_light, generate_revision_number, get_order_sort_clause, get_order_number_sort_expr
from .config import (
    SECRET_KEY, JWT_SECRET_KEY, JWT_EXPIRATION_DELTA, BLUEPRINT_NAME, URL_PREFIX,
    CUSTOMER_REPORT_WORKERS, CUSTOMER_REPORT_MAX_QUEUED, CUSTOMER_REPORT_MAX_ACTIVE_PER_USER,
    DEVELOPER_TOOLS_ENABLED, CLOUD_MODE, CLOUD_READ_ONLY,
    LOCAL_GUEST_SHARING_ENABLED, LOCAL_GUEST_PERMANENT_ENABLED,
    RENDER_PUBLIC_GUEST_ENABLED, PERMANENT_PUBLIC_GUEST_ENABLED, PUBLIC_SHARE_PROVIDER_READY,
    PUBLIC_SHARE_BACKGROUND_SYNC_ENABLED, PUBLIC_SHARE_SYNC_START_DELAY_SECONDS, PUBLIC_SHARE_SYNC_INTERVAL_SECONDS,
    TIDB_PROVIDER_ENABLED, B2_PROVIDER_ENABLED, DEPLOYMENT_PROFILE, IS_CHINA_DEPLOYMENT,
    MEDIA_PREVIEW_CACHE_DIR, MEDIA_PREVIEW_CACHE_MAX_BYTES, MEDIA_PREVIEW_CACHE_RETENTION_DAYS,
    MEDIA_PREVIEW_CACHE_CLEANUP_INTERVAL_SECONDS, MEDIA_PREVIEW_WORKERS,
)
from .status_config import STATUS, STAGE_GROUPS, STATUS_MAP, get_stage_group, get_statuses_by_stage_group  # 向后兼容
from .status_definitions import STATUS_KEYS, QUICK_ACTIONS_MAP, get_status_label, STATUS_LABELS, STAGE_GROUPS as STATUS_STAGE_GROUPS
from .permissions_config import PERMISSION_MATRIX, OWNERSHIP_RULES, STATUS_RULES
from .permissions import check_permission, get_resource_info, require_permission, get_filtered_resources, ensure_order_lock_columns, get_current_user_context, can_manage_by_owner, can_access_visibility, get_visibility_where_clause, is_admin
from .cloud_mode import cloud_mode_enabled, cloud_read_only_enabled, effective_permission_matrix, enforce_cloud_read_only_request
from .data_provider import register_order_data_provider, get_order_data_provider, provider_ready, provider_last_synced_at
from .db_backend import register_cloud_db_connection_factory, get_tracking_db_connection
from .public_share_provider import public_share_provider_ready, create_public_share, list_public_shares, revoke_public_share, update_public_share
from .customer_report import (
    validate_report_options as validate_customer_report_options,
    load_report_entries, attach_image_metadata, prepare_entries as prepare_customer_report_entries,
    estimate_report as estimate_customer_report, build_report_files as build_customer_report_files,
    cache_report_file, get_cached_report,
)

# Customer report jobs run in a small background queue so report generation never blocks
# the user's page. Two workers let a second customer report start while another report is
# generating/downloading; additional jobs remain queued to avoid overloading the local PC.
_customer_report_executor = ThreadPoolExecutor(max_workers=max(1, int(CUSTOMER_REPORT_WORKERS)), thread_name_prefix='customer-report')
# Media preview work is deliberately separate from report generation. It only prepares
# reproducible thumbnails/PDF-page JPEGs and is capped to a small worker count so it
# cannot monopolize the office PC while users are working.
_media_preview_executor = ThreadPoolExecutor(max_workers=max(1, int(MEDIA_PREVIEW_WORKERS)), thread_name_prefix='media-preview')
_media_preview_pending = set()
_media_preview_pending_lock = threading.Lock()
_media_preview_cleanup_lock = threading.Lock()
_media_preview_cleanup_last = 0.0
_customer_report_jobs = {}
_customer_report_jobs_lock = threading.Lock()
_CUSTOMER_REPORT_JOB_KEEP = 30

# Temporary-office guest report jobs share the same bounded report executor so a customer
# can request a large PDF/HTML without blocking the mobile page or saturating the PC.
# Jobs are ephemeral and are always bound to the exact guest token hash + customer.
_guest_report_jobs = {}
_guest_report_jobs_lock = threading.Lock()
_GUEST_REPORT_JOB_KEEP = 40


class CustomerReportQueueError(RuntimeError):
    def __init__(self, message, code, status_code=429):
        super().__init__(message)
        self.code = code
        self.status_code = status_code


def _customer_report_request_key(entries, report_format, language, image_source, image_count, image_order, pdf_attachment_mode='pages'):
    """Stable fingerprint used to suppress duplicate active jobs for the same user."""
    order_keys = []
    for entry in entries or []:
        workflow_number = str(entry.get('workflow_number') or '').strip()
        order_number = str(entry.get('order_number') or '').strip()
        customer_name = str(entry.get('customer_name') or '').strip()
        order_keys.append((customer_name, workflow_number, order_number))
    payload = {
        'orders': sorted(set(order_keys)),
        'format': str(report_format or ''),
        'language': str(language or ''),
        'image_source': str(image_source or ''),
        'image_count': str(image_count or ''),
        'image_order': str(image_order or ''),
        'pdf_attachment_mode': str(pdf_attachment_mode or 'pages'),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(',', ':')).encode('utf-8')
    return hashlib.sha256(raw).hexdigest()


def _customer_report_queue_counts_locked(owner_user_id=None):
    queued = 0
    processing = 0
    owner_active = 0
    for job in _customer_report_jobs.values():
        status = job.get('status')
        if status == 'queued':
            queued += 1
        elif status == 'processing':
            processing += 1
        if owner_user_id is not None and job.get('owner_user_id') == owner_user_id and status in {'queued', 'processing'}:
            owner_active += 1
    return queued, processing, owner_active


def _customer_report_job_snapshot(job):
    """Return a JSON-safe public copy; download URLs are added inside request context."""
    if not job:
        return None
    return {
        'id': job.get('id'),
        'status': job.get('status'),
        'created_at': job.get('created_at'),
        'started_at': job.get('started_at'),
        'finished_at': job.get('finished_at'),
        'order_count': job.get('order_count', 0),
        'customers': list(job.get('customers') or []),
        'format': job.get('format'),
        'language': job.get('language'),
        'pdf_attachment_mode': job.get('pdf_attachment_mode', 'pages'),
        'error': job.get('error'),
        'files': [dict(x) for x in (job.get('files') or [])],
    }


def _customer_report_set_job(job_id, **updates):
    with _customer_report_jobs_lock:
        job = _customer_report_jobs.get(job_id)
        if job:
            job.update(updates)


def _customer_report_trim_jobs():
    with _customer_report_jobs_lock:
        if len(_customer_report_jobs) <= _CUSTOMER_REPORT_JOB_KEEP:
            return
        completed = [j for j in _customer_report_jobs.values() if j.get('status') in {'completed', 'failed'}]
        completed.sort(key=lambda j: j.get('created_at') or '')
        while len(_customer_report_jobs) > _CUSTOMER_REPORT_JOB_KEEP and completed:
            old = completed.pop(0)
            _customer_report_jobs.pop(old.get('id'), None)


def _run_customer_report_job(job_id, entries, report_format, language, image_source, image_count, image_order, pdf_attachment_mode='pages'):
    _customer_report_set_job(job_id, status='processing', started_at=datetime.now().isoformat(timespec='seconds'))
    conn = None
    try:
        conn = get_db()
        prepared = prepare_customer_report_entries(conn, entries, image_source, image_count, language, image_order, pdf_attachment_mode)
        generated = build_customer_report_files(prepared, report_format, language)
        files = []
        for filename, mimetype, content in generated:
            cached = cache_report_file(filename, mimetype, content)
            files.append({
                'file_id': cached['id'],
                'name': cached['filename'],
                'size': cached['size'],
            })
        _customer_report_set_job(
            job_id,
            status='completed',
            finished_at=datetime.now().isoformat(timespec='seconds'),
            files=files,
            error=None,
        )
    except Exception as exc:
        print(f'[ERROR] customer report background job failed: {exc}')
        _customer_report_set_job(
            job_id,
            status='failed',
            finished_at=datetime.now().isoformat(timespec='seconds'),
            error=str(exc) or '产生客户报告失败',
        )
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        _customer_report_trim_jobs()

# Normalize lock flag from DB values to 0/1 int
def _normalize_lock_flag(value):
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if int(value) == 1 else 0
    return 1 if str(value).strip().lower() in {'1', 'true', 'yes'} else 0


def _normalize_customer_name(value):
    """Normalize customer name to uppercase (ASCII) and trimmed.

    - Non-ASCII characters (e.g. 中文) are unaffected by upper().
    - Always returns a string (possibly empty).
    """
    if value is None:
        return ''
    try:
        s = str(value)
    except Exception:
        return ''
    s = s.strip()
    return s.upper()

# ==================== 兼容性辅助函数 ====================
def get_status_for_query(status_key):
    """
    获取用于数据库查询的状态值（兼容旧数据）
    返回 (key, 简体中文) 的元组，用于 IN 查询
    """
    label_zh_cn = get_status_label(status_key, 'zh_cn')
    return (status_key, label_zh_cn)

def get_completed_cancelled_for_query():
    """获取已完成和已取消的状态值（用于查询，兼容旧数据）"""
    return (
        (STATUS_KEYS['COMPLETED'], STATUS['COMPLETED']),
        (STATUS_KEYS['CANCELLED'], STATUS['CANCELLED'])
    )

def get_user_display_name(user_id, conn):
    """根據 user_id 取得顯示名稱（優先 real_name，其次 display_name，再次 username）"""
    if not user_id:
        return None
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COALESCE(real_name, display_name, username) as name
        FROM users
        WHERE id = ?
    ''', (user_id,))
    row = cursor.fetchone()
    return row['name'] if row and row['name'] else None

def get_latest_workflow_history(cursor, workflow_number):
    """获取工作流最新历史记录（用于并发控制）"""
    cursor.execute('''
        SELECT id, to_status
        FROM workflow_status_history
        WHERE workflow_number = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 1
    ''', (workflow_number,))
    row = cursor.fetchone()
    return dict(row) if row else None

def normalize_action_date(value):
    """Validate a user supplied action date and return YYYY-MM-DD.

    Back-dated business events are allowed. A date more than one day in the
    future is rejected so a browser/server timezone difference does not cause
    false validation failures.
    """
    raw = str(value or '').strip()
    if not raw:
        return date.today().isoformat()
    try:
        parsed = datetime.strptime(raw, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        raise ValueError('日期格式错误，请使用 YYYY-MM-DD')
    if parsed > date.today() + timedelta(days=1):
        raise ValueError('操作日期不能晚于明天')
    return parsed.isoformat()


def should_insert_workflow_history(cursor, workflow_number, to_status):
    """避免連續寫入相同狀態（修復重複狀態記錄）"""
    if not workflow_number or not to_status:
        return True
    cursor.execute('''
        SELECT to_status
        FROM workflow_status_history
        WHERE workflow_number = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 1
    ''', (workflow_number,))
    row = cursor.fetchone()
    if not row:
        return True
    last_status = row['to_status'] if isinstance(row, sqlite3.Row) else dict(row).get('to_status')
    return last_status != to_status


def apply_workflow_status_transition(cursor, workflow, workflow_number, new_status,
                                     action_date, operator_id, notes='',
                                     update_current_status=True, force_insert=False):
    """寫入流程狀態；ALL_SHIPPED 會保留歷史後自動落到 COMPLETED。"""
    old_status = workflow.get('current_status')
    order_number = workflow.get('order_number')
    latest_history_id = None
    auto_completed = False
    auto_note = '系統自動：已全部出貨後轉為已完成'

    if force_insert or should_insert_workflow_history(cursor, workflow_number, new_status):
        cursor.execute('''
            INSERT INTO workflow_status_history
            (workflow_number, order_number, from_status, to_status, action_date, operator_id, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            workflow_number,
            order_number,
            old_status,
            new_status,
            action_date,
            operator_id,
            notes or ''
        ))
        latest_history_id = cursor.lastrowid
    else:
        latest = get_latest_workflow_history(cursor, workflow_number)
        latest_history_id = latest.get('id') if latest else None

    final_status = new_status
    if update_current_status and new_status == STATUS_KEYS['ALL_SHIPPED']:
        final_status = STATUS_KEYS['COMPLETED']
        auto_completed = True
        if should_insert_workflow_history(cursor, workflow_number, final_status):
            cursor.execute('''
                INSERT INTO workflow_status_history
                (workflow_number, order_number, from_status, to_status, action_date, operator_id, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                workflow_number,
                order_number,
                new_status,
                final_status,
                action_date,
                operator_id,
                auto_note
            ))
            latest_history_id = cursor.lastrowid

    if update_current_status:
        cursor.execute('''
            UPDATE workflows
            SET current_status = ?,
                status_updated_at = ?,
                status_days = 0,
                updated_at = CURRENT_TIMESTAMP
            WHERE workflow_number = ?
        ''', (final_status, action_date, workflow_number))
    else:
        cursor.execute('''
            UPDATE workflows
            SET updated_at = CURRENT_TIMESTAMP
            WHERE workflow_number = ?
        ''', (workflow_number,))

    return {
        'requested_status': new_status,
        'final_status': final_status,
        'latest_history_id': latest_history_id,
        'auto_completed': auto_completed
    }

# 創建Blueprint
tracking_bp = Blueprint(
    BLUEPRINT_NAME,
    __name__,
    url_prefix=URL_PREFIX,
    template_folder='templates',
    static_folder='static',
    static_url_path='/static/tracking'
)


@tracking_bp.record_once
def _register_snapshot_scheduler(state):
    """Start local SQLite maintenance only outside Render/cloud mode."""
    cloud_mode = bool(state.app.config.get('TRACKING_CLOUD_MODE', CLOUD_MODE))
    if cloud_mode:
        return
    from .snapshot import schedule_snapshot_scheduler_start
    schedule_snapshot_scheduler_start()


@tracking_bp.record_once
def _register_public_share_scheduler(state):
    enabled = bool(state.app.config.get('TRACKING_PUBLIC_SHARE_BACKGROUND_SYNC_ENABLED', PUBLIC_SHARE_BACKGROUND_SYNC_ENABLED))
    if not enabled:
        return
    from .public_share_sync import schedule_public_share_sync
    schedule_public_share_sync(
        state.app,
        state.app.config.get('TRACKING_PUBLIC_SHARE_SYNC_START_DELAY_SECONDS', PUBLIC_SHARE_SYNC_START_DELAY_SECONDS),
        state.app.config.get('TRACKING_PUBLIC_SHARE_SYNC_INTERVAL_SECONDS', PUBLIC_SHARE_SYNC_INTERVAL_SECONDS),
    )


@tracking_bp.before_request
def _enforce_cloud_read_only_mode():
    """Server-side safety net: Render/cloud deployments cannot mutate order data."""
    return enforce_cloud_read_only_request()


def _developer_tools_enabled():
    """Allow the parent Flask app to override config.py, otherwise use this module's config."""
    return bool(current_app.config.get('DEVELOPER_TOOLS_ENABLED', DEVELOPER_TOOLS_ENABLED))


@tracking_bp.app_context_processor
def inject_permission_config():
    """向模板注入权限配置"""
    # 静态资源版本号：用于 cache-busting，避免浏览器长期缓存导致前端逻辑“改了但没生效”
    # 取关键静态文件的最大 mtime（秒），任何一个文件变更都会触发 URL 变化
    base_dir = os.path.dirname(__file__)
    static_candidates = [
        os.path.join(base_dir, 'static', 'css', 'tracking.css'),
        os.path.join(base_dir, 'static', 'css', 'login.css'),
        os.path.join(base_dir, 'static', 'css', 'WORKSPACE_drawer.css'),
        os.path.join(base_dir, 'static', 'css', 'theme.css'),
        os.path.join(base_dir, 'static', 'css', 'guest.css'),
        os.path.join(base_dir, 'static', 'js', 'STATUS_SYSTEM.js'),
        os.path.join(base_dir, 'static', 'js', 'theme.js'),
        os.path.join(base_dir, 'static', 'js', 'tracking.js'),
        os.path.join(base_dir, 'static', 'js', 'new_order.js'),
        os.path.join(base_dir, 'static', 'js', 'notifications.js'),
        os.path.join(base_dir, 'static', 'js', 'WORKSPACE_drawer.js'),
        os.path.join(base_dir, 'static', 'js', 'WORKSPACE_drawer_addon.js'),
    ]
    # 使用纳秒 mtime + 文件大小做 cache-busting。
    # 秒级 mtime 在短时间连续覆盖文件时可能不变，手机浏览器就可能继续拿到旧 tracking.js。
    static_versions = []
    for p in static_candidates:
        try:
            st = os.stat(p)
            static_versions.append(f"{st.st_mtime_ns:x}-{st.st_size:x}")
        except OSError:
            continue
    static_ver = '-'.join(static_versions) if static_versions else '1'

    cloud_mode = cloud_mode_enabled()
    cloud_read_only = cloud_read_only_enabled()
    return {
        'PERMISSION_MATRIX': effective_permission_matrix(),
        'OWNERSHIP_RULES': OWNERSHIP_RULES,
        'STATUS_RULES': STATUS_RULES,
        'current_user_context': get_current_user_context(),
        'STATIC_VER': static_ver,
        'is_overseas': current_app.config.get('IS_OVERSEAS', False),
        'developer_tools_enabled': _developer_tools_enabled(),
        'developer_tools_admin_visible': bool(_developer_tools_enabled() and is_admin()),
        'cloud_mode': cloud_mode,
        'cloud_read_only': cloud_read_only,
        'cloud_provider_ready': bool(provider_ready()) if cloud_mode else True,
        'cloud_last_synced_at': provider_last_synced_at() if cloud_mode else None,
        'local_guest_sharing_enabled': bool(LOCAL_GUEST_SHARING_ENABLED),
        'local_guest_permanent_enabled': bool(LOCAL_GUEST_PERMANENT_ENABLED),
        'render_public_guest_enabled': bool(RENDER_PUBLIC_GUEST_ENABLED),
        'permanent_public_guest_enabled': bool(PERMANENT_PUBLIC_GUEST_ENABLED),
        'public_share_provider_ready': bool(PUBLIC_SHARE_PROVIDER_READY and public_share_provider_ready()),
        'deployment_profile': DEPLOYMENT_PROFILE,
        'is_china_deployment': bool(IS_CHINA_DEPLOYMENT),
    }

# ==================== 工具函數 ====================

def login_required(f):
    """登入驗證裝飾器（Session）"""
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/tracking/api'):
                return jsonify({'success': False, 'error': '未登入', 'code': 'UNAUTHORIZED'}), 401
            return redirect(url_for('tracking_bp.login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """管理員權限裝飾器"""
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/tracking/api'):
                return jsonify({'success': False, 'error': '未登入', 'code': 'UNAUTHORIZED'}), 401
            return redirect(url_for('tracking_bp.login'))
        if not is_admin():
            if request.path.startswith('/tracking/api'):
                return jsonify({'success': False, 'error': '無權限', 'code': 'FORBIDDEN'}), 403
            return redirect(url_for('tracking_bp.index'))
        return f(*args, **kwargs)
    return decorated_function

def api_login_required(f):
    """API登入驗證裝飾器（JWT）"""
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        # 1) 先支援已有的 Session 登入（從網頁呼叫 API）
        if 'user_id' in session:
            g.current_user = {
                'id': session['user_id'],
                'username': session.get('username'),
                'role': session.get('role', 'viewer')
            }
            return f(*args, **kwargs)

        # 2) 若無 Session，改用 JWT Token 驗證（純 API 用途）
        if not HAS_JWT:
            return jsonify({'success': False, 'error': 'JWT未安裝', 'code': 'JWT_NOT_AVAILABLE'}), 500
        
        token = None
        auth_header = request.headers.get('Authorization')
        if auth_header:
            try:
                token = auth_header.split(' ')[1]  # Bearer <token>
            except IndexError:
                return jsonify({'success': False, 'error': 'Token格式錯誤', 'code': 'INVALID_TOKEN'}), 401
        
        if not token:
            return jsonify({'success': False, 'error': '未提供Token或未登入', 'code': 'UNAUTHORIZED'}), 401
        
        try:
            data = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
            g.current_user = {
                'id': data['user_id'],
                'username': data['username'],
                'role': data['role']
            }
        except jwt.ExpiredSignatureError:
            return jsonify({'success': False, 'error': 'Token已過期', 'code': 'TOKEN_EXPIRED'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'success': False, 'error': 'Token無效', 'code': 'INVALID_TOKEN'}), 401
        
        return f(*args, **kwargs)
    return decorated_function

def api_admin_required(f):
    """API管理員權限裝飾器"""
    @functools.wraps(f)
    @api_login_required
    def decorated_function(*args, **kwargs):
        if not is_admin():
            return jsonify({'success': False, 'error': '無權限', 'code': 'FORBIDDEN'}), 403
        return f(*args, **kwargs)
    return decorated_function

# ==================== 管理員系統診斷 ====================

@tracking_bp.route('/admin/diagnostics')
def admin_diagnostics():
    """Read-only system diagnostics, visible only when enabled and to administrators."""
    if not _developer_tools_enabled():
        raise NotFound()
    if 'user_id' not in session:
        return redirect(url_for('tracking_bp.login'))
    if not is_admin():
        raise NotFound()

    from .diagnostics import collect_diagnostics
    diagnostics = collect_diagnostics()
    response = make_response(render_template('tracking/admin_diagnostics.html', diagnostics=diagnostics))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response


# ==================== 認證路由 ====================

@tracking_bp.route('/login', methods=['GET', 'POST'])
def login():
    """登入頁面 - 密碼驗證失敗時在當前頁面顯示錯誤，不跳轉"""
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        # 驗證輸入
        if not username:
            error_msg = '請輸入用戶名'
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg, 'code': 'MISSING_USERNAME'}), 400
            return render_template('tracking/login.html', error=error_msg, username=username)
        
        if not password:
            error_msg = '請輸入密碼'
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg, 'code': 'MISSING_PASSWORD'}), 400
            return render_template('tracking/login.html', error=error_msg, username=username)
        
        # 從資料庫驗證用戶和密碼
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
        user = cursor.fetchone()
        
        # 驗證用戶是否存在和密碼是否正確
        if not user:
            conn.close()
            error_msg = '用戶名或密碼錯誤'
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg, 'code': 'INVALID_CREDENTIALS'}), 401
            return render_template('tracking/login.html', error=error_msg, username=username)
        
        # 檢查用戶狀態
        try:
            user_status = user['status']
        except (KeyError, IndexError):
            user_status = 'active'  # 兼容旧数据
        
        if user_status == 'pending':
            conn.close()
            error_msg = '您的帳號正在等待主管審核，請稍後再試'
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg, 'code': 'PENDING_APPROVAL'}), 403
            return render_template('tracking/login.html', error=error_msg, username=username)
        
        if user_status == 'rejected':
            conn.close()
            error_msg = '您的註冊申請已被拒絕，請聯繫主管'
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg, 'code': 'REJECTED'}), 403
            return render_template('tracking/login.html', error=error_msg, username=username)

        if user_status == 'suspended':
            conn.close()
            error_msg = '您的帳號已被停權，請聯繫主管'
            if request.is_json:
                return jsonify({'success': False, 'error': error_msg, 'code': 'SUSPENDED'}), 403
            return render_template('tracking/login.html', error=error_msg, username=username)
        
        # 檢查是否需要重置密碼
        try:
            needs_reset = user['needs_password_reset']
        except (KeyError, IndexError):
            needs_reset = False
        
        # 如果需要重置密碼，檢查確認密碼
        if needs_reset:
            confirm_password = data.get('confirm_password', '')
            if not confirm_password:
                conn.close()
                error_msg = '您的密碼已重置，請設置新密碼'
                if request.is_json:
                    return jsonify({
                        'success': False, 
                        'error': error_msg, 
                        'code': 'NEEDS_PASSWORD_RESET',
                        'needs_confirm': True
                    }), 400
                return render_template('tracking/login.html', error=error_msg, username=username, needs_password_reset=True)
            
            # 驗證新密碼和確認密碼
            if password != confirm_password:
                conn.close()
                error_msg = '兩次輸入的密碼不一致'
                if request.is_json:
                    return jsonify({'success': False, 'error': error_msg}), 400
                return render_template('tracking/login.html', error=error_msg, username=username, needs_password_reset=True)
            
            if len(password) < 6:
                conn.close()
                error_msg = '密碼至少需要6位'
                if request.is_json:
                    return jsonify({'success': False, 'error': error_msg}), 400
                return render_template('tracking/login.html', error=error_msg, username=username, needs_password_reset=True)
            
            # 更新密碼並清除重置標記
            from werkzeug.security import generate_password_hash
            new_password_hash = generate_password_hash(password)
            cursor.execute('''
                UPDATE users 
                SET password_hash = ?, needs_password_reset = 0 
                WHERE id = ?
            ''', (new_password_hash, user['id']))
            conn.commit()
        else:
            # 正常登入，驗證密碼
            if not check_password_hash(user['password_hash'], password):
                conn.close()
                error_msg = '用戶名或密碼錯誤'
                if request.is_json:
                    return jsonify({'success': False, 'error': error_msg, 'code': 'INVALID_CREDENTIALS'}), 401
                return render_template('tracking/login.html', error=error_msg, username=username)
        
        conn.close()
        
        # 登入成功
        if request.is_json:
            # API登入，返回JWT Token
            if not HAS_JWT:
                return jsonify({'success': False, 'error': 'JWT未安裝', 'code': 'JWT_NOT_AVAILABLE'}), 500
            
            token = jwt.encode({
                'user_id': user['id'],
                'username': user['username'],
                'role': user['role'],
                'exp': datetime.utcnow().timestamp() + JWT_EXPIRATION_DELTA
            }, JWT_SECRET_KEY, algorithm='HS256')
            
            return jsonify({
                'success': True,
                'token': token,
                'user': {
                    'id': user['id'],
                    'username': user['username'],
                    'display_name': user['display_name'],
                    'role': user['role']
                },
                'expires_in': JWT_EXPIRATION_DELTA
            })
        else:
            # 網頁登入，設置Session
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['display_name'] = user['display_name']
            session['role'] = user['role']
            return redirect(url_for('tracking_bp.index'))
    
    # GET請求 - 如果已登入，重定向到主頁
    if 'user_id' in session:
        return redirect(url_for('tracking_bp.index'))
    
    # 檢查是否有 needs_password_reset 參數
    needs_reset = request.args.get('needs_password_reset', 'false') == 'true'
    return render_template('tracking/login.html', needs_password_reset=needs_reset)

@tracking_bp.route('/logout', methods=['GET', 'POST'])
def logout():
    """登出"""
    if request.is_json:
        session.clear()
        return jsonify({'success': True, 'message': '已登出'})
    session.clear()
    return redirect(url_for('tracking_bp.login'))

# ==================== 錯誤處理 ====================

@tracking_bp.errorhandler(404)
def handle_404(e):
    """處理 404 錯誤 - 錯誤的 URL 跳轉回登入頁面或主頁"""
    # 如果是 API 請求，返回 JSON 錯誤
    if request.path.startswith('/tracking/api'):
        return jsonify({'success': False, 'error': '路由不存在', 'code': 'NOT_FOUND'}), 404
    
    # 如果是網頁請求，根據登入狀態跳轉
    if 'user_id' in session:
        # 已登入，跳轉到主頁
        return redirect(url_for('tracking_bp.index'))
    else:
        # 未登入，跳轉到登入頁面
        return redirect(url_for('tracking_bp.login'))

# ==================== 網頁路由 ====================


def _load_home_orders_dataset(conn, current_role, current_user_id):
    """Load the complete home-page dataset in two SQL queries, without per-row history lookups."""
    ensure_order_lock_columns(conn)
    cursor = conn.cursor()
    filter_info = get_filtered_resources('workflow', current_role, current_user_id)

    permission_sql = ''
    params = [
        STATUS_KEYS['DRAFT_CONFIRMING'],
        STATUS_KEYS['PARTIAL_SHIPPED'],
        STATUS_KEYS['PARTIAL_SHIPPED'],
        STATUS_KEYS['ALL_SHIPPED'],
        STATUS_KEYS['COMPLETED'],
        STATUS_KEYS['COMPLETED'],
        '系統自動：已全部出貨後轉為已完成',
    ]
    if not (is_admin() or filter_info['rule'] == 'all'):
        permission_sql = " AND w.handler_id = ? AND o.visibility = 'all_sales'"
        params.extend(filter_info['params'])

    params.extend([STATUS_KEYS['COMPLETED'], STATUS_KEYS['CANCELLED'], STATUS_KEYS['COMPLETED']])

    query = f"""
        WITH ranked_history AS (
            SELECT h.id, h.workflow_number, h.action_date,
                   ROW_NUMBER() OVER (
                       PARTITION BY h.workflow_number
                       ORDER BY h.created_at DESC, h.id DESC
                   ) AS rn
            FROM workflow_status_history h
        ),
        latest_history AS (
            SELECT id, workflow_number, action_date
            FROM ranked_history
            WHERE rn = 1
        ),
        draft_history AS (
            SELECT workflow_number, MIN(action_date) AS draft_date
            FROM workflow_status_history
            WHERE to_status = ?
            GROUP BY workflow_number
        ),
        shipping_ranked AS (
            SELECT h.workflow_number, h.action_date, h.to_status, h.created_at, h.id,
                   SUM(CASE WHEN h.to_status = ? THEN 1 ELSE 0 END) OVER (
                       PARTITION BY h.workflow_number
                   ) AS partial_ship_count,
                   ROW_NUMBER() OVER (
                       PARTITION BY h.workflow_number
                       ORDER BY h.action_date DESC, h.created_at DESC, h.id DESC
                   ) AS rn
            FROM workflow_status_history h
            WHERE h.to_status IN (?, ?, ?)
              AND NOT (h.to_status = ? AND COALESCE(h.notes, '') = ?)
        ),
        shipping_summary AS (
            SELECT workflow_number, action_date AS last_shipping_date,
                   to_status AS last_shipping_status, partial_ship_count
            FROM shipping_ranked
            WHERE rn = 1
        )
        SELECT w.id, w.workflow_number, w.order_number,
               w.product_name, w.product_code, w.quantity, w.factory, w.production_type,
               w.expected_delivery_date, w.current_status, w.status_updated_at,
               w.created_by_id, w.handler_id, w.folder_path, w.notes, w.created_at, w.updated_at,
               o.customer_name, o.order_date, o.status AS order_status, o.visibility, o.is_locked,
               COALESCE(u.real_name, u.display_name, u.username) AS handler_name,
               lh.id AS last_history_id, lh.action_date AS last_action_date,
               dh.draft_date AS draft_date,
               ss.last_shipping_date, ss.last_shipping_status,
               COALESCE(ss.partial_ship_count, 0) AS partial_ship_count
        FROM workflows w
        INNER JOIN orders o ON w.order_number = o.order_number
        LEFT JOIN users u ON w.handler_id = u.id
        LEFT JOIN latest_history lh ON lh.workflow_number = w.workflow_number
        LEFT JOIN draft_history dh ON dh.workflow_number = w.workflow_number
        LEFT JOIN shipping_summary ss ON ss.workflow_number = w.workflow_number
        WHERE o.status = 'ACTIVE'
          {permission_sql}
          AND (
              w.current_status NOT IN (?, ?)
              OR (w.current_status = ? AND w.status_updated_at >= date('now', '-3 months'))
          )
        ORDER BY o.order_date DESC, w.created_at DESC
    """
    cursor.execute(query, params)

    orders_list = []
    today_date = date.today()
    current_id_str = str(current_user_id) if current_user_id is not None else ''
    for row in cursor.fetchall():
        workflow = dict(row)
        last_action_date_str = workflow.get('last_action_date') or ''
        order_obj = {
            'order_number': workflow.get('order_number'),
            'order_date': workflow.get('order_date'),
            'expected_delivery_date': workflow.get('expected_delivery_date'),
            'status_updated_at': workflow.get('status_updated_at'),
            'current_status': workflow.get('current_status'),
            'last_status_change_date': last_action_date_str or workflow.get('status_updated_at')
        }
        workflow['status_light'] = calculate_status_light(order_obj)
        workflow['status_light_hint'] = get_status_light_hint(order_obj)
        if last_action_date_str:
            try:
                action_date_obj = datetime.strptime(str(last_action_date_str).strip().split()[0], '%Y-%m-%d').date()
                workflow['status_days'] = max(0, (today_date - action_date_obj).days)
            except (ValueError, TypeError):
                workflow['status_days'] = 0
        else:
            workflow['status_days'] = 0
        workflow['last_status_change_date'] = last_action_date_str
        workflow['is_locked'] = _normalize_lock_flag(workflow.get('is_locked'))
        owner_id = workflow.get('handler_id') or workflow.get('created_by_id')
        owner_id_str = str(owner_id) if owner_id is not None else ''
        workflow['can_edit_notes'] = (not cloud_read_only_enabled()) and bool(owner_id_str and owner_id_str == current_id_str) and not workflow['is_locked']
        workflow['no_workflow'] = False
        orders_list.append(workflow)

    seen_orders = {w.get('order_number') for w in orders_list}
    no_wf_query = """
        SELECT o.order_number, o.customer_name, o.order_date, o.visibility
        FROM orders o
        LEFT JOIN workflows w ON o.order_number = w.order_number
        WHERE w.id IS NULL
          AND o.status = 'ACTIVE'
          AND o.status != 'CANCELLED'
    """
    if not is_admin():
        no_wf_query += " AND o.visibility = 'all_sales'"
    no_wf_query += " ORDER BY o.order_date DESC, o.order_number DESC"
    cursor.execute(no_wf_query)
    for row in cursor.fetchall():
        o = dict(row)
        on = o.get('order_number')
        if on in seen_orders:
            continue
        orders_list.append({
            'order_number': on,
            'workflow_number': '',
            'customer_name': o.get('customer_name', ''),
            'order_date': o.get('order_date', ''),
            'order_status': 'ACTIVE',
            'visibility': o.get('visibility', ''),
            'current_status': '',
            'status_light': 'grey',
            'status_light_hint': '',
            'handler_name': '',
            'handler_id': None,
            'product_name': '', 'product_code': '',
            'quantity': '', 'factory': '',
            'production_type': '', 'expected_delivery_date': '',
            'status_days': 0, 'is_locked': False,
            'can_edit_notes': False,
            'no_workflow': True,
            'last_history_id': None,
            'last_status_change_date': '',
            'draft_date': None,
            'last_shipping_date': '',
            'last_shipping_status': '',
            'partial_ship_count': 0,
            'notes': '',
        })
    return orders_list


def _load_home_orders_from_active_source(current_role, current_user_id):
    """Load homepage rows from an external provider or the existing local SQLite DB."""
    provider = get_order_data_provider()
    if provider is not None:
        loader = getattr(provider, 'load_home_orders', None)
        if not callable(loader):
            raise RuntimeError('OrderDataProvider must implement load_home_orders(role, user_id)')
        rows = loader(current_role, current_user_id)
        return [dict(row) if not isinstance(row, dict) else row for row in (rows or [])]

    # Cloud mode must never silently fall back to a local SQLite copy. Until the
    # Render TiDB provider is registered, return an empty, valid dataset.
    if cloud_mode_enabled():
        return []

    conn = get_db()
    try:
        return _load_home_orders_dataset(conn, current_role, current_user_id)
    finally:
        conn.close()


def _cloud_provider_call(method_name, *args, **kwargs):
    """Call one Render/TiDB read method without changing LAN/SQLite behavior."""
    provider = get_order_data_provider()
    if provider is None:
        if cloud_mode_enabled():
            raise RuntimeError('Render ORDER data provider is not registered')
        return None
    method = getattr(provider, method_name, None)
    if not callable(method):
        raise RuntimeError(f'OrderDataProvider must implement {method_name}() in cloud mode')
    return method(*args, **kwargs)


@tracking_bp.route('/api/orders/all-for-filter', methods=['GET'])
@api_login_required
def api_orders_all_for_filter():
    """Return the complete home-page dataset as compact JSON for client-side filtering/rendering."""
    current_ctx = get_current_user_context()
    try:
        data = _load_home_orders_from_active_source(current_ctx.get('role', 'viewer'), current_ctx.get('id'))
    except Exception as exc:
        return jsonify({
            'success': False,
            'error': f'雲端訂單資料載入失敗: {exc}' if cloud_mode_enabled() else str(exc),
            'code': 'CLOUD_PROVIDER_ERROR' if cloud_mode_enabled() else 'LOAD_ERROR',
        }), 503
    return jsonify({
        'success': True,
        'data': data,
        'total': len(data),
        'cloud_mode': cloud_mode_enabled(),
        'cloud_read_only': cloud_read_only_enabled(),
        'provider_ready': provider_ready() if cloud_mode_enabled() else True,
        'last_synced_at': provider_last_synced_at() if cloud_mode_enabled() else None,
    })


def _load_customer_history_dataset(conn, customer_name, current_role, current_user_id, history_scope='current', include_cancelled=False):
    history_scope = _normalize_guest_history_scope(history_scope)
    cursor = conn.cursor()
    filter_info = get_filtered_resources('workflow', current_role, current_user_id)
    params = [customer_name]
    permission_sql = ''
    if not (is_admin() or filter_info['rule'] == 'all'):
        permission_sql = " AND w.handler_id = ? AND o.visibility = 'all_sales'"
        params.extend(filter_info['params'])
    scope_sql, scope_params = _guest_workflow_scope_sql(
        workflow_alias='w', history_scope=history_scope, include_cancelled=include_cancelled
    )
    params.extend(scope_params)
    query = f'''
        SELECT w.id, w.workflow_number, w.order_number,
               w.product_name, w.product_code, w.quantity, w.factory, w.production_type,
               w.expected_delivery_date, w.current_status, w.status_updated_at,
               w.created_by_id, w.handler_id, w.folder_path, w.notes, w.created_at, w.updated_at,
               o.customer_name, o.order_date, o.status AS order_status, o.visibility, o.is_locked,
               COALESCE(u.real_name, u.display_name, u.username) AS handler_name,
               (SELECT h.id FROM workflow_status_history h
                WHERE h.workflow_number = w.workflow_number
                ORDER BY h.created_at DESC, h.id DESC LIMIT 1) AS last_history_id,
               (SELECT h.action_date FROM workflow_status_history h
                WHERE h.workflow_number = w.workflow_number
                ORDER BY h.created_at DESC, h.id DESC LIMIT 1) AS last_action_date
        FROM workflows w
        INNER JOIN orders o ON o.order_number = w.order_number
        LEFT JOIN users u ON u.id = w.handler_id
        WHERE o.status = 'ACTIVE' AND o.customer_name = ?
          {permission_sql}
          AND {scope_sql}
        ORDER BY o.order_date DESC, w.created_at DESC, w.workflow_number DESC
    '''
    cursor.execute(query, params)
    today_date = date.today()
    current_id_str = str(current_user_id) if current_user_id is not None else ''
    rows = []
    for raw in cursor.fetchall():
        item = dict(raw)
        last_action = item.get('last_action_date') or item.get('status_updated_at') or ''
        light_obj = {
            'order_number': item.get('order_number'),
            'order_date': item.get('order_date'),
            'expected_delivery_date': item.get('expected_delivery_date'),
            'status_updated_at': item.get('status_updated_at'),
            'current_status': item.get('current_status'),
            'last_status_change_date': last_action,
        }
        item['status_light'] = calculate_status_light(light_obj)
        item['status_light_hint'] = get_status_light_hint(light_obj)
        try:
            d = datetime.strptime(str(last_action).strip().split()[0], '%Y-%m-%d').date() if last_action else None
            item['status_days'] = max(0, (today_date - d).days) if d else 0
        except Exception:
            item['status_days'] = 0
        item['last_status_change_date'] = last_action
        item['is_locked'] = _normalize_lock_flag(item.get('is_locked'))
        owner_id = item.get('handler_id') or item.get('created_by_id')
        item['can_edit_notes'] = (not cloud_read_only_enabled()) and str(owner_id or '') == current_id_str and not item['is_locked']
        item['no_workflow'] = False
        rows.append(item)

    no_wf_sql = '''
        SELECT o.order_number, o.customer_name, o.order_date, o.visibility
        FROM orders o
        WHERE o.status = 'ACTIVE' AND o.customer_name = ?
          AND NOT EXISTS (SELECT 1 FROM workflows w WHERE w.order_number = o.order_number)
    '''
    no_wf_params = [customer_name]
    if not is_admin():
        no_wf_sql += " AND o.visibility = 'all_sales'"
    no_wf_sql += ' ORDER BY o.order_date DESC, o.order_number DESC'
    cursor.execute(no_wf_sql, no_wf_params)
    for raw in cursor.fetchall():
        o = dict(raw)
        rows.append({
            'order_number': o.get('order_number'), 'workflow_number': '',
            'customer_name': o.get('customer_name', ''), 'order_date': o.get('order_date', ''),
            'order_status': 'ACTIVE', 'visibility': o.get('visibility', ''),
            'current_status': '', 'status_light': 'grey', 'status_light_hint': '',
            'handler_name': '', 'handler_id': None, 'product_name': '', 'product_code': '',
            'quantity': '', 'factory': '', 'production_type': '', 'expected_delivery_date': '',
            'status_days': 0, 'is_locked': False, 'can_edit_notes': False, 'no_workflow': True,
            'last_history_id': None, 'last_status_change_date': '', 'notes': '',
        })
    return rows


@tracking_bp.route('/api/customers/history-orders', methods=['GET'])
@api_login_required
def api_customer_history_orders():
    customer_input = ' '.join(str(request.args.get('customer_name') or '').strip().split())
    if not customer_input:
        return jsonify({'success': False, 'error': '缺少客户名称'}), 400
    history_scope = _normalize_guest_history_scope(request.args.get('scope') or 'current')
    include_cancelled = str(request.args.get('include_cancelled') or '').strip().lower() in {'1','true','yes','on'}
    ctx = get_current_user_context()
    if cloud_mode_enabled():
        try:
            result = _cloud_provider_call(
                'get_customer_history', customer_input, history_scope, include_cancelled,
                ctx.get('role', 'viewer'), ctx.get('id')
            ) or {}
        except Exception as exc:
            return jsonify({'success': False, 'error': f'雲端客戶資料載入失敗: {exc}'}), 503
        rows = list(result.get('data') or [])
        customer_name = result.get('customer_name') or customer_input
        response = jsonify({
            'success': True, 'customer_name': customer_name, 'scope': history_scope,
            'include_cancelled': bool(include_cancelled), 'total': len(rows), 'data': rows
        })
        response.headers['Cache-Control'] = 'no-store, private, max-age=0'
        return response

    conn = get_db()
    try:
        access_all, user_id = _customer_report_access_context()
        resolved, missing = _resolve_accessible_customer_names(conn, [customer_input], access_all, user_id)
        customer_name = resolved.get(customer_input)
        if not customer_name or missing:
            return jsonify({'success': False, 'error': '找不到该客户，或当前账号无权查看'}), 403
        rows = _load_customer_history_dataset(
            conn, customer_name, ctx.get('role', 'viewer'), ctx.get('id'),
            history_scope=history_scope, include_cancelled=include_cancelled
        )
    finally:
        conn.close()
    response = jsonify({
        'success': True, 'customer_name': customer_name, 'scope': history_scope,
        'include_cancelled': bool(include_cancelled), 'total': len(rows), 'data': rows
    })
    response.headers['Cache-Control'] = 'no-store, private, max-age=0'
    return response


@tracking_bp.route('/')
@tracking_bp.route('/index')
def index():
    """Lightweight home shell; order rows are loaded asynchronously as JSON."""
    if 'user_id' not in session:
        return render_template('tracking/login.html')

    from .status_definitions import get_statuses_by_stage_group as get_status_keys_by_stage_group

    def make_compatible_status_list(stage_group):
        keys = get_status_keys_by_stage_group(stage_group)
        labels = [get_status_label(key, 'zh_cn') for key in keys]
        return keys + labels

    response = make_response(render_template(
        'tracking/index.html',
        all_orders=[],
        total_orders=0,
        light_stats={'green': 0, 'yellow': 0, 'red': 0},
        new_and_quote_statuses=make_compatible_status_list('new_and_quote'),
        draft_statuses=make_compatible_status_list('draft'),
        sampling_statuses=make_compatible_status_list('sampling'),
        production_statuses=make_compatible_status_list('production'),
        shipping_statuses=make_compatible_status_list('shipping'),
        quote_orders=[],
        available_orders_count=0,
        STATUS=STATUS,
        STATUS_KEYS=STATUS_KEYS,
        STATUS_LABELS=STATUS_LABELS,
    ))
    # 主页面壳不缓存，避免手机继续保留旧模板 / 旧静态资源版本号。
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@tracking_bp.route('/orders')
@login_required
def orders():
    """訂單列表頁 - 與主頁相同，返回所有數據"""
    # 重定向到主頁（使用同一個HTML和數據）
    return redirect(url_for('tracking_bp.index'))

@tracking_bp.route('/orders/new', methods=['GET', 'POST'])
@admin_required
def order_new():
    """新增訂單（支持無訂單號的詢價/修圖）"""
    # GET請求 - 重定向到主頁（使用 modal 新增）
    if request.method == 'GET':
        return redirect(url_for('tracking_bp.index'))
    
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        
        conn = get_db()
        cursor = conn.cursor()
        
        order_number = data.get('order_number', '').strip()
        
        # 如果沒有提供訂單號，生成詢價/修圖編號（YU00001開始）
        if not order_number:
            cursor.execute('''
                SELECT order_number 
                FROM orders 
                WHERE order_number LIKE 'YU%'
                ORDER BY order_number DESC
                LIMIT 1
            ''')
            last_order = cursor.fetchone()
            if last_order and last_order['order_number'].startswith('YU'):
                try:
                    last_num = int(last_order['order_number'].replace('YU', ''))
                    next_num = last_num + 1
                except:
                    next_num = 1
            else:
                next_num = 1
            order_number = f'KC{next_num:05d}'
            initial_status = STATUS_KEYS['NEW_ORDER']  # 使用 key（数据库存储）
        else:
            # 如果提供了訂單號，檢查是否已存在
            cursor.execute('SELECT id FROM orders WHERE order_number = ?', (order_number,))
            if cursor.fetchone():
                conn.close()
                error = '订单号已存在'
                if request.is_json:
                    return jsonify({'success': False, 'error': error, 'code': 'DUPLICATE_ORDER'}), 400
                return jsonify({'success': False, 'error': error, 'code': 'DUPLICATE_ORDER'}), 400
            initial_status = STATUS_KEYS['NEW_ORDER']  # 使用 key（数据库存储）
        
        # 插入订单
        order_date = data.get('order_date', date.today().isoformat())
        today_str = date.today().isoformat() 
        creator_id = session.get('user_id')
        creator_name = session.get('display_name') or session.get('username')
        if creator_id and not creator_name:
            creator_name = get_user_display_name(creator_id, conn)
        
        customer_name = _normalize_customer_name(data.get('customer_name'))
        cursor.execute('''
            INSERT INTO orders (
                order_number, customer_name, created_by_id, created_by_name, order_date, current_status,
                production_type, product_name, product_code, pattern_code, expected_delivery_date, notes, last_status_change_date
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            order_number,
            customer_name,
            creator_id,
            creator_name,
            order_date,              # ← 订单日期（可以是过去、今天、未来）
            initial_status,
            data.get('production_type'),
            data.get('product_name'),
            data.get('product_code'),
            data.get('pattern_code'),
            data.get('expected_delivery_date'),
            data.get('notes'),
            today_str                # ← 关键修改！改用 today_str（永远是今天）
        ))
        
        order_id = cursor.lastrowid
        
        # 更新燈號
        update_status_light(order_id, conn)
        
        conn.commit()
        conn.close()
        
        if request.is_json:
            return jsonify({'success': True, 'message': '訂單創建成功'})
        return redirect(url_for('tracking_bp.index'))

# ==================== API路由 ====================

@tracking_bp.route('/api/auth/me', methods=['GET'])
@api_login_required
def api_auth_me():
    """獲取當前用戶信息"""
    if cloud_mode_enabled():
        ctx = get_current_user_context()
        return jsonify({
            'success': True,
            'data': {
                'id': ctx.get('id'),
                'username': ctx.get('username') or session.get('username'),
                'display_name': session.get('display_name') or ctx.get('username') or '用户',
                'role': ctx.get('role', 'viewer'),
            }
        })
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE id = ?', (g.current_user['id'],))
    user = cursor.fetchone()
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'id': user['id'],
            'username': user['username'],
            'display_name': user['display_name'],
            'role': user['role']
        }
    })

@tracking_bp.route('/api/orders', methods=['GET'])
@api_login_required
def api_orders():
    """獲取流程列表API（支持分頁）- 從 workflows 表讀取，JOIN orders 表獲取客戶信息"""
    tab = request.args.get('tab', 'all')
    stage = request.args.get('stage', 'all')
    light = request.args.get('light', 'all')
    search = request.args.get('search', '')
    locked_filter = request.args.get('locked', '')  # Admin 專用：查看未解鎖訂單
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 重构后：从 workflows 表读取流程，JOIN orders 表获取订单信息，JOIN users 表获取业务员名字
    # 基础查询：从 workflows 表开始，JOIN orders 表，JOIN users 表
    query = """
        SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility,
               COALESCE(u.real_name, u.display_name, u.username) as handler_name
        FROM workflows w
        INNER JOIN orders o ON w.order_number = o.order_number
        LEFT JOIN users u ON w.handler_id = u.id
        WHERE o.status IN ('ACTIVE', 'UNLOCKED', 'COMPLETED', 'CANCELLED')
          AND o.status != 'CANCELLED'
    """
    count_query = "SELECT COUNT(*) as count FROM workflows w INNER JOIN orders o ON w.order_number = o.order_number WHERE o.status = 'ACTIVE'"
    params = []
    
    # M0-4: 權限過濾 - Sales/Viewer 零可見性原則
    current_ctx = get_current_user_context()
    current_role = current_ctx.get('role', 'viewer')
    current_user_id = current_ctx.get('id')
    filter_info = get_filtered_resources('workflow', current_role, current_user_id)
    
    if not is_admin() and filter_info['rule'] != 'all':
        # 非 Admin：只能看到自己的流程（handler_id = current_user_id）且訂單可見
        visibility_clause = get_visibility_where_clause(current_role, table_alias='o')
        visibility_sql = f" AND {visibility_clause}" if visibility_clause else ""
        query += f" AND {filter_info['where_sql']}{visibility_sql}"
        count_query += f" AND {filter_info['where_sql']}{visibility_sql}"
        params.extend(filter_info['params'])
    elif locked_filter == '1':
        # Admin 專用：查看未解鎖訂單的流程（UNLOCKED 状态）
        query = query.replace("WHERE o.status = 'ACTIVE'", "WHERE o.status = 'UNLOCKED'")
        count_query = count_query.replace("WHERE o.status = 'ACTIVE'", "WHERE o.status = 'UNLOCKED'")
    elif locked_filter == '0':
        # Admin 專用：查看已解鎖訂單的流程（ACTIVE 状态）
        pass  # 已经是 ACTIVE
    # 否則 Admin 看到全部（包括 ACTIVE 和 UNLOCKED）
    
    if tab == 'all':
        # 排除已完成和已取消的流程
        query += " AND w.current_status NOT IN (?, ?)"
        count_query += " AND w.current_status NOT IN (?, ?)"
        params.extend([STATUS_KEYS['COMPLETED'], STATUS_KEYS['CANCELLED']])
    elif tab == 'quote':
        # 等国外确认 - 筛选所有需要国外确认的状态
        quote_statuses = get_statuses_by_stage_group('quote')
        placeholders = ','.join(['?'] * len(quote_statuses))
        query += f" AND w.current_status IN ({placeholders})"
        count_query += f" AND w.current_status IN ({placeholders})"
        params.extend(quote_statuses)
    elif tab == 'draft':
        draft_statuses = get_statuses_by_stage_group('draft')
        placeholders = ','.join(['?'] * len(draft_statuses))
        query += f" AND w.current_status IN ({placeholders})"
        count_query += f" AND w.current_status IN ({placeholders})"
        params.extend(draft_statuses)
    elif tab == 'sampling':
        sampling_statuses = get_statuses_by_stage_group('sampling')
        placeholders = ','.join(['?'] * len(sampling_statuses))
        query += f" AND w.current_status IN ({placeholders})"
        count_query += f" AND w.current_status IN ({placeholders})"
        params.extend(sampling_statuses)
    elif tab == 'production':
        production_statuses = get_statuses_by_stage_group('production')
        placeholders = ','.join(['?'] * len(production_statuses))
        query += f" AND w.current_status IN ({placeholders})"
        count_query += f" AND w.current_status IN ({placeholders})"
        params.extend(production_statuses)
    
    if stage != 'all':
        query += " AND w.current_status = ?"
        count_query += " AND w.current_status = ?"
        params.append(stage)
    
    # 灯号筛选暂时移除（需要根据状态计算）
    # if light != 'all':
    #     query += " AND status_light = ?"
    #     count_query += " AND status_light = ?"
    #     params.append(light)
    
    if search:
        # When searching, also include CANCELLED orders so admins can find them by keyword
        query = query.replace("AND o.status != 'CANCELLED'", "")
        query += " AND (w.order_number LIKE ? OR o.customer_name LIKE ? OR w.workflow_number LIKE ?)"
        count_query += " AND (w.order_number LIKE ? OR o.customer_name LIKE ? OR w.workflow_number LIKE ?)"
        search_term = f'%{search}%'
        params.extend([search_term, search_term, search_term])
    
    # 獲取總數
    cursor.execute(count_query, params)
    total = cursor.fetchone()['count']
    
    # 分頁查詢
    query += " ORDER BY o.order_date DESC, w.order_number DESC, w.created_at DESC"
    offset = (page - 1) * page_size
    query += f" LIMIT ? OFFSET ?"
    params.extend([page_size, offset])
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    # 处理查询结果，添加计算字段
    workflows_list = []
    for row in rows:
        workflow = dict(row)
        
        # 計算狀態燈（使用 models.py 中的函數）
        order_obj = {
            'order_number': workflow.get('order_number'),
            'order_date': workflow.get('order_date'),
            'expected_delivery_date': workflow.get('expected_delivery_date'),
            'status_updated_at': workflow.get('status_updated_at'),
            'current_status': workflow.get('current_status'),
            'last_status_change_date': workflow.get('status_updated_at')
        }
        workflow['status_light'] = calculate_status_light(order_obj)
        
        # 確保字段存在
        workflow['current_status'] = workflow.get('current_status', '')
        workflow['status_days'] = workflow.get('status_days', 0)
        workflow['product_name'] = workflow.get('product_name', '')
        workflow['product_code'] = workflow.get('product_code', '')
        workflow['quantity'] = workflow.get('quantity', '')
        workflow['factory'] = workflow.get('factory', '')
        workflow['production_type'] = workflow.get('production_type', '')
        workflow['expected_delivery_date'] = workflow.get('expected_delivery_date', '')
        workflow['is_locked'] = _normalize_lock_flag(workflow.get('is_locked'))
        # 訂單號層級狀態（主管設定，如 ACTIVE / CANCELLED），與業務員的 current_status 分開，
        # 明確帶給前端：訂單被取消時優先顯示「已取消」，復原時 current_status 不受影響直接顯示回來。
        workflow['order_status'] = workflow.get('order_status', '')
        
        workflows_list.append(workflow)
    
    # 使用 workflows_list 作為 orders_list（保持 API 響應格式兼容）
    orders_list = workflows_list

    # ── 補查無流程訂單（只在 tab=all、無 stage/search/locked 過濾時補）──
    # 讓業務員看到「公共池」裡等待認領的訂單
    if tab == 'all' and stage == 'all' and not search and not locked_filter:
        seen_orders = {w.get('order_number') for w in workflows_list}
        no_wf_query = """
            SELECT o.order_number, o.customer_name, o.order_date, o.visibility
            FROM orders o
            LEFT JOIN workflows w ON o.order_number = w.order_number
            WHERE w.id IS NULL
              AND o.status = 'ACTIVE'
              AND o.status != 'CANCELLED'
        """
        no_wf_params = []
        # Sales 只看 visibility = all_sales
        if not is_admin():
            no_wf_query += " AND o.visibility = 'all_sales'"
        no_wf_query += " ORDER BY o.order_date DESC, o.order_number DESC"
        cursor.execute(no_wf_query, no_wf_params)
        for row in cursor.fetchall():
            o = dict(row)
            on = o.get('order_number')
            if on in seen_orders:
                continue
            orders_list.append({
                'order_number': on,
                'workflow_number': '',
                'customer_name': o.get('customer_name', ''),
                'order_date': o.get('order_date', ''),
                'order_status': 'ACTIVE',
                'visibility': o.get('visibility', ''),
                'current_status': '',
                'status_light': 'grey',
                'handler_name': '',
                'handler_id': None,
                'product_name': '',
                'product_code': '',
                'quantity': '',
                'factory': '',
                'production_type': '',
                'expected_delivery_date': '',
                'status_days': 0,
                'is_locked': False,
                'can_edit_notes': False,
                'no_workflow': True,
            })

    conn.close()
    
    return jsonify({
        'success': True,
        'data': orders_list,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': (total + page_size - 1) // page_size
        }
    })


@tracking_bp.route('/api/orders/advanced-search', methods=['POST'])
@api_login_required
def api_orders_advanced_search():
    data = request.get_json() or {}

    order_number = (data.get('orderNumber') or '').strip()
    customer_name = (data.get('customerName') or '').strip()
    date_from = data.get('dateFrom') or None
    date_to = data.get('dateTo') or None
    status_ongoing = bool(data.get('statusOngoing', True))
    status_completed = bool(data.get('statusCompleted', False))
    status_cancelled = bool(data.get('statusCancelled', False))
    salesperson = data.get('salesperson', 'all')

    if not status_ongoing and not status_completed and not status_cancelled:
        status_ongoing = True

    if status_completed and (not date_from or not date_to):
        return jsonify({
            'success': False,
            'error': '查询已完成的订单时，必须指定时间范围'
        }), 400

    if (date_from and not date_to) or (date_to and not date_from):
        return jsonify({'success': False, 'error': '请同时填写开始与结束日期'}), 400

    if date_from and date_to:
        try:
            from_dt = datetime.strptime(date_from, '%Y-%m-%d')
            to_dt = datetime.strptime(date_to, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'error': '日期格式错误'}), 400

        if to_dt < from_dt:
            return jsonify({'success': False, 'error': '结束日期不能早于开始日期'}), 400

        if (to_dt - from_dt).days > 366:
            return jsonify({'success': False, 'error': '时间范围不能超过1年'}), 400
    else:
        from_dt = None
        to_dt = None

    conn = get_db()
    cursor = conn.cursor()

    # Build order status filter: always include ACTIVE; include CANCELLED orders if requested
    if status_cancelled:
        order_status_clause = "o.status IN ('ACTIVE', 'CANCELLED')"
    else:
        order_status_clause = "o.status = 'ACTIVE'"

    query = f"""
        SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility,
               COALESCE(u.real_name, u.display_name, u.username) as handler_name
        FROM workflows w
        INNER JOIN orders o ON w.order_number = o.order_number
        LEFT JOIN users u ON w.handler_id = u.id
        WHERE {order_status_clause}
    """
    params = []

    current_ctx = get_current_user_context()
    current_role = current_ctx.get('role', 'viewer')
    current_user_id = current_ctx.get('id')
    filter_info = get_filtered_resources('workflow', current_role, current_user_id)

    if not is_admin() and filter_info['rule'] != 'all':
        visibility_clause = get_visibility_where_clause(current_role, table_alias='o')
        visibility_sql = f" AND {visibility_clause}" if visibility_clause else ""
        query += f" AND {filter_info['where_sql']}{visibility_sql}"
        params.extend(filter_info['params'])
    elif is_admin():
        if isinstance(salesperson, list):
            if 'all' not in salesperson and len(salesperson) > 0:
                placeholders = ','.join(['?'] * len(salesperson))
                query += f" AND w.handler_id IN ({placeholders})"
                params.extend(salesperson)
        elif salesperson and salesperson != 'all':
            query += " AND w.handler_id = ?"
            params.append(salesperson)

    if order_number:
        query += " AND (w.order_number LIKE ? OR w.workflow_number LIKE ?)"
        like_term = f'%{order_number}%'
        params.extend([like_term, like_term])

    if customer_name:
        query += " AND o.customer_name LIKE ?"
        params.append(f'%{customer_name}%')

    status_keys = list(STATUS_KEYS.values())
    completed_key = STATUS_KEYS['COMPLETED']
    cancelled_key = STATUS_KEYS['CANCELLED']
    ongoing_keys = [key for key in status_keys if key not in (completed_key, cancelled_key)]
    ongoing_labels = [get_status_label(key, 'zh_cn') for key in ongoing_keys]

    selected_statuses = []
    if status_ongoing:
        selected_statuses.extend(ongoing_keys + ongoing_labels)
    if status_completed:
        selected_statuses.extend([completed_key, get_status_label(completed_key, 'zh_cn')])
    # Note: status_cancelled uses o.status = 'CANCELLED' (order-level), not w.current_status

    if selected_statuses:
        placeholders = ','.join(['?'] * len(selected_statuses))
        query += f" AND w.current_status IN ({placeholders})"
    elif status_cancelled and not status_ongoing and not status_completed:
        pass  # already filtered by o.status = 'CANCELLED' in the WHERE clause
        params.extend(selected_statuses)

    if from_dt and to_dt:
        query += " AND o.order_date BETWEEN ? AND ?"
        params.extend([date_from, date_to])

    query += " ORDER BY o.order_date DESC, w.order_number DESC, w.created_at DESC LIMIT 1000"

    cursor.execute(query, params)
    rows = cursor.fetchall()

    orders_list = []
    for row in rows:
        workflow = dict(row)
        order_obj = {
            'order_number': workflow.get('order_number'),
            'order_date': workflow.get('order_date'),
            'expected_delivery_date': workflow.get('expected_delivery_date'),
            'status_updated_at': workflow.get('status_updated_at'),
            'current_status': workflow.get('current_status'),
            'last_status_change_date': workflow.get('status_updated_at')
        }
        workflow['status_light'] = calculate_status_light(order_obj)
        workflow['current_status'] = workflow.get('current_status', '')
        workflow['status_days'] = workflow.get('status_days', 0)
        workflow['product_name'] = workflow.get('product_name', '')
        workflow['product_code'] = workflow.get('product_code', '')
        workflow['quantity'] = workflow.get('quantity', '')
        workflow['factory'] = workflow.get('factory', '')
        workflow['production_type'] = workflow.get('production_type', '')
        workflow['expected_delivery_date'] = workflow.get('expected_delivery_date', '')
        workflow['is_locked'] = _normalize_lock_flag(workflow.get('is_locked'))
        # 訂單號層級狀態（主管在「全部訂單號」頁面設定，如 ACTIVE / CANCELLED）。
        # 與 current_status（業務員填的流程階段）是兩個獨立欄位，這裡明確帶給前端，
        # 讓前端可以在訂單被取消時優先顯示「已取消」，不被 current_status 蓋掉；
        # 若主管把取消復原（改回 ACTIVE），current_status 完全沒被動過，會自動恢復顯示。
        workflow['order_status'] = workflow.get('order_status', '')
        orders_list.append(workflow)

    conn.close()

    return jsonify({
        'success': True,
        'orders': orders_list,
        'count': len(orders_list)
    })


# ==================== 導出 Excel ====================

@tracking_bp.route('/api/workflows/export', methods=['GET', 'POST'])
@api_login_required
def api_export_workflows_xlsx():
    """
    導出工作流數據為 Excel（.xlsx）。
    
    支持的篩選參數（與前端 currentFilter 對應）：
    - search: 搜索關鍵字（訂單號/客戶名稱）
    - stage_group: 階段分組（如 draft, sampling, production, completed, cancelled）
    - substatus: 子狀態（如 DRAFT_MAKING）
    - light: 燈號（red, yellow, green）
    - handler_id: 負責人 ID
    - include_completed: 是否包含已完成（1/0）
    - include_cancelled: 是否包含已取消（1/0）
    
    返回：.xlsx 文件（Sheet1=流程總覽, Sheet2=時間軸明細）
    """
    from io import BytesIO
    from flask import send_file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl 未安裝，無法導出 Excel'}), 500

    request_json = request.get_json(silent=True) if request.method == 'POST' else None
    selected_items = (request_json or {}).get('selected_items') or []

    # --- 解析篩選參數 ---
    search = (request.args.get('search') or '').strip()
    stage_groups_raw = (request.args.get('stage_groups') or '').strip()  # 逗號分隔
    substatus = (request.args.get('substatus') or '').strip()
    light_filter = (request.args.get('light') or '').strip()  # 逗號分隔 如 "red,yellow"
    handler_ids_raw = (request.args.get('handler_ids') or '').strip()
    include_completed = request.args.get('include_completed', '0') == '1'
    include_cancelled = request.args.get('include_cancelled', '0') == '1'

    stage_groups = [s.strip() for s in stage_groups_raw.split(',') if s.strip()] if stage_groups_raw else []
    handler_ids = [s.strip() for s in handler_ids_raw.split(',') if s.strip()] if handler_ids_raw else []

    # --- 查詢工作流數據（與 index() 相同邏輯）---
    conn = get_db()
    ensure_order_lock_columns(conn)
    cursor = conn.cursor()
    current_ctx = get_current_user_context()
    current_user_id = current_ctx.get('id')
    current_role = current_ctx.get('role', 'viewer')

    filter_info = get_filtered_resources('workflow', current_role, current_user_id)
    
    if is_admin() or filter_info['rule'] == 'all':
        query = """
            SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility, o.is_locked,
                   COALESCE(u.real_name, u.display_name, u.username) as handler_name
            FROM workflows w
            INNER JOIN orders o ON w.order_number = o.order_number
            LEFT JOIN users u ON w.handler_id = u.id
            WHERE o.status = 'ACTIVE'
            ORDER BY o.order_date DESC, w.created_at DESC
        """
        cursor.execute(query)
    else:
        query = """
            SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility, o.is_locked,
                   COALESCE(u.real_name, u.display_name, u.username) as handler_name
            FROM workflows w
            INNER JOIN orders o ON w.order_number = o.order_number
            LEFT JOIN users u ON w.handler_id = u.id
            WHERE w.handler_id = ?
              AND o.status = 'ACTIVE'
              AND o.visibility = 'all_sales'
            ORDER BY o.order_date DESC, w.created_at DESC
        """
        cursor.execute(query, filter_info['params'])

    all_workflows = []
    today_date = date.today()
    for row in cursor.fetchall():
        wf = dict(row)
        wf['is_locked'] = _normalize_lock_flag(wf.get('is_locked'))
        
        # 計算狀態燈
        order_obj = {
            'order_number': wf.get('order_number'),
            'order_date': wf.get('order_date'),
            'expected_delivery_date': wf.get('expected_delivery_date'),
            'status_updated_at': wf.get('status_updated_at'),
            'current_status': wf.get('current_status'),
            'last_status_change_date': wf.get('status_updated_at')
        }
        wf['status_light'] = calculate_status_light(order_obj)
        all_workflows.append(wf)

    # Browser-supplied IDs are the source of truth for the current frontend filter result.
    selected_workflow_numbers = {str(x.get('workflow_number') or '').strip() for x in selected_items if isinstance(x, dict) and x.get('workflow_number')}
    selected_order_numbers = {str(x.get('order_number') or '').strip() for x in selected_items if isinstance(x, dict) and x.get('order_number')}
    if selected_items:
        all_workflows = [wf for wf in all_workflows if str(wf.get('workflow_number') or '') in selected_workflow_numbers]
        # Frontend already applied every filter. Do not filter these selected rows a second time.
        search = ''
        stage_groups = []
        substatus = ''
        light_filter = ''
        handler_ids = []
        include_completed = True
        include_cancelled = True

    # --- 在 Python 中執行篩選（與前端 applyFilters 對應）---
    from .status_definitions import get_stage_group as _get_stage_group, get_statuses_by_stage_group as _get_statuses_by_stage_group

    filtered_workflows = []
    for wf in all_workflows:
        status_key = wf.get('current_status', '')
        wf_stage = _get_stage_group(status_key)
        
        # 1. 階段篩選
        if stage_groups and 'all' not in stage_groups:
            stage_match = False
            if wf_stage in stage_groups:
                stage_match = True
            # 「等國外確認」特殊篩選
            if 'waiting_confirm' in stage_groups:
                if status_key in ['QUOTE_CONFIRMING', 'DRAFT_CONFIRMING', 'SAMPLE_CONFIRMING']:
                    stage_match = True
            if not stage_match:
                continue

        # 2. 已完成/已取消 排除（默認排除，除非明確包含）
        if status_key == 'COMPLETED' and not include_completed:
            # 如果 stage_groups 明確包含 completed，則保留
            if 'completed' not in stage_groups:
                continue
        if status_key == 'CANCELLED' and not include_cancelled:
            if 'cancelled' not in stage_groups:
                continue

        # 3. 子狀態篩選
        if substatus and substatus != 'all':
            if status_key != substatus:
                continue

        # 4. 搜索篩選
        if search:
            order_num = (wf.get('order_number') or '').upper()
            cust_name = (wf.get('customer_name') or '').upper()
            search_upper = search.upper()
            if search_upper not in order_num and search_upper not in cust_name:
                # 也做模糊匹配（如 26500 匹配 KC26500）
                wf_number = (wf.get('workflow_number') or '').upper()
                if search_upper not in wf_number:
                    continue

        # 5. 燈號篩選
        if light_filter:
            active_lights = [l.strip() for l in light_filter.split(',') if l.strip()]
            if active_lights and wf.get('status_light') not in active_lights:
                continue

        # 6. 負責人篩選
        if handler_ids:
            wf_handler = str(wf.get('handler_id', ''))
            if wf_handler not in handler_ids:
                continue

        filtered_workflows.append(wf)

    # --- 查詢所有相關的時間軸數據 ---
    workflow_numbers = [wf['workflow_number'] for wf in filtered_workflows]
    
    # 批量查詢所有時間軸記錄
    timeline_map = {}  # workflow_number -> [history records]
    if workflow_numbers:
        placeholders = ','.join(['?'] * len(workflow_numbers))
        cursor.execute(f'''
            SELECT h.*, 
                   COALESCE(u.real_name, u.display_name, u.username) as operator_name
            FROM workflow_status_history h
            LEFT JOIN users u ON h.operator_id = u.id
            WHERE h.workflow_number IN ({placeholders})
            ORDER BY h.workflow_number, h.action_date ASC, h.created_at ASC, h.id ASC
        ''', workflow_numbers)
        for row in cursor.fetchall():
            rec = dict(row)
            wn = rec['workflow_number']
            if wn not in timeline_map:
                timeline_map[wn] = []
            timeline_map[wn].append(rec)

    # --- 圖片統計資料（主管參考圖=order_files；業務員附件圖=workflow_files）---
    if selected_items:
        wf_map = {str(wf.get('workflow_number') or ''): wf for wf in all_workflows}
        source_rows = []
        for item in selected_items:
            if not isinstance(item, dict):
                continue
            wf_no = str(item.get('workflow_number') or '').strip()
            order_no = str(item.get('order_number') or '').strip()
            wf = wf_map.get(wf_no, {})
            source_rows.append({
                'workflow_number': wf_no,
                'order_number': order_no or wf.get('order_number', ''),
                'customer_name': item.get('customer_name') or wf.get('customer_name', ''),
                'product_code': wf.get('product_code', ''),
                'production_type': wf.get('production_type', '') or wf.get('product_name', ''),
            })
    else:
        source_rows = [{
            'workflow_number': wf.get('workflow_number', ''),
            'order_number': wf.get('order_number', ''),
            'customer_name': wf.get('customer_name', ''),
            'product_code': wf.get('product_code', ''),
            'production_type': wf.get('production_type', '') or wf.get('product_name', ''),
        } for wf in filtered_workflows]

    source_order_numbers = sorted({r['order_number'] for r in source_rows if r.get('order_number')})
    source_workflow_numbers = sorted({r['workflow_number'] for r in source_rows if r.get('workflow_number')})
    order_image_counts = {}
    workflow_image_counts = {}
    if source_order_numbers:
        placeholders = ','.join(['?'] * len(source_order_numbers))
        sql = f"""
            SELECT order_number, COUNT(*) AS cnt FROM order_files
            WHERE order_number IN ({placeholders})
              AND (COALESCE(mime_type, '') LIKE 'image/%'
                   OR LOWER(COALESCE(original_filename, '')) GLOB '*.jpg'
                   OR LOWER(COALESCE(original_filename, '')) GLOB '*.jpeg'
                   OR LOWER(COALESCE(original_filename, '')) GLOB '*.png'
                   OR LOWER(COALESCE(original_filename, '')) GLOB '*.webp'
                   OR LOWER(COALESCE(original_filename, '')) GLOB '*.bmp'
                   OR LOWER(COALESCE(original_filename, '')) GLOB '*.gif')
            GROUP BY order_number
        """
        cursor.execute(sql, source_order_numbers)
        order_image_counts = {str(r['order_number']): int(r['cnt'] or 0) for r in cursor.fetchall()}
    if source_workflow_numbers:
        placeholders = ','.join(['?'] * len(source_workflow_numbers))
        sql = f"""
            SELECT workflow_number, COUNT(*) AS cnt FROM workflow_files
            WHERE workflow_number IN ({placeholders}) AND COALESCE(is_deleted, 0) = 0
              AND (COALESCE(file_type, '') LIKE 'image/%'
                   OR LOWER(COALESCE(file_name, '')) GLOB '*.jpg'
                   OR LOWER(COALESCE(file_name, '')) GLOB '*.jpeg'
                   OR LOWER(COALESCE(file_name, '')) GLOB '*.png'
                   OR LOWER(COALESCE(file_name, '')) GLOB '*.webp'
                   OR LOWER(COALESCE(file_name, '')) GLOB '*.bmp'
                   OR LOWER(COALESCE(file_name, '')) GLOB '*.gif')
            GROUP BY workflow_number
        """
        cursor.execute(sql, source_workflow_numbers)
        workflow_image_counts = {str(r['workflow_number']): int(r['cnt'] or 0) for r in cursor.fetchall()}

    image_stats_rows = []
    for r in source_rows:
        supervisor_count = order_image_counts.get(str(r.get('order_number') or ''), 0)
        sales_count = workflow_image_counts.get(str(r.get('workflow_number') or ''), 0)
        total_count = supervisor_count + sales_count
        if supervisor_count == 0 and sales_count == 0:
            result_text = '无图片'
        elif supervisor_count == 0:
            result_text = '缺主管图'
        elif sales_count == 0:
            result_text = '缺业务图'
        else:
            result_text = '完整'
        item = dict(r)
        item.update({'supervisor_count': supervisor_count, 'sales_count': sales_count, 'total_count': total_count, 'result': result_text})
        image_stats_rows.append(item)

    conn.close()

    # --- 構建 STATUS_FLOW_ORDER 用於 Sheet1 欄位 ---
    from .status_definitions import STATUS_FLOW_ORDER as _STATUS_FLOW_ORDER

    # Sheet1 的狀態欄順序（排除 CANCELLED）
    status_columns = list(_STATUS_FLOW_ORDER)  # ['NEW_ORDER', 'QUOTE_CONFIRMING', ..., 'COMPLETED']

    # --- 创建 Excel ---
    wb = Workbook()
    
    # ========== Sheet 1: 流程总览 ==========
    ws1 = wb.active
    ws1.title = "流程总览"

    # 样式定义
    header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='000000')
    header_fill = PatternFill(start_color='D5D8DC', end_color='D5D8DC', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Microsoft YaHei', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC')
    )
    
    light_fills = {
        'green': PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid'),
        'yellow': PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid'),
        'red': PatternFill(start_color='FDEDEC', end_color='FDEDEC', fill_type='solid'),
    }

    # 基础栏位（简体中文）
    # 注：两个「历时」栏位与首页表格保持一致 —— 「距下单日」= 今天-订单日期；
    # 「当前阶段等待天数」= 今天-最后一次状态变更日期（对应首页 status_days）
    base_headers = ['流程号', '订单号', '客户名称', '负责人', '产品名称', '产品代码',
                     '数量', '工厂', '生产类型', '当前状态', '历时（当前阶段等待天数）',
                     '灯号', '订单日期', '历时（距下单日）', '交货日期']
    # 状态栏位
    status_headers = [get_status_label(s, 'zh_cn') for s in status_columns]
    # 最后一栏
    tail_headers = ['最新备注']
    
    all_headers = base_headers + status_headers + tail_headers

    # 寫入表頭
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 寫入數據
    for row_idx, wf in enumerate(filtered_workflows, 2):
        wf_number = wf.get('workflow_number', '')
        status_key = wf.get('current_status', '')
        light = wf.get('status_light', '')
        
        light_display = {'green': '🟢 正常', 'yellow': '🟡 注意', 'red': '🔴 逾期'}.get(light, '-')

        # 計算每個狀態欄的日期 + 停留天數（提前到這裡，因為下面的「历时」也需要用到 timeline）
        timeline = timeline_map.get(wf_number, [])

        # 歷時（距下單日）= 今天 - 訂單日期，與首頁 order-age 邏輯一致
        order_age_value = None
        order_date_str = wf.get('order_date', '')
        if order_date_str:
            try:
                order_date_obj = datetime.strptime(str(order_date_str).strip().split()[0], '%Y-%m-%d').date()
                # Keep the underlying Excel value numeric for correct sorting.
                # The visible "天" suffix is applied via Excel number format below.
                order_age_value = max(0, (today_date - order_date_obj).days)
            except (ValueError, TypeError):
                pass

        # 歷時（當前階段等待天數）= 今天 - 最後一次狀態變更日期，與首頁 status_days 邏輯一致
        # 最後一次狀態變更 = 時間軸（按 action_date/created_at/id 排序）中的最後一條記錄
        status_days_value = None
        if timeline:
            last_action_date_str = timeline[-1].get('action_date', '')
            if last_action_date_str:
                try:
                    last_action_date_obj = datetime.strptime(
                        str(last_action_date_str).strip().split()[0], '%Y-%m-%d'
                    ).date()
                    # Keep the underlying Excel value numeric for correct sorting.
                    # The visible "天" suffix is applied via Excel number format below.
                    status_days_value = max(0, (today_date - last_action_date_obj).days)
                except (ValueError, TypeError):
                    pass

        # 基礎欄位值
        base_values = [
            wf_number,
            wf.get('order_number', ''),
            wf.get('customer_name', ''),
            wf.get('handler_name', ''),
            wf.get('product_name', ''),
            wf.get('product_code', ''),
            wf.get('quantity', ''),
            wf.get('factory', ''),
            wf.get('production_type', ''),
            get_status_label(status_key, 'zh_cn'),
            status_days_value,
            light_display,
            wf.get('order_date', ''),
            order_age_value,
            wf.get('expected_delivery_date', ''),
        ]
        status_values = []
        for s_idx, s_key in enumerate(status_columns):
            # 找到到達這個狀態的記錄
            matching = [t for t in timeline if t.get('to_status') == s_key]
            if matching:
                record = matching[0]  # 取第一次到達的
                action_date = record.get('action_date', '')
                
                # 計算停留天數
                # 下一個狀態的到達日期 - 當前狀態的到達日期
                days_text = ''
                # 找到時間軸中此記錄之後的下一條
                timeline_idx = None
                for ti, t in enumerate(timeline):
                    if t['id'] == record['id']:
                        timeline_idx = ti
                        break
                
                if timeline_idx is not None and timeline_idx + 1 < len(timeline):
                    next_record = timeline[timeline_idx + 1]
                    next_date_str = next_record.get('action_date', '')
                    if action_date and next_date_str:
                        try:
                            d1 = datetime.strptime(action_date.strip().split()[0], '%Y-%m-%d').date()
                            d2 = datetime.strptime(next_date_str.strip().split()[0], '%Y-%m-%d').date()
                            days = (d2 - d1).days
                            days_text = f" ({days}天)"
                        except (ValueError, TypeError):
                            pass
                elif s_key == status_key and action_date:
                    # 當前狀態，計算至今天
                    try:
                        d1 = datetime.strptime(action_date.strip().split()[0], '%Y-%m-%d').date()
                        days = (today_date - d1).days
                        days_text = f" ({days}天,进行中)"
                    except (ValueError, TypeError):
                        pass
                
                status_values.append(f"{action_date}{days_text}")
            else:
                status_values.append('')
        
        # 最新備註：取時間軸中最後一條有備註的
        latest_note = ''
        for t in reversed(timeline):
            t_notes = t.get('notes') or ''
            if t_notes.strip():
                latest_note = t_notes.strip()
                break
        # 如果沒有，用 workflow.notes
        if not latest_note:
            latest_note = (wf.get('notes') or '').strip()
        # 截斷到 50 字
        if len(latest_note) > 50:
            latest_note = latest_note[:50] + '...'
        
        tail_values = [latest_note]
        
        row_values = base_values + status_values + tail_values
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            # 狀態欄居中
            if col_idx > len(base_headers) and col_idx <= len(base_headers) + len(status_columns):
                cell.alignment = center_align
            else:
                cell.alignment = cell_align

            # K「当前阶段等待天数」与 N「距下单日」都保存为纯数值，
            # 仅通过 Excel 数字格式显示「天」，确保排序/筛选按数值处理。
            if col_idx in (11, 14) and isinstance(val, (int, float)):
                cell.number_format = '0"天"'
                # N「距下单日」在首页是左对齐；即使底层改成数值，导出仍保持相同视觉排列。
                if col_idx == 14:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 整行燈號底色
        if light in light_fills:
            for col_idx in range(1, len(row_values) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = light_fills[light]

    # 調整欄寬
    col_widths_base = [16, 14, 14, 10, 16, 12, 10, 12, 12, 14, 10, 12, 12]
    col_widths_status = [18] * len(status_columns)
    col_widths_tail = [30]
    all_widths = col_widths_base + col_widths_status + col_widths_tail
    for i, w in enumerate(all_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # 凍結首行 + 前 3 欄
    ws1.freeze_panes = 'D2'

    # 自動篩選
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(filtered_workflows) + 1}"

    # ========== Sheet 2: 时间轴明细 ==========
    ws2 = wb.create_sheet(title="时间轴明细")
    
    detail_headers = ['流程号', '订单号', '客户名称', '负责人', '状态', '到达日期', '停留天数', '操作人', '备注']
    
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    detail_row = 2
    for wf in filtered_workflows:
        wf_number = wf.get('workflow_number', '')
        timeline = timeline_map.get(wf_number, [])
        
        for ti, record in enumerate(timeline):
            to_status = record.get('to_status') or ''
            action_date = record.get('action_date') or ''
            operator_name = record.get('operator_name') or ''
            notes = record.get('notes') or ''
            
            # 計算停留天數
            days_display = ''
            if ti + 1 < len(timeline):
                next_date_str = timeline[ti + 1].get('action_date', '')
                if action_date and next_date_str:
                    try:
                        d1 = datetime.strptime(action_date.strip().split()[0], '%Y-%m-%d').date()
                        d2 = datetime.strptime(next_date_str.strip().split()[0], '%Y-%m-%d').date()
                        days_display = str((d2 - d1).days)
                    except (ValueError, TypeError):
                        pass
            elif to_status == wf.get('current_status') and action_date:
                try:
                    d1 = datetime.strptime(action_date.strip().split()[0], '%Y-%m-%d').date()
                    days_display = f"{(today_date - d1).days} (进行中)"
                except (ValueError, TypeError):
                    pass

            detail_values = [
                wf_number,
                wf.get('order_number', ''),
                wf.get('customer_name', ''),
                wf.get('handler_name', ''),
                get_status_label(to_status, 'zh_cn'),
                action_date,
                days_display,
                operator_name,
                notes,
            ]
            
            for col_idx, val in enumerate(detail_values, 1):
                cell = ws2.cell(row=detail_row, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = cell_align
            
            detail_row += 1

    # 調整 Sheet2 欄寬
    detail_widths = [16, 14, 14, 10, 14, 12, 12, 10, 40]
    for i, w in enumerate(detail_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    
    # 凍結首行
    ws2.freeze_panes = 'A2'
    # 自動篩選
    if detail_row > 2:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{detail_row - 1}"

    # ========== Sheet 3: 图片统计 ==========
    ws3 = wb.create_sheet(title="图片统计")
    image_headers = ['流程号', '订单号', '客户名称', '花型 / 产品', '产品编号', '主管参考图', '业务员附件图', '图片合计', '检查结果']
    for col_idx, header in enumerate(image_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    zero_fill = PatternFill(start_color='FDE2E2', end_color='FDE2E2', fill_type='solid')
    zero_font = Font(name='Microsoft YaHei', size=10, bold=True, color='C62828')
    ok_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    warn_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    for row_idx, item in enumerate(image_stats_rows, 2):
        vals = [item.get('workflow_number', ''), item.get('order_number', ''), item.get('customer_name', ''),
                item.get('production_type', ''), item.get('product_code', ''), item.get('supervisor_count', 0),
                item.get('sales_count', 0), item.get('total_count', 0), item.get('result', '')]
        for col_idx, val in enumerate(vals, 1):
            c = ws3.cell(row=row_idx, column=col_idx, value=val)
            c.font = cell_font
            c.alignment = center_align if col_idx >= 6 else cell_align
            c.border = thin_border
        for col_idx in (6, 7, 8):
            c = ws3.cell(row=row_idx, column=col_idx)
            if isinstance(c.value, (int, float)) and c.value == 0:
                c.fill = zero_fill
                c.font = zero_font
        result_cell = ws3.cell(row=row_idx, column=9)
        if item.get('result') == '完整':
            result_cell.fill = ok_fill
        else:
            result_cell.fill = warn_fill
            result_cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color='C62828')

    for i, width in enumerate([16, 14, 20, 22, 16, 12, 14, 10, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = width
    ws3.freeze_panes = 'A2'
    if image_stats_rows:
        ws3.auto_filter.ref = f"A1:I{len(image_stats_rows) + 1}"

    # --- 输出 ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 文件名：简体中文 + 日期时间（避免重复）
    now = datetime.now()
    filename = f"流程数据导出_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== PyQt / 外部客户端整合 API ====================

def _normalize_customer_lookup_name(value):
    """Normalize a customer name for exact logical matching without changing DB data."""
    return ' '.join(str(value or '').strip().split()).casefold()


def _api_bool(value, default=False):
    if value is None:
        return bool(default)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in {'1', 'true', 'yes', 'y', 'on'}:
        return True
    if text in {'0', 'false', 'no', 'n', 'off', ''}:
        return False
    return bool(default)


def _resolve_accessible_customer_names(conn, requested_names, access_all, current_user_id):
    """Resolve caller-supplied names to the exact DB spelling, respecting permissions."""
    requested = []
    seen = set()
    for raw in requested_names or []:
        text = ' '.join(str(raw or '').strip().split())
        key = _normalize_customer_lookup_name(text)
        if not key or key in seen:
            continue
        seen.add(key)
        requested.append((text, key))
    if not requested:
        return {}, []

    cursor = conn.cursor()
    sql = """
        SELECT DISTINCT o.customer_name
        FROM orders o
        LEFT JOIN workflows w ON w.order_number = o.order_number
        WHERE o.status = 'ACTIVE'
          AND o.customer_name IS NOT NULL
          AND TRIM(o.customer_name) != ''
    """
    params = []
    if not access_all:
        sql += " AND o.visibility = 'all_sales' AND (w.handler_id = ? OR w.id IS NULL)"
        params.append(current_user_id)
    cursor.execute(sql, params)

    available = {}
    for row in cursor.fetchall():
        db_name = ' '.join(str(row['customer_name'] or '').strip().split())
        key = _normalize_customer_lookup_name(db_name)
        if key and key not in available:
            available[key] = db_name

    resolved = {}
    missing = []
    for original, key in requested:
        db_name = available.get(key)
        if db_name:
            resolved[original] = db_name
        else:
            missing.append(original)
    return resolved, missing


def _customer_report_items_for_name(conn, customer_name, access_all, current_user_id, include_completed=False):
    """Return report item IDs for one exact customer, preserving newest orders first."""
    cursor = conn.cursor()
    sql = """
        SELECT w.workflow_number, w.order_number, w.current_status,
               o.order_date, w.created_at
        FROM workflows w
        INNER JOIN orders o ON o.order_number = w.order_number
        WHERE o.status = 'ACTIVE'
          AND o.customer_name = ?
          AND w.current_status != ?
    """
    params = [customer_name, STATUS_KEYS['CANCELLED']]
    if not include_completed:
        sql += " AND w.current_status != ?"
        params.append(STATUS_KEYS['COMPLETED'])
    if not access_all:
        sql += " AND w.handler_id = ? AND o.visibility = 'all_sales'"
        params.append(current_user_id)
    sql += " ORDER BY o.order_date DESC, w.order_number DESC, w.created_at DESC"
    cursor.execute(sql, params)
    items = [
        {'workflow_number': str(row['workflow_number']), 'order_number': str(row['order_number'])}
        for row in cursor.fetchall()
        if row['workflow_number']
    ]

    no_wf_sql = """
        SELECT o.order_number, o.order_date
        FROM orders o
        WHERE o.status = 'ACTIVE'
          AND o.customer_name = ?
          AND NOT EXISTS (SELECT 1 FROM workflows w WHERE w.order_number = o.order_number)
    """
    no_wf_params = [customer_name]
    if not access_all:
        no_wf_sql += " AND o.visibility = 'all_sales'"
    no_wf_sql += " ORDER BY o.order_date DESC, o.order_number DESC"
    cursor.execute(no_wf_sql, no_wf_params)
    items.extend(
        {'workflow_number': '', 'order_number': str(row['order_number'])}
        for row in cursor.fetchall()
        if row['order_number']
    )
    return items


def _public_customer_order(entry):
    """Return fields useful to an external client; internal sticky-note text is omitted."""
    status_key = str(entry.get('current_status') or '')
    return {
        'workflow_number': str(entry.get('workflow_number') or ''),
        'order_number': str(entry.get('order_number') or ''),
        'customer_name': str(entry.get('customer_name') or ''),
        'order_date': entry.get('order_date') or '',
        'product_name': entry.get('product_name') or '',
        'product_code': entry.get('product_code') or '',
        'quantity': entry.get('quantity') or '',
        'factory': entry.get('factory') or '',
        'production_type': entry.get('production_type') or '',
        'expected_delivery_date': entry.get('expected_delivery_date') or '',
        'current_status': status_key,
        'status_zh_cn': get_status_label(status_key, 'zh_cn') if status_key else '',
        'status_es': get_status_label(status_key, 'es') if status_key else '',
        'handler_name': entry.get('handler_name') or '',
        'no_workflow': bool(entry.get('no_workflow')),
        'report_item': {
            'workflow_number': str(entry.get('workflow_number') or ''),
            'order_number': str(entry.get('order_number') or ''),
        },
    }


def _load_exact_customer_orders(conn, requested_names, include_completed=False):
    access_all, user_id = _customer_report_access_context()
    resolved, missing = _resolve_accessible_customer_names(conn, requested_names, access_all, user_id)
    groups = []
    for requested_name, db_name in resolved.items():
        items = _customer_report_items_for_name(
            conn, db_name, access_all, user_id, include_completed=include_completed
        )
        entries = load_report_entries(conn, items, access_all, user_id) if items else []
        groups.append({
            'requested_name': requested_name,
            'customer_name': db_name,
            'order_count': len(entries),
            'orders': [_public_customer_order(e) for e in entries],
            'report_items': [
                {'workflow_number': str(e.get('workflow_number') or ''), 'order_number': str(e.get('order_number') or '')}
                for e in entries
            ],
        })
    return groups, missing


def _job_with_download_urls(job):
    item = _customer_report_job_snapshot(job)
    if not item:
        return None
    for f in item.get('files') or []:
        file_id = f.pop('file_id', None)
        if file_id:
            f['url'] = url_for('tracking_bp.api_customer_report_download', file_id=file_id)
    return item


def _enqueue_customer_report(entries, report_format, language, image_source, image_count, image_order, pdf_attachment_mode, owner_user_id):
    customers = _validate_customer_report_customer_limit(entries)

    request_key = _customer_report_request_key(
        entries, report_format, language, image_source, image_count, image_order, pdf_attachment_mode
    )
    job_id = os.urandom(16).hex()
    job = {
        'id': job_id,
        'owner_user_id': owner_user_id,
        'status': 'queued',
        'created_at': datetime.now().isoformat(timespec='seconds'),
        'started_at': None,
        'finished_at': None,
        'order_count': len(entries),
        'customers': customers,
        'format': report_format,
        'language': language,
        'pdf_attachment_mode': pdf_attachment_mode,
        'files': [],
        'error': None,
        'request_key': request_key,
    }

    # Capacity check + duplicate suppression must be atomic. The job is inserted while
    # holding the same lock, before submit(), so simultaneous web/mobile/PyQt requests
    # cannot all pass the limit check at once.
    with _customer_report_jobs_lock:
        for existing in _customer_report_jobs.values():
            if (
                existing.get('owner_user_id') == owner_user_id
                and existing.get('request_key') == request_key
                and existing.get('status') in {'queued', 'processing'}
            ):
                return existing, True

        queued_count, _processing_count, owner_active = _customer_report_queue_counts_locked(owner_user_id)
        if owner_active >= max(1, int(CUSTOMER_REPORT_MAX_ACTIVE_PER_USER)):
            raise CustomerReportQueueError(
                f'你的报告队列已有 {owner_active} 个任务，请等其中一个完成后再加入',
                'USER_REPORT_QUEUE_LIMIT',
                429,
            )
        if queued_count >= max(1, int(CUSTOMER_REPORT_MAX_QUEUED)):
            raise CustomerReportQueueError(
                '目前报告生成队列较忙，请稍后再试',
                'REPORT_QUEUE_FULL',
                429,
            )
        _customer_report_jobs[job_id] = job

    try:
        _customer_report_executor.submit(
            _run_customer_report_job,
            job_id,
            [dict(x) for x in entries],
            report_format,
            language,
            image_source,
            image_count,
            image_order,
            pdf_attachment_mode,
        )
    except Exception:
        _customer_report_set_job(
            job_id,
            status='failed',
            finished_at=datetime.now().isoformat(timespec='seconds'),
            error='无法启动报告生成任务',
        )
        raise
    return job, False


@tracking_bp.route('/api/customers/orders', methods=['GET'])
@api_login_required
def api_customer_exact_orders():
    """Exact customer -> current orders. Intended for PyQt/automation clients."""
    customer_name = request.args.get('name', '').strip()
    if not customer_name:
        return jsonify({'success': False, 'error': '缺少客户名称 name', 'code': 'MISSING_CUSTOMER_NAME'}), 400
    include_completed = _api_bool(request.args.get('include_completed'), False)
    conn = get_db()
    try:
        groups, _ = _load_exact_customer_orders(conn, [customer_name], include_completed=include_completed)
    finally:
        conn.close()
    if not groups:
        return jsonify({
            'success': True,
            'found': False,
            'requested_name': customer_name,
            'customer': None,
            'order_count': 0,
            'orders': [],
            'report_items': [],
        })
    group = groups[0]
    return jsonify({
        'success': True,
        'found': True,
        'requested_name': customer_name,
        'customer': group['customer_name'],
        'order_count': group['order_count'],
        'orders': group['orders'],
        'report_items': group['report_items'],
    })


@tracking_bp.route('/api/customers/orders/batch', methods=['POST'])
@api_login_required
def api_customer_exact_orders_batch():
    """Resolve and return current orders for many exact customer names in one call."""
    data = request.get_json(silent=True) or {}
    customers = data.get('customers') or []
    if not isinstance(customers, list) or not customers:
        return jsonify({'success': False, 'error': 'customers 必须是非空数组', 'code': 'INVALID_CUSTOMERS'}), 400
    if len(customers) > 200:
        return jsonify({'success': False, 'error': '一次最多查询 200 个客户', 'code': 'TOO_MANY_CUSTOMERS'}), 400
    include_completed = _api_bool(data.get('include_completed'), False)
    conn = get_db()
    try:
        groups, missing = _load_exact_customer_orders(conn, customers, include_completed=include_completed)
    finally:
        conn.close()
    return jsonify({
        'success': True,
        'customer_count': len(groups),
        'order_count': sum(g['order_count'] for g in groups),
        'customers': groups,
        'not_found': missing,
    })


@tracking_bp.route('/api/customer-reports/by-customers', methods=['POST'])
@api_login_required
def api_customer_reports_by_customers():
    """Create one independent report job per exact customer name for PyQt batch sending."""
    data = request.get_json(silent=True) or {}
    customers = data.get('customers') or []
    if not isinstance(customers, list) or not customers:
        return jsonify({'success': False, 'error': 'customers 必须是非空数组', 'code': 'INVALID_CUSTOMERS'}), 400
    if len(customers) > 100:
        return jsonify({'success': False, 'error': '一次最多生成 100 个客户报告', 'code': 'TOO_MANY_CUSTOMERS'}), 400

    report_format = str(data.get('format') or 'pdf').strip().lower()
    language = str(data.get('language') or 'es').strip().lower()
    image_source = str(data.get('image_source') or 'both').strip().lower()
    image_count = str(data.get('image_count') or 'all').strip().lower()
    image_order = str(data.get('image_order') or 'order_first').strip().lower()
    pdf_attachment_mode = str(data.get('pdf_attachment_mode') or 'pages').strip().lower()
    include_completed = _api_bool(data.get('include_completed'), False)
    try:
        validate_customer_report_options(report_format, language, image_source, image_count, image_order, pdf_attachment_mode)
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc), 'code': 'INVALID_REPORT_OPTIONS'}), 400

    access_all, user_id = _customer_report_access_context()
    conn = get_db()
    try:
        resolved, missing = _resolve_accessible_customer_names(conn, customers, access_all, user_id)
        jobs = []
        empty = []
        queue_rejected = []
        for requested_name, db_name in resolved.items():
            items = _customer_report_items_for_name(
                conn, db_name, access_all, user_id, include_completed=include_completed
            )
            entries = load_report_entries(conn, items, access_all, user_id) if items else []
            if not entries:
                empty.append(requested_name)
                continue
            try:
                job, deduplicated = _enqueue_customer_report(
                    entries, report_format, language, image_source, image_count, image_order, pdf_attachment_mode, user_id
                )
            except CustomerReportQueueError as exc:
                queue_rejected.append({
                    'requested_name': requested_name,
                    'customer_name': db_name,
                    'error': str(exc),
                    'code': exc.code,
                })
                # Keep scanning the batch so the client receives an explicit rejection
                # for every customer it asked for. No additional job is submitted while
                # the limit remains full.
                continue
            jobs.append({
                'requested_name': requested_name,
                'customer_name': db_name,
                'deduplicated': deduplicated,
                'job': _customer_report_job_snapshot(job),
            })
    finally:
        conn.close()

    return jsonify({
        'success': True,
        'job_count': len(jobs),
        'jobs': jobs,
        'not_found': missing,
        'no_current_orders': empty,
        'queue_rejected': queue_rejected,
        'queue_limits': {
            'workers': int(CUSTOMER_REPORT_WORKERS),
            'max_queued': int(CUSTOMER_REPORT_MAX_QUEUED),
            'max_active_per_user': int(CUSTOMER_REPORT_MAX_ACTIVE_PER_USER),
        },
    }), 202 if jobs else (429 if queue_rejected else 200)


# ==================== 客戶訂單報告（PDF / Word / Excel）====================

def _parse_customer_report_request():
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []
    if not isinstance(items, list) or not items:
        raise ValueError('目前筛选结果没有可导出的订单')
    if len(items) > 5000:
        raise ValueError('一次导出的订单数量过多')
    report_format = str(data.get('format') or 'pdf').strip().lower()
    language = str(data.get('language') or 'zh_cn').strip().lower()
    image_source = str(data.get('image_source') or 'both').strip().lower()
    image_count = str(data.get('image_count') or 'representative').strip().lower()
    image_order = str(data.get('image_order') or 'order_first').strip().lower()
    pdf_attachment_mode = str(data.get('pdf_attachment_mode') or 'pages').strip().lower()
    validate_customer_report_options(report_format, language, image_source, image_count, image_order, pdf_attachment_mode)
    return data, items, report_format, language, image_source, image_count, image_order, pdf_attachment_mode


_CUSTOMER_REPORT_MAX_CUSTOMERS_PER_JOB = 3

def _customer_report_customer_names(entries):
    names = []
    seen = set()
    for entry in entries or []:
        name = ' '.join(str(entry.get('customer_name') or '').strip().split())
        if not name:
            continue
        key = name.upper()
        if key in seen:
            continue
        seen.add(key)
        names.append(name)
    return names

def _validate_customer_report_customer_limit(entries):
    names = _customer_report_customer_names(entries)
    if len(names) > _CUSTOMER_REPORT_MAX_CUSTOMERS_PER_JOB:
        raise ValueError(
            f'一次最多只能生成 {_CUSTOMER_REPORT_MAX_CUSTOMERS_PER_JOB} 个客户的订单报告；'
            f'目前包含 {len(names)} 个客户，请先缩小筛选范围'
        )
    return names


def _customer_report_access_context():
    ctx = get_current_user_context()
    user_id = ctx.get('id')
    role = ctx.get('role', 'viewer')
    filter_info = get_filtered_resources('workflow', role, user_id)
    access_all = bool(is_admin() or filter_info.get('rule') == 'all')
    return access_all, user_id



# ==================== 手机图片墙 / PDF 附件逐页预览 ====================
# PDF 原文件不修改；这里只按需读取页数并将指定页渲染成 JPEG。
# 内部手机页面和办公室 Guest 页面共用这套解析逻辑，但权限仍由各自路由单独检查。
_PDF_PREVIEW_DPI = 118
_PDF_PREVIEW_MAX_EDGE = 1500
_PDF_PREVIEW_JPEG_QUALITY = 84
_PDF_CARD_PREVIEW_DPI = 82
_PDF_CARD_PREVIEW_MAX_EDGE = 920
_PDF_CARD_PREVIEW_JPEG_QUALITY = 76
_IMAGE_CARD_PREVIEW_MAX_EDGE = 920
_IMAGE_CARD_PREVIEW_JPEG_QUALITY = 78


def _media_is_pdf(name='', file_type='', mime_type=''):
    name = str(name or '').lower()
    file_type = str(file_type or '').lower().split(';', 1)[0].strip()
    mime_type = str(mime_type or '').lower().split(';', 1)[0].strip()
    return (
        mime_type == 'application/pdf' or file_type == 'application/pdf' or
        file_type == 'pdf' or name.endswith('.pdf')
    )


def _local_guest_image_file(name='', file_type='', mime_type=''):
    """Return True only for image attachments (never PDF).

    This helper is shared by the authenticated mobile preview routes and the
    temporary guest viewer.  It intentionally uses only lightweight metadata /
    extension checks; Pillow performs the final decode validation when a preview
    is actually generated.
    """
    name = str(name or '').strip().lower()
    file_type = str(file_type or '').strip().lower().split(';', 1)[0]
    mime_type = str(mime_type or '').strip().lower().split(';', 1)[0]
    if _media_is_pdf(name, file_type, mime_type):
        return False
    if mime_type.startswith('image/') or file_type.startswith('image/'):
        return True
    if file_type in {'jpg','jpeg','png','webp','bmp','gif','tif','tiff','heic','heif'}:
        return True
    ext = os.path.splitext(name)[1].lower()
    return ext in {'.jpg','.jpeg','.png','.webp','.bmp','.gif','.tif','.tiff','.heic','.heif'}


def _resolve_order_file_full_path(info):
    """Mirror the legacy order-file download path rules without changing DB rows."""
    from .config import UPLOAD_FOLDER
    info = dict(info or {})
    stored = info.get('stored_filename') or os.path.basename(info.get('file_path') or '')
    path_dir = info.get('file_path') or ''
    if '/' in str(stored) or '\\' in str(stored):
        path_dir = os.path.dirname(info.get('file_path') or '')
        stored = os.path.basename(info.get('file_path') or '')
    elif stored and os.path.basename(str(path_dir)) == str(stored):
        path_dir = os.path.dirname(str(path_dir))
    return os.path.join(UPLOAD_FOLDER, str(path_dir), str(stored))


def _resolve_workflow_file_full_path(info):
    """Mirror the workflow-file download compatibility rules for old directory rows."""
    from .config import UPLOAD_FOLDER
    info = dict(info or {})
    original_name = info.get('file_name') or os.path.basename(info.get('file_path') or '')
    file_path = str(info.get('file_path') or '')
    full_path = os.path.join(UPLOAD_FOLDER, file_path)
    if os.path.isfile(full_path):
        return full_path

    candidate_dir = full_path if os.path.isdir(full_path) else os.path.join(UPLOAD_FOLDER, file_path.rstrip('/\\'))
    if os.path.isdir(candidate_dir):
        ext = os.path.splitext(str(original_name))[1].lower()
        try:
            candidates = [
                f for f in os.listdir(candidate_dir)
                if os.path.isfile(os.path.join(candidate_dir, f)) and (not ext or f.lower().endswith(ext))
            ]
        except Exception:
            candidates = []
        if candidates:
            candidates.sort(key=lambda f: os.path.getmtime(os.path.join(candidate_dir, f)), reverse=True)
            return os.path.join(candidate_dir, candidates[0])
    return full_path


def _file_cache_stamp(path):
    """Cheap source-version stamp used only to validate reproducible preview cache files.

    Business/file lists are NEVER read from this cache. Every page visit still queries the
    latest SQLite attachment rows first.  The cache is considered only after the source file
    is confirmed to exist, and a replaced file gets a different stamp.
    """
    try:
        stat = os.stat(path)
        mtime_ns = int(getattr(stat, 'st_mtime_ns', int(stat.st_mtime * 1_000_000_000)))
        ctime_ns = int(getattr(stat, 'st_ctime_ns', int(stat.st_ctime * 1_000_000_000)))
        return mtime_ns, ctime_ns, int(stat.st_size)
    except Exception:
        return 0, 0, 0


def _file_version_token(path):
    """Short query-string version so browser cache can never hide a replaced source file."""
    mtime_ns, ctime_ns, file_size = _file_cache_stamp(path)
    raw = f'{mtime_ns}:{ctime_ns}:{file_size}'.encode('utf-8')
    return hashlib.sha256(raw).hexdigest()[:16]


def _media_preview_cache_path(kind, path, extra='', ext='.jpg'):
    mtime_ns, ctime_ns, file_size = _file_cache_stamp(path)
    source = os.path.abspath(str(path or ''))
    raw = f'{kind}|{source}|{mtime_ns}|{ctime_ns}|{file_size}|{extra}'.encode('utf-8')
    digest = hashlib.sha256(raw).hexdigest()
    folder = os.path.join(MEDIA_PREVIEW_CACHE_DIR, digest[:2])
    return os.path.join(folder, digest + ext)


def _media_preview_atomic_write(cache_path, payload):
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    tmp_path = cache_path + f'.tmp-{os.getpid()}-{threading.get_ident()}-{secrets.token_hex(4)}'
    try:
        with open(tmp_path, 'wb') as handle:
            handle.write(payload)
        os.replace(tmp_path, cache_path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
    _media_preview_schedule_cleanup()


def _media_preview_read(cache_path):
    if not os.path.isfile(cache_path):
        return None
    try:
        with open(cache_path, 'rb') as handle:
            payload = handle.read()
        # mtime of the generated cache file doubles as LRU last-access time.
        try:
            os.utime(cache_path, None)
        except Exception:
            pass
        return payload
    except Exception:
        return None


def _media_preview_cleanup_cache(force=False):
    """Bound preview cache by age and total bytes; originals are never touched."""
    global _media_preview_cleanup_last
    if not _media_preview_cleanup_lock.acquire(blocking=False):
        return
    try:
        now = time.time()
        if (not force and _media_preview_cleanup_last and
                now - _media_preview_cleanup_last < max(60, int(MEDIA_PREVIEW_CACHE_CLEANUP_INTERVAL_SECONDS))):
            return
        _media_preview_cleanup_last = now
        root = str(MEDIA_PREVIEW_CACHE_DIR or '')
        if not root or not os.path.isdir(root):
            return
        retention_seconds = max(1, int(MEDIA_PREVIEW_CACHE_RETENTION_DAYS)) * 86400
        max_bytes = max(64 * 1024 * 1024, int(MEDIA_PREVIEW_CACHE_MAX_BYTES))
        entries = []
        total = 0
        for dirpath, _, filenames in os.walk(root):
            for filename in filenames:
                if '.tmp-' in filename:
                    continue
                full_path = os.path.join(dirpath, filename)
                try:
                    stat = os.stat(full_path)
                except OSError:
                    continue
                age = now - float(stat.st_mtime)
                if age > retention_seconds:
                    try:
                        os.remove(full_path)
                    except OSError:
                        pass
                    continue
                size = int(stat.st_size)
                total += size
                entries.append((float(stat.st_mtime), size, full_path))

        if total > max_bytes:
            # Leave a little headroom so every new preview does not trigger another cleanup.
            target = int(max_bytes * 0.90)
            for _, size, full_path in sorted(entries, key=lambda item: item[0]):
                try:
                    os.remove(full_path)
                    total -= size
                except OSError:
                    pass
                if total <= target:
                    break
        # Empty shard folders are harmless; remove them opportunistically.
        for dirpath, dirnames, filenames in os.walk(root, topdown=False):
            if dirpath == root:
                continue
            try:
                if not os.listdir(dirpath):
                    os.rmdir(dirpath)
            except OSError:
                pass
    finally:
        _media_preview_cleanup_lock.release()


def _media_preview_schedule_cleanup():
    now = time.time()
    if (_media_preview_cleanup_last and
            now - _media_preview_cleanup_last < max(60, int(MEDIA_PREVIEW_CACHE_CLEANUP_INTERVAL_SECONDS))):
        return
    _media_preview_submit('cache-cleanup', _media_preview_cleanup_cache, False)


def _media_preview_submit(job_key, func, *args):
    """Deduplicate best-effort background preview work."""
    key = str(job_key or '')
    if not key:
        return
    with _media_preview_pending_lock:
        if key in _media_preview_pending:
            return
        _media_preview_pending.add(key)

    def runner():
        try:
            func(*args)
        except Exception as exc:
            print(f'[WARN] media preview background job failed: {key}: {exc}')
        finally:
            with _media_preview_pending_lock:
                _media_preview_pending.discard(key)

    try:
        _media_preview_executor.submit(runner)
    except Exception:
        with _media_preview_pending_lock:
            _media_preview_pending.discard(key)


def _pdf_preview_page_count(path):
    if not path or not os.path.isfile(path):
        return 0
    count_path = _media_preview_cache_path('pdf-count', path, ext='.count')
    cached = _media_preview_read(count_path)
    if cached is not None:
        try:
            return max(0, int(cached.decode('ascii', errors='ignore').strip()))
        except Exception:
            pass
    try:
        import fitz  # PyMuPDF
        with fitz.open(path) as doc:
            count = max(0, int(len(doc)))
        _media_preview_atomic_write(count_path, str(count).encode('ascii'))
        return count
    except Exception as exc:
        print(f'[WARN] PDF preview page count failed: {path}: {exc}')
        return 0


def _pdf_preview_render_page(path, page_number, compact=False):
    """Return one 1-based PDF page as JPEG bytes from a bounded, source-versioned disk cache."""
    if not path or not os.path.isfile(path):
        raise FileNotFoundError(path or '')
    page_number = int(page_number)
    compact = bool(compact)
    kind = 'pdf-card' if compact else 'pdf-detail'
    cache_path = _media_preview_cache_path(kind, path, extra=f'page={page_number}')
    cached = _media_preview_read(cache_path)
    total = _pdf_preview_page_count(path)
    if cached is not None:
        if page_number < 1 or page_number > total:
            raise ValueError('page out of range')
        return cached, total

    try:
        import fitz  # PyMuPDF
    except ImportError as exc:
        raise RuntimeError('PDF 预览需要 PyMuPDF，请执行: python -m pip install PyMuPDF') from exc
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError('PDF 预览需要 Pillow') from exc

    dpi = _PDF_CARD_PREVIEW_DPI if compact else _PDF_PREVIEW_DPI
    max_edge = _PDF_CARD_PREVIEW_MAX_EDGE if compact else _PDF_PREVIEW_MAX_EDGE
    quality = _PDF_CARD_PREVIEW_JPEG_QUALITY if compact else _PDF_PREVIEW_JPEG_QUALITY
    with fitz.open(path) as doc:
        if page_number < 1 or page_number > len(doc):
            raise ValueError('page out of range')
        total = len(doc)
        page = doc.load_page(page_number - 1)
        pix = page.get_pixmap(dpi=dpi, colorspace=fitz.csRGB, alpha=False)
        image = Image.frombytes('RGB', (pix.width, pix.height), pix.samples)
        image.thumbnail((max_edge, max_edge), Image.Resampling.LANCZOS)
        output = BytesIO()
        image.save(output, format='JPEG', quality=quality, optimize=False)
        payload = output.getvalue()
    _media_preview_atomic_write(cache_path, payload)
    return payload, total


def _image_card_preview(path):
    if not path or not os.path.isfile(path):
        raise FileNotFoundError(path or '')
    cache_path = _media_preview_cache_path('image-card', path)
    cached = _media_preview_read(cache_path)
    if cached is not None:
        return cached
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:
        raise RuntimeError('图片预览需要 Pillow') from exc
    with Image.open(path) as source:
        image = ImageOps.exif_transpose(source)
        if image.mode not in {'RGB', 'L'}:
            if 'A' in image.getbands():
                base = Image.new('RGB', image.size, 'white')
                base.paste(image, mask=image.getchannel('A'))
                image = base
            else:
                image = image.convert('RGB')
        elif image.mode != 'RGB':
            image = image.convert('RGB')
        image.thumbnail((_IMAGE_CARD_PREVIEW_MAX_EDGE, _IMAGE_CARD_PREVIEW_MAX_EDGE), Image.Resampling.LANCZOS)
        output = BytesIO()
        image.save(output, format='JPEG', quality=_IMAGE_CARD_PREVIEW_JPEG_QUALITY, optimize=False)
        payload = output.getvalue()
    _media_preview_atomic_write(cache_path, payload)
    return payload


def _warm_pdf_preview(path, page_number):
    _pdf_preview_render_page(path, page_number, compact=True)


def _warm_image_preview(path):
    _image_card_preview(path)


def _schedule_visual_preview_batch(specs):
    """Prepare visible media progressively: all first previews first, then remaining PDF pages.

    specs items: {'type':'image','path':...} or {'type':'pdf','path':...,'pages':N}.
    This never supplies the attachment list; it only warms reproducible previews after a fresh scan.
    """
    unique = []
    seen = set()
    for spec in specs or []:
        path = str((spec or {}).get('path') or '')
        if not path or not os.path.isfile(path):
            continue
        typ = str((spec or {}).get('type') or '')
        pages = max(0, int((spec or {}).get('pages') or 0))
        version = _file_version_token(path)
        marker = (typ, os.path.abspath(path), version)
        if marker in seen:
            continue
        seen.add(marker)
        unique.append((typ, path, pages, version))

    # Phase 1: make every card capable of showing something as soon as possible.
    for typ, path, pages, version in unique:
        if typ == 'image':
            _media_preview_submit(f'img:{path}:{version}', _warm_image_preview, path)
        elif typ == 'pdf' and pages > 0:
            _media_preview_submit(f'pdf:{path}:{version}:1', _warm_pdf_preview, path, 1)

    # Phase 2: continue quietly until every PDF page has a compact preview.
    for typ, path, pages, version in unique:
        if typ != 'pdf' or pages <= 1:
            continue
        for page_number in range(2, pages + 1):
            _media_preview_submit(
                f'pdf:{path}:{version}:{page_number}', _warm_pdf_preview, path, page_number
            )


def _image_preview_response(path, filename='preview.jpg', no_store=False, cache_max_age=None):
    from flask import send_file
    payload = _image_card_preview(path)
    response = send_file(
        BytesIO(payload), mimetype='image/jpeg', as_attachment=False,
        download_name=str(filename or 'preview.jpg'), conditional=False,
        max_age=0,
    )
    if no_store:
        return _local_guest_no_store(response)
    if cache_max_age is not None:
        response.headers['Cache-Control'] = f'private, max-age={max(0, int(cache_max_age))}'
    else:
        response.headers['Cache-Control'] = 'private, max-age=3600'
    return response


def _pdf_page_response(payload, filename='pdf-page.jpg', no_store=False, cache_max_age=None):
    from flask import send_file
    response = send_file(
        BytesIO(payload), mimetype='image/jpeg', as_attachment=False,
        download_name=str(filename or 'pdf-page.jpg'), conditional=False,
        max_age=0,
    )
    if no_store:
        return _local_guest_no_store(response)
    if cache_max_age is not None:
        response.headers['Cache-Control'] = f'private, max-age={max(0, int(cache_max_age))}'
    else:
        response.headers['Cache-Control'] = 'private, max-age=3600'
    return response


# ==================== 本地办公室临时客户查看链接 ====================
# 只在本地/LAN 使用；匿名、短时效、只读、严格限定单一客户。
_LOCAL_GUEST_ALLOWED_MINUTES = {30, 60, 240, 1440}
_LOCAL_GUEST_MAX_IMAGES_PER_CARD = 6


def _local_guest_token_hash(token):
    return hashlib.sha256(str(token or '').encode('utf-8')).hexdigest()


def _local_guest_feature_enabled():
    return bool(current_app.config.get('TRACKING_LOCAL_GUEST_SHARING_ENABLED', LOCAL_GUEST_SHARING_ENABLED))


def _local_guest_permanent_enabled():
    return bool(current_app.config.get('TRACKING_LOCAL_GUEST_PERMANENT_ENABLED', LOCAL_GUEST_PERMANENT_ENABLED))


def _local_guest_password_session_key(token):
    return 'guest_unlock_' + _local_guest_token_hash(token)[:24]


def _local_guest_password_authorized(link, token):
    if not str((link or {}).get('password_hash') or '').strip():
        return True
    return bool(session.get(_local_guest_password_session_key(token)))


def _local_guest_client_allowed():
    """Cloud/Render never serves office guest pages; local clients must use a private address."""
    if not _local_guest_feature_enabled() or cloud_mode_enabled():
        return False
    raw = str(request.remote_addr or '').strip()
    if not raw:
        return False
    try:
        addr = ipaddress.ip_address(raw.split('%', 1)[0])
        if getattr(addr, 'ipv4_mapped', None):
            addr = addr.ipv4_mapped
        return bool(addr.is_private or addr.is_loopback or addr.is_link_local)
    except ValueError:
        return raw == 'localhost'


def _local_guest_get_link(conn, token, touch=False, allow_locked=False):
    ensure_local_guest_link_tables(conn)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM local_guest_links WHERE token_hash = ?', (_local_guest_token_hash(token),))
    row = cursor.fetchone()
    if not row:
        return None, 'not_found'
    item = dict(row)
    now = int(time.time())
    if item.get('revoked_at_epoch'):
        return item, 'revoked'
    is_permanent = bool(int(item.get('is_permanent') or 0))
    if not is_permanent and int(item.get('expires_at_epoch') or 0) <= now:
        return item, 'expired'
    if str(item.get('password_hash') or '').strip() and not _local_guest_password_authorized(item, token):
        if allow_locked:
            return item, 'locked'
        return item, 'locked'
    if touch:
        cursor.execute(
            'UPDATE local_guest_links SET last_accessed_at_epoch = ?, access_count = COALESCE(access_count, 0) + 1 WHERE id = ?',
            (now, item['id'])
        )
        conn.commit()
    return item, 'active'


def _local_guest_no_store(response):
    response.headers['Cache-Control'] = 'no-store, private, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Referrer-Policy'] = 'no-referrer'
    response.headers['X-Robots-Tag'] = 'noindex, nofollow, noarchive'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    return response


def _local_guest_public_url(token):
    """Build a LAN-usable guest URL even when the admin opened Flask via localhost.

    An explicit TRACKING_LOCAL_GUEST_BASE_URL wins. Otherwise a non-loopback request
    host is preserved; localhost/127.0.0.1 is replaced with the machine's primary LAN
    IPv4 when it can be detected. This changes only the address shown in desktop share
    management and does not alter guest authorization rules.
    """
    path = url_for('tracking_bp.local_guest_customer', token=token)
    explicit = str(current_app.config.get('TRACKING_LOCAL_GUEST_BASE_URL') or os.environ.get('TRACKING_LOCAL_GUEST_BASE_URL') or '').strip().rstrip('/')
    if explicit:
        return explicit + path
    host = str(request.host or '').strip()
    host_only = host
    port = ''
    if host.startswith('['):
        # IPv6 literal; keep the request host as-is rather than guessing another interface.
        return f'{request.scheme}://{host}{path}'
    if ':' in host:
        host_only, port = host.rsplit(':', 1)
        if not port.isdigit():
            host_only, port = host, ''
    try:
        addr = ipaddress.ip_address(host_only)
        is_loopback = addr.is_loopback
    except ValueError:
        is_loopback = host_only.lower() in {'localhost', 'localhost.localdomain'}
    if not is_loopback:
        return f'{request.scheme}://{host}{path}'
    lan_ip = ''
    try:
        probe = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        probe.settimeout(0.25)
        probe.connect(('10.255.255.255', 1))
        lan_ip = str(probe.getsockname()[0] or '').strip()
        probe.close()
    except Exception:
        try:
            lan_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            lan_ip = ''
    try:
        if not lan_ip or ipaddress.ip_address(lan_ip).is_loopback:
            return f'{request.scheme}://{host}{path}'
    except ValueError:
        return f'{request.scheme}://{host}{path}'
    port_part = f':{port}' if port else ''
    return f'{request.scheme}://{lan_ip}{port_part}{path}'


def _local_guest_qr_data_uri(value):
    """Return a small PNG QR as a data URI for the office guest-link dialog.

    Keep QR generation local and stateless. Link validation still uses the SHA-256
    token hash; the admin share manager may additionally retain the generated share URL
    so an authorized administrator can reopen/copy active links later. If qrcode is
    unavailable, the link itself still succeeds and copy/open actions remain usable.
    """
    raw = str(value or '').strip()
    if not raw:
        return ''
    try:
        import qrcode
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=6,
            border=3,
        )
        qr.add_data(raw)
        qr.make(fit=True)
        image = qr.make_image(fill_color='black', back_color='white')
        output = BytesIO()
        image.save(output, format='PNG')
        return 'data:image/png;base64,' + base64.b64encode(output.getvalue()).decode('ascii')
    except Exception as exc:
        print(f'[WARN] local guest QR generation failed: {exc}')
        return ''


def _local_guest_error_page(state, status_code=404):
    """Render the same safe Spanish error page for every guest failure state."""
    state = str(state or 'not_found').strip().lower()
    messages = {
        'expired': ('Acceso caducado', 'Este enlace temporal ya caducó. Solicite un nuevo acceso.'),
        'revoked': ('Acceso revocado', 'Este enlace temporal ya no está disponible.'),
        'network': ('Solo red local', 'Este acceso solo funciona dentro de la red local autorizada.'),
        'cloud': ('No disponible', 'Esta vista temporal no está disponible en el modo nube.'),
        'locked': ('Acceso protegido', 'Este enlace requiere contraseña.'),
        'not_found': ('Acceso no disponible', 'El enlace no existe o el contenido ya no está disponible.'),
    }
    title, subtitle = messages.get(state, messages['not_found'])
    response = make_response(render_template(
        'tracking/guest_expired.html', title=title, subtitle=subtitle, error_state=state
    ), int(status_code or 404))
    return _local_guest_no_store(response)


def _local_guest_media_cache_seconds(link):
    """Short private browser cache for already-authorized visual bytes only.

    Attachment/order manifests are never cached. The URL also carries a source-version
    token, so replacing a file cannot be hidden by an old browser preview.
    """
    try:
        remaining = max(0, int((link or {}).get('expires_at_epoch') or 0) - int(time.time()))
    except Exception:
        remaining = 0
    # Keep the speed benefit without letting a long 4-hour token pin media forever.
    return max(0, min(900, remaining))


def _local_guest_cache_media_response(response, link):
    max_age = _local_guest_media_cache_seconds(link)
    response.headers['Cache-Control'] = f'private, max-age={max_age}' if max_age > 0 else 'no-store, private, max-age=0'
    response.headers['Referrer-Policy'] = 'no-referrer'
    response.headers['X-Robots-Tag'] = 'noindex, nofollow, noarchive'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    return response


def _local_guest_media_for_order(conn, token, order_number, workflow_number='', limit=None,
                                 prewarm_specs=None, show_pdf_pages=True, build_urls=True):
    """Fresh-scan visual attachments for one already-authorized Guest order.

    The attachment rows and source-file existence are checked every call. Only generated
    image/PDF-page previews are reusable cache artifacts; the media manifest itself is
    never taken from cache.
    """
    order_number = str(order_number or '').strip()
    workflow_number = str(workflow_number or '').strip()
    prewarm_specs = prewarm_specs if isinstance(prewarm_specs, list) else []
    show_pdf_pages = bool(show_pdf_pages)
    try:
        max_items = None if limit is None else max(0, int(limit))
    except Exception:
        max_items = None

    sources = []
    cursor = conn.cursor()
    if order_number:
        cursor.execute('''
            SELECT * FROM order_files
            WHERE order_number = ?
            ORDER BY uploaded_at DESC, id DESC
        ''', (order_number,))
        for row in cursor.fetchall():
            info = dict(row)
            name = (info.get('original_filename') or info.get('file_name') or
                    info.get('stored_filename') or os.path.basename(str(info.get('file_path') or '')))
            path = _resolve_order_file_full_path(info)
            if os.path.isfile(path):
                sources.append(('order', int(info.get('id') or 0), str(name or ''), info, path))

    if workflow_number:
        cursor.execute('''
            SELECT * FROM workflow_files
            WHERE workflow_number = ? AND COALESCE(is_deleted, 0) = 0
            ORDER BY uploaded_at DESC, id DESC
        ''', (workflow_number,))
        for row in cursor.fetchall():
            info = dict(row)
            name = info.get('file_name') or os.path.basename(str(info.get('file_path') or ''))
            path = _resolve_workflow_file_full_path(info)
            if os.path.isfile(path):
                sources.append(('workflow', int(info.get('id') or 0), str(name or ''), info, path))

    result = []
    seen = set()
    for kind, file_id, name, info, path in sources:
        if not file_id:
            continue
        absolute = os.path.normcase(os.path.abspath(path))
        marker = (kind, file_id, absolute)
        if marker in seen:
            continue
        seen.add(marker)
        version = _file_version_token(path)
        file_type = info.get('file_type') or ''
        mime_type = info.get('mime_type') or file_type

        if _media_is_pdf(name, file_type, mime_type):
            if not show_pdf_pages:
                continue
            page_count = _pdf_preview_page_count(path)
            if page_count <= 0:
                continue
            prewarm_specs.append({'type': 'pdf', 'path': path, 'pages': page_count})
            for page_number in range(1, page_count + 1):
                if max_items is not None and len(result) >= max_items:
                    return result
                if build_urls:
                    url = url_for(
                        'tracking_bp.local_guest_pdf_page', token=token, kind=kind,
                        file_id=file_id, page_number=page_number, v=version
                    )
                    preview_url = url_for(
                        'tracking_bp.local_guest_pdf_page', token=token, kind=kind,
                        file_id=file_id, page_number=page_number, preview=1, v=version
                    )
                else:
                    url = preview_url = ''
                result.append({
                    'name': name,
                    'media_type': 'pdf_page',
                    'pdf_page_number': page_number,
                    'pdf_page_count': page_count,
                    'url': url,
                    'preview_url': preview_url,
                    'media_version': version,
                    '_source_path': path,
                    '_source_kind': kind,
                    '_source_file_id': file_id,
                })
            continue

        if not _local_guest_image_file(name, file_type, mime_type):
            continue
        prewarm_specs.append({'type': 'image', 'path': path})
        if max_items is not None and len(result) >= max_items:
            return result
        if build_urls:
            url = url_for(
                'tracking_bp.local_guest_media', token=token, kind=kind, file_id=file_id, v=version
            )
            preview_url = url_for(
                'tracking_bp.local_guest_media', token=token, kind=kind, file_id=file_id, preview=1, v=version
            )
        else:
            url = preview_url = ''
        result.append({
            'name': name,
            'media_type': 'image',
            'pdf_page_number': 0,
            'pdf_page_count': 0,
            'url': url,
            'preview_url': preview_url,
            'media_version': version,
            '_source_path': path,
            '_source_kind': kind,
            '_source_file_id': file_id,
        })
    return result


_LOCAL_GUEST_HISTORY_SCOPES = {'current', '6m', '12m', 'all'}


def _normalize_guest_history_scope(value):
    scope = str(value or 'current').strip().lower()
    aliases = {
        'default': 'current', 'same': 'current', '3m': 'current',
        '6': '6m', '6months': '6m', 'halfyear': '6m',
        '12': '12m', '1y': '12m', 'year': '12m',
        'all_history': 'all', 'history': 'all',
    }
    scope = aliases.get(scope, scope)
    return scope if scope in _LOCAL_GUEST_HISTORY_SCOPES else 'current'


def _guest_history_scope_months(scope):
    scope = _normalize_guest_history_scope(scope)
    if scope == '6m':
        return 6
    if scope == '12m':
        return 12
    if scope == 'all':
        return None
    return 3


def _guest_history_scope_from_link(link):
    return _normalize_guest_history_scope((link or {}).get('history_scope') or 'current')


def _guest_include_cancelled_from_link(link):
    try:
        return bool(int((link or {}).get('include_cancelled') or 0))
    except Exception:
        return False


def _guest_workflow_scope_sql(link=None, workflow_alias='w', history_scope=None, include_cancelled=None):
    scope = _normalize_guest_history_scope(history_scope if history_scope is not None else _guest_history_scope_from_link(link))
    include_cancelled = _guest_include_cancelled_from_link(link) if include_cancelled is None else bool(include_cancelled)
    completed = STATUS_KEYS['COMPLETED']
    cancelled = STATUS_KEYS['CANCELLED']
    months = _guest_history_scope_months(scope)
    if months is None:
        if include_cancelled:
            return '(1 = 1)', []
        return f'({workflow_alias}.current_status <> ?)', [cancelled]

    interval = f'-{int(months)} months'
    parts = [
        f'{workflow_alias}.current_status NOT IN (?, ?)',
        f"({workflow_alias}.current_status = ? AND {workflow_alias}.status_updated_at >= date('now', ?))",
    ]
    params = [completed, cancelled, completed, interval]
    if include_cancelled:
        parts.append(f"({workflow_alias}.current_status = ? AND {workflow_alias}.status_updated_at >= date('now', ?))")
        params.extend([cancelled, interval])
    return '(' + ' OR '.join(parts) + ')', params


def _guest_order_scope_sql(link=None, order_alias='o', history_scope=None, include_cancelled=None):
    workflow_sql, params = _guest_workflow_scope_sql(
        link=link, workflow_alias='wa', history_scope=history_scope, include_cancelled=include_cancelled
    )
    sql = (
        f'(NOT EXISTS (SELECT 1 FROM workflows wx WHERE wx.order_number = {order_alias}.order_number) '
        f'OR EXISTS (SELECT 1 FROM workflows wa WHERE wa.order_number = {order_alias}.order_number AND {workflow_sql}))'
    )
    return sql, params


def _guest_scope_label(scope, language='zh_cn'):
    scope = _normalize_guest_history_scope(scope)
    if language == 'es':
        return {'current':'Vista actual', '6m':'Últimos 6 meses', '12m':'Último año', 'all':'Todo el historial'}.get(scope, 'Vista actual')
    return {'current':'目前', '6m':'近半年', '12m':'近一年', 'all':'全部历史'}.get(scope, '目前')


def _local_guest_visible_items_for_customer(conn, customer_name, link=None, history_scope=None, include_cancelled=None):
    cursor = conn.cursor()
    workflow_sql, workflow_params = _guest_workflow_scope_sql(
        link=link, workflow_alias='w', history_scope=history_scope, include_cancelled=include_cancelled
    )
    cursor.execute(f"""
        SELECT w.workflow_number, w.order_number
        FROM workflows w
        INNER JOIN orders o ON o.order_number = w.order_number
        WHERE o.status = 'ACTIVE' AND o.customer_name = ?
          AND {workflow_sql}
        ORDER BY o.order_date DESC, w.created_at DESC, w.workflow_number DESC
    """, [customer_name, *workflow_params])
    items = [
        {'workflow_number': str(row['workflow_number'] or ''), 'order_number': str(row['order_number'] or '')}
        for row in cursor.fetchall()
        if str(row['workflow_number'] or '').strip()
    ]
    cursor.execute("""
        SELECT o.order_number
        FROM orders o
        WHERE o.status = 'ACTIVE' AND o.customer_name = ?
          AND NOT EXISTS (SELECT 1 FROM workflows w WHERE w.order_number = o.order_number)
        ORDER BY o.order_date DESC, o.order_number DESC
    """, (customer_name,))
    items.extend(
        {'workflow_number': '', 'order_number': str(row['order_number'] or '')}
        for row in cursor.fetchall()
        if str(row['order_number'] or '').strip()
    )
    return items


def _local_guest_load_orders(conn, token, customer_name, show_pdf_pages=True, build_urls=True, link=None, history_scope=None, include_cancelled=None):
    cursor = conn.cursor()
    workflow_sql, workflow_params = _guest_workflow_scope_sql(
        link=link, workflow_alias='w', history_scope=history_scope, include_cancelled=include_cancelled
    )
    cursor.execute(f"""
        SELECT w.workflow_number, w.order_number, w.product_name, w.product_code,
               w.quantity, w.factory, w.production_type, w.expected_delivery_date,
               w.current_status, w.status_updated_at, o.order_date
        FROM workflows w
        INNER JOIN orders o ON o.order_number = w.order_number
        WHERE o.status = 'ACTIVE' AND o.customer_name = ?
          AND {workflow_sql}
        ORDER BY o.order_date DESC, w.created_at DESC, w.workflow_number DESC
    """, [customer_name, *workflow_params])
    rows = [dict(row) for row in cursor.fetchall()]
    cursor.execute("""
        SELECT o.order_number, o.order_date FROM orders o
        WHERE o.status = 'ACTIVE' AND o.customer_name = ?
          AND NOT EXISTS (SELECT 1 FROM workflows w WHERE w.order_number = o.order_number)
        ORDER BY o.order_date DESC, o.order_number DESC
    """, (customer_name,))
    for row in cursor.fetchall():
        rows.append({
            'workflow_number': '', 'order_number': row['order_number'], 'order_date': row['order_date'],
            'product_name': '', 'product_code': '', 'quantity': '', 'factory': '', 'production_type': '',
            'expected_delivery_date': '', 'current_status': '', 'status_updated_at': ''
        })

    result = []
    prewarm_specs = []
    for item in rows:
        status_key = str(item.get('current_status') or '')
        item['status_es'] = get_status_label(status_key, 'es') if status_key else 'Pedido'
        item['status_zh'] = get_status_label(status_key, 'zh_cn') if status_key else '订单'
        item['images'] = _local_guest_media_for_order(
            conn, token, str(item.get('order_number') or ''), str(item.get('workflow_number') or ''),
            limit=None, prewarm_specs=prewarm_specs, show_pdf_pages=show_pdf_pages, build_urls=build_urls
        )
        item['detail_url'] = (url_for(
            'tracking_bp.local_guest_order_detail', token=token,
            item_kind='w' if item.get('workflow_number') else 'o',
            item_key=item.get('workflow_number') or item.get('order_number')
        ) if build_urls else '')
        result.append(item)
    if prewarm_specs and build_urls:
        _schedule_visual_preview_batch(prewarm_specs)
    return result


def _local_guest_safe_zip_name(value, fallback='file'):
    """Keep ZIP entry names readable while preventing path traversal / invalid path chars."""
    text = str(value or '').strip()
    forbidden = set('/\\:*?"<>|')
    text = ''.join('_' if (ch in forbidden or ord(ch) < 32) else ch for ch in text)
    text = text.strip(' .')[:120]
    return text or fallback


def _local_guest_collect_pdf_files(conn, customer_name, link=None):
    cursor = conn.cursor()
    customer_name = str(customer_name or '').strip()
    if not customer_name:
        return []

    visible_items = _local_guest_visible_items_for_customer(conn, customer_name, link=link)
    visible_orders = {
        str(item.get('order_number') or '').strip()
        for item in visible_items if str(item.get('order_number') or '').strip()
    }
    visible_workflows = {
        str(item.get('workflow_number') or '').strip()
        for item in visible_items if str(item.get('workflow_number') or '').strip()
    }
    if not visible_orders:
        return []

    pdfs = []
    placeholders = ','.join('?' for _ in visible_orders)
    cursor.execute(f'''
        SELECT f.*, o.customer_name
        FROM order_files f
        INNER JOIN orders o ON o.order_number = f.order_number
        WHERE f.order_number IN ({placeholders})
          AND o.customer_name = ? AND o.status = 'ACTIVE'
        ORDER BY f.order_number, f.uploaded_at, f.id
    ''', (*list(visible_orders), customer_name))
    for row in cursor.fetchall():
        item = dict(row)
        name = (item.get('original_filename') or item.get('stored_filename') or
                os.path.basename(str(item.get('file_path') or '')) or 'document.pdf')
        if not _media_is_pdf(name, item.get('file_type'), item.get('mime_type')):
            continue
        path = _resolve_order_file_full_path(item)
        if not os.path.isfile(path):
            continue
        order_number = _local_guest_safe_zip_name(item.get('order_number'), 'order')
        pdfs.append({
            'path': path,
            'arcname': f'{order_number}/ORDER/{_local_guest_safe_zip_name(name, "document.pdf")}',
        })

    if visible_workflows:
        wf_placeholders = ','.join('?' for _ in visible_workflows)
        cursor.execute(f'''
            SELECT wf.*, w.order_number, o.customer_name
            FROM workflow_files wf
            INNER JOIN workflows w ON w.workflow_number = wf.workflow_number
            INNER JOIN orders o ON o.order_number = w.order_number
            WHERE wf.workflow_number IN ({wf_placeholders})
              AND o.customer_name = ? AND o.status = 'ACTIVE'
              AND COALESCE(wf.is_deleted, 0) = 0
            ORDER BY w.order_number, wf.workflow_number, wf.uploaded_at, wf.id
        ''', (*list(visible_workflows), customer_name))
        for row in cursor.fetchall():
            item = dict(row)
            name = item.get('file_name') or os.path.basename(str(item.get('file_path') or '')) or 'document.pdf'
            if not _media_is_pdf(name, item.get('file_type'), item.get('file_type')):
                continue
            path = _resolve_workflow_file_full_path(item)
            if not os.path.isfile(path):
                continue
            order_number = _local_guest_safe_zip_name(item.get('order_number'), 'order')
            workflow_number = _local_guest_safe_zip_name(item.get('workflow_number'), 'workflow')
            pdfs.append({
                'path': path,
                'arcname': f'{order_number}/{workflow_number}/{_local_guest_safe_zip_name(name, "document.pdf")}',
            })

    used = {}
    for item in pdfs:
        original = item['arcname']
        count = used.get(original, 0)
        used[original] = count + 1
        if count:
            folder, filename = os.path.split(original)
            stem, ext = os.path.splitext(filename)
            item['arcname'] = f'{folder}/{stem}_{count}{ext}' if folder else f'{stem}_{count}{ext}'
    return pdfs


@tracking_bp.route('/api/local-guest-links', methods=['GET'])
@login_required
def api_list_active_local_guest_links():
    """List active/revocable LAN shares. Expired/revoked rows never appear."""
    if not _local_guest_feature_enabled() or cloud_mode_enabled():
        return jsonify({'success': False, 'error': '本机客户分享功能在此部署已关闭'}), 403
    if not is_admin():
        return jsonify({'success': False, 'error': '客户分享仅限 ADMIN 使用'}), 403
    compact = str(request.args.get('compact') or '').strip().lower() in {'1','true','yes','on'}

    conn = get_db()
    try:
        ensure_local_guest_link_tables(conn)
        now = int(time.time())
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, customer_name, created_at_epoch, expires_at_epoch,
                   created_by_id, created_by_name, access_count,
                   show_pdf_pages, allow_report_pdf_download,
                   history_scope, include_cancelled, share_mode, is_permanent,
                   password_kind, share_url, qr_data_uri, last_synced_at_epoch, snapshot_version
            FROM local_guest_links
            WHERE revoked_at_epoch IS NULL
              AND (COALESCE(is_permanent, 0) = 1 OR expires_at_epoch > ?)
            ORDER BY created_at_epoch DESC, id DESC
        """, (now,))
        rows = []
        for row in cursor.fetchall():
            item = dict(row)
            created = int(item.get('created_at_epoch') or 0)
            expires = int(item.get('expires_at_epoch') or 0)
            permanent = bool(int(item.get('is_permanent') or 0))
            item['is_permanent'] = permanent
            item['duration_minutes'] = None if permanent else max(1, int(round(max(0, expires - created) / 60.0)))
            item['remaining_seconds'] = None if permanent else max(0, expires - now)
            item['show_pdf_pages'] = bool(int(item.get('show_pdf_pages') or 0))
            item['allow_report_pdf_download'] = bool(int(item.get('allow_report_pdf_download') or 0))
            item['include_cancelled'] = bool(int(item.get('include_cancelled') or 0))
            item['share_mode'] = str(item.get('share_mode') or 'lan').lower()
            item['password_protected'] = str(item.get('password_kind') or 'none') != 'none'
            share_url = str(item.get('share_url') or '').strip()
            if compact:
                item.pop('share_url', None)
                item.pop('qr_data_uri', None)
                item['has_url'] = bool(share_url)
            elif share_url:
                item['url'] = share_url
                if not str(item.get('qr_data_uri') or '').strip():
                    item['qr_data_uri'] = _local_guest_qr_data_uri(share_url)
            rows.append(item)
    finally:
        conn.close()

    response = jsonify({'success': True, 'data': rows, 'total': len(rows)})
    response.headers['Cache-Control'] = 'no-store, private, max-age=0'
    return response


@tracking_bp.route('/api/local-guest-links', methods=['POST'])
@login_required
def api_create_local_guest_link():
    if not _local_guest_feature_enabled() or cloud_mode_enabled():
        return jsonify({'success': False, 'error': '本机客户分享功能在此部署已关闭'}), 403
    if not is_admin():
        return jsonify({'success': False, 'error': '客户分享仅限 ADMIN 使用'}), 403
    data = request.get_json(silent=True) or {}
    customer_input = ' '.join(str(data.get('customer_name') or '').strip().split())
    is_permanent = bool(data.get('is_permanent') is True or str(data.get('is_permanent','')).lower() in {'1','true','yes','on'})
    if is_permanent and not _local_guest_permanent_enabled():
        return jsonify({'success': False, 'error': '此部署未开放永久客户链接'}), 403
    try:
        duration_minutes = int(data.get('duration_minutes') or 60)
    except (TypeError, ValueError):
        duration_minutes = 60
    if not is_permanent and duration_minutes not in _LOCAL_GUEST_ALLOWED_MINUTES:
        return jsonify({'success': False, 'error': '有效时间只支持 30 分、1 小时、4 小时或 1 天'}), 400

    allow_pdf_download = 1 if (data.get('allow_pdf_download') is True or str(data.get('allow_pdf_download','')).strip().lower() in {'1','true','yes','on'}) else 0
    allow_report_pdf_download = 1 if (data.get('allow_report_pdf_download') is True or str(data.get('allow_report_pdf_download','')).strip().lower() in {'1','true','yes','on'}) else 0
    show_pdf_pages = 0 if (data.get('show_pdf_pages') is False or str(data.get('show_pdf_pages','')).strip().lower() in {'0','false','no','off'}) else 1
    history_scope = _normalize_guest_history_scope(data.get('history_scope') or 'current')
    include_cancelled = 1 if (data.get('include_cancelled') is True or str(data.get('include_cancelled','')).strip().lower() in {'1','true','yes','on'}) else 0
    password_kind = str(data.get('password_kind') or 'none').strip().lower()
    if password_kind not in {'none','phone','custom'}:
        password_kind = 'none'
    raw_password = str(data.get('password') or '').strip()
    if password_kind != 'none' and len(raw_password) < 4:
        return jsonify({'success': False, 'error': '分享密码至少需要 4 个字符'}), 400
    password_hash = generate_password_hash(raw_password) if password_kind != 'none' else None
    if not customer_input:
        return jsonify({'success': False, 'error': '缺少客户名称'}), 400

    conn = get_db()
    try:
        ensure_local_guest_link_tables(conn)
        access_all, user_id = _customer_report_access_context()
        resolved, missing = _resolve_accessible_customer_names(conn, [customer_input], access_all, user_id)
        customer_name = resolved.get(customer_input)
        if not customer_name or missing:
            return jsonify({'success': False, 'error': '找不到该客户，或当前账号无权查看该客户'}), 403
        if not _local_guest_visible_items_for_customer(conn, customer_name, history_scope=history_scope, include_cancelled=bool(include_cancelled)):
            return jsonify({'success': False, 'error': '该客户目前没有可查看订单'}), 404

        now = int(time.time())
        expires = now + (duration_minutes * 60 if not is_permanent else 36500 * 86400)
        raw_token = secrets.token_urlsafe(32)
        guest_url = _local_guest_public_url(raw_token)
        guest_qr_data_uri = _local_guest_qr_data_uri(guest_url)
        creator = get_current_user_context()
        creator_name = creator.get('display_name') or creator.get('username') or session.get('username') or ''
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM local_guest_links
            WHERE COALESCE(is_permanent, 0) = 0
              AND (expires_at_epoch < ? OR revoked_at_epoch IS NOT NULL) AND created_at_epoch < ?
        """, (now - 7 * 86400, now - 30 * 86400))
        cursor.execute("""
            INSERT INTO local_guest_links (
                token_hash, customer_name, created_at_epoch, expires_at_epoch,
                created_by_id, created_by_name, access_count, allow_pdf_download,
                allow_report_pdf_download, show_pdf_pages, history_scope, include_cancelled,
                share_mode, is_permanent, password_hash, password_kind, share_url, qr_data_uri
            ) VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, 'lan', ?, ?, ?, ?, ?)
        """, (
            _local_guest_token_hash(raw_token), customer_name, now, expires, user_id, creator_name,
            allow_pdf_download, allow_report_pdf_download, show_pdf_pages, history_scope, include_cancelled,
            1 if is_permanent else 0, password_hash, password_kind, guest_url, guest_qr_data_uri
        ))
        link_id = cursor.lastrowid
        conn.commit()
    finally:
        conn.close()

    return jsonify({'success': True, 'data': {
        'id': link_id, 'customer_name': customer_name, 'duration_minutes': None if is_permanent else duration_minutes,
        'is_permanent': is_permanent, 'share_mode':'lan', 'password_kind': password_kind,
        'allow_pdf_download': bool(allow_pdf_download), 'allow_report_pdf_download': bool(allow_report_pdf_download),
        'show_pdf_pages': bool(show_pdf_pages), 'history_scope': history_scope,
        'include_cancelled': bool(include_cancelled), 'expires_at_epoch': expires,
        'url': guest_url, 'qr_data_uri': guest_qr_data_uri
    }})

def _share_epoch(value, default=0):
    if value is None or value == '':
        return int(default or 0)
    try:
        return int(float(value))
    except (TypeError, ValueError):
        try:
            text = str(value).strip().replace('Z', '+00:00')
            dt = datetime.fromisoformat(text)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return int(dt.timestamp())
        except Exception:
            return int(default or 0)


def _public_share_registry_upsert(conn, item, creator=None):
    """Mirror public-share provider metadata for scalable admin management/history."""
    row = dict(item or {})
    creator = dict(creator or {})
    share_id = str(row.get('id') or row.get('share_id') or row.get('token') or '').strip()
    if not share_id:
        return None
    cursor = conn.cursor()
    existing = cursor.execute('SELECT * FROM public_guest_share_registry WHERE share_id = ?', (share_id,)).fetchone()
    if existing:
        existing = dict(existing)
        for key in ('customer_name','created_at_epoch','expires_at_epoch','created_by_id','created_by_name',
                    'revoked_at_epoch','last_accessed_at_epoch','access_count','show_pdf_pages',
                    'allow_report_pdf_download','history_scope','include_cancelled','is_permanent','password_kind'):
            if key not in row or row.get(key) is None:
                row[key] = existing.get(key)
        if not any(str(row.get(k) or '').strip() for k in ('url','public_url','share_url','guest_url')):
            row['share_url'] = existing.get('share_url') or ''
        if not str(row.get('qr_data_uri') or '').strip():
            row['qr_data_uri'] = existing.get('qr_data_uri') or ''
    customer_name = str(row.get('customer_name') or row.get('customer') or '').strip() or '—'
    now = int(time.time())
    created = _share_epoch(row.get('created_at_epoch') or row.get('created_at'), now)
    permanent = bool(row.get('is_permanent'))
    expires = _share_epoch(row.get('expires_at_epoch') or row.get('expires_at'), 0)
    if permanent and not expires:
        expires = created + 36500 * 86400
    revoked = _share_epoch(row.get('revoked_at_epoch') or row.get('revoked_at'), 0) or None
    last_accessed = _share_epoch(row.get('last_accessed_at_epoch') or row.get('last_accessed_at'), 0) or None
    url = str(row.get('url') or row.get('public_url') or row.get('share_url') or row.get('guest_url') or '').strip()
    qr = str(row.get('qr_data_uri') or '').strip()
    creator_id = row.get('created_by_id') or creator.get('id')
    creator_name = str(row.get('created_by_name') or row.get('creator_name') or creator.get('display_name') or creator.get('username') or '').strip()
    cursor.execute("""
        INSERT INTO public_guest_share_registry (
            share_id, customer_name, created_at_epoch, expires_at_epoch,
            created_by_id, created_by_name, revoked_at_epoch, last_accessed_at_epoch,
            access_count, show_pdf_pages, allow_report_pdf_download, history_scope,
            include_cancelled, is_permanent, password_kind, share_url, qr_data_uri,
            updated_at_epoch
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(share_id) DO UPDATE SET
            customer_name=CASE WHEN excluded.customer_name <> '—' AND excluded.customer_name <> '' THEN excluded.customer_name ELSE public_guest_share_registry.customer_name END,
            created_at_epoch=COALESCE(public_guest_share_registry.created_at_epoch, excluded.created_at_epoch),
            expires_at_epoch=CASE WHEN excluded.expires_at_epoch > 0 THEN excluded.expires_at_epoch ELSE public_guest_share_registry.expires_at_epoch END,
            created_by_id=COALESCE(public_guest_share_registry.created_by_id, excluded.created_by_id),
            created_by_name=CASE WHEN excluded.created_by_name <> '' THEN excluded.created_by_name ELSE public_guest_share_registry.created_by_name END,
            revoked_at_epoch=COALESCE(excluded.revoked_at_epoch, public_guest_share_registry.revoked_at_epoch),
            last_accessed_at_epoch=COALESCE(excluded.last_accessed_at_epoch, public_guest_share_registry.last_accessed_at_epoch),
            access_count=CASE WHEN excluded.access_count > public_guest_share_registry.access_count THEN excluded.access_count ELSE public_guest_share_registry.access_count END,
            show_pdf_pages=excluded.show_pdf_pages,
            allow_report_pdf_download=excluded.allow_report_pdf_download,
            history_scope=excluded.history_scope,
            include_cancelled=excluded.include_cancelled,
            is_permanent=excluded.is_permanent,
            password_kind=excluded.password_kind,
            share_url=CASE WHEN excluded.share_url <> '' THEN excluded.share_url ELSE public_guest_share_registry.share_url END,
            qr_data_uri=CASE WHEN excluded.qr_data_uri <> '' THEN excluded.qr_data_uri ELSE public_guest_share_registry.qr_data_uri END,
            updated_at_epoch=excluded.updated_at_epoch
    """, (
        share_id, customer_name, created, expires, creator_id, creator_name, revoked, last_accessed,
        int(row.get('access_count') or 0), 0 if row.get('show_pdf_pages') is False else 1,
        1 if row.get('allow_report_pdf_download') else 0, str(row.get('history_scope') or 'current'),
        1 if row.get('include_cancelled') else 0, 1 if permanent else 0,
        str(row.get('password_kind') or ('custom' if row.get('password_protected') else 'none')),
        url, qr, now
    ))
    return share_id


def _refresh_public_share_registry(ensure_qr=False):
    if not (bool(current_app.config.get('TRACKING_RENDER_PUBLIC_GUEST_ENABLED', RENDER_PUBLIC_GUEST_ENABLED))
            and bool(current_app.config.get('TRACKING_PUBLIC_SHARE_PROVIDER_READY', PUBLIC_SHARE_PROVIDER_READY))
            and public_share_provider_ready()):
        return []
    rows = [dict(x or {}) for x in list_public_shares()]
    conn = get_db()
    try:
        ensure_public_guest_share_registry(conn)
        for item in rows:
            item.setdefault('share_mode', 'render')
            url = str(item.get('url') or item.get('public_url') or item.get('share_url') or item.get('guest_url') or '').strip()
            if url:
                item['url'] = url
                if ensure_qr and not str(item.get('qr_data_uri') or '').strip():
                    item['qr_data_uri'] = _local_guest_qr_data_uri(url)
            _public_share_registry_upsert(conn, item)
        conn.commit()
    finally:
        conn.close()
    return rows

_SHARE_ADMIN_PUBLIC_REFRESH_LOCK = threading.Lock()
_SHARE_ADMIN_PUBLIC_REFRESH_AT = 0


def _admin_maybe_refresh_public_registry(force=False):
    global _SHARE_ADMIN_PUBLIC_REFRESH_AT
    now = int(time.time())
    if not force and now - int(_SHARE_ADMIN_PUBLIC_REFRESH_AT or 0) < 15:
        return
    with _SHARE_ADMIN_PUBLIC_REFRESH_LOCK:
        now = int(time.time())
        if not force and now - int(_SHARE_ADMIN_PUBLIC_REFRESH_AT or 0) < 15:
            return
        try:
            _refresh_public_share_registry()
        finally:
            _SHARE_ADMIN_PUBLIC_REFRESH_AT = now


def _admin_share_status(revoked_at_epoch, expires_at_epoch, is_permanent, now=None):
    now = int(now or time.time())
    if _share_epoch(revoked_at_epoch, 0) > 0:
        return 'revoked'
    if bool(is_permanent):
        return 'active'
    return 'active' if _share_epoch(expires_at_epoch, 0) > now else 'expired'


def _admin_share_common(row, mode, now=None, include_qr=False):
    now = int(now or time.time())
    item = dict(row or {})
    permanent = bool(int(item.get('is_permanent') or 0)) if not isinstance(item.get('is_permanent'), bool) else bool(item.get('is_permanent'))
    created = _share_epoch(item.get('created_at_epoch'), 0)
    expires = _share_epoch(item.get('expires_at_epoch'), 0)
    revoked = _share_epoch(item.get('revoked_at_epoch'), 0) or None
    status = _admin_share_status(revoked, expires, permanent, now)
    share_url = str(item.get('share_url') or item.get('url') or '').strip()
    result = {
        'id': str(item.get('id') if mode == 'lan' else item.get('share_id') or ''),
        'share_mode': mode,
        'customer_name': str(item.get('customer_name') or ''),
        'created_at_epoch': created,
        'expires_at_epoch': expires,
        'created_by_id': item.get('created_by_id'),
        'created_by_name': str(item.get('created_by_name') or ''),
        'revoked_at_epoch': revoked,
        'last_accessed_at_epoch': _share_epoch(item.get('last_accessed_at_epoch'), 0) or None,
        'access_count': int(item.get('access_count') or 0),
        'show_pdf_pages': bool(int(item.get('show_pdf_pages') or 0)),
        'allow_report_pdf_download': bool(int(item.get('allow_report_pdf_download') or 0)),
        'history_scope': str(item.get('history_scope') or 'current'),
        'include_cancelled': bool(int(item.get('include_cancelled') or 0)),
        'is_permanent': permanent,
        'password_kind': str(item.get('password_kind') or 'none'),
        'password_protected': str(item.get('password_kind') or 'none') != 'none',
        'url': share_url,
        'has_url': bool(share_url),
        'status': status,
        'remaining_seconds': None if permanent else max(0, expires - now) if status == 'active' else 0,
    }
    if include_qr:
        qr = str(item.get('qr_data_uri') or '').strip()
        if share_url and not qr:
            qr = _local_guest_qr_data_uri(share_url)
        result['qr_data_uri'] = qr
    return result


def _admin_load_all_shares(refresh_public=True, force_public=False):
    if refresh_public:
        try:
            _admin_maybe_refresh_public_registry(force=force_public)
        except Exception as exc:
            print(f'[WARN] share registry refresh failed: {exc}')
    conn = get_db()
    now = int(time.time())
    try:
        ensure_local_guest_link_tables(conn)
        ensure_public_guest_share_registry(conn)
        local_rows = conn.execute("""
            SELECT id, customer_name, created_at_epoch, expires_at_epoch,
                   created_by_id, created_by_name, revoked_at_epoch, last_accessed_at_epoch,
                   access_count, show_pdf_pages, allow_report_pdf_download, history_scope,
                   include_cancelled, is_permanent, password_kind, share_url, qr_data_uri
            FROM local_guest_links
        """).fetchall()
        public_rows = conn.execute("""
            SELECT share_id, customer_name, created_at_epoch, expires_at_epoch,
                   created_by_id, created_by_name, revoked_at_epoch, last_accessed_at_epoch,
                   access_count, show_pdf_pages, allow_report_pdf_download, history_scope,
                   include_cancelled, is_permanent, password_kind, share_url, qr_data_uri
            FROM public_guest_share_registry
        """).fetchall()
        items = [_admin_share_common(row, 'lan', now) for row in local_rows]
        items.extend(_admin_share_common(row, 'render', now) for row in public_rows)
        return items
    finally:
        conn.close()


@tracking_bp.route('/admin/customer-shares')
@admin_required
def admin_customer_shares():
    """Desktop ADMIN workspace for scalable customer-share management."""
    return render_template('tracking/customer_share_admin.html')


@tracking_bp.route('/api/admin/customer-shares', methods=['GET'])
@admin_required
def api_admin_customer_shares():
    try:
        page = max(1, int(request.args.get('page', 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(100, max(10, int(request.args.get('page_size', 30))))
    except (TypeError, ValueError):
        page_size = 30
    keyword = str(request.args.get('q') or '').strip().casefold()
    mode = str(request.args.get('mode') or 'all').strip().lower()
    status_filter = str(request.args.get('status') or 'all').strip().lower()
    duration = str(request.args.get('duration') or 'all').strip().lower()
    sort_key = str(request.args.get('sort') or 'created_desc').strip().lower()

    force_public = str(request.args.get('refresh') or '').strip().lower() in {'1','true','yes','on'}
    all_items = _admin_load_all_shares(refresh_public=True, force_public=force_public)
    now = int(time.time())
    stats = {
        'active': sum(1 for x in all_items if x['status'] == 'active'),
        'permanent': sum(1 for x in all_items if x['status'] == 'active' and x['is_permanent']),
        'expiring_24h': sum(1 for x in all_items if x['status'] == 'active' and not x['is_permanent'] and 0 < int(x.get('remaining_seconds') or 0) <= 86400),
        'lan': sum(1 for x in all_items if x['status'] == 'active' and x['share_mode'] == 'lan'),
        'render': sum(1 for x in all_items if x['status'] == 'active' and x['share_mode'] == 'render'),
        'total': len(all_items),
    }

    items = all_items
    if keyword:
        items = [x for x in items if keyword in str(x.get('customer_name') or '').casefold() or keyword in str(x.get('created_by_name') or '').casefold() or keyword in str(x.get('id') or '').casefold()]
    if mode in {'lan', 'render'}:
        items = [x for x in items if x['share_mode'] == mode]
    if status_filter in {'active', 'expired', 'revoked'}:
        items = [x for x in items if x['status'] == status_filter]
    if duration == 'permanent':
        items = [x for x in items if x['is_permanent']]
    elif duration == 'temporary':
        items = [x for x in items if not x['is_permanent']]

    if sort_key == 'expires_asc':
        items.sort(key=lambda x: (0 if x['status'] == 'active' else 1, 1 if x['is_permanent'] else 0, int(x.get('expires_at_epoch') or 0), str(x.get('customer_name') or '').casefold()))
    elif sort_key == 'customer_asc':
        items.sort(key=lambda x: (str(x.get('customer_name') or '').casefold(), -int(x.get('created_at_epoch') or 0)))
    elif sort_key == 'access_desc':
        items.sort(key=lambda x: (-int(x.get('access_count') or 0), -int(x.get('created_at_epoch') or 0)))
    else:
        items.sort(key=lambda x: -int(x.get('created_at_epoch') or 0))

    total = len(items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    page_items = []
    for source in items[start:start + page_size]:
        summary = dict(source)
        # Keep raw guest URLs out of the bulk list response. They are returned only
        # from the admin-only detail endpoint when the user explicitly opens a row.
        summary.pop('url', None)
        page_items.append(summary)
    response = jsonify({'success': True, 'data': page_items, 'stats': stats, 'page': page, 'page_size': page_size, 'total': total, 'total_pages': total_pages, 'server_epoch': now})
    response.headers['Cache-Control'] = 'no-store, private, max-age=0'
    return response


@tracking_bp.route('/api/admin/customer-shares/<mode>/<share_id>', methods=['GET'])
@admin_required
def api_admin_customer_share_detail(mode, share_id):
    mode = str(mode or '').strip().lower()
    if mode not in {'lan', 'render'}:
        return jsonify({'success': False, 'error': '分享类型错误'}), 400
    conn = get_db()
    try:
        if mode == 'lan':
            ensure_local_guest_link_tables(conn)
            try:
                numeric_id = int(share_id)
            except (TypeError, ValueError):
                return jsonify({'success': False, 'error': '分享编号错误'}), 400
            row = conn.execute("""
                SELECT id, customer_name, created_at_epoch, expires_at_epoch,
                       created_by_id, created_by_name, revoked_at_epoch, last_accessed_at_epoch,
                       access_count, show_pdf_pages, allow_report_pdf_download, history_scope,
                       include_cancelled, is_permanent, password_kind, share_url, qr_data_uri
                FROM local_guest_links WHERE id = ?
            """, (numeric_id,)).fetchone()
        else:
            ensure_public_guest_share_registry(conn)
            row = conn.execute("""
                SELECT share_id, customer_name, created_at_epoch, expires_at_epoch,
                       created_by_id, created_by_name, revoked_at_epoch, last_accessed_at_epoch,
                       access_count, show_pdf_pages, allow_report_pdf_download, history_scope,
                       include_cancelled, is_permanent, password_kind, share_url, qr_data_uri
                FROM public_guest_share_registry WHERE share_id = ?
            """, (str(share_id),)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': '找不到该分享'}), 404
        item = _admin_share_common(row, mode, include_qr=True)
        response = jsonify({'success': True, 'data': item})
        response.headers['Cache-Control'] = 'no-store, private, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        return response
    finally:
        conn.close()


@tracking_bp.route('/api/admin/customer-shares/<mode>/<share_id>/password', methods=['PATCH'])
@admin_required
def api_admin_update_customer_share_password(mode, share_id):
    """Add, replace, or remove password protection on an existing active share.

    LAN links are updated in the local SQLite row. Render links are delegated to the
    registered provider via its optional ``update_share`` capability so the real public
    share — not merely the local mirror — is changed.
    """
    mode = str(mode or '').strip().lower()
    if mode not in {'lan', 'render'}:
        return jsonify({'success': False, 'error': '分享类型错误'}), 400
    data = request.get_json(silent=True) or {}
    enabled = bool(data.get('enabled') is True or str(data.get('enabled', '')).strip().lower() in {'1','true','yes','on'})
    raw_password = str(data.get('password') or '').strip()
    if enabled and len(raw_password) < 4:
        return jsonify({'success': False, 'error': '分享密码至少需要 4 个字符'}), 400

    if mode == 'lan':
        if not _local_guest_feature_enabled() or cloud_mode_enabled():
            return jsonify({'success': False, 'error': '本机客户分享功能在此部署已关闭'}), 403
        try:
            numeric_id = int(share_id)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '分享编号错误'}), 400
        conn = get_db()
        try:
            ensure_local_guest_link_tables(conn)
            row = conn.execute('SELECT id, revoked_at_epoch, expires_at_epoch, is_permanent FROM local_guest_links WHERE id = ?', (numeric_id,)).fetchone()
            if not row:
                return jsonify({'success': False, 'error': '找不到该分享'}), 404
            row = dict(row)
            now = int(time.time())
            if row.get('revoked_at_epoch') or (not bool(int(row.get('is_permanent') or 0)) and int(row.get('expires_at_epoch') or 0) <= now):
                return jsonify({'success': False, 'error': '只有有效分享可以修改密码'}), 409
            password_hash = generate_password_hash(raw_password) if enabled else None
            password_kind = 'custom' if enabled else 'none'
            conn.execute('UPDATE local_guest_links SET password_hash = ?, password_kind = ? WHERE id = ?', (password_hash, password_kind, numeric_id))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'success': True, 'data': {'id': str(numeric_id), 'share_mode': 'lan', 'password_kind': password_kind, 'password_protected': enabled}})

    # Render/public links must be updated at the provider so the public page enforces
    # the change.  Do not pretend a local mirror-only change is sufficient.
    if not bool(current_app.config.get('TRACKING_RENDER_PUBLIC_GUEST_ENABLED', RENDER_PUBLIC_GUEST_ENABLED)):
        return jsonify({'success': False, 'error': '此部署未开放 Render 公網分享'}), 403
    if not public_share_provider_ready():
        return jsonify({'success': False, 'error': 'Render/B2 Provider 尚未配置'}), 503
    try:
        update_public_share(str(share_id), {
            'password_kind': 'custom' if enabled else 'none',
            'password': raw_password if enabled else '',
        })
        conn = get_db()
        try:
            ensure_public_guest_share_registry(conn)
            conn.execute('UPDATE public_guest_share_registry SET password_kind = ?, updated_at_epoch = ? WHERE share_id = ?',
                         ('custom' if enabled else 'none', int(time.time()), str(share_id)))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'success': True, 'data': {'id': str(share_id), 'share_mode': 'render', 'password_kind': 'custom' if enabled else 'none', 'password_protected': enabled}})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc) or '无法更新 Render 分享密码'}), 503


@tracking_bp.route('/api/public-guest-links', methods=['GET'])
@login_required
def api_list_public_guest_links():
    if not is_admin():
        return jsonify({'success':False,'error':'客户公網分享仅限 ADMIN 使用'}), 403
    enabled = bool(current_app.config.get('TRACKING_RENDER_PUBLIC_GUEST_ENABLED', RENDER_PUBLIC_GUEST_ENABLED))
    ready = bool(current_app.config.get('TRACKING_PUBLIC_SHARE_PROVIDER_READY', PUBLIC_SHARE_PROVIDER_READY))
    if not enabled:
        return jsonify({'success':True,'data':[],'total':0,'disabled':True})
    if not ready or not public_share_provider_ready():
        return jsonify({'success':True,'data':[],'total':0,'provider_ready':False})
    compact = str(request.args.get('compact') or '').strip().lower() in {'1','true','yes','on'}
    try:
        rows = _refresh_public_share_registry(ensure_qr=not compact)
        if compact:
            compact_rows = []
            for source in rows:
                item = dict(source or {})
                url = str(item.get('url') or item.get('public_url') or item.get('share_url') or item.get('guest_url') or '').strip()
                for key in ('url','public_url','share_url','guest_url','qr_data_uri'):
                    item.pop(key, None)
                item['has_url'] = bool(url)
                compact_rows.append(item)
            rows = compact_rows
        return jsonify({'success':True,'data':rows,'total':len(rows),'provider_ready':True})
    except Exception as exc:
        return jsonify({'success':False,'error':str(exc) or '无法读取 Render 分享'}), 503


@tracking_bp.route('/api/public-guest-links', methods=['POST'])
@login_required
def api_create_public_guest_link_placeholder():
    if not is_admin():
        return jsonify({'success':False,'error':'客户公網分享仅限 ADMIN 使用'}), 403
    enabled = bool(current_app.config.get('TRACKING_RENDER_PUBLIC_GUEST_ENABLED', RENDER_PUBLIC_GUEST_ENABLED))
    ready = bool(current_app.config.get('TRACKING_PUBLIC_SHARE_PROVIDER_READY', PUBLIC_SHARE_PROVIDER_READY))
    if not enabled:
        return jsonify({'success':False,'error':'此部署未开放 Render 公網分享','code':'PUBLIC_SHARE_DISABLED'}), 403
    if not ready or not public_share_provider_ready():
        return jsonify({'success':False,'error':'Render/B2 Provider 尚未配置；未上传任何客户资料','code':'PUBLIC_SHARE_PROVIDER_NOT_READY'}), 503
    data = request.get_json(silent=True) or {}
    if bool(data.get('is_permanent')) and not bool(current_app.config.get('TRACKING_PERMANENT_PUBLIC_GUEST_ENABLED', PERMANENT_PUBLIC_GUEST_ENABLED)):
        return jsonify({'success':False,'error':'此部署未开放永久 Render 链接'}), 403
    creator = get_current_user_context()
    try:
        result = create_public_share(data, creator)
        if not isinstance(result, dict):
            raise RuntimeError('Public provider returned invalid result')
        public_url = str(result.get('url') or result.get('public_url') or result.get('share_url') or result.get('guest_url') or '').strip()
        if not public_url:
            raise RuntimeError('Render 分享已建立，但 Provider 未返回可访问网址')
        result['url'] = public_url
        result.setdefault('share_mode', 'render')
        result.setdefault('customer_name', str(data.get('customer_name') or '').strip())
        result.setdefault('history_scope', str(data.get('history_scope') or 'current'))
        result.setdefault('include_cancelled', bool(data.get('include_cancelled')))
        result.setdefault('show_pdf_pages', data.get('show_pdf_pages') is not False)
        result.setdefault('allow_report_pdf_download', bool(data.get('allow_report_pdf_download')))
        result.setdefault('password_kind', str(data.get('password_kind') or 'none'))
        result.setdefault('is_permanent', bool(data.get('is_permanent')))
        result.setdefault('created_at_epoch', int(time.time()))
        if not result.get('expires_at_epoch') and not bool(result.get('is_permanent')):
            try:
                result['expires_at_epoch'] = int(result['created_at_epoch']) + int(data.get('duration_minutes') or 1440) * 60
            except Exception:
                pass
        if not str(result.get('qr_data_uri') or '').strip():
            result['qr_data_uri'] = _local_guest_qr_data_uri(public_url)
        conn = get_db()
        try:
            ensure_public_guest_share_registry(conn)
            _public_share_registry_upsert(conn, result, creator)
            conn.commit()
        finally:
            conn.close()
        return jsonify({'success':True,'data':result})
    except Exception as exc:
        return jsonify({'success':False,'error':str(exc) or 'Render 分享建立失败'}), 503


@tracking_bp.route('/api/public-guest-links/<share_id>', methods=['DELETE'])
@login_required
def api_revoke_public_guest_link(share_id):
    if not is_admin():
        return jsonify({'success':False,'error':'客户公網分享仅限 ADMIN 使用'}), 403
    if not bool(current_app.config.get('TRACKING_RENDER_PUBLIC_GUEST_ENABLED', RENDER_PUBLIC_GUEST_ENABLED)):
        return jsonify({'success':False,'error':'此部署未开放 Render 公網分享'}), 403
    if not public_share_provider_ready():
        return jsonify({'success':False,'error':'Render/B2 Provider 尚未配置'}), 503
    try:
        revoke_public_share(share_id)
        conn = get_db()
        try:
            ensure_public_guest_share_registry(conn)
            conn.execute('UPDATE public_guest_share_registry SET revoked_at_epoch = ?, updated_at_epoch = ? WHERE share_id = ?', (int(time.time()), int(time.time()), str(share_id)))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'success':True})
    except Exception as exc:
        return jsonify({'success':False,'error':str(exc) or 'Render 分享撤销失败'}), 503


@tracking_bp.route('/api/local-guest-links/<int:link_id>', methods=['DELETE'])
@login_required
def api_revoke_local_guest_link(link_id):
    if not _local_guest_feature_enabled() or cloud_mode_enabled():
        return jsonify({'success': False, 'error': '本机客户分享功能在此部署已关闭'}), 403
    if not is_admin():
        return jsonify({'success': False, 'error': '临时客户查看仅限 ADMIN 使用'}), 403
    conn = get_db()
    try:
        ensure_local_guest_link_tables(conn)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM local_guest_links WHERE id = ?', (link_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': '临时链接不存在'}), 404
        item = dict(row)
        ctx = get_current_user_context()
        if not is_admin() and str(item.get('created_by_id') or '') != str(ctx.get('id') or ''):
            return jsonify({'success': False, 'error': '只能撤销自己建立的临时链接'}), 403
        cursor.execute('UPDATE local_guest_links SET revoked_at_epoch = ? WHERE id = ?', (int(time.time()), link_id))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True})


@tracking_bp.route('/guest/<token>', methods=['GET'])
def local_guest_customer(token):
    if cloud_mode_enabled() or not _local_guest_feature_enabled():
        return _local_guest_error_page('cloud', 404)
    if not _local_guest_client_allowed():
        return _local_guest_error_page('network', 403)
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=True)
        if state == 'locked':
            response = make_response(render_template(
                'tracking/guest_unlock.html', token=token,
                customer_name=str((link or {}).get('customer_name') or ''),
                error_message='', password_kind=str((link or {}).get('password_kind') or 'custom')
            ), 200)
            return _local_guest_no_store(response)
        if state != 'active':
            return _local_guest_error_page(state, 410 if state in {'expired', 'revoked'} else 404)
        show_pdf_pages = bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1))
        orders = _local_guest_load_orders(conn, token, link['customer_name'], show_pdf_pages=show_pdf_pages, link=link)
        allow_pdf_download = bool(int(link.get('allow_pdf_download') or 0))
        allow_report_pdf_download = bool(int(link.get('allow_report_pdf_download') or 0))
        pdf_count = len(_local_guest_collect_pdf_files(conn, link['customer_name'], link=link)) if allow_pdf_download else 0
        is_permanent = bool(int(link.get('is_permanent') or 0))
    finally:
        conn.close()
    response = make_response(render_template(
        'tracking/guest_customer.html', customer_name=link['customer_name'], orders=orders,
        token=token, expires_at_epoch=int(link['expires_at_epoch']), is_permanent=is_permanent,
        allow_pdf_download=allow_pdf_download, pdf_count=pdf_count,
        allow_report_pdf_download=allow_report_pdf_download,
        show_pdf_pages=show_pdf_pages,
        history_scope=_guest_history_scope_from_link(link), include_cancelled=_guest_include_cancelled_from_link(link)
    ))
    return _local_guest_no_store(response)


@tracking_bp.route('/guest/<token>/unlock', methods=['POST'])
def local_guest_unlock(token):
    if cloud_mode_enabled() or not _local_guest_feature_enabled() or not _local_guest_client_allowed():
        return _local_guest_error_page('network', 403)
    password = str(request.form.get('password') or '').strip()
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=False, allow_locked=True)
        if not link:
            return _local_guest_error_page('not_found', 404)
        if state in {'expired','revoked'}:
            return _local_guest_error_page(state, 410)
        password_hash = str(link.get('password_hash') or '').strip()
        if not password_hash or check_password_hash(password_hash, password):
            session[_local_guest_password_session_key(token)] = True
            session.modified = True
            return redirect(url_for('tracking_bp.local_guest_customer', token=token))
        response = make_response(render_template(
            'tracking/guest_unlock.html', token=token,
            customer_name=str(link.get('customer_name') or ''),
            error_message='密码不正确 / Contraseña incorrecta',
            password_kind=str(link.get('password_kind') or 'custom')
        ), 401)
        return _local_guest_no_store(response)
    finally:
        conn.close()


@tracking_bp.route('/guest/<token>/download-pdfs', methods=['GET'])
def local_guest_pdf_bundle(token):
    """Download only this customer's original PDF attachments as one ZIP when explicitly allowed."""
    if cloud_mode_enabled():
        return _local_guest_error_page('cloud', 404)
    if not _local_guest_client_allowed():
        return _local_guest_error_page('network', 403)

    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=True)
        if state != 'active':
            return _local_guest_error_page(state, 410 if state in {'expired', 'revoked'} else 404)
        if not bool(int(link.get('allow_pdf_download') or 0)):
            return _local_guest_error_page('not_found', 404)
        pdfs = _local_guest_collect_pdf_files(conn, link['customer_name'], link=link)
    finally:
        conn.close()

    if not pdfs:
        response = make_response('No hay archivos PDF disponibles para descargar.', 404)
        return _local_guest_no_store(response)

    spool = tempfile.SpooledTemporaryFile(max_size=32 * 1024 * 1024, mode='w+b')
    try:
        with zipfile.ZipFile(spool, 'w', compression=zipfile.ZIP_STORED) as archive:
            for item in pdfs:
                archive.write(item['path'], item['arcname'])
        spool.seek(0)

        conn = get_db()
        try:
            latest, state = _local_guest_get_link(conn, token, touch=False)
            latest_allowed = bool(int((latest or {}).get('allow_pdf_download') or 0))
            if state != 'active' or not latest_allowed:
                spool.close()
                return _local_guest_error_page(
                    state if state != 'active' else 'not_found',
                    410 if state in {'expired', 'revoked'} else 404,
                )
        finally:
            conn.close()

        from flask import send_file
        customer_part = _local_guest_safe_zip_name(link['customer_name'], 'customer')
        response = send_file(
            spool, mimetype='application/zip', as_attachment=True,
            download_name=f'{customer_part}_PDF.zip', conditional=False, max_age=0,
        )
        response = _local_guest_no_store(response)
        response.headers['X-Guest-PDF-Count'] = str(len(pdfs))
        response.call_on_close(spool.close)
        return response
    except Exception:
        spool.close()
        raise



def _guest_report_job_snapshot(job):
    if not job:
        return None
    return {
        'id': job.get('id'),
        'status': job.get('status'),
        'format': job.get('format'),
        'created_at': job.get('created_at'),
        'started_at': job.get('started_at'),
        'finished_at': job.get('finished_at'),
        'order_count': int(job.get('order_count') or 0),
        'estimated_bytes': int(job.get('estimated_bytes') or 0),
        'estimated_seconds_low': int(job.get('estimated_seconds_low') or 0),
        'estimated_seconds_high': int(job.get('estimated_seconds_high') or 0),
        'pdf_page_count': int(job.get('pdf_page_count') or 0),
        'image_count': int(job.get('image_count') or 0),
        'queue_ahead': int(job.get('queue_ahead') or 0),
        'error': job.get('error'),
        'files': [
            {
                'name': f.get('name'),
                'size': int(f.get('size') or 0),
                'mimetype': f.get('mimetype'),
                'file_index': idx,
            }
            for idx, f in enumerate(job.get('files') or [])
        ],
    }


def _guest_report_trim_jobs():
    with _guest_report_jobs_lock:
        if len(_guest_report_jobs) <= _GUEST_REPORT_JOB_KEEP:
            return
        done = [j for j in _guest_report_jobs.values() if j.get('status') in {'completed', 'failed'}]
        done.sort(key=lambda x: x.get('created_at') or '')
        while len(_guest_report_jobs) > _GUEST_REPORT_JOB_KEEP and done:
            old = done.pop(0)
            _guest_report_jobs.pop(old.get('id'), None)


def _guest_report_time_range(order_count, image_count, pdf_page_count):
    """Deliberately approximate CPU/render time shown to the customer before queueing."""
    # PDF pages dominate. Normal images are much cheaper, and order layout has a small fixed cost.
    seconds = 2.5 + (max(0, int(order_count or 0)) * 0.16) + (max(0, int(image_count or 0)) * 0.09) + (max(0, int(pdf_page_count or 0)) * 0.34)
    low = max(3, int(seconds * 0.75))
    high = max(low + 2, int(seconds * 1.65) + 2)
    return low, high


def _guest_report_estimate(conn, customer_name, report_format, show_pdf_pages=True, link=None):
    items = _local_guest_visible_items_for_customer(conn, customer_name, link=link)
    entries = load_report_entries(conn, items, True, None) if items else []
    if not entries:
        return None
    attach_image_metadata(conn, entries, 'both', 'all', 'order_first', 'pages' if show_pdf_pages else 'skip')

    normal_images = 0
    pdf_pages = 0
    seen_pdf_paths = set()
    estimated_media_bytes = 0
    for entry in entries:
        for meta in entry.get('images') or []:
            path = str(meta.get('path') or '')
            if _media_is_pdf(meta.get('display_name') or path, meta.get('mime_type'), meta.get('mime_type')):
                if not show_pdf_pages:
                    continue
                norm = os.path.normcase(os.path.abspath(path)) if path else ''
                if not norm or norm in seen_pdf_paths:
                    continue
                seen_pdf_paths.add(norm)
                try:
                    pages = int(_pdf_preview_page_count(path))
                except Exception:
                    pages = 1
                page_total = max(1, pages)
                pdf_pages += page_total
                estimated_media_bytes += page_total * 190_000
            else:
                normal_images += 1
                raw = int(meta.get('file_size') or 0)
                estimated_media_bytes += min(max(int(raw * 0.32), 85_000), 460_000) if raw else 220_000

    estimated_bytes = 110_000 + len(entries) * 7_000 + estimated_media_bytes
    low, high = _guest_report_time_range(len(entries), normal_images, pdf_pages)
    return {
        'order_count': len(entries),
        'image_count': normal_images,
        'pdf_page_count': pdf_pages,
        'estimated_bytes': int(estimated_bytes),
        'estimated_seconds_low': low,
        'estimated_seconds_high': high,
    }


def _guest_report_set_job(job_id, **updates):
    with _guest_report_jobs_lock:
        job = _guest_report_jobs.get(job_id)
        if job:
            job.update(updates)




def _run_guest_report_job(job_id, token, customer_name, report_format):
    _guest_report_set_job(job_id, status='processing', started_at=datetime.now().isoformat(timespec='seconds'))
    conn = None
    try:
        # Validate again at worker start. A revoked/expired token must not continue as if authorized.
        conn = get_db()
        link, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active' or not _local_guest_report_permission(link, report_format):
            raise RuntimeError('El acceso temporal ya no está disponible.')
        show_pdf_pages = bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1))
        items = _local_guest_visible_items_for_customer(conn, customer_name, link=link)
        entries = load_report_entries(conn, items, True, None) if items else []
        if not entries:
            raise RuntimeError('No hay pedidos disponibles para generar el informe.')

        prepared = prepare_customer_report_entries(
            conn, entries, 'both', 'all', 'es', 'order_first',
            'pages' if show_pdf_pages else 'skip'
        )
        generated = build_customer_report_files(prepared, 'pdf', 'es')
        conn.close()
        conn = None

        # Validate once more after expensive rendering and before publishing any cached result.
        conn = get_db()
        latest, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active' or not _local_guest_report_permission(latest, report_format):
            raise RuntimeError('El acceso temporal caducó o fue revocado antes de terminar.')

        files = []
        for filename, mimetype, content in generated:
            cached = cache_report_file(filename, mimetype, content)
            files.append({
                'file_id': cached['id'],
                'name': cached['filename'],
                'size': cached['size'],
                'mimetype': mimetype,
            })
        _guest_report_set_job(
            job_id,
            status='completed',
            finished_at=datetime.now().isoformat(timespec='seconds'),
            files=files,
            order_count=len(entries),
            error=None,
        )
    except Exception as exc:
        print(f'[ERROR] guest report background job failed: {exc}')
        _guest_report_set_job(
            job_id,
            status='failed',
            finished_at=datetime.now().isoformat(timespec='seconds'),
            error=str(exc) or 'No se pudo generar el informe.',
        )
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
        _guest_report_trim_jobs()


def _guest_report_queue_ahead():
    with _guest_report_jobs_lock:
        guest_queued = sum(1 for j in _guest_report_jobs.values() if j.get('status') == 'queued')
    # The authenticated report queue shares the same executor, so include its queued jobs as an
    # approximation of how many report tasks may be ahead of the customer.
    with _customer_report_jobs_lock:
        internal_queued = sum(1 for j in _customer_report_jobs.values() if j.get('status') == 'queued')
    return guest_queued + internal_queued


@tracking_bp.route('/guest/<token>/report-estimate/<report_format>', methods=['GET'])
def local_guest_report_estimate(token, report_format):
    report_format = str(report_format or '').strip().lower()
    if report_format != 'pdf' or cloud_mode_enabled() or not _local_guest_client_allowed():
        return jsonify({'success': False, 'error': 'No disponible'}), 404
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active':
            return jsonify({'success': False, 'error': state}), 410 if state in {'expired', 'revoked'} else 404
        if not _local_guest_report_permission(link, report_format):
            return jsonify({'success': False, 'error': 'No disponible'}), 404
        estimate = _guest_report_estimate(conn, link['customer_name'], report_format, bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1)), link=link)
        if not estimate:
            return jsonify({'success': False, 'error': 'No hay pedidos disponibles'}), 404
        return _local_guest_no_store(jsonify({'success': True, **estimate, 'queue_ahead': _guest_report_queue_ahead()}))
    finally:
        conn.close()


@tracking_bp.route('/guest/<token>/report-jobs', methods=['POST'])
def local_guest_report_create_job(token):
    if cloud_mode_enabled() or not _local_guest_client_allowed():
        return jsonify({'success': False, 'error': 'No disponible'}), 404
    data = request.get_json(silent=True) or {}
    report_format = str(data.get('format') or '').strip().lower()
    if report_format != 'pdf':
        return jsonify({'success': False, 'error': 'Formato no válido'}), 400

    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=True)
        if state != 'active':
            return jsonify({'success': False, 'error': state}), 410 if state in {'expired', 'revoked'} else 404
        if not _local_guest_report_permission(link, report_format):
            return jsonify({'success': False, 'error': 'No disponible'}), 404
        estimate = _guest_report_estimate(conn, link['customer_name'], report_format, bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1)), link=link)
        if not estimate:
            return jsonify({'success': False, 'error': 'No hay pedidos disponibles'}), 404
        customer_name = link['customer_name']
        token_hash = _local_guest_token_hash(token)
    finally:
        conn.close()

    queue_ahead = _guest_report_queue_ahead()
    with _guest_report_jobs_lock:
        # Reuse an active identical task instead of making the customer wait twice.
        for existing in _guest_report_jobs.values():
            if (
                existing.get('token_hash') == token_hash
                and existing.get('format') == report_format
                and existing.get('status') in {'queued', 'processing'}
            ):
                return _local_guest_no_store(jsonify({'success': True, 'deduplicated': True, 'job': _guest_report_job_snapshot(existing)}))
        active_for_token = sum(
            1 for j in _guest_report_jobs.values()
            if j.get('token_hash') == token_hash and j.get('status') in {'queued', 'processing'}
        )
        if active_for_token >= 2:
            return jsonify({'success': False, 'error': 'Ya hay dos informes en preparación. Espere a que termine uno.'}), 429
        job_id = secrets.token_hex(16)
        job = {
            'id': job_id,
            'token_hash': token_hash,
            'customer_name': customer_name,
            'format': report_format,
            'status': 'queued',
            'created_at': datetime.now().isoformat(timespec='seconds'),
            'started_at': None,
            'finished_at': None,
            'files': [],
            'error': None,
            'queue_ahead': queue_ahead,
            **estimate,
        }
        _guest_report_jobs[job_id] = job

    try:
        _customer_report_executor.submit(_run_guest_report_job, job_id, token, customer_name, report_format)
    except Exception:
        _guest_report_set_job(job_id, status='failed', finished_at=datetime.now().isoformat(timespec='seconds'), error='No se pudo iniciar el informe.')
        return jsonify({'success': False, 'error': 'No se pudo iniciar el informe.'}), 500
    return _local_guest_no_store(jsonify({'success': True, 'deduplicated': False, 'job': _guest_report_job_snapshot(job)}))


@tracking_bp.route('/guest/<token>/report-jobs/<job_id>', methods=['GET'])
def local_guest_report_get_job(token, job_id):
    if cloud_mode_enabled() or not _local_guest_client_allowed():
        return jsonify({'success': False, 'error': 'No disponible'}), 404
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active':
            return jsonify({'success': False, 'error': state}), 410 if state in {'expired', 'revoked'} else 404
    finally:
        conn.close()
    with _guest_report_jobs_lock:
        job = _guest_report_jobs.get(str(job_id or ''))
        if not job or job.get('token_hash') != _local_guest_token_hash(token):
            return jsonify({'success': False, 'error': 'No encontrado'}), 404
        if not _local_guest_report_permission(link, job.get('format')):
            return jsonify({'success': False, 'error': 'No disponible'}), 404
        snapshot = _guest_report_job_snapshot(job)
    return _local_guest_no_store(jsonify({'success': True, 'job': snapshot}))


@tracking_bp.route('/guest/<token>/report-jobs/<job_id>/files/<int:file_index>', methods=['GET'])
def local_guest_report_job_file(token, job_id, file_index):
    from flask import send_file
    if cloud_mode_enabled() or not _local_guest_client_allowed():
        return _local_guest_error_page('not_found', 404)
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=True)
        if state != 'active':
            return _local_guest_error_page(state, 410 if state in {'expired', 'revoked'} else 404)
    finally:
        conn.close()
    with _guest_report_jobs_lock:
        job = _guest_report_jobs.get(str(job_id or ''))
        if not job or job.get('token_hash') != _local_guest_token_hash(token):
            return _local_guest_error_page('not_found', 404)
        if job.get('status') != 'completed' or not _local_guest_report_permission(link, job.get('format')):
            return _local_guest_error_page('not_found', 404)
        files = list(job.get('files') or [])
        if file_index < 0 or file_index >= len(files):
            return _local_guest_error_page('not_found', 404)
        file_info = dict(files[file_index])
    cached = get_cached_report(file_info.get('file_id'))
    if not cached:
        return _local_guest_error_page('not_found', 404)
    inline = str(request.args.get('inline') or '').lower() in {'1', 'true', 'yes', 'on'}
    response = send_file(
        cached['path'],
        mimetype=cached.get('mimetype') or file_info.get('mimetype') or 'application/octet-stream',
        as_attachment=not inline,
        download_name=cached.get('filename') or file_info.get('name') or 'report.bin',
        conditional=False,
        max_age=0,
    )
    return _local_guest_no_store(response)


def _local_guest_report_permission(link, report_format):
    if report_format == 'pdf':
        return bool(int((link or {}).get('allow_report_pdf_download') or 0))
    return False


def _local_guest_report_response_files(files, customer_name, report_format):
    """Return one report file directly, or ZIP when the report engine had to split it."""
    from flask import send_file
    files = list(files or [])
    if not files:
        response = make_response('No hay informe disponible.', 404)
        return _local_guest_no_store(response)

    if len(files) == 1:
        filename, mimetype, content = files[0]
        response = send_file(
            BytesIO(content), mimetype=mimetype, as_attachment=True,
            download_name=filename, conditional=False, max_age=0,
        )
        return _local_guest_no_store(response)

    spool = tempfile.SpooledTemporaryFile(max_size=32 * 1024 * 1024, mode='w+b')
    try:
        with zipfile.ZipFile(spool, 'w', compression=zipfile.ZIP_STORED) as archive:
            for filename, _mimetype, content in files:
                archive.writestr(filename, content)
        spool.seek(0)
        customer_part = _local_guest_safe_zip_name(customer_name, 'customer')
        suffix = 'PDF'
        response = send_file(
            spool, mimetype='application/zip', as_attachment=True,
            download_name=f'{customer_part}_{suffix}_Report.zip', conditional=False, max_age=0,
        )
        response = _local_guest_no_store(response)
        response.call_on_close(spool.close)
        return response
    except Exception:
        spool.close()
        raise


@tracking_bp.route('/guest/<token>/download-report/<report_format>', methods=['GET'])
def local_guest_download_report(token, report_format):
    """Generate a customer-safe PDF snapshot on demand.

    The token is checked before and after rendering. Only orders belonging to the
    token's exact customer are loaded. Internal notes are not rendered verbatim by
    customer_report.py. A downloaded file is a snapshot and remains on the customer's
    device after the temporary link later expires.
    """
    report_format = str(report_format or '').strip().lower()
    if report_format != 'pdf':
        return _local_guest_error_page('not_found', 404)
    if cloud_mode_enabled():
        return _local_guest_error_page('cloud', 404)
    if not _local_guest_client_allowed():
        return _local_guest_error_page('network', 403)

    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=True)
        if state != 'active':
            return _local_guest_error_page(state, 410 if state in {'expired', 'revoked'} else 404)
        if not _local_guest_report_permission(link, report_format):
            return _local_guest_error_page('not_found', 404)

        items = _local_guest_visible_items_for_customer(conn, link['customer_name'], link=link)
        entries = load_report_entries(conn, items, True, None) if items else []
        if not entries:
            response = make_response('No hay pedidos disponibles para generar el informe.', 404)
            return _local_guest_no_store(response)
        show_pdf_pages = bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1))
        prepared = prepare_customer_report_entries(
            conn, entries, 'both', 'all', 'es', 'order_first',
            'pages' if show_pdf_pages else 'skip'
        )
        generated = build_customer_report_files(prepared, 'pdf', 'es')
    finally:
        conn.close()

    # Rendering may take time for many PDF pages. Re-check the same token/permission
    # immediately before returning bytes so an expired/revoked link cannot finish later.
    conn = get_db()
    try:
        latest, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active' or not _local_guest_report_permission(latest, report_format):
            return _local_guest_error_page(
                state if state != 'active' else 'not_found',
                410 if state in {'expired', 'revoked'} else 404,
            )
    finally:
        conn.close()

    return _local_guest_report_response_files(generated, link['customer_name'], report_format)


def _local_guest_timeline_stage(status_key):
    """Map an exact workflow status to the 5 customer-facing progress stages."""
    key = str(status_key or '').strip()
    if not key:
        return 'order'
    # Check the real business stage groups before the overlapping waiting_confirm group.
    for stage in ('draft', 'sampling', 'production', 'shipping'):
        if key in (STATUS_STAGE_GROUPS.get(stage, {}).get('status_keys') or []):
            return stage
    if key in (STATUS_STAGE_GROUPS.get('completed', {}).get('status_keys') or []):
        return 'shipping'
    return 'order'


def _local_guest_horizontal_timeline(order, history):
    stage_defs = [
        ('order', 'Pedido'),
        ('draft', 'Diseño'),
        ('sampling', 'Muestra'),
        ('production', 'Producción'),
        ('shipping', 'Envío'),
    ]
    stage_index = {key: idx for idx, (key, _label) in enumerate(stage_defs)}
    dates = {'order': str((order or {}).get('order_date') or '').strip()}
    for item in history or []:
        stage = _local_guest_timeline_stage(item.get('status_key') or item.get('to_status'))
        action_date = str(item.get('date') or item.get('action_date') or '').strip()
        if stage != 'order' and action_date and not dates.get(stage):
            dates[stage] = action_date

    current_status = str((order or {}).get('current_status') or '').strip()
    current_stage = _local_guest_timeline_stage(current_status)
    current_idx = stage_index.get(current_stage, 0)
    is_completed = current_status == STATUS_KEYS.get('COMPLETED')
    # A legacy row may have no stage history. Use the current status timestamp as a
    # conservative fallback only for the current stage, never to invent past dates.
    if current_stage != 'order' and not dates.get(current_stage):
        dates[current_stage] = str((order or {}).get('status_updated_at') or '').strip()

    result = []
    for idx, (key, label) in enumerate(stage_defs):
        if is_completed:
            state = 'done' if idx <= current_idx else 'future'
        elif idx < current_idx:
            state = 'done'
        elif idx == current_idx:
            state = 'current'
        else:
            state = 'future'
        result.append({
            'key': key,
            'label_es': label,
            'date': dates.get(key) or '',
            'state': state,
        })
    return result


@tracking_bp.route('/guest/<token>/order/<item_kind>/<path:item_key>', methods=['GET'])
def local_guest_order_detail(token, item_kind, item_key):
    if cloud_mode_enabled():
        return _local_guest_error_page('cloud', 404)
    if not _local_guest_client_allowed():
        return _local_guest_error_page('network', 403)
    if item_kind not in {'w', 'o'}:
        return _local_guest_error_page('not_found', 404)
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=True)
        if state != 'active':
            return _local_guest_error_page(state, 410 if state in {'expired', 'revoked'} else 404)
        cursor = conn.cursor()
        if item_kind == 'w':
            workflow_sql, workflow_params = _guest_workflow_scope_sql(link=link, workflow_alias='w')
            cursor.execute(f'''
                SELECT w.workflow_number, w.order_number, w.product_name, w.product_code,
                       w.quantity, w.factory, w.production_type, w.expected_delivery_date,
                       w.current_status, w.status_updated_at, o.order_date, o.customer_name
                FROM workflows w INNER JOIN orders o ON o.order_number = w.order_number
                WHERE w.workflow_number = ? AND o.customer_name = ? AND o.status = 'ACTIVE'
                  AND {workflow_sql}
            ''', [item_key, link['customer_name'], *workflow_params])
        else:
            cursor.execute('''
                SELECT '' AS workflow_number, o.order_number, '' AS product_name, '' AS product_code,
                       '' AS quantity, '' AS factory, '' AS production_type, '' AS expected_delivery_date,
                       '' AS current_status, '' AS status_updated_at, o.order_date, o.customer_name
                FROM orders o
                WHERE o.order_number = ? AND o.customer_name = ? AND o.status = 'ACTIVE'
                  AND NOT EXISTS (SELECT 1 FROM workflows w WHERE w.order_number = o.order_number)
            ''', (item_key, link['customer_name']))
        row = cursor.fetchone()
        if not row:
            return _local_guest_error_page('not_found', 404)
        order = dict(row)
        status_key = str(order.get('current_status') or '')
        order['status_es'] = get_status_label(status_key, 'es') if status_key else 'Pedido'
        order['status_zh'] = get_status_label(status_key, 'zh_cn') if status_key else '订单'
        detail_prewarm_specs = []
        show_pdf_pages = bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1))
        order['images'] = _local_guest_media_for_order(
            conn, token, str(order.get('order_number') or ''), str(order.get('workflow_number') or ''),
            limit=None, prewarm_specs=detail_prewarm_specs, show_pdf_pages=show_pdf_pages
        )
        if detail_prewarm_specs:
            _schedule_visual_preview_batch(detail_prewarm_specs)
        history = []
        if order.get('workflow_number'):
            cursor.execute('''
                SELECT to_status, action_date FROM workflow_status_history
                WHERE workflow_number = ? ORDER BY action_date ASC, created_at ASC, id ASC
            ''', (order['workflow_number'],))
            for h in cursor.fetchall():
                key = str(h['to_status'] or '')
                history.append({'date': h['action_date'] or '',
                                'status_key': key,
                                'status_es': get_status_label(key, 'es') if key else '',
                                'status_zh': get_status_label(key, 'zh_cn') if key else ''})
        timeline_stages = _local_guest_horizontal_timeline(order, history)
    finally:
        conn.close()
    response = make_response(render_template(
        'tracking/guest_order.html', customer_name=link['customer_name'], order=order, history=history,
        timeline_stages=timeline_stages, token=token, expires_at_epoch=int(link['expires_at_epoch']),
        is_permanent=bool(int(link.get('is_permanent') or 0))
    ))
    return _local_guest_no_store(response)


@tracking_bp.route('/guest/<token>/media/<kind>/<int:file_id>', methods=['GET'])
def local_guest_media(token, kind, file_id):
    if cloud_mode_enabled() or not _local_guest_client_allowed():
        return '', 403
    if kind not in {'order', 'workflow'}:
        return '', 404
    from flask import send_file
    from .config import UPLOAD_FOLDER
    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active':
            return '', 410 if state in {'expired', 'revoked'} else 404
        cursor = conn.cursor()
        if kind == 'order':
            order_sql, order_params = _guest_order_scope_sql(link=link, order_alias='o')
            cursor.execute(f'''
                SELECT f.*, o.customer_name FROM order_files f
                INNER JOIN orders o ON o.order_number = f.order_number
                WHERE f.id = ? AND o.customer_name = ? AND o.status = 'ACTIVE'
                  AND {order_sql}
            ''', [file_id, link['customer_name'], *order_params])
        else:
            workflow_sql, workflow_params = _guest_workflow_scope_sql(link=link, workflow_alias='w')
            cursor.execute(f'''
                SELECT wf.*, o.customer_name FROM workflow_files wf
                INNER JOIN workflows w ON w.workflow_number = wf.workflow_number
                INNER JOIN orders o ON o.order_number = w.order_number
                WHERE wf.id = ? AND COALESCE(wf.is_deleted, 0) = 0
                  AND o.customer_name = ? AND o.status = 'ACTIVE'
                  AND {workflow_sql}
            ''', [file_id, link['customer_name'], *workflow_params])
        row = cursor.fetchone()
        if not row:
            return '', 404
        info = dict(row)
    finally:
        conn.close()

    if kind == 'order':
        original_name = info.get('original_filename') or info.get('file_name') or info.get('stored_filename') or ''
        if not _local_guest_image_file(original_name, info.get('file_type'), info.get('mime_type')):
            return '', 403
        stored = info.get('stored_filename') or os.path.basename(info.get('file_path') or '')
        path_dir = info.get('file_path') or ''
        # Mirror the existing order-file download compatibility rules for older rows.
        if '/' in str(stored) or '\\' in str(stored):
            path_dir = os.path.dirname(info.get('file_path') or '')
            stored = os.path.basename(info.get('file_path') or '')
        elif os.path.basename(path_dir) == stored:
            path_dir = os.path.dirname(path_dir)
        full_path = os.path.join(UPLOAD_FOLDER, path_dir, stored)
    else:
        original_name = info.get('file_name') or os.path.basename(info.get('file_path') or '')
        if not _local_guest_image_file(original_name, info.get('file_type'), info.get('file_type')):
            return '', 403
        full_path = os.path.join(UPLOAD_FOLDER, info.get('file_path') or '')
        if os.path.isdir(full_path) or not os.path.exists(full_path):
            candidate_dir = os.path.join(UPLOAD_FOLDER, str(info.get('file_path') or '').rstrip('/\\'))
            if os.path.isdir(candidate_dir):
                ext = os.path.splitext(original_name)[1].lower()
                candidates = [f for f in os.listdir(candidate_dir)
                              if os.path.isfile(os.path.join(candidate_dir, f)) and (not ext or f.lower().endswith(ext))]
                if candidates:
                    candidates.sort(key=lambda f: os.path.getmtime(os.path.join(candidate_dir, f)), reverse=True)
                    full_path = os.path.join(candidate_dir, candidates[0])
    if not os.path.isfile(full_path):
        return '', 404
    preview_requested = str(request.args.get('preview') or '').strip().lower() in {'1', 'true', 'yes'}
    if preview_requested:
        try:
            response = _image_preview_response(
                full_path, original_name or 'preview.jpg',
                cache_max_age=_local_guest_media_cache_seconds(link)
            )
            return _local_guest_cache_media_response(response, link)
        except Exception as exc:
            print(f'[WARN] guest image preview fallback to original: {exc}')
    return _local_guest_cache_media_response(
        send_file(full_path, as_attachment=False, conditional=True), link
    )


@tracking_bp.route('/guest/<token>/pdf-page/<kind>/<int:file_id>/<int:page_number>', methods=['GET'])
def local_guest_pdf_page(token, kind, file_id, page_number):
    # Same short-lived token/customer boundary as the guest image route.
    if cloud_mode_enabled() or not _local_guest_client_allowed():
        return '', 403
    if kind not in {'order', 'workflow'}:
        return '', 404

    conn = get_db()
    try:
        link, state = _local_guest_get_link(conn, token, touch=False)
        if state != 'active':
            return '', 410 if state in {'expired', 'revoked'} else 404
        if not bool(int(link.get('show_pdf_pages') if link.get('show_pdf_pages') is not None else 1)):
            return '', 404
        cursor = conn.cursor()
        if kind == 'order':
            order_sql, order_params = _guest_order_scope_sql(link=link, order_alias='o')
            cursor.execute(f'''
                SELECT f.*, o.customer_name
                FROM order_files f
                INNER JOIN orders o ON o.order_number = f.order_number
                WHERE f.id = ? AND o.customer_name = ? AND o.status = 'ACTIVE'
                  AND {order_sql}
            ''', [file_id, link['customer_name'], *order_params])
        else:
            workflow_sql, workflow_params = _guest_workflow_scope_sql(link=link, workflow_alias='w')
            cursor.execute(f'''
                SELECT wf.*, o.customer_name
                FROM workflow_files wf
                INNER JOIN workflows w ON w.workflow_number = wf.workflow_number
                INNER JOIN orders o ON o.order_number = w.order_number
                WHERE wf.id = ? AND COALESCE(wf.is_deleted, 0) = 0
                  AND o.customer_name = ? AND o.status = 'ACTIVE'
                  AND {workflow_sql}
            ''', [file_id, link['customer_name'], *workflow_params])
        row = cursor.fetchone()
    finally:
        conn.close()

    if not row:
        return '', 404
    info = dict(row)
    if kind == 'order':
        name = info.get('original_filename') or info.get('file_name') or info.get('stored_filename') or ''
        if not _media_is_pdf(name, info.get('file_type'), info.get('mime_type')):
            return '', 415
        path = _resolve_order_file_full_path(info)
    else:
        name = info.get('file_name') or os.path.basename(str(info.get('file_path') or ''))
        if not _media_is_pdf(name, info.get('file_type'), info.get('file_type')):
            return '', 415
        path = _resolve_workflow_file_full_path(info)

    try:
        payload, total = _pdf_preview_render_page(path, page_number, compact=str(request.args.get('preview') or '').lower() in {'1','true','yes'})
    except FileNotFoundError:
        return '', 404
    except ValueError:
        return '', 404
    except RuntimeError:
        return '', 500
    stem = os.path.splitext(os.path.basename(str(name or 'document.pdf')))[0]
    response = _pdf_page_response(
        payload, f'{stem}-page-{page_number}-of-{total}.jpg',
        cache_max_age=_local_guest_media_cache_seconds(link)
    )
    return _local_guest_cache_media_response(response, link)


@tracking_bp.route('/api/customer-reports/estimate', methods=['POST'])
@api_login_required
def api_customer_report_estimate():
    """Estimate report size/parts for the currently filtered rows."""
    try:
        _, items, report_format, language, image_source, image_count, image_order, pdf_attachment_mode = _parse_customer_report_request()
        conn = get_db()
        try:
            access_all, user_id = _customer_report_access_context()
            entries = load_report_entries(conn, items, access_all, user_id)
            if not entries:
                return jsonify({'success': False, 'error': '没有可导出的订单或无权限'}), 403
            _validate_customer_report_customer_limit(entries)
            attach_image_metadata(conn, entries, image_source, image_count, image_order, pdf_attachment_mode)
            estimate = estimate_customer_report(entries, report_format)
            image_total = sum(len(e.get('images') or []) for e in entries)
            pdf_attachment_total = sum(int(e.get('_pdf_attachment_count') or 0) for e in entries)
        finally:
            conn.close()
        return jsonify({
            'success': True,
            'order_count': len(entries),
            'image_count': image_total,
            'pdf_attachment_count': pdf_attachment_total,
            'estimated_bytes': estimate['estimated_bytes'],
            'estimated_parts': estimate['estimated_parts'],
            'estimated_files': estimate.get('estimated_files', estimate['estimated_parts']),
            'customer_count': estimate.get('customer_count', 1),
        })
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        print(f'[ERROR] customer report estimate failed: {exc}')
        return jsonify({'success': False, 'error': '预估报告大小失败'}), 500


@tracking_bp.route('/api/customer-reports/jobs', methods=['POST'])
@api_login_required
def api_customer_report_create_job():
    """Queue a customer report and return immediately."""
    try:
        _, items, report_format, language, image_source, image_count, image_order, pdf_attachment_mode = _parse_customer_report_request()
        conn = get_db()
        try:
            access_all, user_id = _customer_report_access_context()
            entries = load_report_entries(conn, items, access_all, user_id)
            if not entries:
                return jsonify({'success': False, 'error': '没有可导出的订单或无权限'}), 403
            _validate_customer_report_customer_limit(entries)
        finally:
            conn.close()

        job, deduplicated = _enqueue_customer_report(
            entries, report_format, language, image_source, image_count, image_order, pdf_attachment_mode, user_id
        )
        return jsonify({
            'success': True,
            'job': _customer_report_job_snapshot(job),
            'deduplicated': deduplicated,
            'queue_limits': {
                'workers': int(CUSTOMER_REPORT_WORKERS),
                'max_queued': int(CUSTOMER_REPORT_MAX_QUEUED),
                'max_active_per_user': int(CUSTOMER_REPORT_MAX_ACTIVE_PER_USER),
            },
        }), 200 if deduplicated else 202
    except CustomerReportQueueError as exc:
        return jsonify({'success': False, 'error': str(exc), 'code': exc.code}), exc.status_code
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        print(f'[ERROR] customer report queue failed: {exc}')
        return jsonify({'success': False, 'error': '加入客户报告队列失败'}), 500


@tracking_bp.route('/api/customer-reports/jobs', methods=['GET'])
@api_login_required
def api_customer_report_list_jobs():
    """List the current user's latest report jobs."""
    ctx = get_current_user_context()
    user_id = ctx.get('id')
    with _customer_report_jobs_lock:
        jobs = [dict(j) for j in _customer_report_jobs.values() if j.get('owner_user_id') == user_id]
    jobs.sort(key=lambda j: j.get('created_at') or '', reverse=True)
    public_jobs = []
    for job in jobs[:20]:
        item = _customer_report_job_snapshot(job)
        for f in item.get('files') or []:
            file_id = f.pop('file_id', None)
            if file_id:
                f['url'] = url_for('tracking_bp.api_customer_report_download', file_id=file_id)
        public_jobs.append(item)
    with _customer_report_jobs_lock:
        queued_count, processing_count, owner_active = _customer_report_queue_counts_locked(user_id)
    return jsonify({
        'success': True,
        'jobs': public_jobs,
        'queue': {
            'queued': queued_count,
            'processing': processing_count,
            'your_active': owner_active,
            'workers': int(CUSTOMER_REPORT_WORKERS),
            'max_queued': int(CUSTOMER_REPORT_MAX_QUEUED),
            'max_active_per_user': int(CUSTOMER_REPORT_MAX_ACTIVE_PER_USER),
        },
    })


@tracking_bp.route('/api/customer-reports/jobs/<job_id>', methods=['GET'])
@api_login_required
def api_customer_report_get_job(job_id):
    """Get one report job by id. This is the preferred polling endpoint for PyQt."""
    ctx = get_current_user_context()
    user_id = ctx.get('id')
    with _customer_report_jobs_lock:
        job = _customer_report_jobs.get(job_id)
        if not job or job.get('owner_user_id') != user_id:
            return jsonify({'success': False, 'error': '报告任务不存在或无权限', 'code': 'JOB_NOT_FOUND'}), 404
        public_job = _job_with_download_urls(job)
    return jsonify({'success': True, 'job': public_job})


@tracking_bp.route('/api/customer-reports/export', methods=['POST'])
@api_login_required
def api_customer_report_export():
    """Generate customer-safe reports from the exact rows selected in the UI."""
    try:
        _, items, report_format, language, image_source, image_count, image_order, pdf_attachment_mode = _parse_customer_report_request()
        conn = get_db()
        try:
            access_all, user_id = _customer_report_access_context()
            entries = load_report_entries(conn, items, access_all, user_id)
            if not entries:
                return jsonify({'success': False, 'error': '没有可导出的订单或无权限'}), 403
            _validate_customer_report_customer_limit(entries)
            prepared = prepare_customer_report_entries(conn, entries, image_source, image_count, language, image_order, pdf_attachment_mode)
            generated = build_customer_report_files(prepared, report_format, language)
        finally:
            conn.close()

        files = []
        for filename, mimetype, content in generated:
            cached = cache_report_file(filename, mimetype, content)
            files.append({
                'name': cached['filename'],
                'size': cached['size'],
                'url': url_for('tracking_bp.api_customer_report_download', file_id=cached['id']),
            })
        return jsonify({
            'success': True,
            'order_count': len(prepared),
            'file_count': len(files),
            'files': files,
        })
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except ImportError as exc:
        print(f'[ERROR] customer report dependency missing: {exc}')
        return jsonify({'success': False, 'error': '缺少报告组件，请先安装 requirements.txt'}), 500
    except Exception as exc:
        print(f'[ERROR] customer report export failed: {exc}')
        return jsonify({'success': False, 'error': '产生客户报告失败'}), 500


@tracking_bp.route('/api/customer-reports/download/<file_id>', methods=['GET'])
@api_login_required
def api_customer_report_download(file_id):
    """Download a previously generated report part from the short-lived cache."""
    from flask import send_file
    cached = get_cached_report(file_id)
    if not cached:
        return jsonify({'success': False, 'error': '报告不存在或已过期，请重新产生'}), 404
    mimetype = cached.get('mimetype') or 'application/octet-stream'
    filename = cached.get('filename') or 'report.bin'
    inline_requested = str(request.args.get('inline', '')).strip().lower() in {'1', 'true', 'yes'}
    is_pdf = mimetype == 'application/pdf' or str(filename).lower().endswith('.pdf')
    return send_file(
        cached['path'],
        mimetype=mimetype,
        as_attachment=not (inline_requested and is_pdf),
        download_name=filename,
    )


@tracking_bp.route('/api/orders/<order_number>', methods=['GET'])
@api_login_required
def api_order_detail(order_number):
    """獲取訂單詳情API"""
    if cloud_mode_enabled():
        try:
            order = _cloud_provider_call('get_order_detail', order_number)
        except Exception as exc:
            return jsonify({'success': False, 'error': f'雲端訂單資料載入失敗: {exc}'}), 503
        if not order:
            return jsonify({'success': False, 'error': '訂單不存在', 'code': 'NOT_FOUND'}), 404
        return jsonify({'success': True, 'data': order})

    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '訂單不存在', 'code': 'NOT_FOUND'}), 404
    
    order = dict(order)
    
    # 獲取狀態歷史（兼容：status_history 表可能不存在）
    try:
        cursor.execute('''
            SELECT * FROM status_history 
            WHERE order_number = ?
            ORDER BY action_date ASC, created_at ASC
        ''', (order_number,))
        history = [dict(row) for row in cursor.fetchall()]
    except Exception:
        history = []
    
    order['history'] = history
    
    conn.close()
    
    return jsonify({'success': True, 'data': order})

@tracking_bp.route('/api/workflows', methods=['POST'])
@api_login_required
def api_create_workflow():
    """建立业务流程"""
    data = request.get_json()
    order_number = data.get('order_number', '').strip()
    
    if not order_number:
        return jsonify({'success': False, 'error': '缺少订单号'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # ===== 檢查訂單號是否已解鎖 =====
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单号不存在，请先由管理员建立'}), 404
    
    if order['status'] != 'ACTIVE':
        conn.close()
        return jsonify({'success': False, 'error': '订单号未解锁，请联系管理员'}), 400
    
    # 檢查訂單是否已完成或已取消（鎖定）
    if _normalize_lock_flag(order['is_locked']):
        conn.close()
        if order['status'] == 'CANCELLED':
            return jsonify({'success': False, 'error': '該訂單已取消，無法再建立流程。如需操作請先解除取消。'}), 400
        return jsonify({'success': False, 'error': '该订单已标记完成，无法再创建流程。如需操作请先解除锁定。'}), 400
    
    # 業務員權限檢查
    if not can_access_visibility(order['visibility']):
        conn.close()
        return jsonify({'success': False, 'error': '无权限操作此订单号'}), 403
    # ===== 檢查結束 =====
    
    # ===== 生成 workflow_number（帶併發控制） =====
    import os
    max_retries = 3
    retry_count = 0
    workflow_number = None
    
    while retry_count < max_retries:
        # 查詢當前最大序號
        cursor.execute('''
            SELECT MAX(CAST(SUBSTR(workflow_number, LENGTH(?) + 2) AS INTEGER)) as max_seq
            FROM workflows
            WHERE order_number = ?
        ''', (order_number, order_number))
        
        row = cursor.fetchone()
        max_seq = row['max_seq'] if row and row['max_seq'] else 0
        workflow_number = f"{order_number}-{max_seq + 1}"
        
        # 檢查是否已存在（防止併發重複）
        cursor.execute('SELECT workflow_number FROM workflows WHERE workflow_number = ?', (workflow_number,))
        if cursor.fetchone():
            # 如果已存在，重試
            retry_count += 1
            continue
        
        # 創建文件夾（使用 UPLOAD_FOLDER 確保路徑正確）
        from .config import UPLOAD_FOLDER
        folder_path = f'workflows/{workflow_number}/'
        full_folder_path = os.path.join(UPLOAD_FOLDER, folder_path)
        os.makedirs(full_folder_path, exist_ok=True)
        
        # 嘗試插入（使用 UNIQUE 約束防止重複）
        user_id = session.get('user_id')
        try:
            cursor.execute('''
                INSERT INTO workflows (
                    workflow_number,
                    order_number,
                    product_name,
                    product_code,
                    quantity,
                    factory,
                    production_type,
                    expected_delivery_date,
                    current_status,
                    created_by_id,
                    handler_id,
                    folder_path,
                    notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                workflow_number,
                order_number,
                data.get('product_name'),
                data.get('product_code'),
                data.get('quantity'),
                data.get('factory'),
                data.get('production_type'),
                data.get('expected_delivery_date'),
                STATUS_KEYS['NEW_ORDER'],
                user_id,
                user_id,  # 預設自己是負責人
                folder_path,
                data.get('notes', '')
            ))
            # 插入成功，跳出循環
            break
        except sqlite3.IntegrityError:
            # 如果違反 UNIQUE 約束，重試
            retry_count += 1
            if retry_count >= max_retries:
                conn.close()
                return jsonify({'success': False, 'error': '流程号已存在，请重试'}), 409
            continue
    
    if not workflow_number:
        conn.close()
        return jsonify({'success': False, 'error': '生成流程号失败，请重试'}), 500
    # ===== 生成結束 =====
    # ===== INSERT 結束 =====
    
    # ===== 記錄到 workflow_status_history =====
    if should_insert_workflow_history(cursor, workflow_number, STATUS_KEYS['NEW_ORDER']):
        cursor.execute('''
            INSERT INTO workflow_status_history (
                workflow_number,
                order_number,
                to_status,
                action_date,
                operator_id
            ) VALUES (?, ?, ?, ?, ?)
        ''', (
            workflow_number,
            order_number,
            STATUS_KEYS['NEW_ORDER'],
            date.today().isoformat(),
            user_id
        ))
    # ===== 記錄結束 =====
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': '业务流程已建立',
        'data': {
            'workflow_number': workflow_number,
            'order_number': order_number
        }
    })

@tracking_bp.route('/api/workflows/<workflow_number>', methods=['GET'])
@api_login_required
def api_workflow_detail(workflow_number):
    """獲取流程詳情API（包含時間軸）"""
    if cloud_mode_enabled():
        try:
            workflow = _cloud_provider_call('get_workflow_detail', workflow_number)
        except Exception as exc:
            return jsonify({'success': False, 'error': f'雲端流程資料載入失敗: {exc}'}), 503
        if not workflow:
            return jsonify({'success': False, 'error': '流程不存在', 'code': 'NOT_FOUND'}), 404
        return jsonify({'success': True, 'data': workflow})

    conn = get_db()
    cursor = conn.cursor()
    
    # 獲取流程基本信息（包含訂單備註和流程備註）
    cursor.execute('''
        SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility,
               o.notes as order_notes,
               COALESCE(u.real_name, u.display_name, u.username) as handler_name
        FROM workflows w
        INNER JOIN orders o ON w.order_number = o.order_number
        LEFT JOIN users u ON w.handler_id = u.id
        WHERE w.workflow_number = ?
    ''', (workflow_number,))
    workflow = cursor.fetchone()
    
    if not workflow:
        conn.close()
        return jsonify({'success': False, 'error': '流程不存在', 'code': 'NOT_FOUND'}), 404
    
    workflow = dict(workflow)
    # 將流程備註和訂單備註分開
    workflow['workflow_notes'] = workflow.get('notes', '')  # 流程備註（業務員的）
    workflow['order_notes'] = workflow.get('order_notes', '')  # 訂單備註（主管的）
    
    # 獲取狀態歷史（從 workflow_status_history 表）
    cursor.execute('''
        SELECT h.*, 
               COALESCE(u.real_name, u.display_name, u.username) as operator_name
        FROM workflow_status_history h
        LEFT JOIN users u ON h.operator_id = u.id
        WHERE h.workflow_number = ?
        ORDER BY h.action_date ASC, h.created_at ASC
    ''', (workflow_number,))
    history = []
    for row in cursor.fetchall():
        hist_item = dict(row)
        # 為了兼容前端，將字段名映射為與 status_history 相同的格式
        hist_item['to_status'] = hist_item.get('to_status', '')
        hist_item['from_status'] = hist_item.get('from_status', '')
        hist_item['action_date'] = hist_item.get('action_date', '')
        hist_item['notes'] = hist_item.get('notes', '')
        hist_item['operator'] = hist_item.get('operator_name', '')
        history.append(hist_item)
    
    # 若流程表狀態與最後一次「實際操作」不一致，以 created_at/id 最後一筆修正。
    latest_for_state = None
    if history:
        cursor.execute('''
            SELECT * FROM workflow_status_history
            WHERE workflow_number = ?
            ORDER BY created_at DESC, id DESC
            LIMIT 1
        ''', (workflow_number,))
        latest_row = cursor.fetchone()
        latest_for_state = dict(latest_row) if latest_row else None
    if latest_for_state:
        last_history = latest_for_state
        last_status = last_history.get('to_status')
        if last_status and workflow.get('current_status') != last_status:
            try:
                cursor.execute('''
                    UPDATE workflows
                    SET current_status = ?,
                        status_updated_at = ?,
                        status_days = 0,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE workflow_number = ?
                ''', (last_status, last_history.get('action_date'), workflow_number))
                workflow['current_status'] = last_status
                conn.commit()
            except Exception as e:
                print(f'警告：同步流程狀態失敗: {e}')

    workflow['history'] = history
    latest_history = get_latest_workflow_history(cursor, workflow_number)
    workflow['last_history_id'] = latest_history.get('id') if latest_history else None
    workflow['order_number'] = workflow.get('order_number', '')
    
    # 计算状态灯（供前端 createOrderRow 使用）
    light_obj = {
        'order_number': workflow.get('order_number'),
        'order_date': workflow.get('order_date'),
        'expected_delivery_date': workflow.get('expected_delivery_date'),
        'status_updated_at': workflow.get('status_updated_at'),
        'current_status': workflow.get('current_status'),
        'last_status_change_date': workflow.get('status_updated_at')
    }
    workflow['status_light'] = calculate_status_light(light_obj)
    
    conn.close()
    
    return jsonify({'success': True, 'data': workflow})

@tracking_bp.route('/api/workflows/<workflow_number>/history/<int:history_id>', methods=['PUT'])
@api_login_required
def api_update_workflow_history(workflow_number, history_id):
    """更新工作流歷史備註"""
    try:
        data = request.get_json() or {}
        notes_provided = 'notes' in data
        notes = data.get('notes', '')
        action_date = data.get('action_date')
        normalized_action_date = None
        if action_date is not None:
            try:
                normalized_action_date = normalize_action_date(action_date)
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc), 'code': 'INVALID_DATE'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
        workflow = cursor.fetchone()
        if not workflow:
            conn.close()
            return jsonify({'success': False, 'error': '工作流不存在', 'code': 'NOT_FOUND'}), 404
        workflow = dict(workflow)

        has_permission, error_msg = check_permission('workflow', workflow_number, 'edit')
        if not has_permission:
            conn.close()
            return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403

        current_ctx = get_current_user_context()
        user_id = current_ctx.get('id')
        if not can_manage_by_owner(workflow['handler_id'], user_id=user_id):
            conn.close()
            return jsonify({'success': False, 'error': '無權限', 'code': 'FORBIDDEN'}), 403

        cursor.execute('SELECT * FROM workflow_status_history WHERE id = ? AND workflow_number = ?', (history_id, workflow_number))
        history = cursor.fetchone()
        if not history:
            conn.close()
            return jsonify({'success': False, 'error': '歷史記錄不存在', 'code': 'NOT_FOUND'}), 404
        history = dict(history)

        # 若該步驟為管理員操作，僅管理員可修改備註（權限配置驅動）
        operator_id = history.get('operator_id')
        if operator_id:
            cursor.execute('SELECT role FROM users WHERE id = ?', (operator_id,))
            operator_row = cursor.fetchone()
            operator_role = operator_row['role'] if operator_row else None
            if is_admin(role=operator_role) and not is_admin():
                conn.close()
                return jsonify({'success': False, 'error': '此步驟為管理員操作，無法修改備註', 'code': 'FORBIDDEN'}), 403

        update_fields = []
        params = []
        if normalized_action_date is not None:
            update_fields.append('action_date = ?')
            params.append(normalized_action_date)
        if notes_provided:
            update_fields.append('notes = ?')
            params.append(notes)
        if update_fields:
            params.extend([history_id, workflow_number])
            update_sql = f"UPDATE workflow_status_history SET {', '.join(update_fields)} WHERE id = ? AND workflow_number = ?"
            cursor.execute(update_sql, params)

        paired_auto_history_id = None
        if normalized_action_date is not None:
            auto_complete_note = '系統自動：已全部出貨後轉為已完成'
            history_status = history.get('to_status')

            # 「全部出貨」和系統自動產生的「已完成」是同一次業務操作。
            # 修改實際出貨日期時，兩筆歷史必須保持同一天。
            if history_status == STATUS_KEYS['ALL_SHIPPED']:
                cursor.execute('''
                    SELECT id
                    FROM workflow_status_history
                    WHERE workflow_number = ?
                      AND id > ?
                      AND to_status = ?
                      AND notes = ?
                    ORDER BY id ASC
                    LIMIT 1
                ''', (workflow_number, history_id, STATUS_KEYS['COMPLETED'], auto_complete_note))
                paired_row = cursor.fetchone()
                if paired_row:
                    paired_auto_history_id = paired_row['id'] if isinstance(paired_row, sqlite3.Row) else paired_row[0]
                    cursor.execute('''
                        UPDATE workflow_status_history
                        SET action_date = ?
                        WHERE id = ? AND workflow_number = ?
                    ''', (normalized_action_date, paired_auto_history_id, workflow_number))

            # 只有當被修改的是目前最後一次實際操作（或其自動完成配對）時，
            # 才同步 workflows.status_updated_at；修改較早的部分出貨不能改掉目前狀態日期。
            cursor.execute('''
                SELECT id
                FROM workflow_status_history
                WHERE workflow_number = ?
                ORDER BY created_at DESC, id DESC
                LIMIT 1
            ''', (workflow_number,))
            latest_row = cursor.fetchone()
            latest_id = (latest_row['id'] if isinstance(latest_row, sqlite3.Row) else latest_row[0]) if latest_row else None
            if latest_id in {history_id, paired_auto_history_id}:
                try:
                    parsed_date = datetime.strptime(normalized_action_date, '%Y-%m-%d').date()
                    status_days = max(0, (date.today() - parsed_date).days)
                except Exception:
                    status_days = 0
                cursor.execute('''
                    UPDATE workflows
                    SET status_updated_at = ?, status_days = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE workflow_number = ?
                ''', (normalized_action_date, status_days, workflow_number))

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'data': {
                'history_id': history_id,
                'workflow_number': workflow_number,
                'notes': notes if notes_provided else history.get('notes', ''),
                'action_date': normalized_action_date if normalized_action_date is not None else history.get('action_date', ''),
                'paired_auto_history_id': paired_auto_history_id
            }
        })
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@tracking_bp.route('/api/workflows', methods=['GET'])
@api_login_required
def api_list_workflows():
    """獲取同訂單下的所有工作流（用於工作流卡片顯示）"""
    order_number = request.args.get('order')
    if not order_number:
        return jsonify({'success': False, 'error': '缺少訂單號參數'}), 400

    if cloud_mode_enabled():
        current_ctx = get_current_user_context()
        try:
            workflows = _cloud_provider_call(
                'get_workflows_for_order', order_number,
                current_ctx.get('role', 'viewer'), current_ctx.get('id')
            ) or []
        except Exception as exc:
            return jsonify({'success': False, 'error': f'雲端流程列表載入失敗: {exc}'}), 503
        return jsonify({
            'success': True,
            'data': {'order_number': order_number, 'workflows': list(workflows)}
        })
    
    conn = get_db()
    cursor = conn.cursor()
    current_ctx = get_current_user_context()
    current_user_id = current_ctx.get('id')
    current_role = current_ctx.get('role', 'viewer')
    
    # 檢查訂單是否存在
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '訂單不存在'}), 404
    
    # 根據權限查詢工作流
    if is_admin():
        # 主管：顯示該訂單下的所有工作流
        cursor.execute('''
            SELECT w.*, 
                   COALESCE(u.real_name, u.display_name, u.username) as handler_name
            FROM workflows w
            LEFT JOIN users u ON w.handler_id = u.id
            WHERE w.order_number = ?
            ORDER BY w.workflow_number ASC
        ''', (order_number,))
    else:
        # 業務員：只顯示自己負責的工作流
        cursor.execute('''
            SELECT w.*, 
                   COALESCE(u.real_name, u.display_name, u.username) as handler_name
            FROM workflows w
            LEFT JOIN users u ON w.handler_id = u.id
            WHERE w.order_number = ? AND w.handler_id = ?
            ORDER BY w.workflow_number ASC
        ''', (order_number, current_user_id))
    
    workflows = []
    for row in cursor.fetchall():
        wf = dict(row)
        workflows.append({
            'workflow_number': wf['workflow_number'],
            'order_number': wf['order_number'],
            'handler_id': wf['handler_id'],
            'handler_name': wf['handler_name'],
            'current_status': wf['current_status'],
            'status_days': wf.get('status_days', 0)
        })
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'order_number': order_number,
            'workflows': workflows
        }
    })

@tracking_bp.route('/api/workflows/<workflow_number>', methods=['PUT'])
@api_login_required
@require_permission('workflow', 'edit', resource_id_param='workflow_number')
def api_update_workflow(workflow_number):
    """更新流程信息API（支持备注更新）"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '無效的請求數據'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 檢查流程是否存在
        cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
        workflow = cursor.fetchone()
        
        if not workflow:
            conn.close()
            return jsonify({'success': False, 'error': '流程不存在'}), 404
        
        workflow = dict(workflow)
        current_ctx = get_current_user_context()
        current_user_id = current_ctx.get('id')
        current_role = current_ctx.get('role', 'viewer')
        
        # 權限檢查：僅流程負責人或 ADMIN 可更新備註，且鎖單時僅 ADMIN 可改
        if 'notes' in data:
            if not can_manage_by_owner(workflow.get('handler_id'), user_id=current_user_id):
                conn.close()
                return jsonify({'success': False, 'error': '無權限修改此流程的備註'}), 403
            is_locked = False
            try:
                cursor.execute('SELECT is_locked FROM orders WHERE order_number = ?', (workflow.get('order_number'),))
                order_row = cursor.fetchone()
                if order_row:
                    is_locked = bool(dict(order_row).get('is_locked'))
            except Exception:
                is_locked = False
            if is_locked and not is_admin(current_role):
                conn.close()
                return jsonify({'success': False, 'error': '訂單已被鎖定，無法修改備註'}), 403
        
        # 支持部分更新：只更新传入的字段
        update_fields = []
        update_values = []
        
        if 'notes' in data:
            update_fields.append('notes = ?')
            update_values.append(data['notes'])
        
        if 'product_name' in data:
            update_fields.append('product_name = ?')
            update_values.append(data['product_name'])
        if 'product_code' in data:
            update_fields.append('product_code = ?')
            update_values.append(data['product_code'])
        if 'quantity' in data:
            update_fields.append('quantity = ?')
            update_values.append(data['quantity'])
        if 'factory' in data:
            update_fields.append('factory = ?')
            update_values.append(data['factory'])
        if 'production_type' in data:
            update_fields.append('production_type = ?')
            update_values.append(data['production_type'])
        if 'expected_delivery_date' in data:
            update_fields.append('expected_delivery_date = ?')
            update_values.append(data['expected_delivery_date'])
        
        if not update_fields:
            conn.close()
            return jsonify({'success': False, 'error': '沒有需要更新的字段'}), 400
        
        # 添加 updated_at
        update_fields.append('updated_at = CURRENT_TIMESTAMP')
        
        # 執行更新
        update_values.append(workflow_number)
        update_sql = f"UPDATE workflows SET {', '.join(update_fields)} WHERE workflow_number = ?"
        cursor.execute(update_sql, update_values)
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '流程已更新',
            'data': {
                'workflow_number': workflow_number
            }
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"API Error: {error_detail}")
        
        if 'conn' in locals():
            conn.close()
        
        return jsonify({
            'success': False,
            'error': f'更新失敗：{str(e)}'
        }), 500


@tracking_bp.route('/api/workflows/<workflow_number>/transfer', methods=['POST'])
@admin_required
def api_transfer_workflow_handler(workflow_number):
    """轉移流程負責人（主管專用）"""
    data = request.get_json() or {}
    to_user_id = data.get('to_user_id')
    if not to_user_id:
        return jsonify({'success': False, 'error': '缺少轉移目標'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
    workflow = cursor.fetchone()
    if not workflow:
        conn.close()
        return jsonify({'success': False, 'error': '流程不存在'}), 404

    has_permission, error_msg = check_permission('workflow', workflow_number, 'edit')
    if not has_permission:
        conn.close()
        return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403

    allowed_roles = [
        role_key for role_key, perms in PERMISSION_MATRIX.items()
        if isinstance(perms, dict) and 'workflow' in perms and 'edit' in (perms.get('workflow') or [])
    ]
    if not allowed_roles:
        conn.close()
        return jsonify({'success': False, 'error': '未配置可轉移的角色'}), 400
    role_placeholders = ','.join(['?'] * len(allowed_roles))
    cursor.execute(
        f'SELECT * FROM users WHERE id = ? AND role IN ({role_placeholders}) AND status = ?',
        [to_user_id, *allowed_roles, 'active']
    )
    target_user = cursor.fetchone()
    if not target_user:
        conn.close()
        return jsonify({'success': False, 'error': '目标用戶不存在或不可轉移'}), 400

    cursor.execute('UPDATE workflows SET handler_id = ?, updated_at = CURRENT_TIMESTAMP WHERE workflow_number = ?', (to_user_id, workflow_number))

    # === 通知：三方通知 ===
    try:
        order_num = workflow['order_number'] or ''
        admin_user_id = session.get('user_id')
        admin_name = session.get('display_name') or session.get('username', '管理员')
        old_handler_id = workflow['handler_id']
        new_handler_name = target_user['real_name'] or target_user['display_name'] or target_user['username']

        # 1) 新负责人：流程已转让给你
        _create_notification(
            cursor, to_user_id, 'workflow_transferred',
            f'流程 {workflow_number} 已转让给你',
            f'订单：{order_num}，由 {admin_name} 转让',
            order_num, workflow_number, 'normal', 30
        )

        # 2) 原负责人：流程已转让给别人（排除：新旧相同 / 原负责人为空）
        if old_handler_id and str(old_handler_id) != str(to_user_id):
            _create_notification(
                cursor, old_handler_id, 'workflow_transferred',
                f'流程 {workflow_number} 已转让给 {new_handler_name}',
                f'订单：{order_num}，由 {admin_name} 操作',
                order_num, workflow_number, 'normal', 30
            )

        # 3) 执行转让的管理员：确认记录（排除：管理员自己就是新/旧负责人）
        if admin_user_id and str(admin_user_id) != str(to_user_id) and str(admin_user_id) != str(old_handler_id or ''):
            _create_notification(
                cursor, admin_user_id, 'workflow_transferred',
                f'您已将流程 {workflow_number} 转让给 {new_handler_name}',
                f'订单：{order_num}',
                order_num, workflow_number, 'low', 14
            )
    except Exception as e:
        print(f"[WARN] 创建转让通知失败: {e}")

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': '流程已轉移',
        'data': {
            'workflow_number': workflow_number,
            'handler_id': to_user_id,
            'handler_name': target_user['real_name'] or target_user['display_name'] or target_user['username']
        }
    })

@tracking_bp.route('/api/workflows/<workflow_number>/files', methods=['GET'])
@login_required
def api_get_workflow_files(workflow_number):
    # Render uses the same UI but has no LAN file system. Published cloud media is
    # served by the public-share/B2 path, not by this local attachment API.
    if cloud_mode_enabled():
        return jsonify({'success': True, 'data': {
            'workflow_number': workflow_number, 'files': [], 'total': 0
        }})
    # visual=1 always re-queries current SQLite rows and current source-file versions.
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
        workflow = cursor.fetchone()
        if not workflow:
            conn.close()
            return jsonify({'success': False, 'error': '工作流不存在', 'code': 'NOT_FOUND'}), 404
        workflow = dict(workflow)
        current_ctx = get_current_user_context()
        user_id = current_ctx.get('id')
        if not can_manage_by_owner(workflow['handler_id'], user_id=user_id):
            conn.close()
            return jsonify({'success': False, 'error': '无权限', 'code': 'FORBIDDEN'}), 403

        cursor.execute("""
            SELECT id, file_name, file_path, file_size, file_type, uploaded_by_id, uploaded_at
            FROM workflow_files
            WHERE workflow_number = ? AND is_deleted = 0
            ORDER BY uploaded_at DESC
        """, (workflow_number,))
        files = []
        preview_specs = []
        include_visual_meta = str(request.args.get('visual') or '').strip().lower() in {'1', 'true', 'yes'}
        for row in cursor.fetchall():
            uploader_name = get_user_display_name(row['uploaded_by_id'], conn) if row['uploaded_by_id'] else None
            row_info = dict(row)
            is_pdf = _media_is_pdf(row['file_name'], row_info.get('file_type'), row_info.get('file_type'))
            pdf_page_count = 0
            media_version = ''
            if include_visual_meta:
                visual_path = _resolve_workflow_file_full_path(row_info)
                if os.path.isfile(visual_path):
                    media_version = _file_version_token(visual_path)
                    if is_pdf:
                        pdf_page_count = _pdf_preview_page_count(visual_path)
                        if pdf_page_count:
                            preview_specs.append({'type': 'pdf', 'path': visual_path, 'pages': pdf_page_count})
                    elif _local_guest_image_file(row['file_name'], row_info.get('file_type'), row_info.get('file_type')):
                        preview_specs.append({'type': 'image', 'path': visual_path})
            files.append({
                'id': row['id'], 'file_name': row['file_name'], 'file_size': row['file_size'],
                'file_type': row['file_type'], 'uploaded_by_id': row['uploaded_by_id'],
                'uploaded_by_name': uploader_name, 'uploaded_at': row['uploaded_at'],
                'media_type': 'pdf' if is_pdf else 'file', 'pdf_page_count': pdf_page_count,
                'media_version': media_version,
            })
        conn.close()
        if include_visual_meta and preview_specs:
            _schedule_visual_preview_batch(preview_specs)
        response = jsonify({'success': True, 'data': {
            'workflow_number': workflow_number, 'files': files, 'total': len(files)
        }})
        if include_visual_meta:
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        return response
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        print(f"[ERROR] 获取工作流文件失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API: 上传工作流文件 ====================

@tracking_bp.route('/api/workflows/<workflow_number>/files/upload', methods=['POST'])
@login_required
@require_permission('workflow_file', 'upload', resource_id_param='workflow_number')
def api_upload_workflow_files(workflow_number):
    """上传工作流文件（业务附件）"""
    import os
    import uuid
    from datetime import datetime
    from .config import UPLOAD_FOLDER

    conn = get_db()
    cursor = conn.cursor()

    # 检查工作流是否存在
    cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
    workflow = cursor.fetchone()
    if not workflow:
        conn.close()
        return jsonify({'success': False, 'error': '工作流不存在'}), 404

    workflow = dict(workflow)

    # 权限检查：只有负责人和主管可以上传
    current_ctx = get_current_user_context()
    user_id = current_ctx.get('id')
    if not can_manage_by_owner(workflow.get('handler_id'), user_id=user_id):
        conn.close()
        return jsonify({'success': False, 'error': '无权限'}), 403

    if 'files' not in request.files:
        conn.close()
        return jsonify({'success': False, 'error': '没有选择文件'}), 400

    files = request.files.getlist('files')
    if not files or all(not f.filename for f in files):
        conn.close()
        return jsonify({'success': False, 'error': '没有选择文件'}), 400

    workflow_dir = os.path.join(UPLOAD_FOLDER, 'workflows', workflow_number)
    os.makedirs(workflow_dir, exist_ok=True)

    uploaded_files_info = []
    uploaded_by_id = session.get('user_id')

    for file in files:
        if file and file.filename:
            original_filename = file.filename
            file_ext = os.path.splitext(original_filename)[1].lower()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_id = str(uuid.uuid4())[:8]
            stored_filename = f"{timestamp}_{unique_id}{file_ext}"

            file_path = os.path.join(workflow_dir, stored_filename)
            file.save(file_path)

            file_size = os.path.getsize(file_path)
            relative_path = f"workflows/{workflow_number}/{stored_filename}"
            file_type = file.content_type or 'application/octet-stream'

            cursor.execute('''
                INSERT INTO workflow_files (
                    workflow_number,
                    file_name,
                    file_path,
                    file_size,
                    file_type,
                    uploaded_by_id,
                    uploaded_at
                ) VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (
                workflow_number,
                original_filename,
                relative_path,
                file_size,
                file_type,
                uploaded_by_id
            ))

            uploaded_files_info.append({
                'file_name': original_filename,
                'stored_name': stored_filename,
                'size': file_size
            })

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': f'已上传 {len(uploaded_files_info)} 个文件',
        'data': {
            'workflow_number': workflow_number,
            'uploaded_files': uploaded_files_info
        }
    })


@tracking_bp.route('/api/workflows/files/<int:file_id>', methods=['DELETE'])
@login_required
@require_permission('workflow_file', 'delete', resource_id_param='file_id')
def api_delete_workflow_file(file_id):
    """软删除工作流文件"""
    conn = get_db()
    cursor = conn.cursor()

    # 查询文件
    cursor.execute('SELECT * FROM workflow_files WHERE id = ?', (file_id,))
    file_info = cursor.fetchone()
    if not file_info:
        conn.close()
        return jsonify({'success': False, 'error': '文件不存在'}), 404

    file_info = dict(file_info)

    # 查询工作流
    cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (file_info['workflow_number'],))
    workflow = cursor.fetchone()
    if not workflow:
        conn.close()
        return jsonify({'success': False, 'error': '工作流不存在'}), 404

    workflow = dict(workflow)

    # 权限检查：只有负责人和主管可以删除
    current_ctx = get_current_user_context()
    user_id = current_ctx.get('id')
    if not can_manage_by_owner(workflow.get('handler_id'), user_id=user_id):
        conn.close()
        return jsonify({'success': False, 'error': '无权限'}), 403

    # 软删除
    cursor.execute('''
        UPDATE workflow_files
        SET is_deleted = 1,
            deleted_by_id = ?,
            deleted_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (user_id, file_id))

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': '文件已删除'
    })


# ==================== API 2: 更新工作流状态（快速操作）====================

@tracking_bp.route('/api/workflows/<workflow_number>/status', methods=['PUT'])
@login_required
@require_permission('workflow', 'edit', resource_id_param='workflow_number')
def api_update_workflow_status(workflow_number):
    """更新工作流状态（底部快速操作按钮使用）"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '无效的请求数据'}), 400
        
        action = data.get('action')
        expected_status = data.get('expected_status')
        notes = data.get('notes', '')
        expected_history_id = data.get('expected_history_id')
        try:
            action_date = normalize_action_date(data.get('action_date'))
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc), 'code': 'INVALID_ACTION_DATE'}), 400
        
        if not action:
            return jsonify({'success': False, 'error': '缺少 action 参数'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 查询工作流
        cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
        workflow = cursor.fetchone()
        
        if not workflow:
            conn.close()
            return jsonify({'success': False, 'error': '工作流不存在', 'code': 'NOT_FOUND'}), 404
        
        workflow = dict(workflow)

        # 檢查訂單是否鎖定（已完成或已取消）
        cursor.execute('SELECT is_locked, status FROM orders WHERE order_number = ?', (workflow.get('order_number'),))
        order_row = cursor.fetchone()
        if order_row and _normalize_lock_flag(order_row['is_locked']):
            conn.close()
            if order_row['status'] == 'CANCELLED':
                return jsonify({'success': False, 'error': '該訂單已取消，無法更新流程狀態', 'code': 'ORDER_CANCELLED'}), 403
            return jsonify({'success': False, 'error': '該訂單已完成鎖定，無法更新流程狀態', 'code': 'ORDER_LOCKED'}), 403

        if not expected_status:
            conn.close()
            return jsonify({'success': False, 'error': '缺少 expected_status 参数', 'code': 'MISSING_EXPECTED_STATUS'}), 400
        
        # 特殊處理：出貨一部份可在 PARTIAL_SHIPPED 狀態下重複執行，用來記錄每次出貨
        is_repeat_partial_ship = (action == 'ship_partial' and workflow.get('current_status') == 'PARTIAL_SHIPPED')
        
        if not is_repeat_partial_ship and workflow.get('current_status') != expected_status:
            conn.close()
            return jsonify({
                'success': False,
                'error': '当前状态已变更，请刷新后重试',
                'code': 'STATUS_CHANGED',
                'current_status': workflow.get('current_status')
            }), 409
        if not is_repeat_partial_ship and expected_history_id:
            latest_history = get_latest_workflow_history(cursor, workflow_number)
            latest_id = latest_history.get('id') if latest_history else None
            if str(latest_id) != str(expected_history_id):
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '状态已被更新，请刷新后重试',
                    'code': 'HISTORY_CHANGED',
                    'latest_history_id': latest_id
                }), 409
        
        # 权限检查：只有负责人和主管可以操作
        current_ctx = get_current_user_context()
        user_id = current_ctx.get('id')
        
        if not can_manage_by_owner(workflow['handler_id'], user_id=user_id):
            conn.close()
            return jsonify({'success': False, 'error': '无权限', 'code': 'FORBIDDEN'}), 403
        
        # 从 QUICK_ACTIONS_MAP 获取下一个状态
        from .status_definitions import QUICK_ACTIONS_MAP
        
        if action not in QUICK_ACTIONS_MAP:
            conn.close()
            return jsonify({'success': False, 'error': f'无效的 action: {action}'}), 400
        
        action_config = QUICK_ACTIONS_MAP[action]
        if isinstance(action_config, dict):
            new_status = action_config.get('next', '')
            action_label = action_config.get('label', '')
        else:
            new_status = action_config
            action_label = action
        if not new_status:
            conn.close()
            return jsonify({'success': False, 'error': '無效的狀態轉換'}), 400
        
        old_status = workflow['current_status']
        
        # 获取状态的显示标签（简体中文）
        old_status_label = get_status_label(old_status, 'zh_cn')
        new_status_label = get_status_label(new_status, 'zh_cn')
        # action_label already set above
        
        # 插入状态历史（不自动填备注）
        history_notes = notes if notes else ''

        # ship_partial 強制每次都寫歷史，用來記錄每次出貨
        force_insert = (action == 'ship_partial')
        transition = apply_workflow_status_transition(
            cursor,
            workflow,
            workflow_number,
            new_status,
            action_date,
            user_id,
            history_notes,
            update_current_status=not is_repeat_partial_ship,
            force_insert=force_insert
        )
        final_status = transition['final_status']
        latest_history_id = transition['latest_history_id']
        
        # 记录操作日志（可选，如果有 operation_logs 表）
        try:
            cursor.execute('''
                INSERT INTO operation_logs 
                (user_id, operation_type, operation_desc, order_number, workflow_number)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                user_id,
                'workflow_status_update',
                f'工作流状态更新：{old_status_label} → {get_status_label(final_status, "zh_cn")}',
                workflow['order_number'],
                workflow_number
            ))
        except:
            # 如果没有 operation_logs 表，跳过
            pass
        
        conn.commit()
        conn.close()
        
        final_status_label = get_status_label(final_status, 'zh_cn')
        if transition['auto_completed']:
            response_message = f'状态已更新：{old_status_label} → {new_status_label}，并自动转为{final_status_label}'
        else:
            response_message = f'状态已更新：{old_status_label} → {final_status_label}'

        return jsonify({
            'success': True,
            'message': response_message,
            'data': {
                'workflow_number': workflow_number,
                'order_number': workflow['order_number'],
                'old_status': old_status,
                'new_status': final_status,
                'requested_status': new_status,
                'old_status_label': old_status_label,
                'new_status_label': final_status_label,
                'requested_status_label': new_status_label,
                'auto_completed': transition['auto_completed'],
                'action': action,
                'action_label': action_label,
                'latest_history_id': latest_history_id
            }
        })
        
    except Exception as e:
        import traceback
        print(f"[ERROR] 更新工作流状态失败: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@tracking_bp.route('/api/workflows/<workflow_number>/status-direct', methods=['POST'])
@api_login_required
def api_update_workflow_status_direct(workflow_number):
    """直接更新工作流状态（用于跳过阶段）"""
    try:
        data = request.get_json() or {}
        new_status = data.get('new_status')
        expected_status = data.get('expected_status')
        try:
            action_date = normalize_action_date(data.get('action_date'))
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc), 'code': 'INVALID_ACTION_DATE'}), 400
        notes = data.get('notes', '')
        expected_history_id = data.get('expected_history_id')
        
        if not new_status:
            return jsonify({'success': False, 'error': '缺少新狀態參數'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 查询工作流
        cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
        workflow = cursor.fetchone()
        
        if not workflow:
            conn.close()
            return jsonify({'success': False, 'error': '工作流不存在'}), 404
        
        workflow = dict(workflow)
        if not expected_status:
            conn.close()
            return jsonify({'success': False, 'error': '缺少 expected_status 参数', 'code': 'MISSING_EXPECTED_STATUS'}), 400
        if workflow.get('current_status') != expected_status:
            conn.close()
            return jsonify({
                'success': False,
                'error': '当前状态已变更，请刷新后重试',
                'code': 'STATUS_CHANGED',
                'current_status': workflow.get('current_status')
            }), 409
        if expected_history_id:
            latest_history = get_latest_workflow_history(cursor, workflow_number)
            latest_id = latest_history.get('id') if latest_history else None
            if str(latest_id) != str(expected_history_id):
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '状态已被更新，请刷新后重试',
                    'code': 'HISTORY_CHANGED',
                    'latest_history_id': latest_id
                }), 409
        has_permission, error_msg = check_permission('workflow', workflow_number, 'edit')
        if not has_permission:
            conn.close()
            return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
        current_ctx = get_current_user_context()
        current_user_id = current_ctx.get('id')
        if not can_manage_by_owner(workflow.get('handler_id'), user_id=current_user_id):
            conn.close()
            return jsonify({'success': False, 'error': '無權限操作此流程'}), 403
        old_status = workflow['current_status']
        
        # 获取操作者
        operator_id = current_user_id
        
        force_insert = (new_status == STATUS_KEYS['PARTIAL_SHIPPED'] and old_status == STATUS_KEYS['PARTIAL_SHIPPED'])
        transition = apply_workflow_status_transition(
            cursor,
            workflow,
            workflow_number,
            new_status,
            action_date,
            operator_id,
            notes,
            force_insert=force_insert
        )
        final_status = transition['final_status']
        latest_history_id = transition['latest_history_id']
        
        conn.commit()
        conn.close()
        
        old_status_label = get_status_label(old_status, 'zh_cn')
        new_status_label = get_status_label(new_status, 'zh_cn')
        final_status_label = get_status_label(final_status, 'zh_cn')
        if transition['auto_completed']:
            response_message = f'流程已更新為「{new_status_label}」，並自動轉為「{final_status_label}」'
        else:
            response_message = f'流程已更新為「{final_status_label}」'
        
        return jsonify({
            'success': True,
            'message': response_message,
            'data': {
                'workflow_number': workflow_number,
                'order_number': workflow['order_number'],
                'old_status': old_status,
                'new_status': final_status,
                'requested_status': new_status,
                'old_status_label': old_status_label,
                'new_status_label': final_status_label,
                'requested_status_label': new_status_label,
                'auto_completed': transition['auto_completed'],
                'action_date': action_date,
                'latest_history_id': latest_history_id
            }
        })
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@tracking_bp.route('/api/orders/quick-update', methods=['POST'])
@api_login_required
def api_quick_update():
    """快速更新流程狀態API（支持 workflow_number 和 order_number）"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '無效的請求數據'}), 400
            
        workflow_number = data.get('workflow_number')
        order_number = data.get('order_number')  # 向后兼容
        action = data.get('action')
        expected_status = data.get('expected_status')
        try:
            action_date = normalize_action_date(data.get('date'))
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc), 'code': 'INVALID_ACTION_DATE'}), 400
        notes = data.get('notes', '')
        expected_history_id = data.get('expected_history_id')
        
        if not action:
            return jsonify({'success': False, 'error': '缺少必要參數'}), 400
        
        # 优先使用 workflow_number，如果没有则使用 order_number（向后兼容）
        if not workflow_number and not order_number:
            return jsonify({'success': False, 'error': '缺少 workflow_number 或 order_number'}), 400
        
        # 状态映射（统一使用 status_definitions.py 中的 QUICK_ACTIONS_MAP）
        # 返回的是 key，直接存入数据库
        status_map = QUICK_ACTIONS_MAP.copy()
        
        # 兼容旧版本（保留，但建议逐步迁移）
        status_map.update({
            'quote_to_order': STATUS_KEYS['NEW_ORDER'],
            'quote_complete': STATUS_KEYS['COMPLETED'],
            'draft_sent': STATUS_KEYS['DRAFT_CONFIRMING'],
            'draft_revise': STATUS_KEYS['DRAFT_REVISING'],
            'draft_modified': STATUS_KEYS['DRAFT_CONFIRMING'],
            'sample_start': STATUS_KEYS['SAMPLING'],
            'sample_done': STATUS_KEYS['SAMPLE_CONFIRMING'],
            'sample_confirm': STATUS_KEYS['PENDING_PRODUCTION'],
            'sample_revise': STATUS_KEYS['SAMPLE_REVISING'],
            'sample_modified': STATUS_KEYS['SAMPLE_CONFIRMING'],
            # 舊的 production_complete 向後兼容（直接完成）
            'production_complete': STATUS_KEYS['COMPLETED'],
            'complete': STATUS_KEYS['COMPLETED']
        })
        
        new_status = status_map.get(action)  # 返回 key
        if not new_status:
            return jsonify({'success': False, 'error': f'无效的操作：{action}'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 取得当前用户信息
        current_ctx = get_current_user_context()
        current_user_id = current_ctx.get('id')

        # 如果提供了 workflow_number，使用 workflows 表
        if workflow_number:
            cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
            workflow = cursor.fetchone()
            
            if not workflow:
                conn.close()
                return jsonify({'success': False, 'error': '流程不存在'}), 404
            
            workflow = dict(workflow)
            if not expected_status:
                conn.close()
                return jsonify({'success': False, 'error': '缺少 expected_status 参数', 'code': 'MISSING_EXPECTED_STATUS'}), 400
            if workflow.get('current_status') != expected_status:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '当前状态已变更，请刷新后重试',
                    'code': 'STATUS_CHANGED',
                    'current_status': workflow.get('current_status')
                }), 409
            if expected_history_id:
                latest_history = get_latest_workflow_history(cursor, workflow_number)
                latest_id = latest_history.get('id') if latest_history else None
                if str(latest_id) != str(expected_history_id):
                    conn.close()
                    return jsonify({
                        'success': False,
                        'error': '状态已被更新，请刷新后重试',
                        'code': 'HISTORY_CHANGED',
                        'latest_history_id': latest_id
                    }), 409
            has_permission, error_msg = check_permission('workflow', workflow_number, 'edit')
            if not has_permission:
                conn.close()
                return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
            if not can_manage_by_owner(workflow.get('handler_id'), user_id=current_user_id):
                conn.close()
                return jsonify({'success': False, 'error': '無權限操作此流程'}), 403
            # 檢查訂單是否鎖定（已完成或已取消）
            cursor.execute('SELECT is_locked, status FROM orders WHERE order_number = ?', (workflow.get('order_number'),))
            order_row = cursor.fetchone()
            if order_row and _normalize_lock_flag(order_row['is_locked']):
                conn.close()
                if order_row['status'] == 'CANCELLED':
                    return jsonify({'success': False, 'error': '該訂單已取消，無法更新流程狀態', 'code': 'ORDER_CANCELLED'}), 403
                return jsonify({'success': False, 'error': '該訂單已完成鎖定，無法更新流程狀態', 'code': 'ORDER_LOCKED'}), 403
            old_status = workflow['current_status']
            order_number = workflow['order_number']  # 从 workflow 获取 order_number
            
            # 獲取操作者
            operator_id = current_user_id
            
            # 部分出货可以在 PARTIAL_SHIPPED 状态重复执行；每次都必须留下独立历史。
            force_insert = (action == 'ship_partial')
            is_repeat_partial_ship = (force_insert and old_status == STATUS_KEYS['PARTIAL_SHIPPED'])
            transition = apply_workflow_status_transition(
                cursor,
                workflow,
                workflow_number,
                new_status,
                action_date,
                operator_id,
                notes,
                update_current_status=not is_repeat_partial_ship,
                force_insert=force_insert
            )
            final_status = transition['final_status']
            latest_history_id = transition['latest_history_id']
            
            conn.commit()
            conn.close()

            final_status_label = get_status_label(final_status, 'zh_cn')
            requested_status_label = get_status_label(new_status, 'zh_cn')
            if transition['auto_completed']:
                response_message = f'流程已更新為「{requested_status_label}」，並自動轉為「{final_status_label}」'
            else:
                response_message = f'流程已更新為「{final_status_label}」'
            
            return jsonify({
                'success': True,
                'message': response_message,
                'data': {
                    'workflow_number': workflow_number,
                    'order_number': order_number,
                    'old_status': old_status,
                    'new_status': final_status,
                    'requested_status': new_status,
                    'new_status_label': final_status_label,
                    'requested_status_label': requested_status_label,
                    'auto_completed': transition['auto_completed'],
                    'action_date': action_date,
                    'latest_history_id': latest_history_id
                }
            })
        
        # 向后兼容：如果没有 workflow_number，使用旧的 orders 表逻辑
        else:
            if not is_admin():
                conn.close()
                return jsonify({'success': False, 'error': '無權限操作此訂單'}), 403
            cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
            order = cursor.fetchone()
            
            if not order:
                conn.close()
                return jsonify({'success': False, 'error': '訂單不存在'}), 404
            
            order = dict(order)
            has_permission, error_msg = check_permission('order', order_number, 'edit')
            if not has_permission:
                conn.close()
                return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
            old_status = order.get('current_status', '')
            
            # 特殊處理：詢價轉為訂單
            if action == 'quote_to_order':
                conn.close()
                return jsonify({'success': False, 'error': '請使用轉為訂單功能'}), 400
            
            # 獲取操作者
            operator = 'system'
            if hasattr(g, 'current_user') and g.current_user:
                operator = g.current_user.get('username', 'system')

            requested_status = new_status
            final_status = STATUS_KEYS['COMPLETED'] if requested_status == STATUS_KEYS['ALL_SHIPPED'] else requested_status
            auto_completed = final_status != requested_status
            
            # 記錄狀態變更（如果 orders 表有 status_history 字段）
            # 注意：这可能需要根据实际数据库结构调整
            if 'id' in order:
                try:
                    cursor.execute('''
                        INSERT INTO status_history (order_id, order_number, from_status, to_status, action_date, operator, notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        order['id'],
                        order_number,
                        old_status,
                        requested_status,
                        action_date,
                        operator,
                        notes
                    ))
                    if auto_completed:
                        cursor.execute('''
                            INSERT INTO status_history (order_id, order_number, from_status, to_status, action_date, operator, notes)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            order['id'],
                            order_number,
                            requested_status,
                            final_status,
                            action_date,
                            operator,
                            '系統自動：已全部出貨後轉為已完成'
                        ))
                except:
                    pass  # 如果 status_history 表不存在，跳过
            
            # 更新訂單
            cursor.execute('''
                UPDATE orders 
                SET current_status = ?,
                    last_status_change_date = ?,
                    status_days = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE order_number = ?
            ''', (final_status, action_date, order_number))
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': f'訂單已更新為「{get_status_label(final_status, "zh_cn")}」',
                'data': {
                    'order_number': order_number,
                    'old_status': old_status,
                    'new_status': final_status,
                    'requested_status': requested_status,
                    'auto_completed': auto_completed,
                    'action_date': action_date
                }
            })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"API Error: {error_detail}")  # 打印到控制台
        
        if 'conn' in locals():
            conn.close()
            
        return jsonify({
            'success': False,
            'error': f'更新失敗：{str(e)}'
        }), 500


@tracking_bp.route('/api/revisions', methods=['GET'])
@api_login_required
def api_revisions():
    """獲取修圖列表API"""
    status = request.args.get('status', 'all')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = "SELECT * FROM revisions WHERE 1=1"
    params = []
    
    if status != 'all':
        query += " AND current_status = ?"
        params.append(status)
    
    query += " ORDER BY request_date DESC"
    
    cursor.execute(query, params)
    revisions_list = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': revisions_list,
        'total': len(revisions_list)
    })

@tracking_bp.route('/api/customers/search', methods=['GET'])
@login_required
def api_search_customers():
    """搜索客戶名稱（用於自動完成）"""
    query = request.args.get('q', '').strip()
    
    if not query or len(query) < 1:
        return jsonify({'success': True, 'data': []})

    if cloud_mode_enabled():
        try:
            customers = _cloud_provider_call('search_customers', query, 10) or []
            return jsonify({'success': True, 'data': list(customers)})
        except Exception as exc:
            return jsonify({'success': False, 'error': f'雲端客戶搜尋失敗: {exc}'}), 503
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT DISTINCT UPPER(customer_name) AS name
        FROM orders 
        WHERE customer_name LIKE ? AND customer_name IS NOT NULL AND customer_name != ''
        ORDER BY name
        LIMIT 10
    ''', (f'%{query}%',))
    
    customers = [row['name'] for row in cursor.fetchall() if row['name']]
    conn.close()
    
    return jsonify({'success': True, 'data': customers})


# ===================================================================
# 系统设置 helpers
# ===================================================================

def _get_setting(key, default=None):
    """从 system_settings 表读取一个设置值"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT value FROM system_settings WHERE key = ?', (key,))
    row = cursor.fetchone()
    conn.close()
    return row['value'] if row else default

def _get_setting_int(key, default=0):
    val = _get_setting(key)
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


# ===================================================================
# 通知系统 API
# ===================================================================

def _create_notification(cursor, user_id, ntype, title, message='', order_number=None, workflow_number=None, priority='normal', expires_days=90):
    """内部辅助：向 notifications 表插入一条通知"""
    from datetime import timedelta
    expires_at = (datetime.now() + timedelta(days=expires_days)).strftime('%Y-%m-%d %H:%M:%S') if expires_days else None
    cursor.execute('''
        INSERT INTO notifications (user_id, type, title, message, order_number, workflow_number, priority, expires_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (user_id, ntype, title, message, order_number, workflow_number, priority, expires_at))


def _create_broadcast_notification(cursor, ntype, title, message='', order_number=None, workflow_number=None, priority='normal', role_filter='sales', exclude_user_id=None, expires_days=90):
    """内部辅助：给符合条件的全部用户各插入一条通知"""
    # viewer 角色永远不接收通知
    viewer_exclude = "AND role NOT IN ('viewer', '观察者')"
    if role_filter == 'sales':
        cursor.execute(f"SELECT id FROM users WHERE role NOT IN ('administrator', 'admin', 'root', 'superuser', '管理员', '主管') {viewer_exclude} AND status = 'active'")
    elif role_filter == 'all':
        cursor.execute(f"SELECT id FROM users WHERE status = 'active' {viewer_exclude}")
    else:
        cursor.execute(f"SELECT id FROM users WHERE status = 'active' {viewer_exclude}")
    
    users = cursor.fetchall()
    for user in users:
        uid = user['id']
        if exclude_user_id and uid == exclude_user_id:
            continue
        _create_notification(cursor, uid, ntype, title, message, order_number, workflow_number, priority, expires_days)


@tracking_bp.route('/api/notifications/unread-count', methods=['GET'])
@api_login_required
def api_notification_unread_count():
    """获取当前用户未读通知数量"""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': True, 'count': 0})
    
    visible_days = _get_setting_int('notification_visible_days', 30)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COUNT(*) as count FROM notifications 
        WHERE user_id = ? AND is_read = 0 
        AND type NOT LIKE '\\_%' ESCAPE '\\'
        AND created_at >= datetime('now', ?)
    ''', (user_id, f'-{visible_days} days'))
    count = cursor.fetchone()['count']
    conn.close()
    return jsonify({'success': True, 'count': count})


@tracking_bp.route('/api/notifications', methods=['GET'])
@api_login_required
def api_list_notifications():
    """获取当前用户的通知列表"""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': True, 'data': [], 'total': 0})
    
    tab = request.args.get('tab', 'unread')  # unread / all
    page = max(1, int(request.args.get('page', 1)))
    page_size = min(50, int(request.args.get('page_size', 20)))
    
    conn = get_db()
    cursor = conn.cursor()
    
    visible_days = _get_setting_int('notification_visible_days', 30)
    where = "user_id = ? AND type NOT LIKE '\\_%' ESCAPE '\\' AND created_at >= datetime('now', ?)"
    params = [user_id, f'-{visible_days} days']
    if tab == 'unread':
        where += " AND is_read = 0"
    
    # 总数
    cursor.execute(f'SELECT COUNT(*) as count FROM notifications WHERE {where}', params)
    total = cursor.fetchone()['count']
    
    # 分页查询
    offset = (page - 1) * page_size
    cursor.execute(f'''
        SELECT * FROM notifications 
        WHERE {where}
        ORDER BY created_at DESC
        LIMIT ? OFFSET ?
    ''', params + [page_size, offset])
    
    items = []
    for row in cursor.fetchall():
        items.append({
            'id': row['id'],
            'type': row['type'],
            'title': row['title'],
            'message': row['message'] or '',
            'order_number': row['order_number'],
            'workflow_number': row['workflow_number'],
            'priority': row['priority'] or 'normal',
            'is_read': bool(row['is_read']),
            'created_at': row['created_at'],
        })
    
    conn.close()
    return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'page': page,
        'page_size': page_size,
    })


@tracking_bp.route('/api/notifications/<int:notif_id>/read', methods=['PUT'])
@api_login_required
def api_mark_notification_read(notif_id):
    """标记单条通知为已读"""
    user_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE notifications SET is_read = 1, read_at = datetime('now')
        WHERE id = ? AND user_id = ?
    ''', (notif_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@tracking_bp.route('/api/notifications/read-all', methods=['PUT'])
@api_login_required
def api_mark_all_notifications_read():
    """标记当前用户所有通知为已读"""
    user_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE notifications SET is_read = 1, read_at = datetime('now')
        WHERE user_id = ? AND is_read = 0
    ''', (user_id,))
    affected = cursor.rowcount
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated': affected})


@tracking_bp.route('/api/notifications/check-alerts', methods=['POST'])
@api_login_required
def api_check_notification_alerts():
    """
    检查交货日期预警 + 红灯预警，生成通知（幂等，每用户每天最多触发一次）
    前端在用户首次打开页面时调用
    """
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': True, 'created': 0})
    
    today_str = date.today().isoformat()
    conn = get_db()
    cursor = conn.cursor()
    
    # ---- 并发锁：同一用户同一天只允许一个请求执行检查 ----
    # 用 INSERT 一条临时标记来实现幂等
    try:
        cursor.execute('''
            SELECT COUNT(*) as c FROM notifications
            WHERE user_id = ? AND type = '_alert_check_lock'
            AND DATE(created_at) = ?
        ''', (user_id, today_str))
        if cursor.fetchone()['c'] > 0:
            conn.close()
            return jsonify({'success': True, 'created': 0, 'reason': 'already_checked_today'})
        # 插入锁标记（这条记录不会显示给用户，type 以 _ 开头）
        cursor.execute('''
            INSERT INTO notifications (user_id, type, title, message, is_read, created_at)
            VALUES (?, '_alert_check_lock', '', '', 1, datetime('now'))
        ''', (user_id,))
        conn.commit()
    except Exception:
        pass
    
    # 查询今天已有的告警（防止重复）
    cursor.execute('''
        SELECT user_id, type, workflow_number FROM notifications
        WHERE type IN ('delivery_warning', 'delivery_overdue', 'red_light')
        AND DATE(created_at) = ?
    ''', (today_str,))
    existing_alerts = set()
    for row in cursor.fetchall():
        existing_alerts.add((row['user_id'], row['type'], row['workflow_number']))
    
    # 获取当前用户角色
    current_ctx = get_current_user_context()
    current_role = current_ctx.get('role', 'viewer')
    
    # 获取完成/取消状态值（兼容中英文）
    from .status_definitions import STATUS_KEYS as _SK
    _completed = _SK.get('COMPLETED', 'COMPLETED') if _SK else 'COMPLETED'
    _cancelled = _SK.get('CANCELLED', 'CANCELLED') if _SK else 'CANCELLED'
    exclude_statuses = list(set(filter(None, [_completed, _cancelled, '已完成', '已取消'])))
    placeholders = ','.join(['?'] * len(exclude_statuses))
    
    # 获取该用户负责的流程
    wf_sql_base = f'''
        SELECT w.workflow_number, w.order_number, w.expected_delivery_date,
               w.current_status, w.handler_id,
               w.status_updated_at,
               o.customer_name,
               COALESCE(u.real_name, u.display_name, u.username) AS handler_name,
               (SELECT MAX(wsh.action_date) FROM workflow_status_history wsh
                WHERE wsh.workflow_number = w.workflow_number) AS last_action_date
        FROM workflows w
        LEFT JOIN orders o ON w.order_number = o.order_number
        LEFT JOIN users u ON w.handler_id = u.id
        WHERE w.current_status NOT IN ({placeholders})
    '''
    if current_role == 'admin':
        cursor.execute(wf_sql_base, tuple(exclude_statuses))
    else:
        cursor.execute(wf_sql_base + " AND w.handler_id = ?", tuple(exclude_statuses) + (user_id,))
    
    workflows = [dict(r) for r in cursor.fetchall()]
    created_count = 0
    today_date = date.today()
    
    # 获取所有管理员 ID，用于给管理员也发通知
    cursor.execute("SELECT id FROM users WHERE role IN ('administrator', 'admin', 'root', 'superuser', '管理员', '主管') AND status = 'active'")
    admin_ids = [row['id'] for row in cursor.fetchall()]
    
    for wf in workflows:
        wf_num = wf['workflow_number']
        order_num = wf['order_number'] or ''
        customer = wf['customer_name'] or ''
        delivery = wf.get('expected_delivery_date')
        handler_uid = wf['handler_id']
        
        # 收集需要通知的用户列表（业务员 + 所有管理员，去重）
        notify_uids = set()
        if handler_uid:
            notify_uids.add(handler_uid)
        for aid in admin_ids:
            notify_uids.add(aid)
        
        if not notify_uids:
            continue
        
        # 1) 交货日期预警：3天内到期
        if delivery:
            try:
                delivery_date = datetime.strptime(str(delivery).split()[0], '%Y-%m-%d').date()
                days_left = (delivery_date - today_date).days
                
                if days_left < 0:
                    for uid in notify_uids:
                        key = (uid, 'delivery_overdue', wf_num)
                        if key in existing_alerts:
                            continue
                        _create_notification(
                            cursor, uid, 'delivery_overdue',
                            f'订单 {order_num} 交货已逾期 {abs(days_left)} 天',
                            f'客户：{customer}，交货日期：{delivery}，流程：{wf_num}',
                            order_num, wf_num, 'important', 30
                        )
                        existing_alerts.add(key)
                        created_count += 1
                elif days_left <= 3:
                    for uid in notify_uids:
                        key = (uid, 'delivery_warning', wf_num)
                        if key in existing_alerts:
                            continue
                        _create_notification(
                            cursor, uid, 'delivery_warning',
                            f'订单 {order_num} 还有 {days_left} 天交货',
                            f'客户：{customer}，交货日期：{delivery}，流程：{wf_num}',
                            order_num, wf_num, 'important', 30
                        )
                        existing_alerts.add(key)
                        created_count += 1
            except (ValueError, TypeError):
                pass
        
        # 2) 红灯预警
        try:
            wf['last_status_change_date'] = wf.get('last_action_date') or wf.get('status_updated_at')
            light = calculate_status_light(wf)
            if light == 'red':
                for uid in notify_uids:
                    key = (uid, 'red_light', wf_num)
                    if key in existing_alerts:
                        continue
                    _create_notification(
                        cursor, uid, 'red_light',
                        f'流程 {wf_num} 进入红灯状态',
                        f'订单：{order_num}，客户：{customer}，当前状态停留过长',
                        order_num, wf_num, 'important', 30
                    )
                    existing_alerts.add(key)
                    created_count += 1
        except Exception:
            pass
    
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'created': created_count})


@tracking_bp.route('/api/orders/<order_number>/status', methods=['POST'])
@api_admin_required
def api_update_order_status(order_number):
    """更新訂單狀態API（用於跳過階段、取消訂單等）"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '無效的請求數據'}), 400
        
        new_status = data.get('new_status')
        action_date = data.get('action_date', date.today().isoformat())
        notes = data.get('notes', '')
        
        if not new_status:
            return jsonify({'success': False, 'error': '缺少新狀態參數'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 查詢訂單
        cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
        order = cursor.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'error': '訂單不存在'}), 404
        
        order = dict(order)
        has_permission, error_msg = check_permission('order', order_number, 'edit')
        if not has_permission:
            conn.close()
            return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
        old_status = order['current_status']
        
        # 獲取操作者
        operator = 'system'
        if hasattr(g, 'current_user') and g.current_user:
            operator = g.current_user.get('username', 'system')

        requested_status = new_status
        final_status = STATUS_KEYS['COMPLETED'] if requested_status == STATUS_KEYS['ALL_SHIPPED'] else requested_status
        auto_completed = final_status != requested_status
        
        # 記錄狀態變更（兼容：status_history 表可能不存在）
        try:
            cursor.execute('''
                INSERT INTO status_history (order_id, order_number, from_status, to_status, action_date, operator, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                order['id'],
                order_number,
                old_status,
                requested_status,
                action_date,
                operator,
                notes
            ))
            if auto_completed:
                cursor.execute('''
                    INSERT INTO status_history (order_id, order_number, from_status, to_status, action_date, operator, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    order['id'],
                    order_number,
                    requested_status,
                    final_status,
                    action_date,
                    operator,
                    '系統自動：已全部出貨後轉為已完成'
                ))
        except Exception:
            pass  # status_history 表不存在時跳過
        
        # 更新訂單
        cursor.execute('''
            UPDATE orders 
            SET current_status = ?,
                last_status_change_date = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE order_number = ?
        ''', (final_status, action_date, order_number))
        
        # 更新燈號
        update_status_light(order['id'], conn)
        
        # 記錄操作日誌
        cursor.execute('''
            INSERT INTO audit_log (action_type, order_number, old_status, new_status, operator, reason)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', ('status_update', order_number, old_status, final_status, operator, notes))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'狀態已更新：{old_status} → {get_status_label(final_status, "zh_cn")}',
            'data': {
                'order_number': order_number,
                'old_status': old_status,
                'new_status': final_status,
                'requested_status': requested_status,
                'auto_completed': auto_completed,
                'action_date': action_date
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@tracking_bp.route('/api/orders/check', methods=['GET'])
@api_login_required
def api_check_order():
    """檢查訂單號是否存在且已解鎖"""
    try:
        order_number = request.args.get('order_number', '').strip()
        
        if not order_number:
            return jsonify({'success': False, 'error': '缺少訂單號'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
        order = cursor.fetchone()
        
        conn.close()
        
        if not order:
            return jsonify({
                'success': True,
                'data': {
                    'exists': False,
                    'status': None,
                    'accessible': False
                }
            })
        
        order_dict = dict(order)
        
        # 業務員權限檢查
        current_ctx = get_current_user_context()
        current_role = current_ctx.get('role', 'viewer')
        if not can_access_visibility(order_dict.get('visibility')):
            return jsonify({
                'success': True,
                'data': {
                    'exists': True,
                    'status': order_dict.get('status'),
                    'accessible': False
                }
            })
        
        return jsonify({
            'success': True,
            'data': {
                'exists': True,
                'status': order_dict.get('status'),
                'accessible': True,
                'customer_name': order_dict.get('customer_name'),
                'order_date': order_dict.get('order_date')
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'服务器错误: {str(e)}'
        }), 500

@tracking_bp.route('/api/orders/unlocked', methods=['GET'])
@api_login_required
def api_unlocked_orders():
    """獲取已解鎖且可見的訂單列表（供業務員選擇並創建流程）"""
    conn = get_db()
    cursor = conn.cursor()
    current_ctx = get_current_user_context()
    current_role = current_ctx.get('role', 'viewer')
    
    search = request.args.get('search', '').strip()
    sort = request.args.get('sort', 'unlocked_desc')
    sort_by = (request.args.get('sort_by') or '').strip()
    sort_order = (request.args.get('sort_order') or '').strip().lower()
    include_counts = request.args.get('include_counts', '0') == '1'
    page = request.args.get('page')
    page_size = request.args.get('page_size')
    use_pagination = page is not None and page_size is not None
    page = int(page or 1)
    page_size = int(page_size or 200)

    select_fields = [
        "o.order_number",
        "o.customer_name",
        "o.order_date",
        "o.status",
        "o.visibility",
        "o.notes",
        "COALESCE(o.updated_at, o.created_at) as unlocked_at"
    ]
    join_clause = ""
    group_by = ""
    order_number_expr = get_order_number_sort_expr('o')
    sort_key_map = {
        'order_number': order_number_expr,
        'customer_name': "o.customer_name COLLATE NOCASE",
        'project_count': "project_count",
        'unlocked_at': "unlocked_at"
    }
    if sort and not sort_by:
        sort_value_map = {
            'unlocked_desc': ('unlocked_at', 'desc'),
            'unlocked_asc': ('unlocked_at', 'asc'),
            'order_number_desc': ('order_number', 'desc'),
            'order_number_asc': ('order_number', 'asc'),
            'customer_name_desc': ('customer_name', 'desc'),
            'customer_name_asc': ('customer_name', 'asc'),
            'project_count_desc': ('project_count', 'desc'),
            'project_count_asc': ('project_count', 'asc')
        }
        if sort in sort_value_map:
            sort_by, sort_order = sort_value_map[sort]
    if sort_by not in sort_key_map:
        sort_by = 'unlocked_at'
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'
    needs_counts = include_counts or sort_by == 'project_count'

    where_clauses = ["o.status = 'ACTIVE'", "(o.is_locked IS NULL OR o.is_locked = 0)"]
    join_params = []   # JOIN 子句的參數
    where_params = []  # WHERE 子句的參數
    if needs_counts:
        select_fields.append("COUNT(w.workflow_number) as project_count")
        filter_info = get_filtered_resources('workflow', current_role, current_ctx.get('id'))
        if filter_info['rule'] == 'all':
            join_clause = "LEFT JOIN workflows w ON o.order_number = w.order_number"
        else:
            join_clause = f"LEFT JOIN workflows w ON o.order_number = w.order_number AND {filter_info['where_sql']}"
            join_params.extend(filter_info['params'])
        group_by = "GROUP BY o.order_number"
    visibility_clause = get_visibility_where_clause(current_role, table_alias='o')
    if visibility_clause:
        where_clauses.append(visibility_clause)
    if search:
        where_clauses.append("(o.order_number LIKE ? OR o.customer_name LIKE ?)")
        search_term = f"%{search}%"
        where_params.extend([search_term, search_term])

    where_sql = " AND ".join(where_clauses)
    sort_expr = sort_key_map.get(sort_by, "unlocked_at")
    order_by = f"ORDER BY {sort_expr} {'ASC' if sort_order == 'asc' else 'DESC'}, {order_number_expr} ASC"

    # 主查詢參數 = JOIN 參數 + WHERE 參數
    params = join_params + where_params

    query = f"""
        SELECT {", ".join(select_fields)}
        FROM orders o
        {join_clause}
        WHERE {where_sql}
        {group_by}
        {order_by}
    """

    # 計算總數（只需 WHERE 參數，不含 JOIN 參數）
    count_query = f"SELECT COUNT(*) as count FROM orders o WHERE {where_sql}"
    cursor.execute(count_query, where_params)
    total = cursor.fetchone()['count']

    truncated = False
    if use_pagination:
        offset = (page - 1) * page_size
        query += " LIMIT ? OFFSET ?"
        params.extend([page_size, offset])
    else:
        # 無分頁時，默認上限 5000 條（無搜索）或 1000 條（有搜索）
        default_limit = 1000 if search else 5000
        if total > default_limit:
            truncated = True
        query += f" LIMIT {default_limit}"

    cursor.execute(query, params)
    orders = []
    for row in cursor.fetchall():
        orders.append({
            'order_number': row['order_number'],
            'customer_name': row['customer_name'] or '',
            'order_date': row['order_date'] or '',
            'status': row['status'],
            'visibility': row['visibility'],
            'notes': row['notes'] or '',
            'unlocked_at': row['unlocked_at'],
            'project_count': row['project_count'] if include_counts else None
        })

    conn.close()

    response = {'success': True, 'data': orders, 'total': total, 'truncated': truncated}
    if use_pagination:
        response['pagination'] = {
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': (total + page_size - 1) // page_size
        }
    return jsonify(response)


@tracking_bp.route('/api/workflows/mine', methods=['GET'])
@api_login_required
def api_my_workflows():
    """獲取當前業務員的流程列表（支持篩選/搜尋）"""
    status_filter = request.args.get('status', 'all')
    search = request.args.get('search', '').strip()
    sort = request.args.get('sort', 'created_desc')
    sort_by = (request.args.get('sort_by') or '').strip()
    sort_order = (request.args.get('sort_order') or '').strip().lower()
    created_by_me = request.args.get('created_by_me') == '1'
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))

    current_ctx = get_current_user_context()
    current_role = current_ctx.get('role', 'viewer')
    current_user_id = current_ctx.get('id')

    conn = get_db()
    cursor = conn.cursor()

    query = """
        SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility,
               COALESCE(u.real_name, u.display_name, u.username) as handler_name
        FROM workflows w
        INNER JOIN orders o ON w.order_number = o.order_number
        LEFT JOIN users u ON w.handler_id = u.id
        WHERE o.status = 'ACTIVE'
    """
    count_query = """
        SELECT COUNT(*) as count
        FROM workflows w
        INNER JOIN orders o ON w.order_number = o.order_number
        WHERE o.status = 'ACTIVE'
    """
    params = []

    filter_info = get_filtered_resources('workflow', current_role, current_user_id)
    if filter_info['rule'] != 'all':
        visibility_clause = get_visibility_where_clause(current_role, table_alias='o')
        visibility_sql = f" AND {visibility_clause}" if visibility_clause else ""
        query += f" AND {filter_info['where_sql']}{visibility_sql}"
        count_query += f" AND {filter_info['where_sql']}{visibility_sql}"
        params.extend(filter_info['params'])
    else:
        handler_id = request.args.get('handler_id')
        if handler_id:
            query += " AND w.handler_id = ?"
            count_query += " AND w.handler_id = ?"
            params.append(handler_id)
        if created_by_me:
            query += " AND w.handler_id = ?"
            count_query += " AND w.handler_id = ?"
            params.append(current_user_id)

    if status_filter == 'in_progress':
        query += " AND w.current_status NOT IN (?, ?)"
        count_query += " AND w.current_status NOT IN (?, ?)"
        params.extend([STATUS_KEYS['COMPLETED'], STATUS_KEYS['CANCELLED']])
    elif status_filter == 'completed':
        query += " AND w.current_status = ?"
        count_query += " AND w.current_status = ?"
        params.append(STATUS_KEYS['COMPLETED'])
    elif status_filter == 'cancelled':
        query += " AND w.current_status = ?"
        count_query += " AND w.current_status = ?"
        params.append(STATUS_KEYS['CANCELLED'])

    if search:
        query += " AND (w.workflow_number LIKE ? OR w.order_number LIKE ? OR o.customer_name LIKE ?)"
        count_query += " AND (w.workflow_number LIKE ? OR w.order_number LIKE ? OR o.customer_name LIKE ?)"
        search_term = f"%{search}%"
        params.extend([search_term, search_term, search_term])

    sort_key_map = {
        'workflow_number': "w.workflow_number COLLATE NOCASE",
        'customer_name': "o.customer_name COLLATE NOCASE",
        'product_name': "w.product_name COLLATE NOCASE",
        'status': "w.current_status",
        'created_at': "w.created_at"
    }
    if sort and not sort_by:
        sort_value_map = {
            'status': ('status', 'asc'),
            'status_asc': ('status', 'asc'),
            'status_desc': ('status', 'desc'),
            'created_asc': ('created_at', 'asc'),
            'created_desc': ('created_at', 'desc'),
            'workflow_number_asc': ('workflow_number', 'asc'),
            'workflow_number_desc': ('workflow_number', 'desc'),
            'customer_name_asc': ('customer_name', 'asc'),
            'customer_name_desc': ('customer_name', 'desc'),
            'product_name_asc': ('product_name', 'asc'),
            'product_name_desc': ('product_name', 'desc')
        }
        if sort in sort_value_map:
            sort_by, sort_order = sort_value_map[sort]
    if sort_by not in sort_key_map:
        sort_by = 'created_at'
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'
    sort_expr = sort_key_map.get(sort_by, "w.created_at")
    primary_direction = 'ASC' if sort_order == 'asc' else 'DESC'
    order_by_parts = [f"{sort_expr} {primary_direction}"]
    if sort_by == 'created_at':
        order_by_parts.append("w.workflow_number ASC")
    else:
        order_by_parts.append("w.created_at DESC")
    query += f" ORDER BY {', '.join(order_by_parts)}"

    cursor.execute(count_query, params)
    total = cursor.fetchone()['count']

    offset = (page - 1) * page_size
    query += " LIMIT ? OFFSET ?"
    params.extend([page_size, offset])

    cursor.execute(query, params)
    rows = cursor.fetchall()
    workflows = []
    for row in rows:
        workflow = dict(row)
        order_obj = {
            'order_number': workflow.get('order_number'),
            'order_date': workflow.get('order_date'),
            'expected_delivery_date': workflow.get('expected_delivery_date'),
            'status_updated_at': workflow.get('status_updated_at'),
            'current_status': workflow.get('current_status'),
            'last_status_change_date': workflow.get('status_updated_at')
        }
        workflow['status_light'] = calculate_status_light(order_obj)
        workflow['status_label'] = get_status_label(workflow.get('current_status'))
        workflow['product_name'] = workflow.get('product_name', '')
        workflows.append(workflow)

    conn.close()
    
    return jsonify({
        'success': True,
        'data': workflows,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': (total + page_size - 1) // page_size
        }
    })

@tracking_bp.route('/api/orders/check-number', methods=['GET'])

@login_required
def api_check_order_number():
    """檢查訂單號是否已存在"""
    order_number = request.args.get('order_number', '').strip()
    
    if not order_number:
        return jsonify({'success': False, 'error': '訂單號不能為空'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT id FROM orders WHERE order_number = ?', (order_number,))
    exists = cursor.fetchone() is not None
    
    conn.close()
    
    return jsonify({
        'success': True,
        'exists': exists,
        'message': '訂單號已存在' if exists else '訂單號可用'
    })

@tracking_bp.route('/api/orders/next-quote-number', methods=['GET'])
@login_required
def api_next_quote_number():
    """獲取下一個詢價編號"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT order_number 
        FROM orders 
        WHERE order_number LIKE 'YU%'
        ORDER BY order_number DESC
        LIMIT 1
    ''')
    
    last_order = cursor.fetchone()
    if last_order and last_order['order_number'].startswith('YU'):
        try:
            last_num = int(last_order['order_number'].replace('YU', ''))
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    next_number = f'KC{next_num:05d}'
    conn.close()
    
    return jsonify({
        'success': True,
        'next_number': next_number
    })

@tracking_bp.route('/api/stats', methods=['GET'])
@api_login_required
def api_stats():
    """獲取統計數據API"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) as total FROM orders')
    total = cursor.fetchone()['total']
    
    # 兼容旧数据：同时查询 key 和中文
    completed_key, completed_label = get_status_for_query(STATUS_KEYS['COMPLETED'])
    cancelled_key, cancelled_label = get_status_for_query(STATUS_KEYS['CANCELLED'])
    
    cursor.execute('SELECT COUNT(*) as count FROM orders WHERE current_status NOT IN (?, ?, ?, ?)', 
                   (completed_key, completed_label, cancelled_key, cancelled_label))
    active = cursor.fetchone()['count']
    
    cursor.execute('SELECT COUNT(*) as count FROM orders WHERE status_light = "red" AND current_status NOT IN (?, ?, ?, ?)', 
                   (completed_key, completed_label, cancelled_key, cancelled_label))
    red = cursor.fetchone()['count']
    
    cursor.execute('SELECT COUNT(*) as count FROM orders WHERE status_light = "yellow" AND current_status NOT IN (?, ?, ?, ?)', 
                   (completed_key, completed_label, cancelled_key, cancelled_label))
    yellow = cursor.fetchone()['count']
    
    cursor.execute('SELECT COUNT(*) as count FROM orders WHERE status_light = "green" AND current_status NOT IN (?, ?, ?, ?)', 
                   (completed_key, completed_label, cancelled_key, cancelled_label))
    green = cursor.fetchone()['count']
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'active': active,
            'lights': {
                'red': red,
                'yellow': yellow,
                'green': green
            }
        }
    })

@tracking_bp.route('/api/orders/<order_number>/undo-last-step', methods=['POST'])
@api_admin_required
def api_undo_last_step(order_number):
    """撤銷最後一步（硬刪除）"""
    try:
        data = request.get_json() or {}
        reason = data.get('reason', '')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 1. 獲取訂單
        cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
        order = cursor.fetchone()
        if not order:
            conn.close()
            return jsonify({'success': False, 'error': '訂單不存在', 'code': 'ORDER_NOT_FOUND'}), 404
        
        order = dict(order)
        has_permission, error_msg = check_permission('order', order_number, 'edit')
        if not has_permission:
            conn.close()
            return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
        
        # 2. 獲取最後兩步（兼容：status_history 表可能不存在）
        try:
            cursor.execute('''
                SELECT * FROM status_history 
                WHERE order_number = ? 
                ORDER BY action_date DESC, id DESC 
                LIMIT 2
            ''', (order_number,))
            history = cursor.fetchall()
        except Exception:
            history = []
        
        # 檢查歷史記錄數量
        history_count = len(history)
        if history_count == 0:
            conn.close()
            return jsonify({
                'success': False, 
                'error': '沒有歷史記錄，無法撤銷', 
                'code': 'NO_HISTORY',
                'details': f'訂單 {order_number} 沒有任何狀態歷史記錄'
            }), 400
        
        last_step = dict(history[0])
        
        # 如果只有1條歷史記錄，允許撤銷到初始狀態（NULL）
        if history_count < 2:
            # 記錄到操作日誌
            try:
                cursor.execute('''
                    INSERT INTO audit_log (
                        action_type, order_number, old_status, new_status,
                        operator, reason
                    ) VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    'UNDO_STEP',
                    order_number,
                    last_step['to_status'],
                    None,  # 撤銷到初始狀態（無狀態）
                    g.current_user.get('username', 'system'),
                    reason or '撤銷操作（撤銷到初始狀態）'
                ))
            except Exception as e:
                print(f'警告：無法記錄審計日誌: {e}')
            
            # 硬刪除最後一步（兼容：status_history 表可能不存在）
            try:
                cursor.execute('DELETE FROM status_history WHERE id = ?', (last_step['id'],))
            except Exception:
                pass
            
            # 恢復訂單到初始狀態（NULL）
            cursor.execute('''
                UPDATE orders 
                SET current_status = NULL,
                    last_status_change_date = NULL,
                    status_days = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE order_number = ?
            ''', (order_number,))
            
            # 更新燈號
            try:
                update_status_light(order['id'], conn)
            except Exception as e:
                print(f'警告：無法更新狀態燈號: {e}')
            
            conn.commit()
            conn.close()
            
            # 獲取狀態顯示文字（用於提示信息）
            current_status_label = get_status_label(last_step['to_status'], 'zh_cn') if last_step.get('to_status') else '未知'
            
            return jsonify({
                'success': True,
                'message': f'已撤銷，訂單恢復到初始狀態（已清除「{current_status_label}」狀態）',
                'data': {
                    'order_number': order_number,
                    'restored_status': None
                }
            })
        
        # 如果有2條或以上歷史記錄，撤銷到上一條
        previous_step = dict(history[1])
        
        # 3. 記錄到操作日誌
        try:
            cursor.execute('''
                INSERT INTO audit_log (
                    action_type, order_number, old_status, new_status,
                    operator, reason
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                'UNDO_STEP',
                order_number,
                last_step['to_status'],
                previous_step['to_status'],
                g.current_user.get('username', 'system'),
                reason or '撤銷操作'
            ))
        except Exception as e:
            # 如果審計日誌插入失敗，記錄但不阻止撤銷操作
            print(f'警告：無法記錄審計日誌: {e}')
        
        # 4. 硬刪除最後一步（兼容：status_history 表可能不存在）
        try:
            cursor.execute('DELETE FROM status_history WHERE id = ?', (last_step['id'],))
        except Exception:
            pass
        
        # 5. 恢復訂單到上一個狀態
        cursor.execute('''
            UPDATE orders 
            SET current_status = ?,
                last_status_change_date = ?,
                status_days = 0,
                updated_at = CURRENT_TIMESTAMP
            WHERE order_number = ?
        ''', (previous_step['to_status'], previous_step['action_date'], order_number))
        
        # 6. 更新燈號
        try:
            update_status_light(order['id'], conn)
        except Exception as e:
            print(f'警告：無法更新狀態燈號: {e}')
        
        conn.commit()
        conn.close()
        
        # 獲取狀態顯示文字（用於提示信息）
        previous_status_label = get_status_label(previous_step['to_status'], 'zh_cn') if previous_step.get('to_status') else previous_step.get('to_status', '未知')
        
        return jsonify({
            'success': True,
            'message': f'已撤銷，訂單恢復到「{previous_status_label}」',
            'data': {
                'order_number': order_number,
                'restored_status': previous_step['to_status']
            }
        })
    
    except sqlite3.Error as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({
            'success': False, 
            'error': '數據庫錯誤', 
            'code': 'DATABASE_ERROR',
            'details': str(e)
        }), 500
    
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({
            'success': False, 
            'error': '撤銷操作失敗', 
            'code': 'UNKNOWN_ERROR',
            'details': str(e)
        }), 500

@tracking_bp.route('/api/workflows/<workflow_number>/undo-last-step', methods=['POST'])
@api_login_required
def api_undo_workflow_last_step(workflow_number):
    """撤銷工作流最後一步（硬刪除）"""
    try:
        data = request.get_json() or {}
        reason = data.get('reason', '')
        expected_status = data.get('expected_status')
        expected_history_id = data.get('expected_history_id')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 1. 獲取工作流
        cursor.execute('SELECT * FROM workflows WHERE workflow_number = ?', (workflow_number,))
        workflow = cursor.fetchone()
        if not workflow:
            conn.close()
            return jsonify({'success': False, 'error': '工作流不存在', 'code': 'WORKFLOW_NOT_FOUND'}), 404
        
        workflow = dict(workflow)
        has_permission, error_msg = check_permission('workflow', workflow_number, 'edit')
        if not has_permission:
            conn.close()
            return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
        if not expected_status:
            conn.close()
            return jsonify({'success': False, 'error': '缺少 expected_status 参数', 'code': 'MISSING_EXPECTED_STATUS'}), 400
        if workflow.get('current_status') != expected_status:
            conn.close()
            return jsonify({
                'success': False,
                'error': '当前状态已变更，请刷新后重试',
                'code': 'STATUS_CHANGED',
                'current_status': workflow.get('current_status')
            }), 409
        if expected_history_id:
            latest_history = get_latest_workflow_history(cursor, workflow_number)
            latest_id = latest_history.get('id') if latest_history else None
            if str(latest_id) != str(expected_history_id):
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '状态已被更新，请刷新后重试',
                    'code': 'HISTORY_CHANGED',
                    'latest_history_id': latest_id
                }), 409
        current_ctx = get_current_user_context()
        current_role = current_ctx.get('role', 'viewer')
        current_user_id = current_ctx.get('id')
        if not can_manage_by_owner(workflow.get('handler_id'), user_id=current_user_id):
            conn.close()
            return jsonify({'success': False, 'error': '無權限撤銷此流程', 'code': 'FORBIDDEN'}), 403
        
        # 2. 獲取最後兩步（從 workflow_status_history）
        cursor.execute('''
            SELECT * FROM workflow_status_history 
            WHERE workflow_number = ? 
            ORDER BY created_at ASC, id ASC
        ''', (workflow_number,))
        history = cursor.fetchall()
        
        # 檢查歷史記錄數量
        history_count = len(history)
        if history_count == 0:
            conn.close()
            return jsonify({
                'success': False, 
                'error': '沒有歷史記錄，無法撤銷', 
                'code': 'NO_HISTORY',
                'details': f'工作流 {workflow_number} 沒有任何狀態歷史記錄'
            }), 400
        
        last_step = dict(history[-1])
        previous_step = dict(history[-2]) if history_count >= 2 else None
        if workflow.get('current_status') != last_step.get('to_status'):
            # 兼容歷史順序異常：按當前狀態回溯最近一次匹配的歷史
            cursor.execute('''
                SELECT * FROM workflow_status_history 
                WHERE workflow_number = ? 
                ORDER BY created_at ASC, id ASC
            ''', (workflow_number,))
            full_history = [dict(row) for row in cursor.fetchall()]
            match_index = None
            for idx in range(len(full_history) - 1, -1, -1):
                if full_history[idx].get('to_status') == workflow.get('current_status'):
                    match_index = idx
                    break
            if match_index is None:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '当前状态与最后一步不一致，请刷新后重试',
                    'code': 'STATUS_MISMATCH'
                }), 409
            last_step = full_history[match_index]
            previous_step = full_history[match_index - 1] if match_index > 0 else None

        # 若最後一步是其他人操作，則不允許撤銷（業務員只能撤銷自己操作）
        last_operator_id = last_step.get('operator_id')
        if last_operator_id is None:
            # 兼容旧记录：无操作者时仅允许负责人/管理员撤销
            if not can_manage_by_owner(workflow.get('handler_id'), user_id=current_user_id):
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '此流程最後一步非本人操作，無法撤銷',
                    'code': 'FORBIDDEN'
                }), 403
        else:
            if not can_manage_by_owner(last_operator_id, user_id=current_user_id):
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '此流程最後一步非本人操作，無法撤銷',
                    'code': 'FORBIDDEN'
                }), 403
        
        # NEW_ORDER 不能撤銷
        if last_step.get('to_status') == STATUS_KEYS['NEW_ORDER']:
            conn.close()
            return jsonify({
                'success': False,
                'error': '新訂單狀態不可撤銷',
                'code': 'CANNOT_UNDO_NEW_ORDER'
            }), 400
        
        if not previous_step:
            conn.close()
            return jsonify({
                'success': False,
                'error': '沒有可撤銷的步驟',
                'code': 'INSUFFICIENT_HISTORY'
            }), 400
        
        # 4. 硬刪除最後一步。若「全部出貨」自動產生 COMPLETED，視為同一次操作一起撤回。
        delete_ids = [last_step['id']]
        restore_step = previous_step
        auto_complete_note = '系統自動：已全部出貨後轉為已完成'
        is_auto_completed_pair = (
            last_step.get('to_status') == STATUS_KEYS['COMPLETED']
            and (last_step.get('notes') or '') == auto_complete_note
            and previous_step
            and previous_step.get('to_status') == STATUS_KEYS['ALL_SHIPPED']
            and str(previous_step.get('action_date') or '') == str(last_step.get('action_date') or '')
            and str(previous_step.get('operator_id') or '') == str(last_step.get('operator_id') or '')
        )
        if is_auto_completed_pair:
            delete_ids.append(previous_step['id'])
            # 依實際操作順序找「全部出貨」之前的狀態。
            cursor.execute('''
                SELECT * FROM workflow_status_history
                WHERE workflow_number = ? AND id NOT IN (?, ?)
                ORDER BY created_at DESC, id DESC
                LIMIT 1
            ''', (workflow_number, last_step['id'], previous_step['id']))
            prior = cursor.fetchone()
            restore_step = dict(prior) if prior else None

        if not restore_step:
            conn.close()
            return jsonify({
                'success': False,
                'error': '沒有可撤銷的上一步狀態',
                'code': 'INSUFFICIENT_HISTORY'
            }), 400

        # 3. 記錄到操作日誌：new_status 使用實際恢復的狀態。
        try:
            cursor.execute('''
                INSERT INTO audit_log (
                    action_type, order_number, old_status, new_status,
                    operator, reason
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                'UNDO_STEP',
                workflow['order_number'],
                last_step['to_status'],
                restore_step['to_status'],
                g.current_user.get('username', 'system'),
                reason or ('撤銷全部出貨（含系統自動完成）' if is_auto_completed_pair else '撤銷操作')
            ))
        except Exception as e:
            print(f'警告：無法記錄審計日誌: {e}')

        cursor.executemany('DELETE FROM workflow_status_history WHERE id = ?', [(i,) for i in delete_ids])
        
        # 5. 恢復工作流到上一個有效狀態；日期取該狀態實際 action_date。
        cursor.execute('''
            UPDATE workflows 
            SET current_status = ?,
                status_updated_at = ?,
                status_days = 0,
                updated_at = CURRENT_TIMESTAMP
            WHERE workflow_number = ?
        ''', (restore_step['to_status'], restore_step['action_date'], workflow_number))
        previous_step = restore_step
        
        # 6. 更新燈號
        try:
            cursor.execute('SELECT id FROM orders WHERE order_number = ?', (workflow['order_number'],))
            order_row = cursor.fetchone()
            if order_row:
                update_status_light(order_row['id'], conn)
        except Exception as e:
            print(f'警告：無法更新狀態燈號: {e}')
        
        conn.commit()
        conn.close()
        
        # 獲取狀態顯示文字（用於提示信息）
        previous_status_label = get_status_label(previous_step['to_status'], 'zh_cn') if previous_step.get('to_status') else previous_step.get('to_status', '未知')
        
        return jsonify({
            'success': True,
            'message': f'已撤銷，工作流恢復到「{previous_status_label}」',
            'data': {
                'workflow_number': workflow_number,
                'restored_status': previous_step['to_status']
            }
        })
    
    except sqlite3.Error as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({
            'success': False, 
            'error': '數據庫錯誤', 
            'code': 'DATABASE_ERROR',
            'details': str(e)
        }), 500
    
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({
            'success': False, 
            'error': '撤銷操作失敗', 
            'code': 'UNKNOWN_ERROR',
            'details': str(e)
        }), 500

@tracking_bp.route('/api/orders/<order_number>/history/<int:history_id>', methods=['PUT'])
@api_admin_required
def api_update_history(order_number, history_id):
    """編輯歷史記錄的日期和備註"""
    data = request.get_json() or {}
    action_date = data.get('action_date')
    notes = data.get('notes', '')
    
    if not action_date:
        return jsonify({'success': False, 'error': '日期不能為空'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查歷史記錄是否存在（兼容：status_history 表可能不存在）
    try:
        cursor.execute('''
            SELECT * FROM status_history 
            WHERE id = ? AND order_number = ?
        ''', (history_id, order_number))
        history_record = cursor.fetchone()
    except Exception:
        history_record = None
    
    if not history_record:
        conn.close()
        return jsonify({'success': False, 'error': '歷史記錄不存在'}), 404

    has_permission, error_msg = check_permission('order', order_number, 'edit')
    if not has_permission:
        conn.close()
        return jsonify({'success': False, 'error': error_msg, 'code': 'FORBIDDEN'}), 403
    
    # 更新歷史記錄（兼容：status_history 表可能不存在）
    try:
        cursor.execute('''
            UPDATE status_history 
            SET action_date = ?, notes = ?
            WHERE id = ?
        ''', (action_date, notes, history_id))
    except Exception:
        pass
    
    # 檢查這是否是最後一條歷史記錄（當前狀態）
    try:
        cursor.execute('''
            SELECT id FROM status_history 
            WHERE order_number = ? 
            ORDER BY action_date DESC, id DESC 
            LIMIT 1
        ''', (order_number,))
        latest_history = cursor.fetchone()
    except Exception:
        latest_history = None
    
    # 如果編輯的是最後一條歷史記錄，需要更新訂單的 last_status_change_date
    if latest_history and latest_history['id'] == history_id:
        cursor.execute('''
            UPDATE orders 
            SET last_status_change_date = ?
            WHERE order_number = ?
        ''', (action_date, order_number))
        
        # 重新計算燈號
        cursor.execute('SELECT id FROM orders WHERE order_number = ?', (order_number,))
        order = cursor.fetchone()
        if order:
            from .models import update_status_light
            update_status_light(order['id'], conn)
    
    # 記錄到操作日誌
    cursor.execute('''
        INSERT INTO audit_log (
            action_type, order_number, operator, reason
        ) VALUES (?, ?, ?, ?)
    ''', (
        'EDIT_HISTORY',
        order_number,
        g.current_user.get('username', 'system'),
        f'編輯歷史記錄 #{history_id}'
    ))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': '歷史記錄已更新'
    })

@tracking_bp.route('/api/orders/<order_number>', methods=['PUT'])
@api_login_required
@require_permission('order', 'edit', resource_id_param='order_number')
def api_update_order(order_number):
    """更新訂單信息API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '無效的請求數據'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 檢查訂單是否存在
        cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
        order = cursor.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'error': '訂單不存在'}), 404
        
        order = dict(order)
        
        # 支持部分更新：只更新传入的字段，其他字段保持原值
        update_fields = []
        update_values = []
        
        if 'customer_name' in data:
            data['customer_name'] = _normalize_customer_name(data.get('customer_name'))
            update_fields.append('customer_name = ?')
            update_values.append(data['customer_name'])
        if 'order_date' in data:
            update_fields.append('order_date = ?')
            update_values.append(data['order_date'])
        if 'product_name' in data:
            update_fields.append('product_name = ?')
            update_values.append(data['product_name'])
        if 'product_code' in data:
            update_fields.append('product_code = ?')
            update_values.append(data['product_code'])
        if 'quantity' in data:
            update_fields.append('quantity = ?')
            update_values.append(data['quantity'])
        if 'factory' in data:
            update_fields.append('factory = ?')
            update_values.append(data['factory'])
        if 'production_type' in data:
            update_fields.append('production_type = ?')
            update_values.append(data['production_type'])
        if 'pattern_code' in data:
            update_fields.append('pattern_code = ?')
            update_values.append(data['pattern_code'])
        if 'expected_delivery_date' in data:
            update_fields.append('expected_delivery_date = ?')
            update_values.append(data['expected_delivery_date'])
        if 'notes' in data:
            update_fields.append('notes = ?')
            update_values.append(data['notes'])
        
        # 如果没有要更新的字段，返回错误
        if not update_fields:
            conn.close()
            return jsonify({'success': False, 'error': '沒有提供要更新的字段'}), 400
        
        # 添加 updated_at
        update_fields.append('updated_at = CURRENT_TIMESTAMP')
        update_values.append(order_number)
        
        # 构建并执行更新语句
        update_sql = f'UPDATE orders SET {", ".join(update_fields)} WHERE order_number = ?'
        cursor.execute(update_sql, update_values)
        
        # 只有在更新可能影响状态灯号的字段时才更新燈號
        # 备注、产品信息等字段的更新不影响状态灯号
        fields_affecting_light = ['expected_delivery_date', 'order_date', 'last_status_change_date']
        should_update_light = any(field in data for field in fields_affecting_light)
        
        if should_update_light:
            try:
                update_status_light(order['id'], conn)
            except Exception as light_error:
                # 如果更新灯号失败，记录错误但不影响主更新操作
                print(f"警告: 更新状态灯号失败: {light_error}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '訂單更新成功',
            'data': {'order_number': order_number}
        })
        
    except Exception as e:
        # 确保在异常时关闭数据库连接
        try:
            if 'conn' in locals():
                conn.rollback()
                conn.close()
        except:
            pass
        import traceback
        error_trace = traceback.format_exc()
        print(f"更新订单失败: {error_trace}")
        return jsonify({'success': False, 'error': str(e)}), 500

@tracking_bp.route('/api/orders/<order_number>/change-number', methods=['POST'])
@api_admin_required
def api_change_order_number(order_number):
    """修改訂單號API（同步更新所有相關表）"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '無效的請求數據'}), 400
        
        new_order_number = data.get('new_order_number', '').strip()
        if not new_order_number:
            return jsonify({'success': False, 'error': '新訂單號不能為空'}), 400
        
        if new_order_number == order_number:
            return jsonify({'success': False, 'error': '新訂單號與原訂單號相同'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 檢查原訂單是否存在
        cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
        order = cursor.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'error': '訂單不存在'}), 404
        
        order = dict(order)
        
        # 檢查新訂單號是否已存在
        cursor.execute('SELECT id FROM orders WHERE order_number = ?', (new_order_number,))
        if cursor.fetchone():
            conn.close()
            return jsonify({
                'success': False,
                'error': '新訂單號已存在',
                'code': 'DUPLICATE_ORDER_NUMBER'
            }), 400
        
        # 開始事務：更新所有相關表的訂單號
        try:
            # 1. 更新 orders 表
            cursor.execute('''
                UPDATE orders 
                SET order_number = ?,
                    customer_name = ?,
                    order_date = ?,
                    product_name = ?,
                    product_code = ?,
                    quantity = ?,
                    factory = ?,
                    production_type = ?,
                    pattern_code = ?,
                    expected_delivery_date = ?,
                    notes = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE order_number = ?
            ''', (
                new_order_number,
                _normalize_customer_name(data.get('customer_name')),
                data.get('order_date'),
                data.get('product_name', data.get('product_code', '')),
                data.get('product_code', ''),
                data.get('quantity', ''),
                data.get('factory', ''),
                data.get('production_type', ''),
                data.get('pattern_code', ''),
                data.get('expected_delivery_date'),
                data.get('notes', ''),
                order_number
            ))
            
            # 2. 更新 status_history 表（兼容：status_history 表可能不存在）
            try:
                cursor.execute('''
                    UPDATE status_history 
                    SET order_number = ?
                    WHERE order_number = ?
                ''', (new_order_number, order_number))
            except Exception:
                pass
            
            # 3. 更新 audit_log 表
            cursor.execute('''
                UPDATE audit_log 
                SET order_number = ?
                WHERE order_number = ?
            ''', (new_order_number, order_number))
            
            # 4. 記錄操作日誌
            cursor.execute('''
                INSERT INTO audit_log (
                    action_type, order_number, old_status, new_status,
                    operator, reason
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                'CHANGE_ORDER_NUMBER',
                new_order_number,
                order_number,
                new_order_number,
                g.current_user.get('username', 'system'),
                f'訂單號從 {order_number} 修改為 {new_order_number}'
            ))
            
            # 更新燈號
            update_status_light(order['id'], conn)
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': f'訂單號已從 {order_number} 修改為 {new_order_number}',
                'data': {
                    'old_order_number': order_number,
                    'new_order_number': new_order_number
                }
            })
            
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({
                'success': False,
                'error': f'修改訂單號失敗：{str(e)}',
                'code': 'DATABASE_ERROR'
            }), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@tracking_bp.route('/api/orders/<order_number>', methods=['DELETE'])

@api_admin_required
def api_delete_order(order_number):
    """刪除訂單（僅管理員）"""
    data = request.get_json() or {}
    confirm_order_number = data.get('confirm_order_number', '')
    reason = data.get('reason', '')
    
    # 驗證訂單號確認
    if confirm_order_number != order_number:
        return jsonify({
            'success': False,
            'error': '訂單號不匹配，請重新輸入'
        }), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查訂單是否存在
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({
            'success': False,
            'error': '訂單不存在'
        }), 404
    
    order = dict(order)
    
    # 获取订单状态（从 orders 表的 status 字段，或从 workflows 表获取业务状态）
    order_status = order.get('status', 'UNLOCKED')
    
    # 查询是否有相关的 workflows
    cursor.execute('SELECT current_status FROM workflows WHERE order_number = ? LIMIT 1', (order_number,))
    workflow = cursor.fetchone()
    workflow_status = dict(workflow)['current_status'] if workflow else None
    
    try:
        # 記錄到操作日誌（刪除前）
        cursor.execute('''
            INSERT INTO audit_log (
                action_type, order_number, old_status, new_status,
                operator, reason
            ) VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            'DELETE_ORDER',
            order_number,
            workflow_status or order_status,
            'DELETED',
            g.current_user.get('username', 'system'),
            reason or '刪除訂單'
        ))
        
        # 刪除相關的 workflows（CASCADE 應該自動處理，但為了安全手動刪除）
        cursor.execute('DELETE FROM workflows WHERE order_number = ?', (order_number,))
        
        # 刪除訂單文件
        cursor.execute('DELETE FROM order_files WHERE order_number = ?', (order_number,))
        
        # 刪除訂單備註
        cursor.execute('DELETE FROM order_notes WHERE order_number = ?', (order_number,))
        
        # 刪除狀態歷史（workflow_status_history）
        cursor.execute('DELETE FROM workflow_status_history WHERE order_number = ?', (order_number,))
        
        # 刪除通知
        cursor.execute('DELETE FROM notifications WHERE order_number = ?', (order_number,))
        
        # 刪除操作日誌（可選，通常保留）
        # cursor.execute('DELETE FROM operation_logs WHERE order_number = ?', (order_number,))
        
        # 刪除訂單（最後刪除，因為有外鍵約束）
        cursor.execute('DELETE FROM orders WHERE order_number = ?', (order_number,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'訂單 {order_number} 已刪除',
            'data': {
                'order_number': order_number,
                'deleted_status': workflow_status or order_status
            }
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({
            'success': False,
            'error': f'刪除失敗：{str(e)}'
        }), 500


@tracking_bp.route('/api/orders/batch-delete', methods=['POST'])
@api_admin_required
def api_batch_delete_orders():
    """批量刪除訂單（僅管理員）"""
    data = request.get_json() or {}
    order_numbers = data.get('order_numbers', [])
    reason = data.get('reason', '批量刪除')
    
    if not order_numbers or not isinstance(order_numbers, list):
        return jsonify({
            'success': False,
            'error': '請提供要刪除的訂單號列表'
        }), 400
    
    if len(order_numbers) == 0:
        return jsonify({
            'success': False,
            'error': '訂單號列表為空'
        }), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    deleted_count = 0
    not_found = []
    errors = []
    
    for order_number in order_numbers:
        try:
            # 檢查訂單是否存在
            cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
            order = cursor.fetchone()
            
            if not order:
                not_found.append(order_number)
                continue
            
            order = dict(order)
            
            # 获取订单状态（从 orders 表的 status 字段，或从 workflows 表获取业务状态）
            order_status = order.get('status', 'UNLOCKED')
            
            # 查询是否有相关的 workflows
            cursor.execute('SELECT current_status FROM workflows WHERE order_number = ? LIMIT 1', (order_number,))
            workflow = cursor.fetchone()
            workflow_status = dict(workflow)['current_status'] if workflow else None
            
            # 記錄到操作日誌（刪除前）
            cursor.execute('''
                INSERT INTO audit_log (
                    action_type, order_number, old_status, new_status,
                    operator, reason
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                'DELETE_ORDER',
                order_number,
                workflow_status or order_status,
                'DELETED',
                g.current_user.get('username', 'system'),
                reason
            ))
            
            # 刪除相關的 workflows
            cursor.execute('DELETE FROM workflows WHERE order_number = ?', (order_number,))
            
            # 刪除訂單文件
            cursor.execute('DELETE FROM order_files WHERE order_number = ?', (order_number,))
            
            # 刪除訂單備註
            cursor.execute('DELETE FROM order_notes WHERE order_number = ?', (order_number,))
            
            # 刪除狀態歷史（workflow_status_history）
            cursor.execute('DELETE FROM workflow_status_history WHERE order_number = ?', (order_number,))
            
            # 刪除通知
            cursor.execute('DELETE FROM notifications WHERE order_number = ?', (order_number,))
            
            # 刪除訂單（最後刪除，因為有外鍵約束）
            cursor.execute('DELETE FROM orders WHERE order_number = ?', (order_number,))
            
            deleted_count += 1
            
        except Exception as e:
            errors.append(f'{order_number}: {str(e)}')
            continue
    
    try:
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'成功刪除 {deleted_count} 個訂單',
            'data': {
                'deleted_count': deleted_count,
                'not_found': not_found,
                'errors': errors if errors else None
            }
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({
            'success': False,
            'error': f'批量刪除失敗：{str(e)}'
        }), 500


@tracking_bp.route('/api/search', methods=['GET'])
@login_required
def api_global_search():
    """全局搜索API - 搜索当前权限范围内的流程，并补充无流程订单"""
    keyword = request.args.get('q', '').strip()

    if cloud_mode_enabled():
        ctx = get_current_user_context()
        try:
            rows = _load_home_orders_from_active_source(ctx.get('role', 'viewer'), ctx.get('id'))
        except Exception as exc:
            return jsonify({'success': False, 'error': f'雲端搜尋失敗: {exc}'}), 503
        needle = keyword.casefold()
        if needle:
            rows = [row for row in rows if needle in str(row.get('order_number') or '').casefold()
                    or needle in str(row.get('workflow_number') or '').casefold()
                    or needle in str(row.get('customer_name') or '').casefold()]
        else:
            rows = [row for row in rows if str(row.get('current_status') or '') not in {STATUS_KEYS['COMPLETED'], STATUS_KEYS['CANCELLED']}]
        rows = rows[:200]
        return jsonify({
            'success': True, 'type': 'search' if keyword else 'recent', 'keyword': keyword,
            'orders': rows, 'total': len(rows), 'limit_reached': len(rows) >= 200
        })

    conn = get_db()
    cursor = conn.cursor()

    current_ctx = get_current_user_context()
    current_role = current_ctx.get('role', 'viewer')
    current_user_id = current_ctx.get('id')
    filter_info = get_filtered_resources('workflow', current_role, current_user_id)
    admin = is_admin()

    try:
        # ── Part 1：有流程的訂單（原有邏輯）──
        query = """
            SELECT w.*, o.customer_name, o.order_date, o.status as order_status, o.visibility,
                   COALESCE(u.real_name, u.display_name, u.username) as handler_name
            FROM workflows w
            INNER JOIN orders o ON w.order_number = o.order_number
            LEFT JOIN users u ON w.handler_id = u.id
        """
        params = []

        if keyword:
            query += " WHERE (w.order_number LIKE ? OR w.workflow_number LIKE ? OR o.customer_name LIKE ?)"
            params.extend([f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'])

        if not keyword:
            base_where = " WHERE w.current_status NOT IN (?, ?)"
            query += base_where if "WHERE" not in query else f" AND{base_where[6:]}"
            params.extend([STATUS_KEYS['COMPLETED'], STATUS_KEYS['CANCELLED']])

        if not admin and filter_info['rule'] != 'all':
            visibility_clause = get_visibility_where_clause(current_role, table_alias='o')
            visibility_sql = f" AND {visibility_clause}" if visibility_clause else ""
            query += f" AND {filter_info['where_sql']}{visibility_sql}"
            params.extend(filter_info['params'])

        count_query = f"SELECT COUNT(*) as count FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['count']

        query += " ORDER BY w.created_at DESC, w.order_number DESC LIMIT 200"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        orders_list = []
        seen_order_numbers = set()
        for row in rows:
            workflow = dict(row)
            seen_order_numbers.add(workflow.get('order_number'))
            order_obj = {
                'order_number': workflow.get('order_number'),
                'order_date': workflow.get('order_date'),
                'expected_delivery_date': workflow.get('expected_delivery_date'),
                'status_updated_at': workflow.get('status_updated_at'),
                'current_status': workflow.get('current_status'),
                'last_status_change_date': workflow.get('status_updated_at')
            }
            workflow['status_light'] = calculate_status_light(order_obj)
            workflow['current_status'] = workflow.get('current_status', '')
            workflow['status_days'] = workflow.get('status_days', 0)
            workflow['product_name'] = workflow.get('product_name', '')
            workflow['product_code'] = workflow.get('product_code', '')
            workflow['quantity'] = workflow.get('quantity', '')
            workflow['factory'] = workflow.get('factory', '')
            workflow['production_type'] = workflow.get('production_type', '')
            workflow['expected_delivery_date'] = workflow.get('expected_delivery_date', '')
            workflow['is_locked'] = _normalize_lock_flag(workflow.get('is_locked'))
            # 訂單號層級狀態（主管設定，如 ACTIVE / CANCELLED），與業務員的 current_status 分開，
            # 明確帶給前端：訂單被取消時優先顯示「已取消」，復原時 current_status 不受影響直接顯示回來。
            workflow['order_status'] = workflow.get('order_status', '')
            owner_id = workflow.get('handler_id') or workflow.get('created_by_id') or workflow.get('created_by')
            owner_id_str = str(owner_id) if owner_id is not None else ''
            current_id_str = str(current_user_id) if current_user_id is not None else ''
            workflow['can_edit_notes'] = bool(owner_id_str and owner_id_str == current_id_str) and not workflow['is_locked']
            workflow['no_workflow'] = False
            orders_list.append(workflow)

        # ── Part 2：有關鍵字時，補查無流程 & 別人有流程的訂單 ──
        if keyword:
            # 2a. 無流程訂單
            no_wf_query = """
                SELECT o.order_number, o.customer_name, o.order_date, o.status as order_status,
                       o.visibility
                FROM orders o
                LEFT JOIN workflows w ON o.order_number = w.order_number
                WHERE w.id IS NULL
                  AND o.status = 'ACTIVE'
                  AND o.status != 'CANCELLED'
                  AND (o.order_number LIKE ? OR o.customer_name LIKE ?)
            """
            no_wf_params = [f'%{keyword}%', f'%{keyword}%']
            if not admin:
                no_wf_query += " AND o.visibility = 'all_sales'"

            cursor.execute(no_wf_query, no_wf_params)
            for row in cursor.fetchall():
                o = dict(row)
                on = o.get('order_number')
                if on in seen_order_numbers:
                    continue
                seen_order_numbers.add(on)
                orders_list.append({
                    'order_number': on,
                    'workflow_number': '',
                    'customer_name': o.get('customer_name', ''),
                    'order_date': o.get('order_date', ''),
                    'order_status': o.get('order_status', ''),
                    'visibility': o.get('visibility', ''),
                    'current_status': '',
                    'status_light': 'grey',
                    'handler_name': '',
                    'handler_id': None,
                    'product_name': '', 'product_code': '',
                    'quantity': '', 'factory': '',
                    'production_type': '', 'expected_delivery_date': '',
                    'status_days': 0, 'is_locked': False,
                    'can_edit_notes': False,
                    'no_workflow': True,
                    'workflow_count': 0,
                })

            # 2b. Sales 搜到別人有流程的訂單（自己沒有流程）
            if not admin:
                others_query = """
                    SELECT o.order_number, o.customer_name, o.order_date, o.visibility,
                           COUNT(w.id) as workflow_count
                    FROM orders o
                    INNER JOIN workflows w ON o.order_number = w.order_number
                    WHERE o.status = 'ACTIVE'
                      AND o.visibility = 'all_sales'
                      AND (o.order_number LIKE ? OR o.customer_name LIKE ?)
                      AND w.handler_id != ?
                    GROUP BY o.order_number
                """
                cursor.execute(others_query, [f'%{keyword}%', f'%{keyword}%', current_user_id])
                for row in cursor.fetchall():
                    o = dict(row)
                    on = o.get('order_number')
                    if on in seen_order_numbers:
                        continue  # 自己也有流程的不重複
                    orders_list.append({
                        'order_number': on,
                        'workflow_number': '',
                        'customer_name': o.get('customer_name', ''),
                        'order_date': o.get('order_date', ''),
                        'order_status': 'ACTIVE',
                        'visibility': o.get('visibility', ''),
                        'current_status': '',
                        'status_light': 'grey',
                        'handler_name': '',
                        'handler_id': None,
                        'product_name': '', 'product_code': '',
                        'quantity': '', 'factory': '',
                        'production_type': '', 'expected_delivery_date': '',
                        'status_days': 0, 'is_locked': False,
                        'can_edit_notes': False,
                        'no_workflow': True,
                        'workflow_count': o.get('workflow_count', 0),  # 有幾個別人的流程
                        'others_workflow': True,  # 標記是別人的流程
                    })

        conn.close()

        return jsonify({
            'success': True,
            'type': 'search' if keyword else 'recent',
            'keyword': keyword,
            'orders': orders_list,
            'total': total if total < 200 else 200,
            'limit_reached': total >= 200
        })
    except Exception as e:
        conn.close()
        return jsonify({
            'success': False,
            'error': f'搜索失败：{str(e)}'
        }), 500


# ==================== 新增：用戶管理 API（M1）====================

@tracking_bp.route('/api/users', methods=['GET'])
@admin_required
def get_users_api():
    """獲取所有用戶列表（主管專用）- 包括所有狀態"""
    conn = get_db()
    cursor = conn.cursor()
    
    role = request.args.get('role')
    search = request.args.get('search')
    status_filter = request.args.get('status')
    
    # 查詢所有用戶（包括pending/rejected/active）
    query = 'SELECT * FROM users WHERE 1=1'
    params = []
    
    if role:
        query += ' AND role = ?'
        params.append(role)
    
    if status_filter:
        query += ' AND status = ?'
        params.append(status_filter)
    
    if search:
        query += ' AND (username LIKE ? OR display_name LIKE ? OR real_name LIKE ?)'
        params.extend([f'%{search}%'] * 3)
    
    query += ' ORDER BY created_at DESC'
    
    cursor.execute(query, params)
    users = cursor.fetchall()
    
    result = []
    for user in users:
        # 統計該用戶負責的流程數（從 workflows 表）
        cursor.execute('SELECT COUNT(*) as count FROM workflows WHERE handler_id = ?', (user['id'],))
        product_count = cursor.fetchone()['count']
        
        # 獲取real_name、employee_id和status
        real_name = user['display_name']  # 默认值
        try:
            if user['real_name']:
                real_name = user['real_name']
        except (KeyError, IndexError):
            pass
        
        try:
            employee_id = user['employee_id']
            if not employee_id:
                # 如果沒有員工ID，生成一個（格式：EMP001）
                employee_id = f"EMP{user['id']:03d}"
        except (KeyError, IndexError):
            employee_id = f"EMP{user['id']:03d}"
        
        try:
            user_status = user['status']
            if not user_status:
                user_status = 'active'  # 兼容旧数据
        except (KeyError, IndexError):
            user_status = 'active'
        
        result.append({
            'user_id': user['id'],
            'employee_id': employee_id,
            'username': user['username'],
            'display_name': user['display_name'],
            'real_name': real_name,
            'role': user['role'],
            'status': user_status,
            'created_at': user['created_at'],
            'product_count': product_count,  # 保持向后兼容
            'workflow_count': product_count   # 新增：流程数量
        })
    
    conn.close()
    return jsonify({'success': True, 'data': result})


@tracking_bp.route('/api/users', methods=['POST'])
@admin_required
def create_user_api():
    """創建用戶（主管專用）"""
    data = request.get_json()
    
    if not data.get('username') or not data.get('password'):
        return jsonify({'success': False, 'error': '用戶名和密碼不能為空'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查用戶名是否已存在
    cursor.execute('SELECT id FROM users WHERE username = ?', (data['username'],))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'error': '用戶名已存在'}), 400
    
    # 生成員工ID：查找所有員工ID，找到數字最大的那個，然後+1
    cursor.execute('SELECT employee_id FROM users WHERE employee_id IS NOT NULL AND employee_id LIKE "EMP%"')
    existing_ids = cursor.fetchall()
    
    max_num = 0
    for row in existing_ids:
        emp_id = row['employee_id']
        if emp_id and emp_id.startswith('EMP'):
            try:
                # 提取數字部分（EMP001 -> 001 -> 1）
                num_str = emp_id[3:]  # 去掉 "EMP" 前綴
                num = int(num_str)
                if num > max_num:
                    max_num = num
            except (ValueError, IndexError):
                continue
    
    # 生成新的員工ID
    new_employee_id = f"EMP{max_num + 1:03d}"
    
    # 創建用戶
    from werkzeug.security import generate_password_hash
    password_hash = generate_password_hash(data['password'])
    
    cursor.execute('''
        INSERT INTO users (username, password_hash, display_name, real_name, role, employee_id, status)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['username'],
        password_hash,
        data.get('display_name', data['username']),
        data.get('real_name', data.get('display_name', data['username'])),
        data.get('role', 'sales'),
        new_employee_id,
        data.get('status', 'active')
    ))
    
    user_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'user_id': user_id,
            'username': data['username'],
            'role': data.get('role', 'sales')
        }
    })


@tracking_bp.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user_api(user_id):
    """更新用戶資料（主管專用）"""
    data = request.get_json()
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查用戶是否存在
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用戶不存在'}), 404
    
    # 更新欄位
    updates = []
    params = []
    
    if 'display_name' in data:
        updates.append('display_name = ?')
        params.append(data['display_name'])
    
    if 'real_name' in data:
        updates.append('real_name = ?')
        params.append(data['real_name'])
        # 同步更新 display_name
        updates.append('display_name = ?')
        params.append(data['real_name'])
    
    if 'role' in data:
        if data['role'] not in ['admin', 'sales', 'viewer']:
            conn.close()
            return jsonify({'success': False, 'error': '無效的角色'}), 400
        updates.append('role = ?')
        params.append(data['role'])
    
    if 'status' in data:
        if data['status'] not in ['pending', 'active', 'rejected']:
            conn.close()
            return jsonify({'success': False, 'error': '無效的狀態'}), 400
        updates.append('status = ?')
        params.append(data['status'])
    
    if updates:
        query = f"UPDATE users SET {', '.join(updates)} WHERE id = ?"
        params.append(user_id)
        cursor.execute(query, params)
        conn.commit()
    
    conn.close()
    return jsonify({'success': True, 'message': '用戶資料已更新'})


@tracking_bp.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_user_password_api(user_id):
    """重設用戶密碼（主管專用）- 設置需要用戶確認密碼"""
    data = request.get_json()
    new_password = data.get('new_password')
    
    # 如果提供了新密碼，直接設置；否則設置為需要用戶確認
    conn = get_db()
    cursor = conn.cursor()
    
    if new_password:
        if len(new_password) < 6:
            conn.close()
            return jsonify({'success': False, 'error': '密碼至少6位'}), 400
        
        from werkzeug.security import generate_password_hash
        password_hash = generate_password_hash(new_password)
        cursor.execute('''
            UPDATE users 
            SET password_hash = ?, needs_password_reset = 0 
            WHERE id = ?
        ''', (password_hash, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '密碼已重設'})
    else:
        # 只設置重置標記，讓用戶下次登入時自己設置密碼
        cursor.execute('''
            UPDATE users 
            SET needs_password_reset = 1 
            WHERE id = ?
        ''', (user_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '已設置密碼重置標記，用戶下次登入時需設置新密碼'})


# ==================== 用户自助修改密码 =====================

@tracking_bp.route('/api/auth/change-password', methods=['POST'])
@api_login_required
def api_change_password():
    """用户自己修改密码"""
    data = request.get_json() or {}
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')
    confirm_password = data.get('confirm_password', '')

    if not old_password:
        return jsonify({'success': False, 'error': '请输入旧密码'}), 400
    if not new_password:
        return jsonify({'success': False, 'error': '请输入新密码'}), 400
    if len(new_password) < 6:
        return jsonify({'success': False, 'error': '新密码至少6位'}), 400
    if new_password != confirm_password:
        return jsonify({'success': False, 'error': '两次输入的新密码不一致'}), 400
    if old_password == new_password:
        return jsonify({'success': False, 'error': '新密码不能与旧密码相同'}), 400

    user_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT password_hash FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()

    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用户不存在'}), 404

    if not check_password_hash(user['password_hash'], old_password):
        conn.close()
        return jsonify({'success': False, 'error': '旧密码不正确'}), 400

    from werkzeug.security import generate_password_hash
    new_hash = generate_password_hash(new_password)
    cursor.execute('UPDATE users SET password_hash = ?, needs_password_reset = 0 WHERE id = ?', (new_hash, user_id))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '密码修改成功'})


# ==================== 註冊和審核 API =====================

@tracking_bp.route('/api/register', methods=['POST'])
def register_api():
    """用戶註冊（需主管審核）"""
    data = request.get_json()
    
    if not all([data.get('username'), data.get('password'), data.get('real_name')]):
        return jsonify({'success': False, 'error': '請填寫所有必填欄位'}), 400
    
    if len(data['password']) < 6:
        return jsonify({'success': False, 'error': '密碼至少6位'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查用戶名是否已存在
    cursor.execute('SELECT id FROM users WHERE username = ?', (data['username'],))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'error': '用戶名已存在'}), 400
    
    # 創建用戶（狀態為 pending，等待審核）
    from werkzeug.security import generate_password_hash
    password_hash = generate_password_hash(data['password'])
    
    cursor.execute('''
        INSERT INTO users (username, password_hash, display_name, real_name, role, status)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (
        data['username'],
        password_hash,
        data['real_name'],  # display_name 使用 real_name
        data['real_name'],
        'viewer',  # 默認角色，等待主管審核後決定
        'pending'  # 待審核狀態
    ))
    
    user_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': '註冊成功！請等待主管審核，審核通過後即可登入。',
        'user_id': user_id
    })


@tracking_bp.route('/api/users/pending', methods=['GET'])
@admin_required
def get_pending_users_api():
    """獲取待審核用戶列表（主管專用）"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, username, real_name, display_name, created_at 
        FROM users 
        WHERE status = 'pending'
        ORDER BY created_at DESC
    ''')
    
    users = cursor.fetchall()
    result = []
    for user in users:
        real_name = user['display_name']
        try:
            if user['real_name']:
                real_name = user['real_name']
        except (KeyError, IndexError):
            pass
        
        result.append({
            'user_id': user['id'],
            'username': user['username'],
            'real_name': real_name,
            'created_at': user['created_at']
        })
    
    conn.close()
    return jsonify({'success': True, 'data': result})


@tracking_bp.route('/api/users/<int:user_id>/approve', methods=['POST'])
@admin_required
def approve_user_api(user_id):
    """審核通過用戶（主管專用）- 可從pending或rejected切換到active"""
    data = request.get_json()
    role = data.get('role', 'sales')  # 默認為業務員
    
    if role not in ['admin', 'sales', 'viewer']:
        return jsonify({'success': False, 'error': '無效的角色'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查用戶是否存在
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用戶不存在'}), 404
    
    try:
        user_status = user['status']
        if not user_status:
            user_status = 'active'
    except (KeyError, IndexError):
        user_status = 'active'
    
    # 允許從pending或rejected切換到active
    if user_status not in ['pending', 'rejected']:
        conn.close()
        return jsonify({'success': False, 'error': '該用戶狀態不是待審核或已拒絕'}), 400
    
    # 更新用戶狀態和角色
    cursor.execute('''
        UPDATE users 
        SET status = 'active', role = ?
        WHERE id = ?
    ''', (role, user_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '用戶審核通過'})


@tracking_bp.route('/api/users/<int:user_id>/reject', methods=['POST'])
@admin_required
def reject_user_api(user_id):
    """拒絕用戶註冊（主管專用）- 可從pending或active切換到rejected"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 檢查用戶是否存在
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用戶不存在'}), 404
    
    try:
        user_status = user['status']
        if not user_status:
            user_status = 'active'
    except (KeyError, IndexError):
        user_status = 'active'
    
    # 允許從pending或active切換到rejected
    if user_status == 'rejected':
        conn.close()
        return jsonify({'success': False, 'error': '該用戶已經是拒絕狀態'}), 400
    
    # 更新用戶狀態為 rejected
    cursor.execute('UPDATE users SET status = ? WHERE id = ?', ('suspended', user_id))

    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '用戶狀態已設為拒絕'})


@tracking_bp.route('/users')
@admin_required
def admin_users():
    """用戶管理頁面（主管專用）"""
    return render_template('tracking/users.html')


@tracking_bp.route('/admin/orders')
@login_required
def admin_orders():
    """訂單管理頁面（Admin / 業務員共用）"""
    return render_template('tracking/admin_orders.html')


# 添加到 __init__.py 的 admin API 區域

@tracking_bp.route('/api/users/<int:user_id>/suspend', methods=['POST'])
@admin_required
def suspend_user_api(user_id):
    """停權用戶（主管專用）"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用戶不存在'}), 404
    
    cursor.execute('UPDATE users SET status = ? WHERE id = ?', ('suspended', user_id))

    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '用戶已停權'})


@tracking_bp.route('/api/users/<int:user_id>/restore', methods=['POST'])
@admin_required
def restore_user_api(user_id):
    """恢復用戶（主管專用）"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用戶不存在'}), 404
    
    cursor.execute('UPDATE users SET status = ? WHERE id = ?', ('active', user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '用戶已恢復'})


@tracking_bp.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user_api(user_id):
    """刪除用戶（主管專用）"""
    if session.get('user_id') and int(session.get('user_id')) == user_id:
        return jsonify({'success': False, 'error': '不能删除当前登录用户'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': '用戶不存在'}), 404

    cursor.execute('SELECT COUNT(*) as count FROM workflows WHERE handler_id = ?', (user_id,))
    workflow_count = cursor.fetchone()['count']
    if workflow_count and workflow_count > 0:
        conn.close()
        return jsonify({'success': False, 'error': '該用戶仍有負責流程，請先移轉'}), 400

    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '用戶已刪除'})

# ==================== M0: 訂單號碼池管理 API ====================

def _normalize_order_prefix(prefix):
    return (prefix or '').strip().upper()


def _extract_order_number_parts(order_number):
    text = str(order_number or '').strip()
    prefix = ''.join(filter(str.isalpha, text)).upper()
    digits = ''.join(filter(str.isdigit, text))
    return prefix, digits


def _build_prefix_sql_filter(prefix_value, column_name='order_number'):
    prefix = _normalize_order_prefix(prefix_value)
    if not prefix or prefix == 'ALL':
        return None, []
    if prefix == 'NUMERIC':
        return f"{column_name} GLOB '[0-9]*'", []
    if prefix in ('KC', 'G'):
        return f"{column_name} LIKE ?", [f"{prefix}%"]
    return None, []


def _query_prefixed_max_serial(cursor, prefix, year):
    cursor.execute(
        'SELECT order_number FROM orders WHERE order_number LIKE ?',
        (f'{prefix}{year}%',)
    )
    max_serial = 0
    for row in cursor.fetchall():
        order_number = (row['order_number'] or '').strip().upper()
        suffix = order_number[len(prefix) + 2:]
        if suffix.isdigit():
            max_serial = max(max_serial, int(suffix))
    return max_serial


def _is_complete_order_number_token(token):
    text = str(token or '').strip()
    if not text:
        return False
    prefix, digits = _extract_order_number_parts(text)
    if not digits:
        return False
    return f'{prefix}{digits}'.upper() == text.upper()


def _build_order_pool_search_filter(search_value, order_col='o.order_number', customer_col='o.customer_name'):
    search_text = (search_value or '').strip()
    if not search_text:
        return {'ok': True, 'sql': None, 'params': []}

    if '~' in search_text:
        left_text, right_part = search_text.split('~', 1)
        start_token = left_text.strip()
        right_part = right_part.strip()

        if not start_token or not right_part:
            return {'ok': False, 'error': '范围搜索格式错误，请使用「起始号~结束号」'}

        right_segments = right_part.split(None, 1)
        end_token = right_segments[0].strip()
        customer_keyword = right_segments[1].strip() if len(right_segments) > 1 else ''

        if (not _is_complete_order_number_token(start_token)) or (not _is_complete_order_number_token(end_token)):
            return {'ok': False, 'error': '请输入完整订单号'}

        start_prefix, start_digits = _extract_order_number_parts(start_token)
        end_prefix, end_digits = _extract_order_number_parts(end_token)
        if start_prefix != end_prefix:
            return {'ok': False, 'error': '范围两端类型必须相同'}

        if int(start_digits) > int(end_digits):
            return {'ok': False, 'error': '起始号码不能大于结束号码'}

        start_number = f'{start_prefix}{start_digits}'
        end_number = f'{end_prefix}{end_digits}'
        sql = f'{order_col} BETWEEN ? AND ?'
        params = [start_number, end_number]
        if customer_keyword:
            sql += f' AND {customer_col} LIKE ?'
            params.append(f'%{customer_keyword}%')
        return {'ok': True, 'sql': sql, 'params': params}

    # 无范围符，按前缀或客户名称搜索；若包含空格则支持「订单号 + 客户名」组合搜索
    parts = search_text.split(None, 1)
    first_term = parts[0]
    second_term = parts[1].strip() if len(parts) > 1 else ''
    has_digit = any(ch.isdigit() for ch in first_term)
    upper_text = first_term.upper()

    if first_term.isdigit():
        if second_term:
            return {
                'ok': True,
                'sql': f'{order_col} LIKE ? AND {customer_col} LIKE ?',
                'params': [f'%{first_term}%', f'%{second_term}%']
            }
        # 纯数字输入按模糊匹配，支持命中 KC/G/纯数字三类（如 26500 -> KC26500/G26500/26500）
        return {'ok': True, 'sql': f'{order_col} LIKE ?', 'params': [f'%{first_term}%']}

    if has_digit or upper_text.startswith('KC') or upper_text.startswith('G'):
        if second_term:
            return {
                'ok': True,
                'sql': f'{order_col} LIKE ? AND {customer_col} LIKE ?',
                'params': [f'{upper_text}%', f'%{second_term}%']
            }
        return {'ok': True, 'sql': f'{order_col} LIKE ?', 'params': [f'{upper_text}%']}

    return {'ok': True, 'sql': f'{customer_col} LIKE ?', 'params': [f'%{search_text}%']}


@tracking_bp.route('/api/order-number-pool/next-preview', methods=['GET'])
@admin_required
def api_order_number_next_preview():
    prefix = _normalize_order_prefix(request.args.get('prefix', ''))
    year = (request.args.get('year', '') or '').strip()

    if prefix not in ('KC', 'G'):
        return jsonify({'success': False, 'error': 'prefix 只支持 KC 或 G'}), 400
    if not (len(year) == 2 and year.isdigit()):
        return jsonify({'success': False, 'error': 'year 必须是2位数字'}), 400

    conn = get_db()
    cursor = conn.cursor()
    max_serial = _query_prefixed_max_serial(cursor, prefix, year)
    conn.close()

    return jsonify({
        'success': True,
        'data': {
            'prefix': prefix,
            'year': year,
            'max_serial': max_serial
        }
    })


@tracking_bp.route('/api/order-number-pool/generate', methods=['POST'])
@admin_required
def api_generate_order_numbers():
    """生成订单号段（重构后：使用 orders 表）"""
    data = request.get_json() or {}
    prefix_mode = _normalize_order_prefix(data.get('prefix', ''))
    conn = get_db()
    cursor = conn.cursor()
    created_by = session.get('user_id') or session.get('username', 'admin')
    created_count = 0
    skipped_count = 0
    skipped_numbers = []
    skipped_sample_limit = 200

    if prefix_mode in ('KC', 'G'):
        year = str(data.get('year', '') or '').strip()
        quantity = data.get('quantity', 0)
        try:
            quantity = int(quantity)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'quantity 必须是数字'}), 400

        if not (len(year) == 2 and year.isdigit()):
            return jsonify({'success': False, 'error': 'year 必须是2位数字'}), 400
        if quantity <= 0:
            return jsonify({'success': False, 'error': 'quantity 必须大于 0'}), 400

        max_serial = _query_prefixed_max_serial(cursor, prefix_mode, year)
        start_serial = max_serial + 1
        end_serial = start_serial + quantity - 1
        if end_serial > 999:
            conn.close()
            return jsonify({
                'success': False,
                'error': f'{prefix_mode}{year} 年度号段已不足，当前最大为 {max_serial:03d}，最多只到 999'
            }), 400

        for serial in range(start_serial, end_serial + 1):
            order_number = f'{prefix_mode}{year}{serial:03d}'
            try:
                cursor.execute('''
                    INSERT INTO orders (
                        order_number, status, visibility, created_at, created_by
                    ) VALUES (?, 'UNLOCKED', 'admin_only', CURRENT_TIMESTAMP, ?)
                ''', (order_number, created_by))
                created_count += 1
            except sqlite3.IntegrityError:
                skipped_count += 1
                if len(skipped_numbers) < skipped_sample_limit:
                    skipped_numbers.append(order_number)
    else:
        start_number = str(data.get('start_number', '') or '').strip()
        end_number = str(data.get('end_number', '') or '').strip()
        if not start_number or not end_number:
            conn.close()
            return jsonify({'success': False, 'error': '缺少必要参数'}), 400

        start_prefix, start_digits = _extract_order_number_parts(start_number)
        end_prefix, end_digits = _extract_order_number_parts(end_number)
        if start_prefix != end_prefix or not start_digits or not end_digits:
            conn.close()
            return jsonify({'success': False, 'error': '号码格式错误'}), 400
        try:
            start_num = int(start_digits)
            end_num = int(end_digits)
        except ValueError:
            conn.close()
            return jsonify({'success': False, 'error': '号码格式错误'}), 400

        if start_num > end_num:
            conn.close()
            return jsonify({'success': False, 'error': '起始号码不能大于结束号码'}), 400
        if end_num - start_num > 10000:
            conn.close()
            return jsonify({'success': False, 'error': '单次生成不能超过 10000 个号码'}), 400

        for i in range(start_num, end_num + 1):
            order_number = f"{start_prefix}{i}" if start_prefix else str(i)
            cursor.execute('SELECT order_number FROM orders WHERE order_number = ?', (order_number,))
            if cursor.fetchone():
                skipped_count += 1
                if len(skipped_numbers) < skipped_sample_limit:
                    skipped_numbers.append(order_number)
                continue
            cursor.execute('''
                INSERT INTO orders (
                    order_number, status, visibility, created_at, created_by
                ) VALUES (?, 'UNLOCKED', 'admin_only', CURRENT_TIMESTAMP, ?)
            ''', (order_number, created_by))
            created_count += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'成功生成 {created_count} 个号码，跳过 {skipped_count} 个已存在的号码',
        'data': {
            'created': created_count,
            'skipped': skipped_count,
            'skipped_numbers': skipped_numbers,
            'skipped_truncated': skipped_count > skipped_sample_limit
        }
    })


@tracking_bp.route('/api/order-number-pool', methods=['GET'])
@admin_required
def api_get_order_number_pool():
    """获取订单号列表（重构后：使用 orders 表）"""
    import sys
    import traceback
    
    try:
        page = max(1, int(request.args.get('page', 1)))
        page_size = int(request.args.get('page_size', 50))
        page_size = max(1, min(page_size, 1000))
        status_filter = request.args.get('status', 'all')
        search = request.args.get('search', '').strip()
        sort_key_param = request.args.get('sort_by') or request.args.get('sort')
        sort_key = sort_key_param or ('id' if not search else 'created_at')
        sort_direction = (request.args.get('sort_order') or request.args.get('direction') or 'desc').lower()
        if not sort_key_param and not search:
            sort_direction = 'desc'
        exclude_skip = request.args.get('exclude_skip', 'false') == 'true'
        include_counts = request.args.get('include_counts', '0') == '1'
        project_count_filter = request.args.get('project_count')
        prefix_filter = request.args.get('prefix', 'all')
        cap_limit = 1000 if search else 5000
        
        print(f"[API] 查询订单号 - page={page}, status={status_filter}, exclude_skip={exclude_skip}", file=sys.stderr)
        
        # 检查数据库连接
        try:
            conn = get_db()
            cursor = conn.cursor()
            ensure_order_lock_columns(conn)
            print("[API] 数据库连接成功", file=sys.stderr)
        except Exception as db_error:
            error_msg = f'数据库连接失败: {str(db_error)}'
            print(f"[ERROR] {error_msg}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            return jsonify({
                'success': False,
                'error': error_msg,
                'error_type': 'DATABASE_CONNECTION_ERROR'
            }), 500
        
        # 构建查询条件
        where_clauses = []
        params = []
        
        # 排除跳号
        if exclude_skip:
            where_clauses.append("o.status != ?")
            params.append('SKIPPED')
        
        # 状态筛选：UNLOCKED, ACTIVE, SKIPPED
        if status_filter != 'all':
            if status_filter == 'available':
                where_clauses.append('o.status = ?')
                params.append('UNLOCKED')
            elif status_filter == 'used':
                where_clauses.append('o.status = ?')
                params.append('ACTIVE')
            elif status_filter == 'skip':
                where_clauses.append('o.status = ?')
                params.append('SKIPPED')
            else:
                where_clauses.append('o.status = ?')
                params.append(status_filter)
        
        parsed_search = _build_order_pool_search_filter(search, 'o.order_number', 'o.customer_name')
        if not parsed_search.get('ok'):
            conn.close()
            return jsonify({'success': False, 'error': parsed_search.get('error') or '搜索条件错误'}), 400
        if parsed_search.get('sql'):
            where_clauses.append(parsed_search['sql'])
            params.extend(parsed_search.get('params', []))

        prefix_sql, prefix_params = _build_prefix_sql_filter(prefix_filter, 'o.order_number')
        if prefix_sql:
            where_clauses.append(prefix_sql)
            params.extend(prefix_params)
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # 获取总数
        try:
            if include_counts and project_count_filter is not None:
                count_query = f'''
                    SELECT COUNT(*) as count FROM (
                        SELECT o.order_number
                        FROM orders o
                        LEFT JOIN workflows w ON o.order_number = w.order_number
                        WHERE {where_sql}
                        GROUP BY o.order_number
                        HAVING COUNT(w.workflow_number) = ?
                    ) t
                '''
                count_params = params + [int(project_count_filter)]
            elif include_counts:
                count_query = f'SELECT COUNT(DISTINCT o.order_number) as count FROM orders o WHERE {where_sql}'
                count_params = params
            else:
                count_query = f'SELECT COUNT(*) as count FROM orders o WHERE {where_sql}'
                count_params = params
            print(f"[API] 执行 COUNT 查询: {count_query}", file=sys.stderr)
            print(f"[API] 参数: {count_params}", file=sys.stderr)
            cursor.execute(count_query, count_params)
            total_full = cursor.fetchone()['count']
            total = min(total_full, cap_limit)
            truncated = total_full > cap_limit
            print(f"[API] 总数: {total_full}, 限制后: {total}", file=sys.stderr)
        except Exception as e:
            error_msg = f'查询总数失败: {str(e)}'
            print(f"[ERROR] {error_msg}", file=sys.stderr)
            print(f"[ERROR] 查询: {count_query}", file=sys.stderr)
            print(f"[ERROR] 参数: {count_params}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            conn.close()
            return jsonify({
                'success': False,
                'error': error_msg,
                'error_type': 'COUNT_QUERY_ERROR',
                'query': count_query,
                'params': params
            }), 500
        
        # 获取分页数据
        offset = (page - 1) * page_size
        try:
            sort_expr, sort_direction_sql = get_order_sort_clause(sort_key, sort_direction)
            if sort_key == 'order_number':
                prefix_rank_expr = (
                    "CASE "
                    "WHEN substr(o.order_number, 1, 1) BETWEEN '0' AND '9' THEN 0 "
                    "WHEN upper(o.order_number) LIKE 'G%' THEN 1 "
                    "WHEN upper(o.order_number) LIKE 'KC%' THEN 2 "
                    "ELSE 3 END"
                )
                order_number_expr = get_order_number_sort_expr('o')
                order_by_sql = f"{prefix_rank_expr} ASC, {order_number_expr} {sort_direction_sql}, o.order_number {sort_direction_sql}"
            else:
                order_by_sql = f"{sort_expr} {sort_direction_sql}, o.order_number ASC"

            where_sql_for_cap = where_sql.replace('o.', 'o2.')
            capped_where_sql = (
                f"{where_sql} AND o.id IN ("
                f"SELECT o2.id FROM orders o2 WHERE {where_sql_for_cap} ORDER BY o2.id DESC LIMIT ?"
                f")"
            )
            capped_params = params + params + [cap_limit]
            if include_counts:
                data_query = f'''
                    SELECT 
                        o.*,
                        COUNT(w.workflow_number) as project_count
                    FROM orders o
                    LEFT JOIN workflows w ON o.order_number = w.order_number
                    WHERE {capped_where_sql}
                    GROUP BY o.order_number
                    {('HAVING COUNT(w.workflow_number) = ?' if project_count_filter is not None else '')}
                    ORDER BY {order_by_sql}
                    LIMIT ? OFFSET ?
                '''
            else:
                data_query = f'''
                    SELECT 
                        o.*
                    FROM orders o
                    WHERE {capped_where_sql}
                    ORDER BY {order_by_sql}
                    LIMIT ? OFFSET ?
                '''
            query_params = capped_params + ([int(project_count_filter)] if (include_counts and project_count_filter is not None) else []) + [page_size, offset]
            print(f"[API] 执行数据查询", file=sys.stderr)
            print(f"[API] WHERE: {capped_where_sql}", file=sys.stderr)
            print(f"[API] 参数: {query_params}", file=sys.stderr)
            cursor.execute(data_query, query_params)
            
            numbers = [dict(row) for row in cursor.fetchall()]
            print(f"[API] 查询到 {len(numbers)} 条记录", file=sys.stderr)
        except Exception as e:
            error_msg = f'查询数据失败: {str(e)}'
            print(f"[ERROR] {error_msg}", file=sys.stderr)
            print(f"[ERROR] WHERE: {where_sql}", file=sys.stderr)
            print(f"[ERROR] 参数: {params}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            conn.close()
            return jsonify({
                'success': False,
                'error': error_msg,
                'error_type': 'DATA_QUERY_ERROR',
                'where_sql': where_sql,
                'params': params
            }), 500
        
        # 获取统计
        try:
            stats_query = 'SELECT status, COUNT(*) as count FROM orders GROUP BY status'
            cursor.execute(stats_query)
            stats_raw = {row['status']: row['count'] for row in cursor.fetchall()}
            # 转换为前端期望的格式
            stats = {
                'available': stats_raw.get('UNLOCKED', 0),
                'used': stats_raw.get('ACTIVE', 0),
                'skip': stats_raw.get('SKIPPED', 0)
            }
        except Exception as e:
            # 统计失败不影响主查询，使用空统计
            stats = {'available': 0, 'used': 0, 'skip': 0}
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': {
                'numbers': numbers,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total,
                    'total_pages': (total + page_size - 1) // page_size
                },
                'truncated': truncated,
                'truncate_limit': cap_limit,
                'stats': stats
            }
        })
    except Exception as e:
        import sys
        error_msg = str(e)
        error_type = type(e).__name__
        print(f"[ERROR] 未预期的错误: {error_msg}", file=sys.stderr)
        print(f"[ERROR] 错误类型: {error_type}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        
        # 如果是数据库相关错误，提供更详细的信息
        if 'no such table' in error_msg.lower() or 'no such column' in error_msg.lower():
            return jsonify({
                'success': False,
                'error': f'数据库错误: {error_msg}。请确保已运行数据库初始化。',
                'error_type': error_type
            }), 500
        
        return jsonify({
            'success': False,
            'error': f'服务器错误: {error_msg}',
            'error_type': error_type
        }), 500


@tracking_bp.route('/api/order-number-pool/mark-skip', methods=['POST'])
@admin_required
def api_mark_skip():
    """标记号码为跳号（重构后：使用 orders 表）"""
    data = request.get_json()
    numbers = data.get('numbers', [])
    
    if not numbers:
        return jsonify({'success': False, 'error': '未提供号码'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    updated_count = 0
    created_by = session.get('user_id') or session.get('username', 'admin')
    
    for number in numbers:
        number_str = str(number).strip()
        
        # 检查订单是否存在
        cursor.execute('SELECT * FROM orders WHERE order_number = ?', (number_str,))
        order = cursor.fetchone()
        
        if order:
            # 如果存在且状态是 UNLOCKED，改为 SKIPPED
            if order['status'] == 'UNLOCKED':
                cursor.execute('''
                    UPDATE orders SET
                        status = 'SKIPPED',
                        updated_at = CURRENT_TIMESTAMP,
                        updated_by = ?
                    WHERE order_number = ?
                ''', (created_by, number_str))
                updated_count += 1
        else:
            # 如果不存在，直接建立为 SKIPPED
            cursor.execute('''
                INSERT INTO orders (
                    order_number, status, visibility, created_at, created_by
                ) VALUES (?, 'SKIPPED', 'admin_only', CURRENT_TIMESTAMP, ?)
            ''', (number_str, created_by))
            updated_count += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'已标记 {updated_count} 个号码为跳号',
        'data': {'updated': updated_count}
    })


@tracking_bp.route('/api/skip-numbers', methods=['GET'])
@admin_required
def api_get_skip_numbers():
    """查询所有跳号（重构后：使用 orders 表）"""
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))
    search = request.args.get('search', '').strip()
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 构建查询条件
    where_clauses = ["status = 'SKIPPED'"]
    params = []
    
    if search:
        where_clauses.append('order_number LIKE ?')
        params.append(f'%{search}%')
    
    where_sql = ' AND '.join(where_clauses)
    
    # 获取总数
    count_query = f'SELECT COUNT(*) as count FROM orders WHERE {where_sql}'
    cursor.execute(count_query, params)
    total = cursor.fetchone()['count']
    
    # 获取分页数据
    offset = (page - 1) * page_size
    cursor.execute(f'''
        SELECT *
        FROM orders
        WHERE {where_sql}
        ORDER BY (order_number + 0) ASC, order_number ASC
        LIMIT ? OFFSET ?
    ''', params + [page_size, offset])
    
    numbers = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'numbers': numbers,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total': total,
                'total_pages': (total + page_size - 1) // page_size
            }
        }
    })


@tracking_bp.route('/api/skip-numbers', methods=['POST'])
@admin_required
def api_add_skip_numbers():
    """批量新增跳号（重构后：使用 orders 表）"""
    data = request.get_json()
    numbers = data.get('numbers', [])
    note = data.get('note', '')
    
    if not numbers:
        return jsonify({'success': False, 'error': '请输入号码'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    success_count = 0
    created_by = session.get('user_id') or session.get('username', 'admin')
    
    for number in numbers:
        try:
            number_str = str(number).strip()
            # 检查订单是否存在
            cursor.execute('SELECT * FROM orders WHERE order_number = ?', (number_str,))
            existing = cursor.fetchone()
            
            if existing:
                if existing['status'] == 'ACTIVE':
                    # 已激活的不能标记为跳号
                    continue
                elif existing['status'] == 'UNLOCKED':
                    # 更新为 SKIPPED
                    cursor.execute('''
                        UPDATE orders 
                        SET status = 'SKIPPED',
                            updated_at = CURRENT_TIMESTAMP,
                            updated_by = ?
                        WHERE order_number = ?
                    ''', (created_by, number_str))
            else:
                # 插入新的跳号
                cursor.execute('''
                    INSERT INTO orders (
                        order_number, status, visibility, created_at, created_by
                    ) VALUES (?, 'SKIPPED', 'admin_only', CURRENT_TIMESTAMP, ?)
                ''', (number_str, created_by))
            
            # 如果有备注，记录到 order_notes
            if note:
                cursor.execute('''
                    INSERT INTO order_notes (
                        order_number, note_type, content, created_by
                    ) VALUES (?, 'skip_reason', ?, ?)
                ''', (number_str, note, created_by))
            
            success_count += 1
        except Exception as e:
            print(f'标记 {number} 失败: {e}')
            continue
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'成功标记 {success_count} 个跳号',
        'data': {'added': success_count}
    })


@tracking_bp.route('/api/skip-numbers', methods=['DELETE'])
@admin_required
def api_delete_skip_numbers():
    """批量删除跳号（重构后：使用 orders 表）"""
    data = request.get_json()
    numbers = data.get('numbers', [])
    
    if not numbers:
        return jsonify({'success': False, 'error': '请选择要删除的号码'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 删除跳号（只删除 SKIPPED 状态的）
    placeholders = ','.join(['?' for _ in numbers])
    cursor.execute(f'''
        DELETE FROM orders 
        WHERE order_number IN ({placeholders}) AND status = 'SKIPPED'
    ''', [str(n) for n in numbers])
    
    deleted_count = cursor.rowcount
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'已删除 {deleted_count} 个跳号',
        'data': {'deleted': deleted_count}
    })


@tracking_bp.route('/api/orders/stats', methods=['GET'])
@admin_required
def api_orders_stats():
    """獲取訂單號統計"""
    try:
        prefix_filter = request.args.get('prefix', 'all')
        conn = get_db()
        cursor = conn.cursor()
        where_clauses = []
        params = []
        prefix_sql, prefix_params = _build_prefix_sql_filter(prefix_filter, 'order_number')
        if prefix_sql:
            where_clauses.append(prefix_sql)
            params.extend(prefix_params)
        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ''
        
        # 統計各狀態數量
        cursor.execute(f'''
            SELECT 
                SUM(CASE WHEN status = 'UNLOCKED' THEN 1 ELSE 0 END) as unlocked,
                SUM(CASE WHEN status = 'ACTIVE' THEN 1 ELSE 0 END) as active,
                SUM(CASE WHEN status = 'SKIPPED' THEN 1 ELSE 0 END) as skipped,
                COUNT(*) as total
            FROM orders
            {where_sql}
        ''', params)
        
        row = cursor.fetchone()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': {
                'unlocked': row['unlocked'] or 0,
                'active': row['active'] or 0,
                'skipped': row['skipped'] or 0,
                'total': row['total'] or 0
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'獲取統計失敗: {str(e)}'}), 500

@tracking_bp.route('/api/order-number-pool/stats', methods=['GET'])
@admin_required
def api_order_number_pool_stats():
    """获取订单号统计数据（重构后：使用 orders 表）
    
    返回各状态的数量：
    - available: 可用（UNLOCKED）
    - used: 已使用（ACTIVE）
    - skip: 跳号（SKIPPED）
    """
    prefix_filter = request.args.get('prefix', 'all')
    conn = get_db()
    cursor = conn.cursor()
    where_clauses = []
    params = []
    prefix_sql, prefix_params = _build_prefix_sql_filter(prefix_filter, 'order_number')
    if prefix_sql:
        where_clauses.append(prefix_sql)
        params.extend(prefix_params)
    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ''
    
    # 查询各状态的数量
    cursor.execute(f'''
        SELECT status, COUNT(*) as count
        FROM orders
        {where_sql}
        GROUP BY status
    ''', params)
    
    results = cursor.fetchall()
    conn.close()
    
    # 初始化统计数据
    stats = {
        'available': 0,
        'used': 0,
        'skip': 0,
        'total': 0
    }
    
    # 填充统计数据（映射新状态到旧状态名）
    for row in results:
        status = row['status']
        count = row['count']
        
        if status == 'UNLOCKED':
            stats['available'] = count
        elif status == 'ACTIVE':
            stats['used'] = count
        elif status == 'SKIPPED':
            stats['skip'] = count
        
        stats['total'] += count
    
    return jsonify({
        'success': True,
        'data': stats
    })


MISSING_REMINDER_ENABLED = True
MISSING_REMINDER_START_NUMBER = 1007500
MISSING_REMINDER_LOOKBACK = 1000


def _get_plain_gap_check_start(end_num, min_num=None):
    """Pure numeric orders: start at 1007500, then keep only the latest ~1000 numbers."""
    if end_num is None:
        return MISSING_REMINDER_START_NUMBER
    if end_num < MISSING_REMINDER_START_NUMBER:
        return min_num if min_num is not None else end_num
    return max(MISSING_REMINDER_START_NUMBER, end_num - MISSING_REMINDER_LOOKBACK)


def _parse_order_number_for_gap_check(order_number):
    if not order_number:
        return None
    digits = ''.join(filter(str.isdigit, str(order_number)))
    if not digits:
        return None
    prefix = ''.join(filter(str.isalpha, str(order_number))).upper()
    return prefix, int(digits)



def _get_recent_gap_year_codes(year_count=3):
    """Return rolling recent year codes, e.g. in 2026 -> ['24', '25', '26']."""
    current_year = datetime.now().year
    return [f"{year % 100:02d}" for year in range(current_year - year_count + 1, current_year + 1)]


def _build_recent_prefixed_gap_ranges(all_numbers, prefix, year_count=3):
    """Build annual G/KC ranges for only the rolling recent N years.

    Each annual sequence starts at xx001 and ends at that year's actual
    maximum stored number. Years with no stored number are skipped because
    there is no reliable upper bound to check yet.
    """
    prefix = (prefix or '').upper()
    if prefix not in ('G', 'KC'):
        return []

    recent_years = _get_recent_gap_year_codes(year_count)
    max_by_year = {}

    for raw_number in all_numbers:
        parsed = _parse_order_number_for_gap_check(raw_number)
        if not parsed:
            continue
        alpha_prefix, num = parsed
        if alpha_prefix.upper() != prefix:
            continue
        digits = ''.join(filter(str.isdigit, str(raw_number)))
        if len(digits) < 5:
            continue
        year_code = digits[:2]
        if year_code not in recent_years:
            continue
        max_by_year[year_code] = max(max_by_year.get(year_code, 0), num)

    ranges = []
    for year_code in recent_years:
        end_num = max_by_year.get(year_code)
        if not end_num:
            continue
        start_num = int(f"{year_code}001")
        ranges.append({
            'year_code': year_code,
            'prefix': prefix,
            'start_num': start_num,
            'end_num': end_num,
            'start': f'{prefix}{start_num}',
            'end': f'{prefix}{end_num}',
        })
    return ranges


def _build_order_number_gap_range_options(all_numbers):
    """Build the three user-facing missing-number modes.

    UI intentionally exposes only:
      1) 國外訂單 (plain numeric)
      2) G
      3) KC

    G/KC are internally split by year and only the rolling recent 3 years
    are checked. For example in 2027, 24 is no longer checked; only 25/26/27.
    """
    numeric_values = []
    for raw_number in all_numbers:
        parsed = _parse_order_number_for_gap_check(raw_number)
        if not parsed:
            continue
        alpha_prefix, num = parsed
        if alpha_prefix == '':
            numeric_values.append(num)

    options = []
    if numeric_values:
        numeric_min = min(numeric_values)
        numeric_max = max(numeric_values)
        numeric_start = _get_plain_gap_check_start(numeric_max, numeric_min)
        options.append({
            'key': 'NUMERIC',
            'prefix': '',
            'label': '國外訂單',
            'start': str(numeric_start),
            'end': str(numeric_max),
            'ranges': [],
        })

    for prefix in ('G', 'KC'):
        ranges = _build_recent_prefixed_gap_ranges(all_numbers, prefix, year_count=3)
        options.append({
            'key': prefix,
            'prefix': prefix,
            'label': prefix,
            'start': ranges[0]['start'] if ranges else '',
            'end': ranges[-1]['end'] if ranges else '',
            'ranges': ranges,
            'recent_years': _get_recent_gap_year_codes(3),
        })

    return options


def _build_grouped_order_number_gap_result(prefix, year_count=3):
    """Aggregate recent annual G/KC gap checks into one modal result."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT order_number FROM orders')
    all_numbers = [row['order_number'] for row in cursor.fetchall()]
    conn.close()

    ranges = _build_recent_prefixed_gap_ranges(all_numbers, prefix, year_count=year_count)
    if not ranges:
        return None

    pieces = []
    for item in ranges:
        piece = _build_order_number_gap_result(item['start_num'], item['end_num'], prefix=prefix)
        piece['year_code'] = item['year_code']
        pieces.append(piece)

    missing_numbers = []
    unlocked_numbers = []
    total = existing = excluded = unlocked = missing = 0
    for piece in pieces:
        total += piece['range']['total']
        existing += piece['existing_count']
        excluded += piece['excluded_count']
        unlocked += piece['unlocked_count']
        missing += piece['missing_count']
        missing_numbers.extend(piece['missing_numbers'])
        unlocked_numbers.extend(piece['unlocked_numbers'])

    range_labels = [f"{r['start']}~{r['end']}" for r in ranges]
    signature = f"group:{prefix}|years:{','.join(r['year_code'] for r in ranges)}|missing:{missing}|unlocked:{unlocked}"
    return {
        'range': {
            'start': ranges[0]['start'],
            'end': ranges[-1]['end'],
            'total': total,
            'auto': True,
            'grouped': True,
            'ranges': range_labels,
        },
        'existing_count': existing,
        'excluded_count': excluded,
        'unlocked_count': unlocked,
        'unlocked_numbers': unlocked_numbers,
        'missing_count': missing,
        'missing_numbers': missing_numbers,
        'signature': signature,
    }


def _build_order_number_gap_result(start_num, end_num, prefix=''):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT order_number, status FROM orders')
    rows = cursor.fetchall()
    conn.close()

    existing_numbers = set()
    excluded_numbers = set()
    unlocked_numbers = set()

    for row in rows:
        parsed = _parse_order_number_for_gap_check(row['order_number'])
        if not parsed:
            continue
        order_prefix, num = parsed
        if num < start_num or num > end_num:
            continue
        if order_prefix != prefix:
            continue

        status = row['status']
        if status == 'SKIPPED':
            excluded_numbers.add(num)
        elif status == 'UNLOCKED':
            unlocked_numbers.add(num)
        else:
            existing_numbers.add(num)

    # 缺號口徑：真正不存在 + 尚未解鎖。
    # UNLOCKED 仍另外回傳 unlocked_numbers，讓前端可標示「號碼已有，但正式訂單尚未回來」。
    missing_numbers = [
        f"{prefix}{num}" if prefix else str(num)
        for num in range(start_num, end_num + 1)
        if num not in existing_numbers
        and num not in excluded_numbers
    ]
    unlocked_number_labels = [
        f"{prefix}{num}" if prefix else str(num)
        for num in sorted(unlocked_numbers)
    ]

    signature = (
        f"start:{start_num}|end:{end_num}|"
        f"missing:{len(missing_numbers)}|unlocked:{len(unlocked_number_labels)}"
    )

    return {
        'range': {
            'start': f"{prefix}{start_num}" if prefix else str(start_num),
            'end': f"{prefix}{end_num}" if prefix else str(end_num),
            'total': end_num - start_num + 1,
            'auto': True
        },
        'existing_count': len(existing_numbers),
        'excluded_count': len(excluded_numbers),
        'unlocked_count': len(unlocked_number_labels),
        'unlocked_numbers': unlocked_number_labels,
        'missing_count': len(missing_numbers),
        'missing_numbers': missing_numbers,
        'signature': signature
    }


def _get_max_plain_order_number():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT order_number FROM orders')
    rows = cursor.fetchall()
    conn.close()

    max_num = None
    for row in rows:
        parsed = _parse_order_number_for_gap_check(row['order_number'])
        if not parsed:
            continue
        prefix, num = parsed
        if prefix != '':
            continue
        if max_num is None or num > max_num:
            max_num = num
    return max_num


@tracking_bp.route('/api/order-number-pool/reminder-check', methods=['GET'])
@admin_required
def api_order_number_pool_reminder_check():
    """Daily admin reminder for missing/unlocked order numbers."""
    if not MISSING_REMINDER_ENABLED:
        return jsonify({
            'success': True,
            'data': {
                'enabled': False,
                'has_issues': False,
                'missing_count': 0,
                'unlocked_count': 0,
                'signature': 'disabled'
            }
        })

    end_num = _get_max_plain_order_number()
    start_num = _get_plain_gap_check_start(end_num)

    if end_num is None or end_num < start_num:
        return jsonify({
            'success': True,
            'data': {
                'enabled': True,
                'has_issues': False,
                'start_number': str(start_num),
                'end_number': str(end_num or ''),
                'missing_count': 0,
                'unlocked_count': 0,
                'signature': f"start:{start_num}|end:{end_num or ''}|missing:0|unlocked:0"
            }
        })

    result = _build_order_number_gap_result(start_num, end_num, prefix='')
    has_issues = result['missing_count'] > 0 or result['unlocked_count'] > 0

    return jsonify({
        'success': True,
        'data': {
            'enabled': True,
            'has_issues': has_issues,
            'start_number': result['range']['start'],
            'end_number': result['range']['end'],
            'range': result['range'],
            'existing_count': result['existing_count'],
            'excluded_count': result['excluded_count'],
            'missing_count': result['missing_count'],
            'missing_numbers': result['missing_numbers'][:200],
            'unlocked_count': result['unlocked_count'],
            'unlocked_numbers': result['unlocked_numbers'][:200],
            'signature': result['signature']
        }
    })


@tracking_bp.route('/api/order-number-pool/diff', methods=['POST'])
@admin_required
def api_order_number_pool_diff():
    """缺號檢測（僅 Admin）
    
    檢測指定範圍內的缺失號碼
    Missing = [start..end] − orders（現有編號） − skip（跳號）
    reserved 不影響缺號口徑
    
    Body: {
        start_number: string,  # 起始號碼
        end_number: string     # 結束號碼
    }
    """
    data = request.get_json() or {}

    selected_group = str(data.get('group') or '').strip().upper()
    start_number = str(data.get('start_number') or '').strip()
    end_number = str(data.get('end_number') or '').strip()

    # Backward compatibility for older cached JS: if G/KC arrives without a
    # manual range, keep the old rolling-3-year grouped check. New UI always
    # sends the selected year's editable start/end range instead.
    if selected_group in ('G', 'KC') and not start_number and not end_number:
        grouped_result = _build_grouped_order_number_gap_result(selected_group, year_count=3)
        if not grouped_result:
            recent_years = ' / '.join(_get_recent_gap_year_codes(3))
            return jsonify({
                'success': False,
                'error': f'最近 3 年（{recent_years}）沒有 {selected_group} 訂單號可供檢測'
            }), 400
        return jsonify({'success': True, 'data': grouped_result})
    
    # 若未提供範圍，改用系統內最小/最大號碼
    if not start_number or not end_number:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT order_number FROM orders')
        all_numbers = [row['order_number'] for row in cursor.fetchall()]
        conn.close()

        if not all_numbers:
            return jsonify({'success': False, 'error': '系統內沒有任何訂單號可供檢測'}), 400

        options = _build_order_number_gap_range_options(all_numbers)
        if not options:
            return jsonify({'success': False, 'error': '系統內訂單號無數字部分，請手動指定範圍'}), 400

        # Use the same grouped range logic as the modal.  Pure numeric remains
        # preferred option is the plain numeric foreign-order range.
        selected = options[0]
        prefix = selected['prefix']
        start_number = selected['start']
        end_number = selected['end']
        start_num = int(''.join(filter(str.isdigit, start_number)))
        end_num = int(''.join(filter(str.isdigit, end_number)))
        auto_range = True
    else:
        try:
            # 嘗試提取數字部分
            start_num = int(''.join(filter(str.isdigit, start_number)))
            end_num = int(''.join(filter(str.isdigit, end_number)))
            
            if start_num > end_num:
                return jsonify({'success': False, 'error': '起始號碼不能大於結束號碼'}), 400
            
            # 起止號碼前綴必須一致
            start_prefix = ''.join(filter(str.isalpha, start_number)).upper()
            end_prefix = ''.join(filter(str.isalpha, end_number)).upper()
            if start_prefix != end_prefix:
                return jsonify({'success': False, 'error': '起始和结束号码前缀必须一致'}), 400
            prefix = start_prefix

            if selected_group in ('G', 'KC') and prefix != selected_group:
                return jsonify({'success': False, 'error': f'目前選擇的是 {selected_group}，請輸入 {selected_group} 開頭的號碼'}), 400
            if selected_group == 'NUMERIC' and prefix:
                return jsonify({'success': False, 'error': '國外訂單請輸入純數字號碼'}), 400

            # G/KC are yearly sequences (G25xxx, G26xxx / KC25xxx, KC26xxx).
            # Never allow a manual check to cross from one year group into another.
            if prefix in ('G', 'KC'):
                start_digits = ''.join(filter(str.isdigit, start_number))
                end_digits = ''.join(filter(str.isdigit, end_number))
                if len(start_digits) >= 5 and len(end_digits) >= 5 and start_digits[:2] != end_digits[:2]:
                    return jsonify({
                        'success': False,
                        'error': f'{prefix} 號碼需按年份分開檢測，例如 {prefix}25xxx 與 {prefix}26xxx 不能跨年連續檢測'
                    }), 400
            
        except ValueError:
            return jsonify({'success': False, 'error': '號碼格式錯誤'}), 400
        auto_range = False
    
    conn = get_db()
    cursor = conn.cursor()
    
    range_size = end_num - start_num + 1
    
    # 獲取已存在的號碼（orders）
    # 注意：不能直接用字符串比較，因為可能有前綴，需要提取數字部分比較
    import sys
    print(f"[DEBUG] 缺號檢測 - 範圍: {start_number} ~ {end_number}, 數字範圍: {start_num} ~ {end_num}, 前綴: '{prefix}'", file=sys.stderr)
    
    cursor.execute('SELECT order_number, status FROM orders')
    all_order_rows = cursor.fetchall()
    print(f"[DEBUG] 查詢到 {len(all_order_rows)} 筆 orders 記錄", file=sys.stderr)
    # 特別檢查 1007001 是否在 orders 表中
    cursor.execute("SELECT order_number FROM orders WHERE order_number LIKE '%1007001%'")
    check_1007001 = cursor.fetchall()
    if check_1007001:
        print(f"[DEBUG] [特別檢查] orders 表中找到包含 1007001 的記錄: {[r['order_number'] for r in check_1007001]}", file=sys.stderr)
    else:
        print(f"[DEBUG] [特別檢查] orders 表中沒有找到包含 1007001 的記錄", file=sys.stderr)
    
    existing_orders = set()
    excluded_numbers = set()
    unlocked_numbers = set()
    for row in all_order_rows:
        order_num = row['order_number']
        order_status = row['status']
        if not order_num:
            continue
        # 提取數字部分
        digits = ''.join(filter(str.isdigit, order_num))
        if digits:
            num = int(digits)
            # 提取前綴
            order_prefix = ''.join(filter(str.isalpha, order_num))
            print(f"[DEBUG] 訂單號: {order_num}, 前綴: '{order_prefix}', 數字: {num}", file=sys.stderr)
            
            # 檢查數字部分是否在範圍內
            if num >= start_num and num <= end_num:
                print(f"[DEBUG]   -> 數字 {num} 在範圍內", file=sys.stderr)
                # 前綴匹配規則：
                # - 如果用戶輸入的是純數字範圍（prefix == ''），則只匹配純數字的訂單號
                # - 如果用戶輸入的是帶前綴的範圍（prefix != ''），則只匹配相同前綴的訂單號
                if prefix == '':
                    # 純數字範圍，只匹配純數字的訂單號
                    if order_prefix == '':
                        if order_status == 'SKIPPED':
                            excluded_numbers.add(num)
                            print(f"[DEBUG]   -> 匹配成功（純數字），SKIPPED 加入 excluded_numbers: {num}", file=sys.stderr)
                        elif order_status == 'UNLOCKED':
                            unlocked_numbers.add(num)
                            print(f"[DEBUG]   -> 匹配成功（純數字），UNLOCKED 加入 unlocked_numbers: {num}", file=sys.stderr)
                        else:
                            existing_orders.add(num)
                            print(f"[DEBUG]   -> 匹配成功（純數字），加入 existing_orders: {num}", file=sys.stderr)
                    else:
                        print(f"[DEBUG]   -> 不匹配（訂單有前綴 '{order_prefix}'，但範圍是純數字）", file=sys.stderr)
                else:
                    # 帶前綴範圍，只匹配相同前綴的訂單號
                    if order_prefix == prefix:
                        if order_status == 'SKIPPED':
                            excluded_numbers.add(num)
                            print(f"[DEBUG]   -> 匹配成功（前綴相同），SKIPPED 加入 excluded_numbers: {num}", file=sys.stderr)
                        elif order_status == 'UNLOCKED':
                            unlocked_numbers.add(num)
                            print(f"[DEBUG]   -> 匹配成功（前綴相同），UNLOCKED 加入 unlocked_numbers: {num}", file=sys.stderr)
                        else:
                            existing_orders.add(num)
                            print(f"[DEBUG]   -> 匹配成功（前綴相同），加入 existing_orders: {num}", file=sys.stderr)
                    else:
                        print(f"[DEBUG]   -> 不匹配（前綴不同：訂單='{order_prefix}', 範圍='{prefix}'）", file=sys.stderr)
            else:
                print(f"[DEBUG]   -> 數字 {num} 不在範圍內（{start_num} ~ {end_num}）", file=sys.stderr)
    
    print(f"[DEBUG] existing_orders 最終結果: {sorted(existing_orders)}", file=sys.stderr)
    
    # 获取所有订单号（重构后：只使用 orders 表）
    # 逻辑：
    # - UNLOCKED: 未解鎖，單獨列出，不算缺號
    # - ACTIVE/COMPLETED/CANCELLED/其他正式狀態: 已存在的订单号，不算缺号（加入 existing_orders）
    # - SKIPPED: 跳号，不算缺号（应加入 excluded_numbers）
    
    # 从 orders 表中获取所有订单号（包括所有状态）
    for row in all_order_rows:
        order_num = row['order_number']
        order_status = row['status']
        if not order_num:
            continue
        digits = ''.join(filter(str.isdigit, order_num))
        if digits:
            num = int(digits)
            order_prefix = ''.join(filter(str.isalpha, order_num))
            
            # 检查数字部分是否在范围内
            if num >= start_num and num <= end_num:
                # 前缀匹配规则：与 existing_orders 相同
                should_process = False
                if prefix == '':
                    # 纯数字范围，只匹配纯数字的订单号
                    if order_prefix == '':
                        should_process = True
                else:
                    # 带前缀范围，只匹配相同前缀的订单号
                    if order_prefix == prefix:
                        should_process = True
                
                if should_process:
                    if order_status == 'SKIPPED':
                        # SKIPPED 状态：跳号，加入 excluded_numbers
                        excluded_numbers.add(num)
                        print(f"[DEBUG] 订单号 {order_num} 状态为 SKIPPED，加入 excluded_numbers: {num}", file=sys.stderr)
                    elif order_status == 'UNLOCKED':
                        # UNLOCKED 状态：号码已占用但未传回流程，单独列出
                        unlocked_numbers.add(num)
                        print(f"[DEBUG] 订单号 {order_num} 状态为 UNLOCKED，加入 unlocked_numbers: {num}", file=sys.stderr)
                    else:
                        # 其他状态：正式存在，加入 existing_orders
                        existing_orders.add(num)
                        print(f"[DEBUG] 订单号 {order_num} 状态为 {order_status}，加入 existing_orders: {num}", file=sys.stderr)
    
    print(f"[DEBUG] 最终 existing_orders: {sorted(existing_orders)}", file=sys.stderr)
    print(f"[DEBUG] excluded_numbers (SKIPPED): {sorted(excluded_numbers)}", file=sys.stderr)
    print(f"[DEBUG] unlocked_numbers (UNLOCKED): {sorted(unlocked_numbers)}", file=sys.stderr)
    
    conn.close()
    
    # 計算缺號：真正不存在 + UNLOCKED 都屬於「尚未收到正式訂單」
    # UNLOCKED 同時保留獨立清單，方便管理員分辨原因。
    missing_numbers = []
    for i in range(start_num, end_num + 1):
        if i not in existing_orders and i not in excluded_numbers:
            missing_numbers.append(f"{prefix}{i}" if prefix else str(i))
        else:
            if i in existing_orders:
                print(f"[DEBUG] {i} 在 existing_orders 中，不計入缺號", file=sys.stderr)
            if i in excluded_numbers:
                print(f"[DEBUG] {i} 在 excluded_numbers 中，不計入缺號", file=sys.stderr)
        if i in unlocked_numbers:
            print(f"[DEBUG] {i} 在 unlocked_numbers 中，計入缺號並另外列為未解鎖", file=sys.stderr)
    
    print(f"[DEBUG] 最終缺號數量: {len(missing_numbers)}", file=sys.stderr)
    
    return jsonify({
        'success': True,
        'data': {
            'range': {
                'start': start_number,
                'end': end_number,
                'total': range_size,
                'auto': auto_range
            },
            'existing_count': len(existing_orders),
            'excluded_count': len(excluded_numbers),
            'unlocked_count': len(unlocked_numbers),
            'unlocked_numbers': [f"{prefix}{i}" if prefix else str(i) for i in sorted(unlocked_numbers)],
            'missing_count': len(missing_numbers),
            'missing_numbers': missing_numbers
        }
    })


@tracking_bp.route('/api/order-number-pool/range', methods=['GET'])
@admin_required
def api_order_number_pool_range():
    """取得系統內最小/最大訂單號（按前綴分組）（重构后：只使用 orders 表）"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT order_number FROM orders')
    all_numbers = [row['order_number'] for row in cursor.fetchall()]
    conn.close()

    if not all_numbers:
        return jsonify({'success': False, 'error': '系統內沒有任何訂單號'}), 400

    options = _build_order_number_gap_range_options(all_numbers)

    if not options:
        return jsonify({'success': False, 'error': '系統內訂單號無數字部分'}), 400

    preferred_key = options[0]['key']

    return jsonify({
        'success': True,
        'data': {
            'options': options,
            'preferred_key': preferred_key,
            # Keep the old field for older cached JS; new JS uses preferred_key.
            'preferred_prefix': options[0]['prefix']
        }
    })


# ==================== 订单号管理 API（重构后）====================

@tracking_bp.route('/api/orders/batch-create', methods=['POST'])
@admin_required
def api_batch_create_orders():
    """批量建立订单号"""
    data = request.get_json()
    start_number = data.get('start', '').strip()
    end_number = data.get('end', '').strip()
    
    if not start_number or not end_number:
        return jsonify({'success': False, 'error': '缺少起始或结束号码'}), 400
    
    # 解析前缀和数字
    prefix = ''.join(filter(str.isalpha, start_number))
    try:
        start_num = int(''.join(filter(str.isdigit, start_number)))
        end_num = int(''.join(filter(str.isdigit, end_number)))
    except ValueError:
        return jsonify({'success': False, 'error': '号码格式错误'}), 400
    
    if start_num > end_num:
        return jsonify({'success': False, 'error': '起始号码不能大于结束号码'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    created_by = session.get('user_id') or session.get('username', 'admin')
    created_count = 0
    skipped_count = 0
    
    for i in range(start_num, end_num + 1):
        order_number = f"{prefix}{i}" if prefix else str(i)
        
        # 检查是否已存在
        cursor.execute('SELECT order_number FROM orders WHERE order_number = ?', (order_number,))
        if cursor.fetchone():
            skipped_count += 1
            continue
        
        # INSERT 为 UNLOCKED 状态
        cursor.execute('''
            INSERT INTO orders (
                order_number, status, visibility, created_at, created_by
            ) VALUES (?, 'UNLOCKED', 'admin_only', CURRENT_TIMESTAMP, ?)
        ''', (order_number, created_by))
        
        created_count += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'成功建立 {created_count} 个订单号',
        'data': {
            'created': created_count,
            'skipped': skipped_count,
            'total': end_num - start_num + 1
        }
    })


@tracking_bp.route('/api/orders/list', methods=['GET'])
@admin_required
def api_list_orders():
    """获取订单号列表"""
    status_filter = request.args.get('status', '')
    prefix_filter = request.args.get('prefix', 'all')
    
    conn = get_db()
    cursor = conn.cursor()
    
    sql = 'SELECT * FROM orders'
    params = []
    
    if status_filter:
        sql += ' WHERE status = ?'
        params.append(status_filter)
    prefix_sql, prefix_params = _build_prefix_sql_filter(prefix_filter, 'order_number')
    if prefix_sql:
        sql += (' AND ' if ' WHERE ' in sql else ' WHERE ') + prefix_sql
        params.extend(prefix_params)
    
    sql += ' ORDER BY order_number DESC'
    
    cursor.execute(sql, params)
    orders = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({
        'success': True,
        'data': orders
    })


@tracking_bp.route('/api/orders/<order_number>/visibility', methods=['PUT'])
@admin_required
def api_update_order_visibility(order_number):
    """修改訂單可見性（流程8：設為僅管理員可見或所有業務員可見）"""
    data = request.get_json()
    new_visibility = data.get('visibility')
    
    if new_visibility not in ['admin_only', 'all_sales']:
        return jsonify({'success': False, 'error': '無效的可見性值'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 检查订单是否存在
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '訂單號不存在'}), 404
    
    # 只有 ACTIVE 状态的订单才能修改可见性
    if order['status'] != 'ACTIVE':
        conn.close()
        return jsonify({'success': False, 'error': '只有已解鎖的訂單才能修改可見性'}), 400
    
    # 更新可见性
    updated_by = session.get('user_id') or session.get('username', 'admin')
    cursor.execute('''
        UPDATE orders SET
            visibility = ?,
            updated_at = CURRENT_TIMESTAMP,
            updated_by = ?
        WHERE order_number = ?
    ''', (new_visibility, updated_by, order_number))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'訂單可見性已更新為 {"僅管理員可見" if new_visibility == "admin_only" else "所有業務員可見"}',
        'data': {
            'order_number': order_number,
            'visibility': new_visibility
        }
    })



@tracking_bp.route('/api/orders/<order_number>/cancel', methods=['PUT'])
@admin_required
def api_cancel_order(order_number):
    """Mark order as cancelled, or uncancel (restore to ACTIVE)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': 'Order not found'}), 404
    current_status = order['status']
    updated_by = session.get('user_id') or session.get('username', 'admin')
    ensure_order_lock_columns(conn)
    if current_status == 'CANCELLED':
        # 解除取消：恢復 ACTIVE，解鎖
        cursor.execute("""
            UPDATE orders SET
                status = 'ACTIVE',
                is_locked = 0,
                locked_at = NULL,
                locked_by_id = NULL,
                locked_by_name = NULL,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE order_number = ?
        """, (updated_by, order_number))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '訂單已恢復為正常狀態', 'new_status': 'ACTIVE'})
    elif current_status == 'ACTIVE':
        # 標記取消：同時鎖定，防止再建流程或編輯
        cursor.execute("""
            UPDATE orders SET
                status = 'CANCELLED',
                is_locked = 1,
                locked_at = CURRENT_TIMESTAMP,
                locked_by_id = ?,
                locked_by_name = ?,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE order_number = ?
        """, (updated_by, updated_by, updated_by, order_number))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '訂單已標記為取消並鎖定', 'new_status': 'CANCELLED'})
    else:
        conn.close()
        return jsonify({'success': False, 'error': f'此訂單狀態（{current_status}）不支援此操作'}), 400

@tracking_bp.route('/api/orders/<order_number>/lock', methods=['PUT'])
@api_login_required
@require_permission('order', 'lock', resource_id_param='order_number')
def api_lock_order(order_number):
    """标记订单为已完成（锁定）"""
    conn = get_db()
    cursor = conn.cursor()
    ensure_order_lock_columns(conn)

    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    if hasattr(g, 'current_user') and g.current_user:
        current_user_id = g.current_user.get('id')
    else:
        current_user_id = session.get('user_id')
    current_user_name = session.get('display_name') or session.get('username')
    if not current_user_name and current_user_id:
        current_user_name = get_user_display_name(current_user_id, conn)

    cursor.execute('''
        UPDATE orders SET
            is_locked = 1,
            locked_at = CURRENT_TIMESTAMP,
            locked_by_id = ?,
            locked_by_name = ?
        WHERE order_number = ?
    ''', (current_user_id, current_user_name, order_number))

    # === 通知：通知该订单下所有流程的负责人 ===
    try:
        cursor.execute('SELECT DISTINCT handler_id FROM workflows WHERE order_number = ? AND handler_id IS NOT NULL', (order_number,))
        handler_ids = [r['handler_id'] for r in cursor.fetchall()]
        customer = order['customer_name'] or ''
        for hid in handler_ids:
            if hid != current_user_id:
                _create_notification(
                    cursor, hid, 'order_locked',
                    f'订单 {order_number} 已标记完成',
                    f'客户：{customer}，该订单已锁定，无法再创建新流程',
                    order_number, None, 'normal', 60
                )
    except Exception as e:
        print(f"[WARN] 创建锁定通知失败: {e}")

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': '订单已标记为完成并锁定',
        'data': {'order_number': order_number}
    })


@tracking_bp.route('/api/orders/<order_number>/unlock', methods=['PUT'])
@api_login_required
@require_permission('order', 'unlock', resource_id_param='order_number')
def api_unlock_locked_order(order_number):
    """解除订单锁定"""
    conn = get_db()
    cursor = conn.cursor()
    ensure_order_lock_columns(conn)

    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    cursor.execute('''
        UPDATE orders SET
            is_locked = 0,
            locked_at = NULL,
            locked_by_id = NULL,
            locked_by_name = NULL
        WHERE order_number = ?
    ''', (order_number,))

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': '订单锁定已解除',
        'data': {'order_number': order_number}
    })


@tracking_bp.route('/api/orders/unlock', methods=['POST'])
@admin_required
def api_unlock_order():
    """解锁订单并上传文件（按照07_FILE_UPLOAD_MANAGEMENT.md规范）"""
    import os
    import uuid
    from datetime import datetime
    from .config import UPLOAD_FOLDER
    
    # 支持表单数据和JSON数据
    if request.content_type and 'multipart/form-data' in request.content_type:
        order_number = request.form.get('order_number')
        customer_name = request.form.get('customer_name')
        notes = request.form.get('notes', '')
        files = request.files.getlist('files')
    else:
        data = request.get_json() or {}
        order_number = data.get('order_number')
        customer_name = data.get('customer_name')
        notes = data.get('notes', '')
        files = []
    
    customer_name = _normalize_customer_name(customer_name)
    if not order_number or not customer_name:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 检查订单是否存在
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单号不存在'}), 404
    
    if order['status'] != 'UNLOCKED':
        conn.close()
        return jsonify({'success': False, 'error': '订单号已被使用或跳号'}), 400
    
    # 解锁订单：订单日期固定为解锁当天
    order_date = date.today().isoformat()
    updated_by = session.get('user_id') or session.get('username', 'admin')
    cursor.execute('''
        UPDATE orders SET
            status = 'ACTIVE',
            visibility = 'all_sales',
            customer_name = ?,
            order_date = ?,
            notes = ?,
            updated_at = CURRENT_TIMESTAMP,
            updated_by = ?
        WHERE order_number = ?
    ''', (customer_name, order_date, notes, updated_by, order_number))
    
    # 处理文件上传（按照文档规范：时间戳+UUID命名）
    uploaded_files_info = []
    if files:
        # 创建文件夹
        order_dir = os.path.join(UPLOAD_FOLDER, 'orders', order_number)
        os.makedirs(order_dir, exist_ok=True)
        
        for file in files:
            if file and file.filename:
                # 1. 保留原始文件名（完整保留，包括中文）
                original_filename = file.filename
                
                # 2. 生成磁盘存储文件名（格式：YYYYMMDD_HHMMSS_[8码英数随机].[ext]）
                file_ext = os.path.splitext(original_filename)[1].lower()  # 扩展名统一小写
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_id = str(uuid.uuid4())[:8]
                stored_filename = f"{timestamp}_{unique_id}{file_ext}"
                
                # 3. 保存到磁盘
                file_path = os.path.join(order_dir, stored_filename)
                file.save(file_path)
                
                # 4. 写入数据库（file_path 只存目录，不包含文件名）
                file_size = os.path.getsize(file_path)
                relative_path = f"orders/{order_number}/"  # 只存目录路径
                
                # 检查表结构，兼容旧字段
                cursor.execute("PRAGMA table_info(order_files)")
                columns = [row[1] for row in cursor.fetchall()]
                has_file_name = 'file_name' in columns
                
                if has_file_name:
                    # 兼容旧结构：同时插入 file_name
                    cursor.execute('''
                        INSERT INTO order_files (
                            order_number,
                            file_name,
                            original_filename,
                            stored_filename,
                            file_path,
                            file_size,
                            file_type,
                            mime_type,
                            uploaded_at,
                            uploaded_by
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                    ''', (
                        order_number,
                        original_filename,  # file_name（兼容旧字段）
                        original_filename,  # original_filename（新字段）
                        stored_filename,   # stored_filename（新字段）
                        relative_path,      # 目录路径（不包含文件名）
                        file_size,
                        file_ext.lstrip('.'),
                        file.content_type or 'application/octet-stream',
                        updated_by
                    ))
                else:
                    # 新结构：只使用新字段
                    cursor.execute('''
                        INSERT INTO order_files (
                            order_number,
                            original_filename,
                            stored_filename,
                            file_path,
                            file_size,
                            file_type,
                            mime_type,
                            uploaded_at,
                            uploaded_by
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                    ''', (
                        order_number,
                        original_filename,  # 原始文件名（UI显示/下载）
                        stored_filename,   # 磁盘存储文件名
                        relative_path,      # 目录路径（不包含文件名）
                        file_size,
                        file_ext.lstrip('.'),
                        file.content_type or 'application/octet-stream',
                        updated_by
                    ))
                
                uploaded_files_info.append({
                    'original_name': original_filename,
                    'stored_name': stored_filename,
                    'size': file_size
                })
    
    # === 通知：广播给所有业务员「有新订单解锁了」 ===
    try:
        _create_broadcast_notification(
            cursor, 'order_unlocked',
            f'新订单 {order_number} 已解锁',
            f'客户：{customer_name}',
            order_number, None, 'normal', 'all',
            exclude_user_id=None,
            expires_days=30
        )
    except Exception as e:
        print(f"[WARN] 创建解锁通知失败: {e}")
    
    conn.commit()
    conn.close()
    
    message = '订单已解锁'
    if uploaded_files_info:
        message += f'，上传了 {len(uploaded_files_info)} 个文件'
    
    return jsonify({
        'success': True,
        'message': message,
        'data': {
            'order_number': order_number,
            'uploaded_files': uploaded_files_info
        }
    })


# ==================== 订单文件管理 API（按照07_FILE_UPLOAD_MANAGEMENT.md）====================

@tracking_bp.route('/api/orders/<order_number>/files/upload', methods=['POST'])
@login_required
@require_permission('order_file', 'upload', resource_id_param='order_number')
def api_upload_order_files(order_number):
    """追加上传订单文件（解锁后也可以上传）"""
    import os
    import uuid
    from datetime import datetime
    from .config import UPLOAD_FOLDER
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 检查订单是否存在
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    
    # 检查是否有文件
    if 'files' not in request.files:
        conn.close()
        return jsonify({'success': False, 'error': '没有选择文件'}), 400
    
    files = request.files.getlist('files')
    if not files or all(not f.filename for f in files):
        conn.close()
        return jsonify({'success': False, 'error': '没有选择文件'}), 400
    
    # 创建文件夹
    order_dir = os.path.join(UPLOAD_FOLDER, 'orders', order_number)
    os.makedirs(order_dir, exist_ok=True)
    
    uploaded_files_info = []
    uploaded_by = session.get('user_id') or session.get('username', 'admin')
    
    for file in files:
        if file and file.filename:
            # 保留原始文件名（完整保留，包括中文）
            original_filename = file.filename
            
            # 生成磁盘存储文件名（格式：YYYYMMDD_HHMMSS_[8码英数随机].[ext]）
            file_ext = os.path.splitext(original_filename)[1].lower()  # 扩展名统一小写
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_id = str(uuid.uuid4())[:8]
            stored_filename = f"{timestamp}_{unique_id}{file_ext}"
            
            # 保存到磁盘
            file_path = os.path.join(order_dir, stored_filename)
            file.save(file_path)
            
            # 写入数据库（file_path 只存目录，不包含文件名）
            file_size = os.path.getsize(file_path)
            relative_path = f"orders/{order_number}/"  # 只存目录路径
            
            # 检查表结构，兼容旧字段
            cursor.execute("PRAGMA table_info(order_files)")
            columns = [row[1] for row in cursor.fetchall()]
            has_file_name = 'file_name' in columns
            
            if has_file_name:
                # 兼容旧结构：同时插入 file_name
                cursor.execute('''
                    INSERT INTO order_files (
                        order_number,
                        file_name,
                        original_filename,
                        stored_filename,
                        file_path,
                        file_size,
                        file_type,
                        mime_type,
                        uploaded_at,
                        uploaded_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                ''', (
                    order_number,
                    original_filename,  # file_name（兼容旧字段）
                    original_filename,  # original_filename（新字段）
                    stored_filename,   # stored_filename（新字段）
                    relative_path,      # 目录路径（不包含文件名）
                    file_size,
                    file_ext.lstrip('.'),
                    file.content_type or 'application/octet-stream',
                    uploaded_by
                ))
            else:
                # 新结构：只使用新字段
                cursor.execute('''
                    INSERT INTO order_files (
                        order_number,
                        original_filename,
                        stored_filename,
                        file_path,
                        file_size,
                        file_type,
                        mime_type,
                        uploaded_at,
                        uploaded_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                ''', (
                    order_number,
                    original_filename,  # 原始文件名（UI显示/下载）
                    stored_filename,   # 磁盘存储文件名
                    relative_path,      # 目录路径（不包含文件名）
                    file_size,
                    file_ext.lstrip('.'),
                    file.content_type or 'application/octet-stream',
                    uploaded_by
                ))
            
            uploaded_files_info.append({
                'original_name': original_filename,
                'stored_name': stored_filename,
                'size': file_size
            })
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'成功上传 {len(uploaded_files_info)} 个文件',
        'data': {
            'order_number': order_number,
            'uploaded_files': uploaded_files_info
        }
    })


@tracking_bp.route('/api/orders/<order_number>/files', methods=['GET'])
@login_required
def api_get_order_files(order_number):
    if cloud_mode_enabled():
        return jsonify({'success': True, 'data': {
            'order_number': order_number, 'files': [], 'total': 0
        }})
    # visual=1 always re-queries current SQLite rows and current source-file versions.
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    if not can_access_visibility(order['visibility']):
        conn.close()
        return jsonify({'success': False, 'error': '无权限'}), 403

    cursor.execute("""
        SELECT id, original_filename, stored_filename, file_path, file_size,
               file_type, mime_type, uploaded_at, uploaded_by, file_name
        FROM order_files
        WHERE order_number = ?
        ORDER BY uploaded_at DESC
    """, (order_number,))
    files = []
    preview_specs = []
    include_visual_meta = str(request.args.get('visual') or '').strip().lower() in {'1', 'true', 'yes'}
    for row in cursor.fetchall():
        try:
            original_name = row['original_filename'] if row['original_filename'] else row['file_name']
        except (KeyError, TypeError):
            try:
                original_name = row['file_name']
            except KeyError:
                original_name = ''
        row_info = dict(row)
        is_pdf = _media_is_pdf(original_name, row_info.get('file_type'), row_info.get('mime_type'))
        pdf_page_count = 0
        media_version = ''
        if include_visual_meta:
            visual_path = _resolve_order_file_full_path(row_info)
            if os.path.isfile(visual_path):
                media_version = _file_version_token(visual_path)
                if is_pdf:
                    pdf_page_count = _pdf_preview_page_count(visual_path)
                    if pdf_page_count:
                        preview_specs.append({'type': 'pdf', 'path': visual_path, 'pages': pdf_page_count})
                elif _local_guest_image_file(original_name, row_info.get('file_type'), row_info.get('mime_type')):
                    preview_specs.append({'type': 'image', 'path': visual_path})
        files.append({
            'id': row['id'], 'file_name': original_name, 'file_size': row['file_size'],
            'file_type': row['file_type'], 'mime_type': row['mime_type'],
            'uploaded_at': row['uploaded_at'], 'uploaded_by': row['uploaded_by'],
            'media_type': 'pdf' if is_pdf else 'file', 'pdf_page_count': pdf_page_count,
            'media_version': media_version,
        })
    conn.close()
    if include_visual_meta and preview_specs:
        _schedule_visual_preview_batch(preview_specs)
    response = jsonify({'success': True, 'data': {
        'order_number': order_number, 'files': files, 'total': len(files)
    }})
    if include_visual_meta:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


@tracking_bp.route('/api/orders/files/<int:file_id>/download', methods=['GET'])
@login_required
def api_download_order_file(file_id):
    """下载文件"""
    import os
    from flask import send_file
    from .config import UPLOAD_FOLDER
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 查询文件信息
    cursor.execute('''
        SELECT f.*, o.visibility
        FROM order_files f
        JOIN orders o ON f.order_number = o.order_number
        WHERE f.id = ?
    ''', (file_id,))
    
    file_info = cursor.fetchone()
    conn.close()
    
    if not file_info:
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    # 权限检查
    if not can_access_visibility(file_info['visibility']):
        return jsonify({'success': False, 'error': '无权限'}), 403
    
    # 组合实际文件路径：UPLOAD_ROOT + file_path + stored_filename
    try:
        stored_filename = file_info['stored_filename'] if file_info['stored_filename'] else os.path.basename(file_info['file_path'])
    except KeyError:
        stored_filename = os.path.basename(file_info['file_path'])
    
    file_path_dir = file_info['file_path']
    
    # 兼容旧数据：如果 file_path 包含文件名，提取目录
    if '/' in stored_filename or '\\' in stored_filename:
        file_path_dir = os.path.dirname(file_info['file_path'])
        stored_filename = os.path.basename(file_info['file_path'])
    
    full_path = os.path.join(UPLOAD_FOLDER, file_path_dir, stored_filename)
    
    if not os.path.exists(full_path):
        return jsonify({'success': False, 'error': '文件已被删除'}), 404
    
    # 下载文件（使用原始文件名）
    try:
        original_filename = file_info['original_filename'] if file_info['original_filename'] else file_info['file_name']
    except KeyError:
        try:
            original_filename = file_info['file_name']
        except KeyError:
            original_filename = stored_filename
    preview_requested = str(request.args.get('preview') or '').strip().lower() in {'1', 'true', 'yes'}
    if preview_requested and _local_guest_image_file(
        original_filename,
        file_info['file_type'] if 'file_type' in file_info.keys() else '',
        file_info['mime_type'] if 'mime_type' in file_info.keys() else ''
    ):
        try:
            return _image_preview_response(full_path, original_filename or 'preview.jpg')
        except Exception as exc:
            print(f'[WARN] order image preview fallback to original: {exc}')

    return send_file(
        full_path,
        as_attachment=True,
        download_name=original_filename
    )


@tracking_bp.route('/api/orders/files/<int:file_id>/pdf-page/<int:page_number>', methods=['GET'])
@login_required
def api_preview_order_pdf_page(file_id, page_number):
    # Authenticated mobile/browser preview of one order PDF page as JPEG.
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT f.*, o.visibility
            FROM order_files f
            JOIN orders o ON o.order_number = f.order_number
            WHERE f.id = ?
        ''', (file_id,))
        row = cursor.fetchone()
    finally:
        conn.close()
    if not row:
        return '', 404
    info = dict(row)
    if not can_access_visibility(info.get('visibility')):
        return '', 403
    name = info.get('original_filename') or info.get('file_name') or info.get('stored_filename') or ''
    if not _media_is_pdf(name, info.get('file_type'), info.get('mime_type')):
        return '', 415
    path = _resolve_order_file_full_path(info)
    try:
        payload, total = _pdf_preview_render_page(path, page_number, compact=str(request.args.get('preview') or '').lower() in {'1','true','yes'})
    except FileNotFoundError:
        return '', 404
    except ValueError:
        return '', 404
    except RuntimeError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500
    stem = os.path.splitext(os.path.basename(str(name or 'document.pdf')))[0]
    return _pdf_page_response(payload, f'{stem}-page-{page_number}-of-{total}.jpg')


@tracking_bp.route('/api/workflows/<workflow_number>/files/<int:file_id>/download', methods=['GET'])
@login_required
def api_download_workflow_file(workflow_number, file_id):
    """下载工作流文件"""
    import os
    from flask import send_file
    from .config import UPLOAD_FOLDER
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 查询工作流文件信息
    cursor.execute('''
        SELECT wf.*, w.handler_id
        FROM workflow_files wf
        JOIN workflows w ON wf.workflow_number = w.workflow_number
        WHERE wf.id = ? AND wf.workflow_number = ? AND wf.is_deleted = 0
    ''', (file_id, workflow_number))
    
    file_info = cursor.fetchone()
    conn.close()
    
    if not file_info:
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    file_info = dict(file_info)
    
    # 权限检查：只有负责人和主管可以下载
    current_ctx = get_current_user_context()
    user_id = current_ctx.get('id')
    
    if not can_manage_by_owner(file_info['handler_id'], user_id=user_id):
        return jsonify({'success': False, 'error': '无权限'}), 403
    
    # 组合实际文件路径
    file_path = file_info['file_path']
    full_path = os.path.join(UPLOAD_FOLDER, file_path)

    # 下载文件（使用原始文件名）
    original_filename = file_info.get('file_name', os.path.basename(file_path))

    # 兼容旧数据：file_path 可能只有目录
    if os.path.isdir(full_path) or not os.path.exists(full_path):
        dir_path = None
        if os.path.isdir(full_path):
            dir_path = full_path
        else:
            candidate_dir = os.path.join(UPLOAD_FOLDER, file_path.rstrip('/\\'))
            if os.path.isdir(candidate_dir):
                dir_path = candidate_dir
        if dir_path:
            ext = os.path.splitext(original_filename)[1].lower()
            try:
                candidates = [f for f in os.listdir(dir_path)
                              if os.path.isfile(os.path.join(dir_path, f)) and (not ext or f.lower().endswith(ext))]
            except Exception:
                candidates = []
            if candidates:
                candidates.sort(key=lambda f: os.path.getmtime(os.path.join(dir_path, f)), reverse=True)
                full_path = os.path.join(dir_path, candidates[0])
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            return jsonify({'success': False, 'error': '文件已被删除'}), 404
    
    preview_requested = str(request.args.get('preview') or '').strip().lower() in {'1', 'true', 'yes'}
    if preview_requested and _local_guest_image_file(
        original_filename,
        file_info.get('file_type', ''),
        file_info.get('mime_type', file_info.get('file_type', ''))
    ):
        try:
            return _image_preview_response(full_path, original_filename or 'preview.jpg')
        except Exception as exc:
            print(f'[WARN] workflow image preview fallback to original: {exc}')

    return send_file(
        full_path,
        as_attachment=True,
        download_name=original_filename
    )


@tracking_bp.route('/api/workflows/<workflow_number>/files/<int:file_id>/pdf-page/<int:page_number>', methods=['GET'])
@login_required
def api_preview_workflow_pdf_page(workflow_number, file_id, page_number):
    # Authenticated mobile/browser preview of one workflow PDF page as JPEG.
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT wf.*, w.handler_id
            FROM workflow_files wf
            JOIN workflows w ON w.workflow_number = wf.workflow_number
            WHERE wf.id = ? AND wf.workflow_number = ? AND COALESCE(wf.is_deleted, 0) = 0
        ''', (file_id, workflow_number))
        row = cursor.fetchone()
    finally:
        conn.close()
    if not row:
        return '', 404
    info = dict(row)
    current_ctx = get_current_user_context()
    if not can_manage_by_owner(info.get('handler_id'), user_id=current_ctx.get('id')):
        return '', 403
    name = info.get('file_name') or os.path.basename(str(info.get('file_path') or ''))
    if not _media_is_pdf(name, info.get('file_type'), info.get('file_type')):
        return '', 415
    path = _resolve_workflow_file_full_path(info)
    try:
        payload, total = _pdf_preview_render_page(path, page_number, compact=str(request.args.get('preview') or '').lower() in {'1','true','yes'})
    except FileNotFoundError:
        return '', 404
    except ValueError:
        return '', 404
    except RuntimeError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500
    stem = os.path.splitext(os.path.basename(str(name or 'document.pdf')))[0]
    return _pdf_page_response(payload, f'{stem}-page-{page_number}-of-{total}.jpg')

@tracking_bp.route('/api/orders/files/<int:file_id>', methods=['DELETE'])
@login_required
@require_permission('order_file', 'delete', resource_id_param='file_id')
def api_delete_order_file(file_id):
    """删除文件"""
    import os
    from .config import UPLOAD_FOLDER
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 查询文件
    cursor.execute('SELECT * FROM order_files WHERE id = ?', (file_id,))
    file_info = cursor.fetchone()
    
    if not file_info:
        conn.close()
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    # 组合实际文件路径：UPLOAD_ROOT + file_path + stored_filename
    try:
        stored_filename = file_info['stored_filename'] if file_info['stored_filename'] else os.path.basename(file_info['file_path'])
    except KeyError:
        stored_filename = os.path.basename(file_info['file_path'])
    
    file_path_dir = file_info['file_path']
    
    # 兼容旧数据：如果 file_path 包含文件名，提取目录
    if '/' in stored_filename or '\\' in stored_filename:
        file_path_dir = os.path.dirname(file_info['file_path'])
        stored_filename = os.path.basename(file_info['file_path'])
    
    full_path = os.path.join(UPLOAD_FOLDER, file_path_dir, stored_filename)
    
    try:
        if os.path.exists(full_path):
            os.remove(full_path)
    except Exception as e:
        print(f"删除文件失败: {e}")
    
    # 删除数据库记录
    cursor.execute('DELETE FROM order_files WHERE id = ?', (file_id,))
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': '文件已删除'
    })


@tracking_bp.route('/api/orders/skip', methods=['POST'])
@admin_required
def api_skip_order():
    """标记订单号为跳号"""
    data = request.get_json()
    
    order_number = data.get('order_number')
    reason = data.get('reason', '')
    
    if not order_number:
        return jsonify({'success': False, 'error': '缺少订单号'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 检查订单
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        # 如果不存在,直接建立为 SKIPPED
        created_by = session.get('user_id') or session.get('username', 'admin')
        cursor.execute('''
            INSERT INTO orders (
                order_number, status, visibility, created_at, created_by
            ) VALUES (?, 'SKIPPED', 'admin_only', CURRENT_TIMESTAMP, ?)
        ''', (order_number, created_by))
    elif order['status'] == 'UNLOCKED':
        # 如果是未解锁,改成 SKIPPED
        updated_by = session.get('user_id') or session.get('username', 'admin')
        cursor.execute('''
            UPDATE orders SET
                status = 'SKIPPED',
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE order_number = ?
        ''', (updated_by, order_number))
    else:
        conn.close()
        return jsonify({'success': False, 'error': '订单号已被使用,无法标记跳号'}), 400
    
    # 如果有原因,记录到 order_notes
    if reason:
        created_by = session.get('user_id') or session.get('username', 'admin')
        cursor.execute('''
            INSERT INTO order_notes (
                order_number, note_type, content, created_by
            ) VALUES (?, 'skip_reason', ?, ?)
        ''', (order_number, reason, created_by))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': '已标记为跳号'
    })


@tracking_bp.route('/api/orders/<order_number>/remove-skip', methods=['PUT'])
@admin_required
def api_remove_skip(order_number):
    """解除跳号（将 SKIPPED 改回 UNLOCKED）"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 检查订单是否存在且状态为 SKIPPED
    cursor.execute('SELECT * FROM orders WHERE order_number = ?', (order_number,))
    order = cursor.fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单号不存在'}), 404
    
    if order['status'] != 'SKIPPED':
        conn.close()
        return jsonify({'success': False, 'error': '订单号不是跳号状态，无法解除'}), 400
    
    # 更新状态为 UNLOCKED
    updated_by = session.get('user_id') or session.get('username', 'admin')
    cursor.execute('''
        UPDATE orders SET
            status = 'UNLOCKED',
            updated_at = CURRENT_TIMESTAMP,
            updated_by = ?
        WHERE order_number = ?
    ''', (updated_by, order_number))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': '已解除跳号',
        'data': {
            'order_number': order_number,
            'old_status': 'SKIPPED',
            'new_status': 'UNLOCKED'
        }
    })


@tracking_bp.route('/api/orders/missing', methods=['GET'])
@admin_required
def api_check_missing_numbers():
    """检测缺号"""
    start_number = request.args.get('start', '').strip()
    end_number = request.args.get('end', '').strip()
    
    if not start_number or not end_number:
        return jsonify({'success': False, 'error': '缺少起始或结束号码'}), 400
    
    # 解析前缀和数字
    prefix = ''.join(filter(str.isalpha, start_number))
    
    try:
        start_num = int(''.join(filter(str.isdigit, start_number)))
        end_num = int(''.join(filter(str.isdigit, end_number)))
    except ValueError:
        return jsonify({'success': False, 'error': '号码格式错误'}), 400
    
    if start_num > end_num:
        return jsonify({'success': False, 'error': '起始号码不能大于结束号码'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 查詢範圍內訂單。正式狀態視為已存在；UNLOCKED 視為缺號但另行標示；SKIPPED 排除。
    cursor.execute('''
        SELECT order_number, status FROM orders
        WHERE order_number BETWEEN ? AND ?
    ''', (start_number, end_number))
    
    existing_numbers = set()
    skipped_numbers = set()
    unlocked_numbers = set()
    for row in cursor.fetchall():
        num_part = ''.join(filter(str.isdigit, row['order_number']))
        if not num_part:
            continue
        num = int(num_part)
        if row['status'] == 'SKIPPED':
            skipped_numbers.add(num)
        elif row['status'] == 'UNLOCKED':
            unlocked_numbers.add(num)
        else:
            existing_numbers.add(num)
    
    conn.close()
    
    # 计算缺号
    missing_numbers = []
    for i in range(start_num, end_num + 1):
        if i not in existing_numbers and i not in skipped_numbers:
            missing_numbers.append(f"{prefix}{i}" if prefix else str(i))
    
    return jsonify({
        'success': True,
        'data': {
            'range': {
                'start': start_number,
                'end': end_number,
                'total': end_num - start_num + 1
            },
            'existing_count': len(existing_numbers),
            'skipped_count': len(skipped_numbers),
            'unlocked_count': len(unlocked_numbers),
            'unlocked_numbers': [f"{prefix}{i}" if prefix else str(i) for i in sorted(unlocked_numbers)],
            'missing_count': len(missing_numbers),
            'missing_numbers': missing_numbers
        }
    })


# ===================================================================
# 管理员：通知记录审计 + 系统设置
# ===================================================================

@tracking_bp.route('/admin/notifications')
@admin_required
def admin_notifications():
    """通知记录审计页面（管理员专用）"""
    return render_template('tracking/notification_audit.html')


@tracking_bp.route('/api/admin/notifications', methods=['GET'])
@admin_required
def api_admin_notification_audit():
    """管理员查询所有通知记录（含已过期、已隐藏）"""
    page = max(1, int(request.args.get('page', 1)))
    page_size = min(100, int(request.args.get('page_size', 30)))

    # 筛选参数
    user_id = request.args.get('user_id', '').strip()
    ntype = request.args.get('type', '').strip()
    keyword = request.args.get('keyword', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    read_status = request.args.get('read_status', '').strip()  # '' / 'read' / 'unread'

    where_clauses = ["n.type NOT LIKE '\\_%' ESCAPE '\\'"]
    params = []

    if user_id:
        where_clauses.append("n.user_id = ?")
        params.append(int(user_id))
    if ntype:
        where_clauses.append("n.type = ?")
        params.append(ntype)
    if keyword:
        where_clauses.append("(n.title LIKE ? OR n.message LIKE ? OR n.order_number LIKE ? OR n.workflow_number LIKE ?)")
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw, kw])
    if date_from:
        where_clauses.append("DATE(n.created_at) >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("DATE(n.created_at) <= ?")
        params.append(date_to)
    if read_status == 'read':
        where_clauses.append("n.is_read = 1")
    elif read_status == 'unread':
        where_clauses.append("n.is_read = 0")

    where = ' AND '.join(where_clauses) if where_clauses else '1=1'

    conn = get_db()
    cursor = conn.cursor()

    # 总数
    cursor.execute(f'SELECT COUNT(*) as count FROM notifications n WHERE {where}', params)
    total = cursor.fetchone()['count']

    # 分页查询（JOIN users 获取用户名）
    offset = (page - 1) * page_size
    cursor.execute(f'''
        SELECT n.*, COALESCE(u.real_name, u.display_name, u.username) AS user_name, u.role AS user_role
        FROM notifications n
        LEFT JOIN users u ON n.user_id = u.id
        WHERE {where}
        ORDER BY n.created_at DESC
        LIMIT ? OFFSET ?
    ''', params + [page_size, offset])

    items = []
    for row in cursor.fetchall():
        items.append({
            'id': row['id'],
            'user_id': row['user_id'],
            'user_name': row['user_name'] or f'ID:{row["user_id"]}',
            'user_role': row['user_role'] or '',
            'type': row['type'],
            'title': row['title'],
            'message': row['message'] or '',
            'order_number': row['order_number'],
            'workflow_number': row['workflow_number'],
            'priority': row['priority'] or 'normal',
            'is_read': bool(row['is_read']),
            'read_at': row['read_at'],
            'created_at': row['created_at'],
        })

    conn.close()
    return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size,
    })


@tracking_bp.route('/api/admin/notification-users', methods=['GET'])
@admin_required
def api_admin_notification_users():
    """获取有通知记录的所有用户列表（供筛选下拉框使用）"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DISTINCT n.user_id, COALESCE(u.real_name, u.display_name, u.username) AS user_name, u.role
        FROM notifications n
        LEFT JOIN users u ON n.user_id = u.id
        WHERE n.type NOT LIKE '\\_%' ESCAPE '\\'
        ORDER BY user_name
    ''')
    users = [{'id': r['user_id'], 'name': r['user_name'] or f'ID:{r["user_id"]}', 'role': r['role'] or ''} for r in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': users})


@tracking_bp.route('/api/admin/settings', methods=['GET'])
@admin_required
def api_admin_get_settings():
    """获取所有系统设置"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT key, value, description, updated_at FROM system_settings ORDER BY key')
    settings = {}
    for row in cursor.fetchall():
        settings[row['key']] = {
            'value': row['value'],
            'description': row['description'],
            'updated_at': row['updated_at'],
        }
    conn.close()
    return jsonify({'success': True, 'data': settings})


@tracking_bp.route('/api/admin/settings', methods=['PUT'])
@admin_required
def api_admin_update_settings():
    """更新系统设置"""
    data = request.get_json() or {}
    key = data.get('key', '').strip()
    value = data.get('value', '').strip()

    if not key or not value:
        return jsonify({'success': False, 'error': '参数不完整'}), 400

    # notification_visible_days 同时作为通知显示与数据库保留天数
    if key == 'notification_visible_days':
        try:
            days = int(value)
            if days < 1 or days > 365:
                return jsonify({'success': False, 'error': '天数须在 1-365 之间'}), 400
            value = str(days)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': '天数必须为正整数'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO system_settings (key, value, updated_at)
        VALUES (?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = CURRENT_TIMESTAMP
    ''', (key, value))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '设置已更新'})


# ==================== 初始化函數 ====================

FACTORY_VISIT_FIELDS = [
    'visit_date',
    'factory_name',
    'main_business',
    'companions',
    'business_status',
    'main_market',
    'price_position',
    'visit_note',
    'analysis',
]


def _current_actor_name(conn):
    return (
        session.get('display_name')
        or session.get('real_name')
        or get_user_display_name(session.get('user_id'), conn)
        or session.get('username')
        or ''
    )


def _normalize_factory_visit_payload(data):
    payload = {}
    for field in FACTORY_VISIT_FIELDS:
        value = data.get(field)
        payload[field] = str(value).strip() if value is not None else ''
    if not payload['factory_name']:
        raise ValueError('factory_name is required')
    return payload


def _factory_visit_row_to_dict(row, current_user_id=None, current_role=None):
    item = dict(row)
    item['created_by'] = item.get('created_by_name') or ''
    role = (current_role or session.get('role') or 'viewer').lower()
    owner_id = item.get('created_by_id')
    is_owner = current_user_id is not None and str(owner_id) == str(current_user_id)
    item['can_edit'] = role == 'admin' or (role == 'sales' and is_owner)
    item['can_delete'] = role == 'admin'
    return item


def _write_factory_visit_audit(cursor, visit_id, action, actor, before_row=None, after_row=None):
    cursor.execute('''
        INSERT INTO factory_visit_audit_logs (
            visit_id, action, actor_id, actor_name, actor_role, before_json, after_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        visit_id,
        action,
        actor.get('id'),
        actor.get('name'),
        actor.get('role'),
        json.dumps(dict(before_row), ensure_ascii=False) if before_row else None,
        json.dumps(dict(after_row), ensure_ascii=False) if after_row else None,
    ))


def _get_factory_visit(cursor, visit_id, include_deleted=False):
    where = 'id = ?' if include_deleted else 'id = ? AND COALESCE(is_deleted, 0) = 0'
    cursor.execute(f'SELECT * FROM factory_visits WHERE {where}', (visit_id,))
    return cursor.fetchone()


def _can_edit_factory_visit(row):
    role = (session.get('role') or 'viewer').lower()
    if role == 'admin':
        return True
    if role != 'sales':
        return False
    return str(row['created_by_id']) == str(session.get('user_id'))


@tracking_bp.route('/factory-visit')
@login_required
def factory_visit():
    return render_template('tracking/factory_visit.html')


@tracking_bp.route('/api/me', methods=['GET'])
@api_login_required
def api_current_user_me():
    if cloud_mode_enabled():
        ctx = get_current_user_context()
        name = session.get('display_name') or ctx.get('username') or '用户'
        return jsonify({
            'id': ctx.get('id'),
            'username': ctx.get('username') or session.get('username'),
            'display_name': name,
            'name': name,
            'role': ctx.get('role', 'viewer'),
        })
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, username, display_name, real_name, role
        FROM users
        WHERE id = ?
    ''', (g.current_user.get('id'),))
    user = cursor.fetchone()
    conn.close()
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 404
    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'display_name': user['real_name'] or user['display_name'] or user['username'],
        'name': user['real_name'] or user['display_name'] or user['username'],
        'role': user['role'],
    })


@tracking_bp.route('/api/factory-visits/recent', methods=['GET'])
@api_login_required
def api_factory_visits_recent():
    limit = min(max(int(request.args.get('limit', 1000) or 1000), 1), 2000)
    conn = get_db()
    ensure_factory_visit_tables(conn)
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) AS count FROM factory_visits WHERE COALESCE(is_deleted, 0) = 0')
    total = cursor.fetchone()['count']
    cursor.execute('''
        SELECT *
        FROM factory_visits
        WHERE COALESCE(is_deleted, 0) = 0
        ORDER BY COALESCE(visit_date, created_at) DESC, id DESC
        LIMIT ?
    ''', (limit,))
    items = [
        _factory_visit_row_to_dict(row, g.current_user.get('id'), g.current_user.get('role'))
        for row in cursor.fetchall()
    ]
    conn.close()
    return jsonify({
        'items': items,
        'total': total,
        'returned': len(items),
        'has_more': total > len(items),
    })


@tracking_bp.route('/api/factory-visits/companions', methods=['GET'])
@api_login_required
def api_factory_visit_companions():
    """Active admin/sales names for factory visit companion picker."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT id, username, display_name, real_name, role
            FROM users
            WHERE COALESCE(status, 'active') = 'active'
              AND role IN ('admin', 'administrator', 'root', 'superuser', '管理员', '主管', 'sales', 'seller', 'biz', 'business', '业务员', '業務員')
            ORDER BY
              CASE WHEN role IN ('admin', 'administrator', 'root', 'superuser', '管理员', '主管') THEN 0 ELSE 1 END,
              COALESCE(real_name, display_name, username)
        ''')
    except sqlite3.OperationalError:
        cursor.execute('''
            SELECT id, username, display_name, NULL AS real_name, role
            FROM users
            WHERE role IN ('admin', 'administrator', 'root', 'superuser', '管理员', '主管', 'sales', 'seller', 'biz', 'business', '业务员', '業務員')
            ORDER BY COALESCE(display_name, username)
        ''')
    items = []
    for row in cursor.fetchall():
        name = row['real_name'] or row['display_name'] or row['username']
        items.append({
            'id': row['id'],
            'name': name,
            'username': row['username'],
            'role': row['role'],
        })
    conn.close()
    return jsonify({'items': items})


@tracking_bp.route('/api/factory-visits/search', methods=['GET'])
@api_login_required
def api_factory_visits_search():
    keyword = (request.args.get('keyword') or '').strip()
    limit = min(max(int(request.args.get('limit', 1000) or 1000), 1), 2000)
    conn = get_db()
    ensure_factory_visit_tables(conn)
    cursor = conn.cursor()
    params = []
    where = ['COALESCE(is_deleted, 0) = 0']
    matched_sql_fields = [
        'visit_date', 'factory_name', 'main_business', 'created_by_name', 'companions',
        'business_status', 'main_market', 'price_position', 'visit_note', 'analysis'
    ]
    if keyword:
        like = f'%{keyword}%'
        where.append('(' + ' OR '.join([f'{field} LIKE ?' for field in matched_sql_fields]) + ')')
        params.extend([like] * len(matched_sql_fields))
    where_sql = ' AND '.join(where)
    cursor.execute(f'SELECT COUNT(*) AS count FROM factory_visits WHERE {where_sql}', params)
    total = cursor.fetchone()['count']
    cursor.execute(f'''
        SELECT *
        FROM factory_visits
        WHERE {where_sql}
        ORDER BY COALESCE(visit_date, created_at) DESC, id DESC
        LIMIT ?
    ''', params + [limit])
    items = []
    for row in cursor.fetchall():
        item = _factory_visit_row_to_dict(row, g.current_user.get('id'), g.current_user.get('role'))
        if keyword:
            low = keyword.lower()
            item['matched_fields'] = [
                field for field in matched_sql_fields
                if low in str(item.get(field) or '').lower()
            ]
            if 'created_by_name' in item['matched_fields']:
                item['matched_fields'].append('created_by')
        items.append(item)
    conn.close()
    return jsonify({
        'items': items,
        'total': total,
        'returned': len(items),
        'has_more': total > len(items),
    })


@tracking_bp.route('/api/factory-visits/<int:visit_id>', methods=['GET'])
@api_login_required
def api_factory_visit_detail(visit_id):
    conn = get_db()
    ensure_factory_visit_tables(conn)
    cursor = conn.cursor()
    row = _get_factory_visit(cursor, visit_id)
    conn.close()
    if not row:
        return jsonify({'success': False, 'error': 'Record not found'}), 404
    return jsonify(_factory_visit_row_to_dict(row, g.current_user.get('id'), g.current_user.get('role')))


@tracking_bp.route('/api/factory-visits', methods=['POST'])
@api_login_required
def api_factory_visit_create():
    role = (g.current_user.get('role') or 'viewer').lower()
    if role not in {'admin', 'sales'}:
        return jsonify({'success': False, 'error': 'No permission to create'}), 403
    try:
        payload = _normalize_factory_visit_payload(request.get_json() or {})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400

    conn = get_db()
    ensure_factory_visit_tables(conn)
    cursor = conn.cursor()
    actor = {
        'id': g.current_user.get('id'),
        'name': _current_actor_name(conn),
        'role': role,
    }
    cursor.execute(f'''
        INSERT INTO factory_visits (
            {', '.join(FACTORY_VISIT_FIELDS)},
            created_by_id, created_by_name, updated_by_id, updated_by_name, updated_at
        )
        VALUES ({', '.join(['?'] * len(FACTORY_VISIT_FIELDS))}, ?, ?, ?, ?, CURRENT_TIMESTAMP)
    ''', [payload[field] for field in FACTORY_VISIT_FIELDS] + [
        actor['id'], actor['name'], actor['id'], actor['name']
    ])
    visit_id = cursor.lastrowid
    after = _get_factory_visit(cursor, visit_id)
    _write_factory_visit_audit(cursor, visit_id, 'create', actor, None, after)
    conn.commit()
    item = _factory_visit_row_to_dict(after, actor['id'], role)
    conn.close()
    return jsonify(item), 201


@tracking_bp.route('/api/factory-visits/<int:visit_id>', methods=['PUT'])
@api_login_required
def api_factory_visit_update(visit_id):
    try:
        payload = _normalize_factory_visit_payload(request.get_json() or {})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    conn = get_db()
    ensure_factory_visit_tables(conn)
    cursor = conn.cursor()
    before = _get_factory_visit(cursor, visit_id)
    if not before:
        conn.close()
        return jsonify({'success': False, 'error': 'Record not found'}), 404
    if not _can_edit_factory_visit(before):
        conn.close()
        return jsonify({'success': False, 'error': 'No permission to edit this record'}), 403

    role = (g.current_user.get('role') or 'viewer').lower()
    actor = {
        'id': g.current_user.get('id'),
        'name': _current_actor_name(conn),
        'role': role,
    }
    set_sql = ', '.join([f'{field} = ?' for field in FACTORY_VISIT_FIELDS])
    cursor.execute(f'''
        UPDATE factory_visits
        SET {set_sql},
            updated_by_id = ?,
            updated_by_name = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', [payload[field] for field in FACTORY_VISIT_FIELDS] + [actor['id'], actor['name'], visit_id])
    after = _get_factory_visit(cursor, visit_id)
    _write_factory_visit_audit(cursor, visit_id, 'update', actor, before, after)
    conn.commit()
    item = _factory_visit_row_to_dict(after, actor['id'], role)
    conn.close()
    return jsonify(item)


@tracking_bp.route('/api/factory-visits/<int:visit_id>', methods=['DELETE'])
@api_login_required
def api_factory_visit_delete(visit_id):
    role = (g.current_user.get('role') or 'viewer').lower()
    if role != 'admin':
        return jsonify({'success': False, 'error': 'Only admin can delete'}), 403
    data = request.get_json() or {}
    if data.get('confirm') != 'DELETE':
        return jsonify({'success': False, 'error': 'Type DELETE to confirm'}), 400
    conn = get_db()
    ensure_factory_visit_tables(conn)
    cursor = conn.cursor()
    before = _get_factory_visit(cursor, visit_id)
    if not before:
        conn.close()
        return jsonify({'success': False, 'error': 'Record not found'}), 404
    actor = {
        'id': g.current_user.get('id'),
        'name': _current_actor_name(conn),
        'role': role,
    }
    cursor.execute('''
        UPDATE factory_visits
        SET is_deleted = 1,
            deleted_by_id = ?,
            deleted_by_name = ?,
            deleted_at = CURRENT_TIMESTAMP,
            delete_reason = ?,
            updated_by_id = ?,
            updated_by_name = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (
        actor['id'], actor['name'], data.get('reason') or 'DELETE confirmed',
        actor['id'], actor['name'], visit_id
    ))
    after = _get_factory_visit(cursor, visit_id, include_deleted=True)
    _write_factory_visit_audit(cursor, visit_id, 'delete', actor, before, after)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@tracking_bp.route('/api/missing-contacts', methods=['GET'])
@api_login_required
def api_missing_contacts():
    """查詢有訂單但在 /contacto 裡沒有電話的客戶（海外模式用）"""
    try:
        # 1. 從 order_tracking 拿所有不重複的客戶名稱
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT DISTINCT customer_name 
            FROM orders 
            WHERE customer_name IS NOT NULL AND customer_name != ''
            ORDER BY customer_name ASC
        ''')
        all_customers = [row['customer_name'].strip().upper() for row in cursor.fetchall()]
        conn.close()

        # 2. 從 Flask /contacto 拿有電話的客戶
        import urllib.request
        import json as json_module
        try:
            with urllib.request.urlopen('http://localhost:8000/contacto', timeout=3) as res:
                contacts = json_module.loads(res.read().decode())
            has_phone = set()
            for c in contacts:
                nombre = (c.get('nombre') or '').strip().upper()
                telefono = (c.get('telefono') or '').strip()
                telefono1 = (c.get('telefono1') or '').strip()
                if nombre and (telefono or telefono1):
                    has_phone.add(nombre)
        except Exception:
            return jsonify({'success': False, 'error': '無法連接 /contacto，請確認 Flask 正在運行'}), 503

        # 3. 對比，找出缺電話的
        missing = [name for name in all_customers if name not in has_phone]

        return jsonify({
            'success': True,
            'total_customers': len(all_customers),
            'missing_count': len(missing),
            'missing': missing
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def init_app(app, data_provider=None):
    """初始化應用。Local 保持原 SQLite；Cloud 可由父 Flask 注入 Provider。"""
    if data_provider is not None:
        register_order_data_provider(app, data_provider)
    app.register_blueprint(tracking_bp)
    cloud_mode = bool(app.config.get('TRACKING_CLOUD_MODE', CLOUD_MODE))
    if not cloud_mode:
        init_db()
